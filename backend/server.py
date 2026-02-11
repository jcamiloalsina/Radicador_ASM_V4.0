from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Query, BackgroundTasks, WebSocket, WebSocketDisconnect, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument, UpdateOne
import os
import logging
import random
import hashlib
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import shutil
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io
import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# URL Base para verificación de certificados (OBLIGATORIO - debe configurarse en producción)
# Esta variable DEBE estar definida en el archivo .env del servidor de producción
# Ejemplo: VERIFICACION_URL="https://certificados.asomunicipios.gov.co"
VERIFICACION_BASE_URL = os.environ.get('VERIFICACION_URL')
if not VERIFICACION_BASE_URL:
    print("⚠️ ADVERTENCIA: VERIFICACION_URL no está configurada. Los certificados usarán URL por defecto.")
    VERIFICACION_BASE_URL = "https://certificados.asomunicipios.gov.co"

# Email Configuration
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.office365.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_FROM = os.environ.get('SMTP_FROM', SMTP_USER)

# File upload configuration
UPLOAD_DIR = Path('/app/uploads')
UPLOAD_DIR.mkdir(exist_ok=True)

security = HTTPBearer()

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ===== BACKUP SCHEDULER =====
backup_scheduler = AsyncIOScheduler()


# ===== WEBSOCKET CONNECTION MANAGER =====
class ConnectionManager:
    """Manages WebSocket connections for real-time notifications"""
    def __init__(self):
        self.active_connections: dict[str, WebSocket] = {}  # user_id -> websocket
    
    async def connect(self, websocket: WebSocket, user_id: str):
        await websocket.accept()
        self.active_connections[user_id] = websocket
        logging.info(f"WebSocket connected for user: {user_id}")
    
    def disconnect(self, user_id: str):
        if user_id in self.active_connections:
            del self.active_connections[user_id]
            logging.info(f"WebSocket disconnected for user: {user_id}")
    
    async def send_personal_message(self, message: dict, user_id: str):
        if user_id in self.active_connections:
            try:
                await self.active_connections[user_id].send_json(message)
                # Log de progreso GDB para debugging
                if message.get('type') == 'gdb_upload_progress':
                    data = message.get('data', {})
                    logging.info(f"[WS] GDB Progress enviado a {user_id}: {data.get('status')} - {data.get('progress')}% - {data.get('message', '')[:50]}")
            except Exception as e:
                logging.error(f"Error sending to {user_id}: {e}")
                self.disconnect(user_id)
        else:
            # Log cuando el usuario no está conectado
            if message.get('type') == 'gdb_upload_progress':
                data = message.get('data', {})
                logging.warning(f"[WS] Usuario {user_id} NO conectado. GDB Progress perdido: {data.get('status')} - {data.get('progress')}%")
    
    async def broadcast(self, message: dict, exclude_user: str = None):
        """Broadcast message to all connected users except the sender"""
        disconnected = []
        for user_id, connection in self.active_connections.items():
            if user_id != exclude_user:
                try:
                    await connection.send_json(message)
                except Exception as e:
                    logging.error(f"Error broadcasting to {user_id}: {e}")
                    disconnected.append(user_id)
        
        # Clean up disconnected users
        for user_id in disconnected:
            self.disconnect(user_id)
    
    async def broadcast_to_roles(self, message: dict, roles: list, exclude_user: str = None):
        """Broadcast message to users with specific roles"""
        # This requires knowing user roles - for simplicity, broadcast to all
        await self.broadcast(message, exclude_user)

ws_manager = ConnectionManager()


# ===== HELPER FUNCTIONS =====

# Diccionario de nombres comunes con tildes correctas
NOMBRES_CON_TILDES = {
    "maria": "María", "jose": "José", "jesus": "Jesús", "angel": "Ángel",
    "andres": "Andrés", "raul": "Raúl", "cesar": "César", "hector": "Héctor",
    "oscar": "Óscar", "nelson": "Nélson", "german": "Germán", "ivan": "Iván",
    "nicolas": "Nicolás", "tomas": "Tomás", "simon": "Simón", "joaquin": "Joaquín",
    "martin": "Martín", "agustin": "Agustín", "sebastian": "Sebastián", "adrian": "Adrián",
    "dario": "Darío", "alvaro": "Álvaro", "ramon": "Ramón", "julian": "Julián",
    "fabian": "Fabián", "maximo": "Máximo", "lazaro": "Lázaro", "moises": "Moisés",
    "isaias": "Isaías", "efrain": "Efraín", "hernan": "Hernán", "ruben": "Rubén",
    "felix": "Félix", "ines": "Inés", "belen": "Belén", "lucia": "Lucía",
    "sofia": "Sofía", "rocio": "Rocío", "monica": "Mónica", "veronica": "Verónica",
    "natalia": "Natalia", "cecilia": "Cecilia", "angela": "Ángela", "barbara": "Bárbara",
    "beatriz": "Beatriz", "dolores": "Dolores", "pilar": "Pilar", "teresa": "Teresa",
    "rosa": "Rosa", "elena": "Elena", "elvira": "Elvira", "esperanza": "Esperanza",
    "eugenia": "Eugenia", "gloria": "Gloria", "graciela": "Graciela", "irene": "Irene",
    "josefa": "Josefa", "julia": "Julia", "juana": "Juana", "lidia": "Lidia",
    "lourdes": "Lourdes", "margarita": "Margarita", "mercedes": "Mercedes", "mercedes": "Mercedes",
    "nuria": "Nuria", "patricia": "Patricia", "raquel": "Raquel", "rebeca": "Rebeca",
    "sara": "Sara", "silvia": "Silvia", "susana": "Susana", "victoria": "Victoria",
    "garcia": "García", "gonzalez": "González", "rodriguez": "Rodríguez", "fernandez": "Fernández",
    "lopez": "López", "martinez": "Martínez", "sanchez": "Sánchez", "perez": "Pérez",
    "gomez": "Gómez", "diaz": "Díaz", "jimenez": "Jiménez", "hernandez": "Hernández",
    "alvarez": "Álvarez", "ruiz": "Ruiz", "ramirez": "Ramírez", "romero": "Romero",
    "suarez": "Suárez", "benitez": "Benítez", "mendez": "Méndez", "gutierrez": "Gutiérrez",
    "nuñez": "Núñez", "ortiz": "Ortiz", "vazquez": "Vázquez", "dominguez": "Domínguez",
    "florez": "Flórez", "ayala": "Ayala", "parra": "Parra", "carrascal": "Carrascal"
}

def format_nombre_propio(nombre: str) -> str:
    """
    Formatea un nombre propio:
    - Convierte a formato Título (inicial mayúscula)
    - Aplica tildes automáticamente a nombres comunes
    """
    if not nombre:
        return nombre
    
    palabras = nombre.strip().split()
    resultado = []
    
    for palabra in palabras:
        palabra_lower = palabra.lower()
        # Buscar en el diccionario de nombres con tildes
        if palabra_lower in NOMBRES_CON_TILDES:
            resultado.append(NOMBRES_CON_TILDES[palabra_lower])
        else:
            # Capitalizar la primera letra
            resultado.append(palabra.capitalize())
    
    return ' '.join(resultado)


# ===== MODELS =====

class UserRole:
    USUARIO = "usuario"  # Usuario externo (antes "usuario")
    ATENCION_USUARIO = "atencion_usuario"
    GESTOR = "gestor"
    COORDINADOR = "coordinador"
    ADMINISTRADOR = "administrador"
    COMUNICACIONES = "comunicaciones"  # Puede consultar predios, ver visor, ver trámites, descargar/subir archivos
    EMPRESA = "empresa"  # Similar a comunicaciones pero sin ver total peticiones ni descargar Excel

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    full_name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class VerifyEmailCode(BaseModel):
    email: EmailStr
    code: str

class ResendVerificationCode(BaseModel):
    email: EmailStr

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    full_name: str
    role: str
    email_verified: bool = False
    verification_code: Optional[str] = None
    verification_code_expires: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserRoleUpdate(BaseModel):
    user_id: str
    new_role: str

# Sistema de Permisos Granulares
class Permission:
    """Permisos disponibles en el sistema"""
    UPLOAD_GDB = "upload_gdb"           # Subir archivos GDB
    IMPORT_R1R2 = "import_r1r2"         # Importar archivos R1/R2
    APPROVE_CHANGES = "approve_changes"  # Aprobar cambios de predios
    ACCESO_ACTUALIZACION = "acceso_actualizacion"  # Acceso al módulo de Actualización
    
    @classmethod
    def all_permissions(cls):
        return [cls.UPLOAD_GDB, cls.IMPORT_R1R2, cls.APPROVE_CHANGES, cls.ACCESO_ACTUALIZACION]
    
    @classmethod
    def get_description(cls, perm):
        descriptions = {
            cls.UPLOAD_GDB: "Subir archivos GDB (Base Gráfica)",
            cls.IMPORT_R1R2: "Importar archivos R1/R2 (Excel)",
            cls.APPROVE_CHANGES: "Aprobar/Rechazar cambios de predios",
            cls.ACCESO_ACTUALIZACION: "Acceso al módulo de Actualización (trabajo de campo)"
        }
        return descriptions.get(perm, perm)

class UserPermissionsUpdate(BaseModel):
    user_id: str
    permissions: List[str]

class PetitionStatus:
    RADICADO = "radicado"
    ASIGNADO = "asignado"
    EN_PROCESO = "en_proceso"  # Gestor(es) trabajando activamente
    REVISION = "revision"      # Enviado a coordinador para revisión
    APROBADO = "aprobado"      # Aprobado por coordinador, pendiente finalizar
    RECHAZADO = "rechazado"
    DEVUELTO = "devuelto"
    FINALIZADO = "finalizado"

class PetitionCreate(BaseModel):
    nombre_completo: str
    correo: EmailStr
    telefono: str
    tipo_tramite: str
    municipio: str

class PetitionUpdate(BaseModel):
    nombre_completo: Optional[str] = None
    correo: Optional[EmailStr] = None
    telefono: Optional[str] = None
    tipo_tramite: Optional[str] = None
    municipio: Optional[str] = None
    estado: Optional[str] = None
    notas: Optional[str] = None
    gestor_id: Optional[str] = None
    enviar_archivos_finalizacion: Optional[bool] = False  # Flag para adjuntar archivos al correo de finalización
    observaciones_devolucion: Optional[str] = None  # Observaciones cuando se devuelve un trámite
    comentario_aprobacion: Optional[str] = None  # Comentario del coordinador al aprobar

class GestorAssignment(BaseModel):
    petition_id: str
    gestor_id: str
    is_auxiliar: bool = False
    comentario: Optional[str] = None  # Comentario/instrucciones al asignar

class Petition(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    radicado: str
    user_id: str
    nombre_completo: str
    correo: str
    telefono: str
    tipo_tramite: str
    municipio: str
    descripcion: str = ""
    estado: str = PetitionStatus.RADICADO
    notas: str = ""
    gestor_id: Optional[str] = None
    gestores_asignados: List[str] = []
    archivos: List[dict] = []
    historial: List[dict] = []
    observaciones_devolucion: str = ""  # Observaciones cuando el staff devuelve el trámite
    devuelto_por_id: Optional[str] = None  # ID del usuario que devolvió el trámite
    devuelto_por_nombre: Optional[str] = None  # Nombre del usuario que devolvió
    # Nuevos campos para control de flujo mejorado
    gestores_finalizados: List[str] = []  # IDs de gestores que marcaron su trabajo como completado
    aprobado_por_id: Optional[str] = None  # ID del coordinador que aprobó
    aprobado_por_nombre: Optional[str] = None  # Nombre del coordinador
    fecha_aprobacion: Optional[str] = None  # Fecha de aprobación
    comentario_aprobacion: Optional[str] = None  # Comentario del coordinador al aprobar
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


# ===== PREDIO MODELS (Codigo Nacional Catastral) =====

# Catalogo de municipios con codigo catastral nacional
MUNICIPIOS_DIVIPOLA = {
    "Ábrego": {"departamento": "54", "municipio": "003"},
    "Bucarasica": {"departamento": "54", "municipio": "109"},
    "Cáchira": {"departamento": "54", "municipio": "128"},
    "Convención": {"departamento": "54", "municipio": "206"},
    "El Carmen": {"departamento": "54", "municipio": "245"},
    "El Tarra": {"departamento": "54", "municipio": "250"},
    "Hacarí": {"departamento": "54", "municipio": "344"},
    "La Playa": {"departamento": "54", "municipio": "398"},
    "Río de Oro": {"departamento": "47", "municipio": "545"},  # Cesar
    "San Calixto": {"departamento": "54", "municipio": "670"},
    "Sardinata": {"departamento": "54", "municipio": "720"},
    "Teorama": {"departamento": "54", "municipio": "800"}
}

# Catálogo de destino económico
DESTINO_ECONOMICO = {
    "A": "Habitacional",
    "B": "Industrial",
    "C": "Comercial",
    "D": "Agropecuario",
    "E": "Minero",
    "F": "Cultural",
    "G": "Recreacional",
    "H": "Salubridad",
    "I": "Institucional",
    "J": "Educativo",
    "K": "Religioso",
    "L": "Agrícola",
    "M": "Pecuario",
    "N": "Agroindustrial",
    "O": "Forestal",
    "P": "Uso Público",
    "Q": "Lote Urbanizable No Urbanizado",
    "R": "Lote Urbanizable No Edificado",
    "S": "Lote No Urbanizable",
    "T": "Servicios Especiales"
}

# Catálogo de tipo de documento
TIPO_DOCUMENTO_PREDIO = {
    "C": "Cédula de Ciudadanía",
    "E": "Cédula de Extranjería",
    "N": "NIT",
    "T": "Tarjeta de Identidad",
    "P": "Pasaporte",
    "X": "Sin documento / Entidad"
}

# Catálogo de estado civil
ESTADO_CIVIL_PREDIO = {
    "S": "Soltero/a",
    "E": "Casado/a con sociedad conyugal",
    "D": "Casado/a sin sociedad conyugal",
    "V": "Separación de bienes",
    "U": "Unión marital de hecho"
}

class PredioR1Create(BaseModel):
    """Registro R1 - Información Jurídica del Predio"""
    municipio: str
    zona: str = "00"  # 00=Rural, 01=Urbano, 02-99=Corregimientos
    sector: str = "01"
    comuna: str = "00"
    barrio: str = "00"
    manzana_vereda: str = "0000"
    terreno: str = "0001"
    condicion_predio: str = "0000"
    predio_horizontal: str = "000000000"
    
    # Propietario
    nombre_propietario: str
    tipo_documento: str
    numero_documento: str
    estado_civil: Optional[str] = None
    
    # Ubicación y características
    direccion: str
    comuna: str = "0"
    destino_economico: str
    area_terreno: float
    area_construida: float = 0
    avaluo: float
    
    # Mutación
    tipo_mutacion: Optional[str] = None
    numero_resolucion: Optional[str] = None
    fecha_resolucion: Optional[str] = None

class PredioR2Create(BaseModel):
    """Registro R2 - Información Física del Predio"""
    matricula_inmobiliaria: Optional[str] = None
    
    # Zona 1
    zona_fisica_1: float = 0
    zona_economica_1: float = 0
    area_terreno_1: float = 0
    
    # Zona 2
    zona_fisica_2: float = 0
    zona_economica_2: float = 0
    area_terreno_2: float = 0
    
    # Construcción 1
    habitaciones_1: int = 0
    banos_1: int = 0
    locales_1: int = 0
    pisos_1: int = 1
    tipificacion_1: float = 0
    uso_1: int = 0
    puntaje_1: float = 0
    area_construida_1: float = 0
    
    # Construcción 2
    habitaciones_2: int = 0
    banos_2: int = 0
    locales_2: int = 0
    pisos_2: int = 0
    tipificacion_2: float = 0
    uso_2: int = 0
    puntaje_2: float = 0
    area_construida_2: float = 0

class PredioCreate(BaseModel):
    r1: PredioR1Create
    r2: Optional[PredioR2Create] = None

class PredioUpdate(BaseModel):
    # R1 fields
    nombre_propietario: Optional[str] = None
    tipo_documento: Optional[str] = None
    numero_documento: Optional[str] = None
    estado_civil: Optional[str] = None
    direccion: Optional[str] = None
    comuna: Optional[str] = None
    destino_economico: Optional[str] = None
    area_terreno: Optional[float] = None
    area_construida: Optional[float] = None
    avaluo: Optional[float] = None
    tipo_mutacion: Optional[str] = None
    numero_resolucion: Optional[str] = None
    fecha_resolucion: Optional[str] = None
    
    # R2 fields
    matricula_inmobiliaria: Optional[str] = None
    zona_fisica_1: Optional[float] = None
    zona_economica_1: Optional[float] = None
    area_terreno_1: Optional[float] = None
    habitaciones_1: Optional[int] = None
    banos_1: Optional[int] = None
    locales_1: Optional[int] = None
    pisos_1: Optional[int] = None
    puntaje_1: Optional[float] = None
    area_construida_1: Optional[float] = None


# ===== SISTEMA DE APROBACIÓN DE PREDIOS =====

class PredioEstadoAprobacion:
    """Estados de aprobación para cambios en predios"""
    APROBADO = "aprobado"  # Cambios aplicados y firmes
    PENDIENTE_CREACION = "pendiente_creacion"  # Nuevo predio esperando aprobación
    PENDIENTE_MODIFICACION = "pendiente_modificacion"  # Modificación esperando aprobación
    PENDIENTE_ELIMINACION = "pendiente_eliminacion"  # Eliminación esperando aprobación
    RECHAZADO = "rechazado"  # Cambio rechazado por coordinador

class CambioPendienteCreate(BaseModel):
    """Modelo para crear un cambio pendiente"""
    predio_id: Optional[str] = None  # None para nuevos predios
    tipo_cambio: str  # creacion, modificacion, eliminacion
    datos_propuestos: dict  # Datos del predio (nuevo o modificado)
    justificacion: Optional[str] = None
    radicado_id: Optional[str] = None  # ID de la petición/radicado asociado
    radicado_numero: Optional[str] = None  # Número del radicado (ej: "RASMGC-0011-02-02-2026")
    # Flujo de Gestor de Apoyo (opcional)
    gestor_apoyo_id: Optional[str] = None  # ID del gestor de apoyo asignado
    observaciones_apoyo: Optional[str] = None  # Observaciones para el gestor de apoyo

class CambioAprobacionRequest(BaseModel):
    """Modelo para aprobar/rechazar un cambio"""
    cambio_id: str
    aprobado: bool
    comentario: Optional[str] = None


# ===== FLUJO DE TRABAJO PARA PREDIOS NUEVOS (CONSERVACIÓN) =====

class PredioNuevoEstado:
    """Estados del flujo de trabajo para predios nuevos"""
    CREADO = "creado"                    # Recién creado por gestor
    DIGITALIZACION = "digitalizacion"    # Asignado a gestor de apoyo
    REVISION = "revision"                # Enviado para revisión
    APROBADO = "aprobado"                # Aprobado e integrado
    DEVUELTO = "devuelto"                # Devuelto para corrección
    RECHAZADO = "rechazado"              # Rechazado definitivamente

class PredioNuevoCreate(BaseModel):
    """Modelo para crear un predio nuevo con flujo de trabajo"""
    r1: PredioR1Create
    r2: Optional[PredioR2Create] = None
    # Flujo de trabajo
    gestor_apoyo_id: str                 # ID del gestor de apoyo asignado
    # Relaciones
    radicado_numero: Optional[str] = None  # Solo el número (ej: "5511")
    peticiones_ids: Optional[List[str]] = []  # IDs de peticiones relacionadas
    # Observaciones
    observaciones: Optional[str] = None

class PredioNuevoAccion(BaseModel):
    """Modelo para acciones en el flujo de trabajo"""
    accion: str  # enviar_revision, aprobar, devolver, rechazar
    observaciones: Optional[str] = None



class ProyectoActualizacionEstado:
    """Estados de un proyecto de actualización"""
    ACTIVO = "activo"           # Proyecto en proceso
    PAUSADO = "pausado"         # Proyecto pausado temporalmente
    COMPLETADO = "completado"   # Trabajo de campo terminado
    ARCHIVADO = "archivado"     # Proyecto archivado (no editable)

class ProyectoActualizacionCreate(BaseModel):
    """Modelo para crear un proyecto de actualización"""
    nombre: str
    municipio: str
    descripcion: Optional[str] = None

class ProyectoActualizacionUpdate(BaseModel):
    """Modelo para actualizar un proyecto"""
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    estado: Optional[str] = None

class ProyectoActualizacion(BaseModel):
    """Modelo completo de proyecto de actualización"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nombre: str
    municipio: str
    descripcion: Optional[str] = None
    estado: str = ProyectoActualizacionEstado.ACTIVO
    # Archivos asociados - Unificados
    base_grafica_archivo: Optional[str] = None
    base_grafica_cargado_en: Optional[datetime] = None
    base_grafica_total_predios: int = 0
    info_alfanumerica_archivo: Optional[str] = None  # R1/R2 unificado
    info_alfanumerica_cargado_en: Optional[datetime] = None
    info_alfanumerica_total_registros: int = 0
    # Fechas del proyecto
    fecha_inicio: Optional[datetime] = None
    fecha_fin_planificada: Optional[datetime] = None
    fecha_fin_real: Optional[datetime] = None
    # Estadísticas
    predios_actualizados: int = 0
    predios_no_identificados: int = 0
    # Metadatos
    creado_por: str
    creado_por_nombre: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


# ===== MODELOS PARA CRONOGRAMA DE ACTUALIZACIÓN =====

class EtapaProyectoTipo:
    """Tipos de etapas fijas del proyecto"""
    PREOPERATIVA = "preoperativa"
    OPERATIVA = "operativa"
    POSTOPERATIVA = "postoperativa"

class ActividadEstado:
    """Estados de una actividad"""
    PENDIENTE = "pendiente"
    EN_PROGRESO = "en_progreso"
    COMPLETADA = "completada"
    BLOQUEADA = "bloqueada"

class ActividadPrioridad:
    """Prioridades de actividad"""
    ALTA = "alta"
    MEDIA = "media"
    BAJA = "baja"

class EtapaProyectoCreate(BaseModel):
    """Modelo para crear una etapa"""
    tipo: str  # preoperativa, operativa, postoperativa
    nombre: str
    fecha_inicio: Optional[str] = None
    fecha_fin_planificada: Optional[str] = None
    duracion_dias: Optional[int] = None

class EtapaProyectoUpdate(BaseModel):
    """Modelo para actualizar una etapa"""
    nombre: Optional[str] = None
    fecha_inicio: Optional[str] = None
    fecha_fin_planificada: Optional[str] = None
    fecha_fin_real: Optional[str] = None
    estado: Optional[str] = None

class ActividadCreate(BaseModel):
    """Modelo para crear una actividad"""
    nombre: str
    descripcion: Optional[str] = None
    fase: Optional[str] = None  # Para agrupar dentro de la etapa
    fecha_inicio: Optional[str] = None
    fecha_fin_planificada: Optional[str] = None
    prioridad: str = ActividadPrioridad.MEDIA
    responsables_ids: Optional[List[str]] = None
    actividad_padre_id: Optional[str] = None  # Para jerarquía de actividades

class ActividadUpdate(BaseModel):
    """Modelo para actualizar una actividad"""
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    fase: Optional[str] = None
    fecha_inicio: Optional[str] = None
    fecha_fin_planificada: Optional[str] = None
    fecha_fin_real: Optional[str] = None
    estado: Optional[str] = None
    prioridad: Optional[str] = None
    porcentaje_avance: Optional[int] = None
    notas: Optional[str] = None
    actividad_padre_id: Optional[str] = None


# ===== UTILITY FUNCTIONS =====

import re

def validate_password(password: str) -> tuple[bool, str]:
    """
    Validate password requirements:
    - Minimum 6 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one digit
    - Special characters are allowed: !@#$%^&*()_+-=[]{}|;':\",./<>?
    Returns: (is_valid, error_message)
    """
    if len(password) < 6:
        return False, "La contraseña debe tener al menos 6 caracteres"
    
    if not re.search(r'[A-Z]', password):
        return False, "La contraseña debe contener al menos una letra mayúscula"
    
    if not re.search(r'[a-z]', password):
        return False, "La contraseña debe contener al menos una letra minúscula"
    
    if not re.search(r'\d', password):
        return False, "La contraseña debe contener al menos un número"
    
    # Allow special characters - password is valid if it passes above checks
    return True, ""

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token inválido")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    token = credentials.credentials
    payload = decode_token(token)
    user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuario no encontrado")
    return user

async def check_permission(user: dict, permission: str) -> bool:
    """
    Verifica si un usuario tiene un permiso específico.
    - Administradores tienen todos los permisos por defecto.
    - Coordinadores tienen todos los permisos por defecto.
    - Otros usuarios necesitan el permiso explícitamente asignado.
    """
    # Admin y Coordinador tienen todos los permisos por defecto
    if user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        return True
    
    # Verificar si el usuario tiene el permiso explícitamente
    user_permissions = user.get('permissions', [])
    return permission in user_permissions

def require_permission(permission: str):
    """Dependency factory para requerir un permiso específico"""
    async def permission_checker(current_user: dict = Depends(get_current_user)):
        has_permission = await check_permission(current_user, permission)
        if not has_permission:
            permission_desc = Permission.get_description(permission)
            raise HTTPException(
                status_code=403,
                detail=f"No tiene el permiso requerido: {permission_desc}"
            )
        return current_user
    return permission_checker

async def generate_radicado() -> str:
    """
    Genera un número de radicado con consecutivo global.
    Formato: RASMGC-XXXX-DD-MM-YYYY
    XXXX es un consecutivo global que NUNCA se reinicia.
    """
    now = datetime.now()
    date_str = now.strftime("%d-%m-%Y")
    
    # Obtener e incrementar el contador global atómicamente
    result = await db.counters.find_one_and_update(
        {"_id": "radicado_counter"},
        {"$inc": {"sequence": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER
    )
    
    sequence = str(result["sequence"]).zfill(4)
    return f"RASMGC-{sequence}-{date_str}"

async def enviar_correo_con_adjunto(destinatario: str, asunto: str, contenido_html: str, adjunto_path: str = None, adjunto_nombre: str = None):
    """Wrapper async para enviar correo con adjunto."""
    await send_email(destinatario, asunto, contenido_html, adjunto_path, adjunto_nombre)

async def send_email(to_email: str, subject: str, body: str, attachment_path: str = None, attachment_name: str = None):
    if not SMTP_USER or not SMTP_PASSWORD:
        logging.warning("SMTP credentials not configured, skipping email")
        return
    
    try:
        from email.mime.base import MIMEBase
        from email import encoders
        
        msg = MIMEMultipart()
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))
        
        # Add attachment if provided
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{attachment_name or os.path.basename(attachment_path)}"')
                msg.attach(part)
        
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        logging.info(f"Email sent to {to_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {str(e)}")


def get_email_template(titulo: str, contenido: str, radicado: str = None, tipo_notificacion: str = "info", boton_texto: str = None, boton_url: str = None) -> str:
    """
    Genera una plantilla HTML profesional para correos electrónicos.
    
    Args:
        titulo: Título principal del correo
        contenido: Contenido HTML del mensaje
        radicado: Número de radicado (opcional)
        tipo_notificacion: "info", "success", "warning", "error"
        boton_texto: Texto del botón CTA (opcional)
        boton_url: URL del botón (opcional)
    """
    frontend_url = os.environ.get('FRONTEND_URL', 'https://land-registry-17.preview.emergentagent.com')
    logo_url = f"{frontend_url}/logo-asomunicipios.png"
    
    # Colores según tipo de notificación
    colores = {
        "info": {"bg": "#009846", "accent": "#10b981", "badge": "#0ea5e9"},
        "success": {"bg": "#009846", "accent": "#10b981", "badge": "#22c55e"},
        "warning": {"bg": "#d97706", "accent": "#f59e0b", "badge": "#f59e0b"},
        "error": {"bg": "#dc2626", "accent": "#ef4444", "badge": "#ef4444"}
    }
    color = colores.get(tipo_notificacion, colores["info"])
    
    radicado_html = ""
    if radicado:
        radicado_html = f'''
        <div style="background: linear-gradient(135deg, {color["badge"]} 0%, {color["accent"]} 100%); 
                    color: white; padding: 8px 16px; border-radius: 20px; 
                    display: inline-block; font-size: 14px; font-weight: bold; margin-bottom: 15px;">
            📋 Radicado: {radicado}
        </div>
        '''
    
    boton_html = ""
    if boton_texto and boton_url:
        boton_html = f'''
        <div style="text-align: center; margin: 25px 0;">
            <a href="{boton_url}" style="background: linear-gradient(135deg, {color["bg"]} 0%, {color["accent"]} 100%); 
                      color: white; padding: 14px 30px; border-radius: 8px; 
                      text-decoration: none; font-weight: bold; display: inline-block;
                      box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                {boton_texto}
            </a>
        </div>
        '''
    
    return f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f1f5f9;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
            <!-- Header con imagen de mapa y gradiente -->
            <!--[if gte mso 9]>
            <v:rect xmlns:v="urn:schemas-microsoft-com:vml" fill="true" stroke="false" style="width:600px;height:180px;">
            <v:fill type="frame" src="https://images.unsplash.com/photo-1662140246046-fc44f41e4362?q=80&w=800" color="{color["bg"]}" />
            <v:textbox inset="0,0,0,0">
            <![endif]-->
            <table cellpadding="0" cellspacing="0" border="0" width="100%" style="border-radius: 16px 16px 0 0; overflow: hidden;">
                <tr>
                    <td background="https://images.unsplash.com/photo-1662140246046-fc44f41e4362?q=80&w=800" 
                        bgcolor="{color["bg"]}" 
                        style="background-color: {color["bg"]}; background-image: url('https://images.unsplash.com/photo-1662140246046-fc44f41e4362?q=80&w=800'); background-size: cover; background-position: center; border-radius: 16px 16px 0 0; padding: 30px; text-align: center;">
                        <!--[if gte mso 9]>
                        <table cellpadding="0" cellspacing="0" border="0" width="100%"><tr><td style="padding: 30px; text-align: center;">
                        <![endif]-->
                        <div style="background: rgba(4, 120, 87, 0.85); padding: 25px; border-radius: 12px;">
                            <img src="{logo_url}" alt="Asomunicipios" style="height: 60px; margin-bottom: 12px; border-radius: 8px; background: white; padding: 6px;">
                            <h1 style="color: white; margin: 0; font-size: 20px; font-weight: 600;">
                                Asociación de Municipios del Catatumbo
                            </h1>
                            <p style="color: #a7f3d0; margin: 6px 0 0 0; font-size: 13px;">
                                Provincia de Ocaña y Sur del Cesar
                            </p>
                        </div>
                        <!--[if gte mso 9]>
                        </td></tr></table>
                        <![endif]-->
                    </td>
                </tr>
            </table>
            <!--[if gte mso 9]>
            </v:textbox>
            </v:rect>
            <![endif]-->
            
            <!-- Contenido principal -->
            <div style="background: white; padding: 35px; border-left: 1px solid #e2e8f0; border-right: 1px solid #e2e8f0;">
                {radicado_html}
                <h2 style="color: #1e293b; margin: 0 0 20px 0; font-size: 20px; font-weight: 600;">
                    {titulo}
                </h2>
                <div style="color: #475569; font-size: 15px; line-height: 1.7;">
                    {contenido}
                </div>
                {boton_html}
            </div>
            
            <!-- Footer -->
            <div style="background: #f8fafc; padding: 25px; border-radius: 0 0 16px 16px; 
                        border: 1px solid #e2e8f0; border-top: none; text-align: center;">
                <p style="color: #64748b; font-size: 13px; margin: 0 0 10px 0;">
                    Este es un mensaje automático del Sistema de Gestión Catastral
                </p>
                <p style="color: #94a3b8; font-size: 12px; margin: 0;">
                    © {datetime.now().year} Asomunicipios - Todos los derechos reservados
                </p>
                <div style="margin-top: 15px;">
                    <a href="{frontend_url}" style="color: {color["bg"]}; text-decoration: none; font-size: 13px;">
                        Acceder al Sistema
                    </a>
                </div>
            </div>
        </div>
    </body>
    </html>
    '''


def get_finalizacion_email(radicado: str, tipo_tramite: str, nombre_solicitante: str, con_archivos: bool = False) -> str:
    """Genera el correo de finalización de trámite."""
    contenido = f'''
    <p>Estimado(a) <strong>{nombre_solicitante}</strong>,</p>
    
    <p>Nos complace informarle que su trámite ha sido <strong style="color: #22c55e;">finalizado exitosamente</strong>.</p>
    
    <div style="background: #f0fdf4; border-left: 4px solid #22c55e; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
        <p style="margin: 0 0 8px 0;"><strong>Tipo de trámite:</strong> {tipo_tramite}</p>
        <p style="margin: 0;"><strong>Estado:</strong> ✅ Finalizado</p>
    </div>
    '''
    
    if con_archivos:
        contenido += '''
    <p>📎 <strong>Documentos adjuntos:</strong> Se han incluido los documentos de respuesta en este correo. 
    Por favor revise los archivos adjuntos.</p>
    '''
    
    contenido += '''
    <p>Si tiene alguna pregunta o requiere información adicional, no dude en contactarnos.</p>
    
    <p style="margin-top: 25px;">Atentamente,<br>
    <strong>Equipo de Gestión Catastral</strong><br>
    <span style="color: #64748b;">Asomunicipios</span></p>
    '''
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://land-registry-17.preview.emergentagent.com')
    
    return get_email_template(
        titulo="¡Su trámite ha sido finalizado!",
        contenido=contenido,
        radicado=radicado,
        tipo_notificacion="success",
        boton_texto="Ver Detalles del Trámite",
        boton_url=f"{frontend_url}/mis-peticiones"
    )


def get_actualizacion_email(radicado: str, estado_nuevo: str, nombre_solicitante: str, observaciones: str = None) -> str:
    """Genera el correo de actualización de estado."""
    estados_info = {
        "radicado": {"texto": "Radicado", "color": "#3b82f6", "icono": "📝", "mensaje": "Su trámite ha sido registrado en el sistema."},
        "asignado": {"texto": "Asignado", "color": "#8b5cf6", "icono": "👤", "mensaje": "Su trámite ha sido asignado a un gestor para su procesamiento."},
        "en_proceso": {"texto": "En Proceso", "color": "#0ea5e9", "icono": "⚙️", "mensaje": "Su trámite está siendo procesado por nuestro equipo."},
        "revision": {"texto": "En Revisión", "color": "#f59e0b", "icono": "🔍", "mensaje": "Su trámite está siendo revisado por el coordinador."},
        "aprobado": {"texto": "Aprobado", "color": "#10b981", "icono": "✓", "mensaje": "Su trámite ha sido aprobado y está en proceso de finalización."},
        "rechazado": {"texto": "Rechazado", "color": "#ef4444", "icono": "❌", "mensaje": "Lamentablemente su trámite ha sido rechazado. Por favor revise las observaciones."},
        "devuelto": {"texto": "Devuelto para Corrección", "color": "#f97316", "icono": "↩️", "mensaje": "Su trámite ha sido devuelto para correcciones. Por favor revise las observaciones e ingrese al sistema para corregir y reenviar."},
        "finalizado": {"texto": "Finalizado", "color": "#22c55e", "icono": "✅", "mensaje": "Su trámite ha sido completado exitosamente."}
    }
    
    info = estados_info.get(estado_nuevo, {"texto": estado_nuevo, "color": "#64748b", "icono": "📋", "mensaje": "El estado de su trámite ha sido actualizado."})
    
    # Agregar sección de observaciones si es devolución
    observaciones_html = ""
    if observaciones and estado_nuevo == "devuelto":
        observaciones_html = f'''
        <div style="background: #fff7ed; border: 1px solid #fdba74; padding: 15px; margin: 20px 0; border-radius: 8px;">
            <p style="margin: 0 0 8px 0; font-weight: 600; color: #c2410c;">📋 Observaciones del Gestor:</p>
            <p style="margin: 0; color: #9a3412; white-space: pre-line;">{observaciones}</p>
        </div>
        <p><strong>¿Qué debe hacer?</strong></p>
        <ol style="color: #64748b; padding-left: 20px;">
            <li>Ingrese al sistema con su cuenta</li>
            <li>Vaya a "Mis Peticiones" y seleccione este trámite</li>
            <li>Revise las observaciones y realice las correcciones solicitadas</li>
            <li>Adjunte los documentos necesarios si aplica</li>
            <li>Haga clic en "Reenviar para Revisión"</li>
        </ol>
        '''
    
    contenido = f'''
    <p>Estimado(a) <strong>{nombre_solicitante}</strong>,</p>
    
    <p>Le informamos que el estado de su trámite ha sido actualizado.</p>
    
    <div style="background: #f8fafc; border-left: 4px solid {info["color"]}; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
        <p style="margin: 0; font-size: 18px;">
            {info["icono"]} <strong style="color: {info["color"]};">{info["texto"]}</strong>
        </p>
        <p style="margin: 10px 0 0 0; color: #64748b;">{info["mensaje"]}</p>
    </div>
    
    {observaciones_html}
    
    <p>Puede consultar el detalle completo de su trámite accediendo al sistema.</p>
    
    <p style="margin-top: 25px;">Atentamente,<br>
    <strong>Equipo de Gestión Catastral</strong><br>
    <span style="color: #64748b;">Asomunicipios</span></p>
    '''
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://land-registry-17.preview.emergentagent.com')
    tipo_noti = "error" if estado_nuevo == "rechazado" else ("warning" if estado_nuevo == "devuelto" else "info")
    
    return get_email_template(
        titulo="Actualización de su Trámite",
        contenido=contenido,
        radicado=radicado,
        tipo_notificacion=tipo_noti,
        boton_texto="Ver Mi Trámite",
        boton_url=f"{frontend_url}/mis-peticiones"
    )


def get_nueva_peticion_email(radicado: str, solicitante: str, tipo_tramite: str, municipio: str) -> str:
    """Genera el correo de nueva petición para staff."""
    contenido = f'''
    <p>Se ha registrado una nueva petición en el sistema.</p>
    
    <div style="background: #eff6ff; border-radius: 8px; padding: 20px; margin: 20px 0;">
        <table style="width: 100%; border-collapse: collapse;">
            <tr>
                <td style="padding: 8px 0; color: #64748b; width: 140px;">👤 Solicitante:</td>
                <td style="padding: 8px 0; font-weight: 600; color: #1e293b;">{solicitante}</td>
            </tr>
            <tr>
                <td style="padding: 8px 0; color: #64748b;">📋 Tipo de Trámite:</td>
                <td style="padding: 8px 0; font-weight: 600; color: #1e293b;">{tipo_tramite}</td>
            </tr>
            <tr>
                <td style="padding: 8px 0; color: #64748b;">📍 Municipio:</td>
                <td style="padding: 8px 0; font-weight: 600; color: #1e293b;">{municipio}</td>
            </tr>
        </table>
    </div>
    
    <p>Por favor, revise y gestione esta solicitud a la brevedad posible.</p>
    '''
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://land-registry-17.preview.emergentagent.com')
    
    return get_email_template(
        titulo="Nueva Petición Registrada",
        contenido=contenido,
        radicado=radicado,
        tipo_notificacion="info",
        boton_texto="Ver Petición",
        boton_url=f"{frontend_url}/todas-peticiones"
    )


def get_asignacion_email(radicado: str, tipo_tramite: str, gestor_nombre: str) -> str:
    """Genera el correo de asignación para gestor."""
    contenido = f'''
    <p>Hola <strong>{gestor_nombre}</strong>,</p>
    
    <p>Se te ha asignado un nuevo trámite para gestionar.</p>
    
    <div style="background: #faf5ff; border-left: 4px solid #8b5cf6; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
        <p style="margin: 0 0 8px 0;"><strong>Tipo de trámite:</strong> {tipo_tramite}</p>
        <p style="margin: 0;"><strong>Estado:</strong> Asignado a ti</p>
    </div>
    
    <p>Por favor, revisa el trámite y procede con su gestión según los procedimientos establecidos.</p>
    
    <p style="margin-top: 25px;">Saludos,<br>
    <strong>Sistema de Gestión Catastral</strong></p>
    '''
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://land-registry-17.preview.emergentagent.com')
    
    return get_email_template(
        titulo="Nuevo Trámite Asignado",
        contenido=contenido,
        radicado=radicado,
        tipo_notificacion="info",
        boton_texto="Ver Trámite Asignado",
        boton_url=f"{frontend_url}/mis-peticiones"
    )


def get_nuevos_archivos_email(radicado: str, es_staff: bool = False) -> str:
    """Genera el correo de notificación de nuevos archivos."""
    if es_staff:
        contenido = '''
        <p>El solicitante ha cargado nuevos documentos en su trámite.</p>
        
        <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
            <p style="margin: 0;">📎 <strong>Nuevos archivos disponibles</strong></p>
            <p style="margin: 8px 0 0 0; color: #92400e;">Por favor revise los documentos adjuntos.</p>
        </div>
        '''
    else:
        contenido = '''
        <p>Se han agregado nuevos documentos a su trámite.</p>
        
        <div style="background: #ecfdf5; border-left: 4px solid #10b981; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
            <p style="margin: 0;">📎 <strong>Documentos disponibles</strong></p>
            <p style="margin: 8px 0 0 0; color: #065f46;">Puede descargarlos desde el sistema.</p>
        </div>
        '''
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://land-registry-17.preview.emergentagent.com')
    
    return get_email_template(
        titulo="Nuevos Documentos en su Trámite",
        contenido=contenido,
        radicado=radicado,
        tipo_notificacion="info",
        boton_texto="Ver Documentos",
        boton_url=f"{frontend_url}/mis-peticiones"
    )


class TestEmailRequest(BaseModel):
    to_email: str

@api_router.post("/admin/test-email")
async def send_test_email(request: TestEmailRequest, current_user: dict = Depends(get_current_user)):
    """Envía un correo de prueba (solo admin)"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden enviar correos de prueba")
    
    contenido = '''
    <p>Este es un <strong>correo de prueba</strong> enviado desde el sistema de gestión catastral de Asomunicipios.</p>
    
    <p>Si estás recibiendo este mensaje, significa que la configuración de correo electrónico 
    con <strong>Office 365</strong> está funcionando correctamente.</p>
    
    <div style="background: #ecfdf5; border-left: 4px solid #009846; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
        <p style="color: #009846; margin: 0; font-weight: bold;">✅ Configuración SMTP verificada:</p>
        <p style="color: #065f46; margin: 8px 0 0 0; font-size: 14px;">
            Servidor: smtp.office365.com<br>
            Remitente: catastro@asomunicipios.gov.co
        </p>
    </div>
    '''
    
    html_body = get_email_template(
        titulo="Correo de Prueba Exitoso",
        contenido=contenido,
        tipo_notificacion="success"
    )
    
    try:
        await send_email(
            request.to_email,
            "✅ Prueba de Correo - Asomunicipios Catastro",
            html_body
        )
        return {"message": f"Correo de prueba enviado a {request.to_email}", "success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo: {str(e)}")


@api_router.post("/admin/send-ficha-tecnica")
async def send_ficha_tecnica_email(request: TestEmailRequest, current_user: dict = Depends(get_current_user)):
    """Envía la ficha técnica por correo con PDF adjunto (solo admin)"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden enviar la ficha técnica")
    
    # Generate PDF first
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, spaceAfter=20, alignment=TA_CENTER, textColor=colors.HexColor('#009846'))
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10, alignment=TA_CENTER, textColor=colors.HexColor('#065f46'))
    section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=14, spaceBefore=20, spaceAfter=10, textColor=colors.HexColor('#009846'))
    body_style = ParagraphStyle('CustomBody', parent=styles['Normal'], fontSize=11, spaceAfter=8, leading=14)
    bullet_style = ParagraphStyle('BulletStyle', parent=styles['Normal'], fontSize=10, leftIndent=20, spaceAfter=5, leading=13)
    
    elements = []
    
    elements.append(Paragraph("ASOMUNICIPIOS", title_style))
    elements.append(Paragraph("Sistema de Gestión Catastral", subtitle_style))
    elements.append(Paragraph("Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar", 
                             ParagraphStyle('Small', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.gray)))
    elements.append(Spacer(1, 20))
    elements.append(Table([['']], colWidths=[6.5*inch], rowHeights=[2], style=TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#009846'))])))
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("1. DESCRIPCIÓN GENERAL", section_style))
    elements.append(Paragraph(
        "El Sistema de Gestión Catastral de Asomunicipios es una plataforma web integral diseñada para modernizar "
        "y optimizar los procesos de gestión catastral de los municipios asociados.",
        body_style))
    
    elements.append(Paragraph("2. FUNCIONALIDADES PRINCIPALES", section_style))
    funcionalidades = [
        "<b>Gestión de Trámites:</b> Radicación con número único consecutivo, seguimiento y notificaciones.",
        "<b>Gestión de Predios:</b> Administración de +174,000 predios con información R1/R2.",
        "<b>Visor Geográfico:</b> Mapa interactivo con geometrías GDB.",
        "<b>Importación de Datos:</b> Carga masiva de Excel y archivos GDB.",
        "<b>Sistema de Roles:</b> 6 roles diferenciados con permisos granulares.",
        "<b>Reportes:</b> Exportación a Excel/PDF con filtros avanzados.",
        "<b>PWA:</b> Acceso móvil con funcionamiento offline."
    ]
    for func in funcionalidades:
        elements.append(Paragraph(f"• {func}", bullet_style))
    
    elements.append(Paragraph("3. ESTADÍSTICAS", section_style))
    stats_data = [["Métrica", "Valor"], ["Total Predios", "174,419"], ["Con Geometría", "143,354 (82%)"], ["Municipios", "25+"]]
    stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#009846')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(stats_table)
    
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, textColor=colors.gray)
    elements.append(Paragraph(f"Documento generado el {datetime.now().strftime('%d/%m/%Y')}", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    # Save PDF temporarily
    pdf_filename = f"Ficha_Tecnica_Asomunicipios_{datetime.now().strftime('%Y%m%d')}.pdf"
    temp_path = UPLOAD_DIR / pdf_filename
    with open(temp_path, 'wb') as f:
        f.write(buffer.getvalue())
    
    # Send email with attachment using standard template
    contenido = f'''
    <p>Adjunto encontrará la <strong>Ficha Técnica</strong> del Sistema de Gestión Catastral de Asomunicipios, 
    la cual incluye:</p>
    <ul style="line-height: 1.8;">
        <li>Descripción general del sistema</li>
        <li>Funcionalidades principales implementadas</li>
        <li>Beneficios clave para la organización</li>
        <li>Estadísticas actuales del sistema</li>
        <li>Stack tecnológico utilizado</li>
        <li>Roadmap de mejoras futuras</li>
    </ul>
    <div style="background: #ecfdf5; border-left: 4px solid #009846; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
        <p style="color: #009846; margin: 0;">
            <strong>📎 Archivo adjunto:</strong> {pdf_filename}
        </p>
    </div>
    '''
    
    html_body = get_email_template(
        titulo="📋 Ficha Técnica del Sistema",
        contenido=contenido,
        tipo_notificacion="info"
    )
    
    try:
        await send_email(
            request.to_email,
            "📋 Ficha Técnica - Sistema de Gestión Catastral Asomunicipios",
            html_body,
            attachment_path=str(temp_path),
            attachment_name=pdf_filename
        )
        return {"message": f"Ficha técnica enviada a {request.to_email}", "success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo: {str(e)}")


@api_router.get("/reports/ficha-tecnica")
async def generate_ficha_tecnica():
    """Genera PDF con ficha técnica del sistema para presentación ejecutiva"""
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#009846')
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#065f46')
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor('#009846'),
        borderPadding=(0, 0, 5, 0)
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=8,
        leading=14
    )
    
    bullet_style = ParagraphStyle(
        'BulletStyle',
        parent=styles['Normal'],
        fontSize=10,
        leftIndent=20,
        spaceAfter=5,
        leading=13
    )
    
    elements = []
    
    # Header
    elements.append(Paragraph("ASOMUNICIPIOS", title_style))
    elements.append(Paragraph("Sistema de Gestión Catastral", subtitle_style))
    elements.append(Paragraph("Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar", 
                             ParagraphStyle('Small', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.gray)))
    elements.append(Spacer(1, 20))
    
    # Horizontal line
    elements.append(Table([['']], colWidths=[6.5*inch], rowHeights=[2], 
                         style=TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#009846'))])))
    elements.append(Spacer(1, 20))
    
    # DESCRIPCIÓN GENERAL
    elements.append(Paragraph("1. DESCRIPCIÓN GENERAL", section_style))
    elements.append(Paragraph(
        "El Sistema de Gestión Catastral de Asomunicipios es una plataforma web integral diseñada para modernizar "
        "y optimizar los procesos de gestión catastral de los municipios asociados. La plataforma permite la administración "
        "completa del ciclo de vida de trámites catastrales, gestión de predios, y visualización geográfica de información territorial.",
        body_style
    ))
    elements.append(Spacer(1, 10))
    
    # FUNCIONALIDADES PRINCIPALES
    elements.append(Paragraph("2. FUNCIONALIDADES PRINCIPALES", section_style))
    
    funcionalidades = [
        ("<b>Gestión de Trámites (PQRS):</b> Sistema completo de radicación con número único consecutivo (RASMCG-XXXX-DD-MM-YYYY), "
         "seguimiento de estados, asignación a gestores, y notificaciones automáticas por correo electrónico."),
        ("<b>Gestión de Predios:</b> Administración de más de 174,000 predios con información alfanumérica (R1/R2), "
         "filtros avanzados por municipio, zona, destino económico y vigencia."),
        ("<b>Visor Geográfico:</b> Visualización de predios en mapa interactivo con geometrías GDB, vinculación automática "
         "entre datos alfanuméricos y gráficos."),
        ("<b>Importación de Datos:</b> Carga masiva de archivos Excel (R1/R2) y archivos GDB para actualización de base gráfica."),
        ("<b>Sistema de Roles y Permisos:</b> Control de acceso granular con 6 roles diferenciados (Usuario, Atención al Usuario, "
         "Gestor, Coordinador, Administrador, Comunicaciones)."),
        ("<b>Histórico y Reportes:</b> Exportación de histórico de trámites a Excel con filtros avanzados, generación de certificados "
         "catastrales en PDF."),
        ("<b>Notificaciones:</b> Sistema de alertas por correo electrónico para asignaciones, cambios de estado y recuperación de contraseña."),
        ("<b>Aplicación Móvil (PWA):</b> Acceso desde dispositivos móviles con capacidad de funcionamiento offline para consulta de predios.")
    ]
    
    for func in funcionalidades:
        elements.append(Paragraph(f"• {func}", bullet_style))
    
    elements.append(Spacer(1, 10))
    
    # BENEFICIOS
    elements.append(Paragraph("3. BENEFICIOS CLAVE", section_style))
    
    beneficios_data = [
        ["Beneficio", "Descripción"],
        ["Eficiencia Operativa", "Reducción de tiempos en gestión de trámites mediante automatización de procesos"],
        ["Trazabilidad", "Seguimiento completo del ciclo de vida de cada trámite y cambio en predios"],
        ["Accesibilidad", "Acceso 24/7 desde cualquier dispositivo, incluyendo modo offline"],
        ["Integración de Datos", "Unificación de información alfanumérica (R1/R2) y geográfica (GDB)"],
        ["Transparencia", "Los usuarios pueden consultar el estado de sus trámites en tiempo real"],
        ["Seguridad", "Autenticación robusta, roles diferenciados y auditoría de acciones"]
    ]
    
    beneficios_table = Table(beneficios_data, colWidths=[2*inch, 4.5*inch])
    beneficios_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#009846')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(beneficios_table)
    elements.append(Spacer(1, 15))
    
    # ESTADÍSTICAS ACTUALES
    elements.append(Paragraph("4. ESTADÍSTICAS DEL SISTEMA", section_style))
    
    stats_data = [
        ["Métrica", "Valor"],
        ["Total de Predios Registrados", "174,419"],
        ["Predios con Geometría", "143,354 (82%)"],
        ["Municipios Cubiertos", "25+"],
        ["Usuarios Activos", "25+"],
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 3.5*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#009846')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 15))
    
    # TECNOLOGÍA
    elements.append(Paragraph("5. STACK TECNOLÓGICO", section_style))
    
    tech_items = [
        "<b>Backend:</b> FastAPI (Python) - Framework moderno de alto rendimiento",
        "<b>Frontend:</b> React.js con Tailwind CSS - Interfaz responsiva y moderna",
        "<b>Base de Datos:</b> MongoDB - Base de datos NoSQL escalable",
        "<b>Mapas:</b> Leaflet + React-Leaflet - Visualización geográfica interactiva",
        "<b>PWA:</b> Service Worker + IndexedDB - Funcionamiento offline",
        "<b>Correo:</b> Integración con Microsoft Office 365",
        "<b>Seguridad:</b> JWT (JSON Web Tokens) para autenticación"
    ]
    
    for tech in tech_items:
        elements.append(Paragraph(f"• {tech}", bullet_style))
    
    elements.append(Spacer(1, 15))
    
    # MEJORAS FUTURAS
    elements.append(Paragraph("6. ROADMAP DE MEJORAS", section_style))
    
    mejoras = [
        ["Prioridad", "Mejora", "Beneficio Esperado"],
        ["Alta", "Generación de archivos XTF (IGAC)", "Cumplimiento normativo Resolución 0301/2025"],
        ["Alta", "Flujo de subsanación de trámites", "Reducción de tiempos en corrección de documentos"],
        ["Media", "App nativa (Android/iOS)", "Mayor alcance y presencia en tiendas de aplicaciones"],
        ["Media", "Dashboard de productividad", "Métricas de desempeño de gestores"],
        ["Media", "Firmas digitales en PDFs", "Validez legal de documentos generados"],
        ["Baja", "Backups automáticos", "Protección de datos ante contingencias"]
    ]
    
    mejoras_table = Table(mejoras, colWidths=[1.2*inch, 2.5*inch, 2.8*inch])
    mejoras_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#009846')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#fef2f2')),
        ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#fef2f2')),
        ('BACKGROUND', (0, 3), (0, 4), colors.HexColor('#fefce8')),
        ('BACKGROUND', (0, 5), (0, 5), colors.HexColor('#fefce8')),
        ('BACKGROUND', (0, 6), (0, 6), colors.HexColor('#f0fdf4')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(mejoras_table)
    elements.append(Spacer(1, 20))
    
    # FOOTER
    elements.append(Table([['']], colWidths=[6.5*inch], rowHeights=[2], 
                         style=TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#009846'))])))
    elements.append(Spacer(1, 10))
    
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, textColor=colors.gray)
    elements.append(Paragraph(f"Documento generado el {datetime.now().strftime('%d de %B de %Y')}", footer_style))
    elements.append(Paragraph("Sistema de Gestión Catastral - Asomunicipios", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Save to file
    temp_path = UPLOAD_DIR / f"Ficha_Tecnica_Asomunicipios_{datetime.now().strftime('%Y%m%d')}.pdf"
    with open(temp_path, 'wb') as f:
        f.write(buffer.getvalue())
    
    return FileResponse(
        path=temp_path,
        filename=f"Ficha_Tecnica_Asomunicipios_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type='application/pdf'
    )


# ===== AUTH ROUTES =====

@api_router.post("/auth/register")
async def register(user_data: UserRegister):
    # Check if user exists (case-insensitive)
    email_escaped = re.escape(user_data.email.lower())
    existing_user = await db.users.find_one(
        {"email": {"$regex": f"^{email_escaped}$", "$options": "i"}}
    )
    if existing_user:
        # Si existe pero no está verificado, permitir re-registro (reenviar código)
        if not existing_user.get('email_verified', False):
            # Generar nuevo código
            verification_code = str(random.randint(100000, 999999))
            expires_at = (datetime.now(timezone.utc) + timedelta(minutes=30)).isoformat()
            
            await db.users.update_one(
                {"id": existing_user['id']},
                {"$set": {
                    "verification_code": verification_code,
                    "verification_code_expires": expires_at,
                    "password": hash_password(user_data.password),
                    "full_name": format_nombre_propio(user_data.full_name)
                }}
            )
            
            # Enviar código por email
            try:
                await enviar_codigo_verificacion(user_data.email, verification_code, format_nombre_propio(user_data.full_name))
            except Exception as e:
                logger.error(f"Error enviando código de verificación: {e}")
            
            return {
                "message": "Se ha enviado un código de verificación a su correo electrónico",
                "requires_verification": True,
                "email": user_data.email
            }
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El correo ya está registrado")
    
    # Validate password
    is_valid, error_msg = validate_password(user_data.password)
    if not is_valid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_msg)
    
    # Formatear el nombre propio con mayúsculas y tildes correctas
    nombre_formateado = format_nombre_propio(user_data.full_name)
    
    # Generar código de verificación (6 dígitos)
    verification_code = str(random.randint(100000, 999999))
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=30)).isoformat()
    
    # Always assign usuario role on self-registration
    user = User(
        email=user_data.email,
        full_name=nombre_formateado,
        role=UserRole.USUARIO,
        email_verified=False,
        verification_code=verification_code,
        verification_code_expires=expires_at
    )
    
    doc = user.model_dump()
    doc['password'] = hash_password(user_data.password)
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    # Enviar código por email
    try:
        await enviar_codigo_verificacion(user_data.email, verification_code, nombre_formateado)
    except Exception as e:
        logger.error(f"Error enviando código de verificación: {e}")
    
    # NO devolver token hasta que el email esté verificado
    return {
        "message": "Se ha enviado un código de verificación a su correo electrónico",
        "requires_verification": True,
        "email": user_data.email
    }


async def enviar_codigo_verificacion(email: str, codigo: str, nombre: str):
    """Envía el código de verificación por email"""
    contenido = f'''
    <h2 style="color: #1e293b; margin-top: 0;">¡Hola {nombre}!</h2>
    <p>Para completar tu registro en el Sistema de Gestión Catastral, ingresa el siguiente código de verificación:</p>
    <div style="background-color: #065f46; color: white; font-size: 32px; font-weight: bold; padding: 20px; text-align: center; border-radius: 8px; letter-spacing: 8px; margin: 20px 0;">
        {codigo}
    </div>
    <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
        <p style="margin: 0; color: #92400e;"><strong>⏰ Importante:</strong> Este código expira en <strong>30 minutos</strong>.</p>
    </div>
    <p style="color: #64748b; font-size: 14px;">Si no solicitaste este registro, puedes ignorar este mensaje.</p>
    '''
    
    html_content = get_email_template(
        titulo="Verificación de Cuenta",
        contenido=contenido,
        tipo_notificacion="info"
    )
    
    await send_email(
        to_email=email,
        subject="Código de Verificación - Asomunicipios",
        body=html_content
    )


@api_router.post("/auth/verify-email")
async def verify_email(data: VerifyEmailCode):
    """Verifica el código de email enviado durante el registro"""
    email_escaped = re.escape(data.email.lower())
    user = await db.users.find_one(
        {"email": {"$regex": f"^{email_escaped}$", "$options": "i"}},
        {"_id": 0}
    )
    
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    if user.get('email_verified', False):
        raise HTTPException(status_code=400, detail="El correo ya está verificado")
    
    # Verificar código
    if user.get('verification_code') != data.code:
        raise HTTPException(status_code=400, detail="Código de verificación inválido")
    
    # Verificar expiración
    expires_str = user.get('verification_code_expires')
    if expires_str:
        expires_at = datetime.fromisoformat(expires_str.replace('Z', '+00:00'))
        if datetime.now(timezone.utc) > expires_at:
            raise HTTPException(status_code=400, detail="El código de verificación ha expirado. Solicita uno nuevo.")
    
    # Marcar como verificado
    await db.users.update_one(
        {"id": user['id']},
        {"$set": {
            "email_verified": True,
            "verification_code": None,
            "verification_code_expires": None
        }}
    )
    
    # Generar token
    token = create_token(user['id'], user['email'], user['role'])
    
    return {
        "message": "Email verificado exitosamente",
        "token": token,
        "user": {
            "id": user['id'],
            "email": user['email'],
            "full_name": user['full_name'],
            "role": user['role']
        }
    }


@api_router.post("/auth/resend-verification")
async def resend_verification_code(data: ResendVerificationCode):
    """Reenvía el código de verificación"""
    email_escaped = re.escape(data.email.lower())
    user = await db.users.find_one(
        {"email": {"$regex": f"^{email_escaped}$", "$options": "i"}},
        {"_id": 0}
    )
    
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    if user.get('email_verified', False):
        raise HTTPException(status_code=400, detail="El correo ya está verificado")
    
    # Generar nuevo código
    verification_code = str(random.randint(100000, 999999))
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=30)).isoformat()
    
    await db.users.update_one(
        {"id": user['id']},
        {"$set": {
            "verification_code": verification_code,
            "verification_code_expires": expires_at
        }}
    )
    
    # Enviar código por email
    try:
        await enviar_codigo_verificacion(data.email, verification_code, user['full_name'])
    except Exception as e:
        logger.error(f"Error reenviando código de verificación: {e}")
        raise HTTPException(status_code=500, detail="Error enviando el código. Intente más tarde.")
    
    return {"message": "Se ha enviado un nuevo código de verificación a su correo"}

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    # Case-insensitive email search using regex (escape special chars)
    email_escaped = re.escape(credentials.email.lower())
    user = await db.users.find_one(
        {"email": {"$regex": f"^{email_escaped}$", "$options": "i"}}, 
        {"_id": 0}
    )
    if not user:
        logger.warning(f"Login fallido: usuario no encontrado para {credentials.email}")
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales inválidas")
    
    logger.info(f"Login intento: {credentials.email} - Usuario encontrado: {user.get('full_name')}")
    
    if not verify_password(credentials.password, user['password']):
        logger.warning(f"Login fallido: contraseña incorrecta para {credentials.email}")
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales inválidas")
    
    logger.info(f"Login exitoso: {credentials.email}")
    
    # Verificar si el email está verificado (excepto admin protegido y usuarios internos)
    is_protected_admin = user.get('email', '').lower() == PROTECTED_ADMIN_EMAIL.lower()
    is_internal_user = user.get('role') in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR, UserRole.ATENCION_USUARIO, UserRole.COMUNICACIONES, UserRole.EMPRESA]
    
    if not user.get('email_verified', True) and not is_protected_admin and not is_internal_user:
        # Reenviar código automáticamente
        verification_code = str(random.randint(100000, 999999))
        expires_at = (datetime.now(timezone.utc) + timedelta(minutes=30)).isoformat()
        
        await db.users.update_one(
            {"id": user['id']},
            {"$set": {
                "verification_code": verification_code,
                "verification_code_expires": expires_at
            }}
        )
        
        try:
            await enviar_codigo_verificacion(user['email'], verification_code, user['full_name'])
        except:
            pass
        
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="email_not_verified"
        )
    
    token = create_token(user['id'], user['email'], user['role'])
    
    # Obtener permisos del usuario
    permissions = user.get('permissions', [])
    
    return {
        "token": token,
        "user": {
            "id": user['id'],
            "email": user['email'],
            "full_name": user['full_name'],
            "role": user['role'],
            "permissions": permissions
        }
    }

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    # Get additional user data from database
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    
    # Obtener permisos del usuario
    permissions = user_db.get('permissions', []) if user_db else []
    
    return {
        "id": current_user['id'],
        "email": current_user['email'],
        "full_name": current_user['full_name'],
        "role": current_user['role'],
        "permissions": permissions,
        "puede_actualizar_gdb": 'upload_gdb' in permissions,
        "puede_cargar_excel": 'import_r1r2' in permissions,
        "puede_aprobar_cambios": 'approve_changes' in permissions
    }


# ===== PASSWORD RECOVERY =====

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Send password reset email"""
    # Case-insensitive email search (escape special chars)
    email_escaped = re.escape(request.email.lower())
    user = await db.users.find_one(
        {"email": {"$regex": f"^{email_escaped}$", "$options": "i"}}, 
        {"_id": 0}
    )
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No existe una cuenta con ese correo")
    
    # Check if SMTP is configured
    if not SMTP_USER or not SMTP_PASSWORD:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="El servicio de correo no está configurado")
    
    # Generate reset token (valid for 1 hour)
    reset_token = str(uuid.uuid4())
    expiration = datetime.now(timezone.utc) + timedelta(hours=1)
    
    # Store reset token in database
    await db.password_resets.delete_many({"email": request.email})  # Remove old tokens
    await db.password_resets.insert_one({
        "email": request.email,
        "token": reset_token,
        "expires_at": expiration.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Get frontend URL from environment or use default
    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    reset_link = f"{frontend_url}/reset-password?token={reset_token}"
    
    # Send email using standard template
    contenido = f'''
    <p>Hola <strong>{user['full_name']}</strong>,</p>
    <p>Hemos recibido una solicitud para restablecer la contraseña de tu cuenta en el Sistema de Gestión Catastral.</p>
    <p>Haz clic en el siguiente botón para crear una nueva contraseña:</p>
    <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
        <p style="margin: 0; color: #92400e;"><strong>⏰ Importante:</strong> Este enlace expirará en 1 hora.</p>
    </div>
    <p style="color: #64748b; font-size: 14px;">Si no solicitaste este cambio, puedes ignorar este correo de forma segura.</p>
    '''
    
    email_body = get_email_template(
        titulo="Recuperación de Contraseña",
        contenido=contenido,
        tipo_notificacion="warning",
        boton_texto="Restablecer Contraseña",
        boton_url=reset_link
    )
    
    try:
        await send_email(request.email, "Recuperación de Contraseña - Asomunicipios", email_body)
    except Exception as e:
        logging.error(f"Failed to send reset email: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al enviar el correo")
    
    return {"message": "Se ha enviado un enlace de recuperación a tu correo"}

@api_router.get("/auth/validate-reset-token")
async def validate_reset_token(token: str):
    """Validate password reset token"""
    reset_record = await db.password_resets.find_one({"token": token}, {"_id": 0})
    
    if not reset_record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Token inválido")
    
    expires_at = datetime.fromisoformat(reset_record['expires_at'])
    if datetime.now(timezone.utc) > expires_at:
        await db.password_resets.delete_one({"token": token})
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El token ha expirado")
    
    return {"valid": True}

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """Reset password with token"""
    reset_record = await db.password_resets.find_one({"token": request.token}, {"_id": 0})
    
    if not reset_record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Token inválido")
    
    expires_at = datetime.fromisoformat(reset_record['expires_at'])
    if datetime.now(timezone.utc) > expires_at:
        await db.password_resets.delete_one({"token": request.token})
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El token ha expirado")
    
    # Validate new password
    is_valid, error_msg = validate_password(request.new_password)
    if not is_valid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_msg)
    
    # Update password
    new_hashed_password = hash_password(request.new_password)
    await db.users.update_one(
        {"email": reset_record['email']},
        {"$set": {"password": new_hashed_password}}
    )
    
    # Delete used token
    await db.password_resets.delete_one({"token": request.token})
    
    return {"message": "Contraseña actualizada exitosamente"}


# ===== USER MANAGEMENT ROUTES =====

@api_router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    # Only admin, coordinador, and atencion can view users
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
        # Asegurar que municipios_asignados siempre esté presente para usuarios empresa
        if user.get('role') == 'empresa' and 'municipios_asignados' not in user:
            user['municipios_asignados'] = []
    
    return users


@api_router.get("/users/gestores-disponibles")
async def get_gestores_disponibles(current_user: dict = Depends(get_current_user)):
    """
    Obtiene lista de gestores disponibles para asignar trabajo.
    Accesible para roles internos incluyendo Gestor.
    """
    # Permitir acceso a roles internos (no ciudadanos)
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Buscar usuarios con roles que pueden recibir trabajo
    roles_asignables = ['gestor', 'gestor_auxiliar', 'atencion_usuario', 'coordinador']
    
    gestores = await db.users.find(
        {"role": {"$in": roles_asignables}},
        {"_id": 0, "id": 1, "full_name": 1, "role": 1, "email": 1}
    ).to_list(100)
    
    # Verificar si el usuario tiene permiso de aprobación
    user_permissions = current_user.get('permissions', [])
    has_approve_permission = 'approve_changes' in user_permissions
    
    # Si tiene permiso de aprobación, puede auto-asignarse
    # Si no tiene permiso, excluirlo de la lista
    if not has_approve_permission:
        gestores = [g for g in gestores if g.get('id') != current_user['id']]
    
    return gestores


# Email del administrador protegido - no se puede cambiar su rol
PROTECTED_ADMIN_EMAIL = "catastro@asomunicipios.gov.co"

@api_router.patch("/users/role")
async def update_user_role(role_update: UserRoleUpdate, current_user: dict = Depends(get_current_user)):
    # Only admin, coordinador, and atencion can change roles
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para cambiar roles")
    
    # Validate new role
    valid_roles = [UserRole.USUARIO, UserRole.ATENCION_USUARIO, UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.COMUNICACIONES, UserRole.EMPRESA]
    if role_update.new_role not in valid_roles:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rol inválido")
    
    user = await db.users.find_one({"id": role_update.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    
    # Proteger al admin principal - nadie puede cambiar su rol
    if user.get('email', '').lower() == PROTECTED_ADMIN_EMAIL.lower():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="No se puede modificar el rol del administrador principal del sistema"
        )
    
    await db.users.update_one(
        {"id": role_update.user_id},
        {"$set": {"role": role_update.new_role}}
    )
    
    return {"message": "Rol actualizado exitosamente", "new_role": role_update.new_role}


# ==================== GESTIÓN DE MUNICIPIOS PARA USUARIOS EMPRESA ====================

@api_router.get("/admin/municipios-disponibles")
async def get_municipios_disponibles(current_user: dict = Depends(get_current_user)):
    """Obtiene la lista de municipios disponibles para asignar a usuarios empresa"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver municipios")
    
    return {
        "municipios": list(MUNICIPIOS_DIVIPOLA.keys())
    }


@api_router.get("/admin/users/{user_id}/municipios")
async def get_user_municipios(user_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene los municipios asignados a un usuario empresa"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver municipios de usuarios")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    return {
        "user_id": user_id,
        "email": user.get("email"),
        "full_name": user.get("full_name"),
        "role": user.get("role"),
        "municipios_asignados": user.get("municipios_asignados", [])
    }


@api_router.patch("/admin/users/{user_id}/municipios")
async def update_user_municipios(
    user_id: str, 
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Actualiza los municipios asignados a un usuario empresa"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para asignar municipios")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Solo permitir asignar municipios a usuarios empresa
    if user.get("role") != UserRole.EMPRESA:
        raise HTTPException(status_code=400, detail="Solo se pueden asignar municipios a usuarios con rol 'Empresa'")
    
    municipios = data.get("municipios_asignados", [])
    
    # Validar que los municipios existan
    municipios_validos = [m for m in municipios if m in MUNICIPIOS_DIVIPOLA]
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"municipios_asignados": municipios_validos}}
    )
    
    return {
        "message": "Municipios asignados correctamente",
        "user_id": user_id,
        "municipios_asignados": municipios_validos
    }


@api_router.post("/admin/migrate-ciudadano-to-usuario")
async def migrate_ciudadano_to_usuario(current_user: dict = Depends(get_current_user)):
    """Migra usuarios con rol 'ciudadano' a 'usuario' (solo admin)"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden ejecutar esta migración")
    
    # Count before
    before_count = await db.users.count_documents({'role': 'ciudadano'})
    
    # Update all ciudadano to usuario
    result = await db.users.update_many(
        {'role': 'ciudadano'},
        {'$set': {'role': 'usuario'}}
    )
    
    # Count after
    usuario_count = await db.users.count_documents({'role': 'usuario'})
    
    return {
        "message": f"Migración completada",
        "users_with_ciudadano_before": before_count,
        "users_migrated": result.modified_count,
        "users_with_usuario_after": usuario_count
    }


@api_router.post("/admin/format-user-names")
async def format_user_names(current_user: dict = Depends(get_current_user)):
    """Formatea los nombres de usuarios existentes con mayúsculas y tildes correctas (solo admin)"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden ejecutar esta operación")
    
    # Get all users
    users = await db.users.find({}, {"_id": 0, "id": 1, "full_name": 1}).to_list(None)
    
    updated_count = 0
    examples = []
    
    for user in users:
        original_name = user['full_name']
        formatted_name = format_nombre_propio(original_name)
        
        if original_name != formatted_name:
            await db.users.update_one(
                {"id": user['id']},
                {"$set": {"full_name": formatted_name}}
            )
            updated_count += 1
            if len(examples) < 5:
                examples.append({"original": original_name, "formatted": formatted_name})
    
    return {
        "message": f"Formateo de nombres completado",
        "total_users": len(users),
        "users_updated": updated_count,
        "examples": examples
    }


@api_router.post("/admin/format-petition-names")
async def format_petition_names(current_user: dict = Depends(get_current_user)):
    """Formatea los nombres de solicitantes en peticiones existentes (solo admin)"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden ejecutar esta operación")
    
    # Get all petitions with nombre_completo
    petitions = await db.petitions.find(
        {"nombre_completo": {"$exists": True}},
        {"_id": 0, "id": 1, "nombre_completo": 1}
    ).to_list(None)
    
    updated_count = 0
    examples = []
    
    for petition in petitions:
        original_name = petition.get('nombre_completo', '')
        if original_name:
            formatted_name = format_nombre_propio(original_name)
            
            if original_name != formatted_name:
                await db.petitions.update_one(
                    {"id": petition['id']},
                    {"$set": {"nombre_completo": formatted_name}}
                )
                updated_count += 1
                if len(examples) < 5:
                    examples.append({"original": original_name, "formatted": formatted_name})
    
    return {
        "message": f"Formateo de nombres en peticiones completado",
        "total_petitions": len(petitions),
        "petitions_updated": updated_count,
        "examples": examples
    }


# ===== PERMISSIONS MANAGEMENT =====

@api_router.get("/permissions/available")
async def get_available_permissions(current_user: dict = Depends(get_current_user)):
    """Obtiene la lista de permisos disponibles en el sistema"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver permisos")
    
    return {
        "permissions": [
            {"key": perm, "description": Permission.get_description(perm)}
            for perm in Permission.all_permissions()
        ]
    }

@api_router.get("/permissions/users")
async def get_users_with_permissions(current_user: dict = Depends(get_current_user)):
    """Obtiene usuarios con sus permisos (para gestores y coordinadores)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para gestionar permisos")
    
    # Obtener usuarios que pueden tener permisos (todos los roles internos)
    roles_internos = [
        UserRole.GESTOR, 
        UserRole.COORDINADOR, 
        UserRole.ATENCION_USUARIO,
        UserRole.COMUNICACIONES,
        UserRole.EMPRESA
    ]
    users = await db.users.find(
        {"role": {"$in": roles_internos}},
        {"_id": 0, "password_hash": 0, "password": 0}
    ).to_list(1000)
    
    # Agregar descripciones de permisos
    for user in users:
        user_permissions = user.get('permissions', [])
        user['permissions_detail'] = [
            {"key": perm, "description": Permission.get_description(perm)}
            for perm in user_permissions
        ]
    
    return {"users": users}

@api_router.patch("/permissions/user")
async def update_user_permissions(
    update: UserPermissionsUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Actualiza los permisos de un usuario"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para modificar permisos")
    
    # Validar que los permisos sean válidos
    valid_permissions = Permission.all_permissions()
    for perm in update.permissions:
        if perm not in valid_permissions:
            raise HTTPException(status_code=400, detail=f"Permiso inválido: {perm}")
    
    # Buscar usuario
    user = await db.users.find_one({"id": update.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # No permitir modificar permisos de administradores
    if user['role'] == UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=400, detail="No se pueden modificar permisos de administradores")
    
    # Actualizar permisos
    await db.users.update_one(
        {"id": update.user_id},
        {"$set": {"permissions": update.permissions}}
    )
    
    # Registrar el cambio en historial
    await db.permissions_history.insert_one({
        "user_id": update.user_id,
        "changed_by": current_user['id'],
        "changed_by_name": current_user['full_name'],
        "old_permissions": user.get('permissions', []),
        "new_permissions": update.permissions,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    # Crear notificación para el usuario si se agregaron nuevos permisos
    old_permissions = set(user.get('permissions', []))
    new_permissions = set(update.permissions)
    added_permissions = new_permissions - old_permissions
    removed_permissions = old_permissions - new_permissions
    
    if added_permissions or removed_permissions:
        # Construir mensaje de notificación
        message_parts = []
        if added_permissions:
            added_desc = [Permission.get_description(p) for p in added_permissions]
            message_parts.append(f"Permisos otorgados: {', '.join(added_desc)}")
        if removed_permissions:
            removed_desc = [Permission.get_description(p) for p in removed_permissions]
            message_parts.append(f"Permisos revocados: {', '.join(removed_desc)}")
        
        await crear_notificacion(
            usuario_id=update.user_id,
            titulo="Actualización de Permisos",
            mensaje=f"{current_user['full_name']} ha actualizado tus permisos. {' | '.join(message_parts)}",
            tipo="permisos",
            enviar_email=True  # Enviar notificación por email
        )
    
    return {
        "message": "Permisos actualizados exitosamente",
        "user_id": update.user_id,
        "permissions": update.permissions
    }

@api_router.get("/permissions/user/{user_id}")
async def get_user_permissions(user_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene los permisos de un usuario específico"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        # Un usuario puede ver sus propios permisos
        if current_user['id'] != user_id:
            raise HTTPException(status_code=403, detail="No tiene permiso para ver permisos de otros usuarios")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    permissions = user.get('permissions', [])
    return {
        "user_id": user_id,
        "permissions": permissions,
        "permissions_detail": [
            {"key": perm, "description": Permission.get_description(perm)}
            for perm in permissions
        ]
    }


# ===== PETITION ROUTES =====

@api_router.post("/petitions")
async def create_petition(
    nombre_completo: str = Form(...),
    correo: str = Form(...),
    telefono: str = Form(...),
    tipo_tramite: str = Form(...),
    municipio: str = Form(...),
    descripcion: str = Form(default=""),
    codigo_predial: str = Form(default=""),
    matricula_inmobiliaria: str = Form(default=""),
    files: List[UploadFile] = File(default=[]),
    current_user: dict = Depends(get_current_user)
):
    radicado = await generate_radicado()
    
    # Si es un certificado catastral, buscar el predio relacionado
    predio_relacionado = None
    if codigo_predial or matricula_inmobiliaria:
        predio = None
        if codigo_predial:
            predio = await db.predios.find_one(
                {"codigo_predial_nacional": codigo_predial.strip()}, 
                {"_id": 0, "id": 1, "codigo_predial_nacional": 1, "r2_registros": 1, "direccion": 1, "municipio": 1}
            )
        elif matricula_inmobiliaria:
            # Buscar la matrícula en r2_registros.matricula_inmobiliaria (fuente R1/R2)
            matricula_limpia = matricula_inmobiliaria.strip()
            predio = await db.predios.find_one(
                {"r2_registros.matricula_inmobiliaria": matricula_limpia}, 
                {"_id": 0, "id": 1, "codigo_predial_nacional": 1, "r2_registros": 1, "direccion": 1, "municipio": 1}
            )
        
        if predio:
            # Extraer la matrícula del R2 si existe
            matricula_r2 = None
            r2_registros = predio.get("r2_registros", [])
            if r2_registros:
                matricula_r2 = r2_registros[0].get("matricula_inmobiliaria")
            
            predio_relacionado = {
                "predio_id": predio.get("id"),
                "codigo_predial": predio.get("codigo_predial_nacional"),
                "matricula": matricula_r2 or matricula_inmobiliaria,
                "direccion": predio.get("direccion"),
                "municipio": predio.get("municipio")
            }
    
    # Save files
    saved_files = []
    for file in files:
        if file.filename:
            file_id = str(uuid.uuid4())
            file_ext = Path(file.filename).suffix
            file_path = UPLOAD_DIR / f"{file_id}{file_ext}"
            
            with file_path.open("wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            saved_files.append({
                "id": file_id,
                "original_name": file.filename,
                "path": str(file_path)
            })
    
    # Initialize historial
    historial = [{
        "accion": "Radicado creado",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": None,
        "estado_nuevo": PetitionStatus.RADICADO,
        "notas": "Petición radicada en el sistema",
        "fecha": datetime.now(timezone.utc).isoformat()
    }]
    
    petition = Petition(
        radicado=radicado,
        user_id=current_user['id'],
        nombre_completo=nombre_completo,
        correo=correo,
        telefono=telefono,
        tipo_tramite=tipo_tramite,
        municipio=municipio,
        descripcion=descripcion,
        archivos=saved_files,
        historial=historial
    )
    
    doc = petition.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    # Agregar información del predio relacionado si existe
    if predio_relacionado:
        doc['predio_relacionado'] = predio_relacionado
    
    # Agregar campos de búsqueda para certificados
    if codigo_predial:
        doc['codigo_predial_buscado'] = codigo_predial.strip()
    if matricula_inmobiliaria:
        doc['matricula_buscada'] = matricula_inmobiliaria.strip()
    
    await db.petitions.insert_one(doc)
    
    # Notificación en plataforma a atención al usuario (NO correo) si la crea un ciudadano
    if current_user['role'] == UserRole.USUARIO:
        atencion_users = await db.users.find({"role": UserRole.ATENCION_USUARIO}, {"_id": 0}).to_list(100)
        for user in atencion_users:
            await crear_notificacion(
                usuario_id=user['id'],
                tipo="info",
                titulo="Nueva petición radicada",
                mensaje=f"Nueva petición {radicado} de {nombre_completo} - {tipo_tramite}",
                enlace=f"/dashboard/peticion/{petition.id}"
            )
    
    return petition

@api_router.get("/petitions")
async def get_petitions(current_user: dict = Depends(get_current_user)):
    # Citizens only see their own petitions
    if current_user['role'] == UserRole.USUARIO:
        query = {"user_id": current_user['id']}
    # Gestores see assigned petitions AND petitions they created
    elif current_user['role'] == UserRole.GESTOR or current_user['role'] == 'gestor_auxiliar':
        query = {
            "$or": [
                {"gestores_asignados": current_user['id']},  # Asignadas a él
                {"user_id": current_user['id']}              # Creadas por él
            ]
        }
    # Empresa solo ve sus propias peticiones (las que ellos crearon)
    elif current_user['role'] == UserRole.EMPRESA:
        query = {"user_id": current_user['id']}
    else:
        # Staff (atencion_usuario, coordinador, administrador, comunicaciones) see all petitions
        query = {}
    
    # Sin límite - retorna TODAS las peticiones
    petitions = await db.petitions.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    
    for petition in petitions:
        if isinstance(petition['created_at'], str):
            petition['created_at'] = datetime.fromisoformat(petition['created_at'])
        if isinstance(petition['updated_at'], str):
            petition['updated_at'] = datetime.fromisoformat(petition['updated_at'])
    
    return petitions

@api_router.get("/petitions/mis-peticiones")
async def get_my_petitions(current_user: dict = Depends(get_current_user)):
    """Obtener peticiones del usuario actual.
    
    Para ciudadanos y empresa: solo peticiones creadas por ellos.
    Para staff (atencion_usuario, gestor, coordinador, etc.): peticiones creadas Y asignadas.
    """
    user_role = current_user['role']
    
    # Ciudadanos y empresa solo ven peticiones que ellos crearon
    if user_role in [UserRole.USUARIO, UserRole.EMPRESA]:
        query = {"user_id": current_user['id']}
    # Staff (gestores, atencion_usuario, coordinadores, admin) ven peticiones creadas Y asignadas
    else:
        query = {
            "$or": [
                {"user_id": current_user['id']},               # Creadas por él
                {"gestores_asignados": current_user['id']}     # Asignadas a él
            ]
        }
    
    petitions = await db.petitions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for petition in petitions:
        if isinstance(petition['created_at'], str):
            petition['created_at'] = datetime.fromisoformat(petition['created_at'])
        if isinstance(petition['updated_at'], str):
            petition['updated_at'] = datetime.fromisoformat(petition['updated_at'])
    
    return petitions


@api_router.get("/petitions/mis-peticiones-completas")
async def get_my_complete_petitions(current_user: dict = Depends(get_current_user)):
    """Obtener vista completa de peticiones y predios del usuario actual.
    
    Incluye:
    - Peticiones creadas por el usuario (radicados propios)
    - Peticiones asignadas al gestor
    - Predios nuevos creados por el gestor
    """
    
    # Peticiones creadas por el usuario
    peticiones_propias = await db.petitions.find(
        {"user_id": current_user['id']}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    for p in peticiones_propias:
        if isinstance(p.get('created_at'), str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
        if isinstance(p.get('updated_at'), str):
            p['updated_at'] = datetime.fromisoformat(p['updated_at'])
    
    # Peticiones asignadas al gestor (solo para gestores y superiores)
    peticiones_asignadas = []
    if current_user['role'] not in [UserRole.USUARIO, UserRole.EMPRESA]:
        peticiones_asignadas = await db.petitions.find(
            {
                "gestores_asignados": current_user['id'],
                "user_id": {"$ne": current_user['id']}  # Excluir las propias para evitar duplicados
            }, {"_id": 0}
        ).sort("created_at", -1).to_list(500)
        
        for p in peticiones_asignadas:
            if isinstance(p.get('created_at'), str):
                p['created_at'] = datetime.fromisoformat(p['created_at'])
            if isinstance(p.get('updated_at'), str):
                p['updated_at'] = datetime.fromisoformat(p['updated_at'])
    
    # Predios nuevos creados por el gestor
    predios_creados = []
    if current_user['role'] not in [UserRole.USUARIO, UserRole.EMPRESA]:
        predios_creados = await db.predios_nuevos.find(
            {"gestor_creador_id": current_user['id']}, {"_id": 0}
        ).sort("created_at", -1).to_list(500)
    
    return {
        "peticiones_propias": peticiones_propias,
        "peticiones_asignadas": peticiones_asignadas,
        "predios_creados": predios_creados
    }


@api_router.get("/petitions/{petition_id}")
async def get_petition(petition_id: str, current_user: dict = Depends(get_current_user)):
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Citizens can only see their own petitions
    if current_user['role'] == UserRole.USUARIO and petition['user_id'] != current_user['id']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para ver esta petición")
    
    # Gestores can see assigned petitions OR petitions they created
    if current_user['role'] in [UserRole.GESTOR]:
        is_assigned = current_user['id'] in petition.get('gestores_asignados', [])
        # Check both user_id (creator) and created_by for backwards compatibility
        is_creator = petition.get('user_id') == current_user['id'] or petition.get('created_by') == current_user['id']
        if not is_assigned and not is_creator:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para ver esta petición")
    
    if isinstance(petition['created_at'], str):
        petition['created_at'] = datetime.fromisoformat(petition['created_at'])
    if isinstance(petition['updated_at'], str):
        petition['updated_at'] = datetime.fromisoformat(petition['updated_at'])
    
    return petition

@api_router.post("/petitions/{petition_id}/upload")
async def upload_petition_files(
    petition_id: str,
    files: List[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user)
):
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Citizens and staff can upload files
    # Citizens: only to their own petitions
    # Staff: to any petition they have access to
    if current_user['role'] == UserRole.USUARIO and petition['user_id'] != current_user['id']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    saved_files = []
    for file in files:
        if file.filename:
            file_id = str(uuid.uuid4())
            file_ext = Path(file.filename).suffix
            file_path = UPLOAD_DIR / f"{file_id}{file_ext}"
            
            with file_path.open("wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            saved_files.append({
                "id": file_id,
                "original_name": file.filename,
                "path": str(file_path),
                "uploaded_by": current_user['id'],
                "uploaded_by_name": current_user['full_name'],
                "uploaded_by_role": current_user['role'],
                "upload_date": datetime.now(timezone.utc).isoformat()
            })
    
    current_files = petition.get('archivos', [])
    updated_files = current_files + saved_files
    
    # Add to historial
    uploader_role = "Usuario" if current_user['role'] == UserRole.USUARIO else current_user['role'].replace('_', ' ').title()
    historial_entry = {
        "accion": f"Archivos cargados por {uploader_role} ({len(saved_files)} archivo(s))",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": petition['estado'],
        "estado_nuevo": petition['estado'],
        "notas": f"Se cargaron {len(saved_files)} archivo(s) adicional(es)",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    
    await db.petitions.update_one(
        {"id": petition_id},
        {"$set": {
            "archivos": updated_files,
            "historial": current_historial,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notificar SOLO si NO está en estado devuelto (evitar spam de notificaciones)
    # Si está devuelto, el usuario debe usar el botón "Reenviar" para notificar al staff
    if current_user['role'] == UserRole.USUARIO and petition.get('estado') != PetitionStatus.DEVUELTO:
        # Crear notificación en plataforma para gestores asignados o atención al usuario
        if petition.get('gestores_asignados'):
            for gestor_id in petition['gestores_asignados']:
                await crear_notificacion(
                    usuario_id=gestor_id,
                    tipo="info",
                    titulo="Nuevos archivos cargados",
                    mensaje=f"El usuario ha cargado nuevos archivos en el trámite {petition['radicado']}",
                    enlace=f"/dashboard/peticion/{petition_id}"
                )
        else:
            atencion_users = await db.users.find({"role": UserRole.ATENCION_USUARIO}, {"_id": 0}).to_list(100)
            for user in atencion_users:
                await crear_notificacion(
                    usuario_id=user['id'],
                    tipo="info",
                    titulo="Nuevos archivos cargados",
                    mensaje=f"El usuario ha cargado nuevos archivos en el trámite {petition['radicado']}",
                    enlace=f"/dashboard/peticion/{petition_id}"
                )
    # Si el staff sube archivos, NO notificar (se finaliza automáticamente y envía correo ahí)
    
    return {"message": "Archivos subidos exitosamente", "files": saved_files}


@api_router.get("/petitions/{petition_id}/download-zip")
async def download_citizen_files_as_zip(petition_id: str, current_user: dict = Depends(get_current_user)):
    """Download all files uploaded by citizen as a ZIP file"""
    # Only staff can download citizen files
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Filter files uploaded by citizen (check string value, not enum)
    citizen_files = []
    for archivo in petition.get('archivos', []):
        uploaded_by_role = archivo.get('uploaded_by_role', 'usuario')
        if uploaded_by_role == 'usuario' or not archivo.get('uploaded_by_role'):
            citizen_files.append(archivo)
    
    if not citizen_files:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No hay archivos del usuario para descargar")
    
    import zipfile
    
    # Create ZIP file
    zip_filename = f"{petition['radicado']}_archivos_ciudadano.zip"
    zip_path = UPLOAD_DIR / zip_filename
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for archivo in citizen_files:
            file_path = Path(archivo['path'])
            if file_path.exists():
                zipf.write(file_path, archivo['original_name'])
    
    return FileResponse(
        path=zip_path,
        filename=zip_filename,
        media_type='application/zip'
    )


@api_router.get("/petitions/{petition_id}/archivo/{file_id}")
async def download_petition_file(
    petition_id: str,
    file_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Descarga un archivo individual de una petición.
    - Staff puede descargar cualquier archivo
    - Usuarios solo pueden descargar archivos de sus propias peticiones
    """
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Verificar permisos
    if current_user['role'] == UserRole.USUARIO:
        if petition.get('user_id') != current_user['id']:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para descargar este archivo")
    
    # Buscar el archivo en la lista de archivos
    archivo_encontrado = None
    for archivo in petition.get('archivos', []):
        # Buscar por id o por filename (para compatibilidad)
        if archivo.get('id') == file_id or archivo.get('filename') == file_id:
            archivo_encontrado = archivo
            break
    
    if not archivo_encontrado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Archivo no encontrado")
    
    file_path = Path(archivo_encontrado['path'])
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El archivo no existe en el servidor")
    
    # Determinar el tipo MIME
    content_type = "application/octet-stream"
    filename_lower = archivo_encontrado['original_name'].lower()
    if filename_lower.endswith('.pdf'):
        content_type = "application/pdf"
    elif filename_lower.endswith(('.jpg', '.jpeg')):
        content_type = "image/jpeg"
    elif filename_lower.endswith('.png'):
        content_type = "image/png"
    elif filename_lower.endswith('.doc'):
        content_type = "application/msword"
    elif filename_lower.endswith('.docx'):
        content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif filename_lower.endswith('.xls'):
        content_type = "application/vnd.ms-excel"
    elif filename_lower.endswith('.xlsx'):
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return FileResponse(
        path=file_path,
        filename=archivo_encontrado['original_name'],
        media_type=content_type
    )


@api_router.post("/petitions/{petition_id}/assign-gestor")
async def assign_gestor(
    petition_id: str,
    assignment: GestorAssignment,
    current_user: dict = Depends(get_current_user)
):
    # Only atencion usuario, gestor, coordinador, and admin can assign
    if current_user['role'] not in [UserRole.ATENCION_USUARIO, UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    gestor = await db.users.find_one({"id": assignment.gestor_id}, {"_id": 0})
    if not gestor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Gestor no encontrado")
    
    # Add gestor to assigned list
    gestores_asignados = petition.get('gestores_asignados', [])
    if assignment.gestor_id not in gestores_asignados:
        gestores_asignados.append(assignment.gestor_id)
    
    update_data = {
        "gestores_asignados": gestores_asignados,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # If first assignment, change status to ASIGNADO
    estado_cambio = False
    if petition['estado'] == PetitionStatus.RADICADO:
        update_data['estado'] = PetitionStatus.ASIGNADO
        estado_cambio = True
    
    # Add to historial
    rol_labels = {
        UserRole.GESTOR: "Gestor",
        UserRole.ATENCION_USUARIO: "Atención al Usuario",
        UserRole.COORDINADOR: "Coordinador",
        UserRole.ADMINISTRADOR: "Administrador"
    }
    gestor_rol = rol_labels.get(gestor['role'], "Staff")
    
    # Construir notas con comentario si existe
    notas_historial = f"Asignado a {gestor['full_name']} ({gestor_rol})"
    if assignment.comentario and assignment.comentario.strip():
        notas_historial += f" - Comentario: {assignment.comentario.strip()}"
    
    historial_entry = {
        "accion": f"{gestor_rol} asignado: {gestor['full_name']}",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": petition['estado'],
        "estado_nuevo": update_data.get('estado', petition['estado']),
        "notas": notas_historial,
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    update_data['historial'] = current_historial
    
    await db.petitions.update_one({"id": petition_id}, {"$set": update_data})
    
    # Notificación en plataforma al gestor asignado (NO correo)
    mensaje_notificacion = f"Se te ha asignado el trámite {petition['radicado']} - {petition['tipo_tramite']}"
    if assignment.comentario and assignment.comentario.strip():
        mensaje_notificacion += f". Instrucciones: {assignment.comentario.strip()}"
    
    await crear_notificacion(
        usuario_id=gestor['id'],
        tipo="info",
        titulo="Nuevo trámite asignado",
        mensaje=mensaje_notificacion,
        enlace=f"/dashboard/peticion/{petition_id}"
    )
    
    return {"message": "Gestor asignado exitosamente"}


@api_router.delete("/petitions/{petition_id}/desasignar/{user_id}")
async def desasignar_staff(petition_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    """Quitar staff de un radicado - Solo admin, coordinador y atención al usuario"""
    # Verificar permisos
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=403, detail="Solo administradores, coordinadores y atención al usuario pueden quitar staff")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    gestores_asignados = petition.get('gestores_asignados', [])
    
    if user_id not in gestores_asignados:
        raise HTTPException(status_code=400, detail="El usuario no está asignado a esta petición")
    
    # Obtener nombre del usuario que se va a quitar
    user_to_remove = await db.users.find_one({"id": user_id}, {"_id": 0, "full_name": 1, "role": 1})
    user_name = user_to_remove.get('full_name', 'Usuario') if user_to_remove else 'Usuario'
    
    # Quitar del listado
    gestores_asignados.remove(user_id)
    
    # Historial
    historial_entry = {
        "accion": f"Staff removido: {user_name}",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "notas": f"Se quitó a {user_name} del trámite",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    
    await db.petitions.update_one(
        {"id": petition_id},
        {"$set": {
            "gestores_asignados": gestores_asignados,
            "historial": current_historial,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"{user_name} ha sido removido del trámite"}


@api_router.post("/petitions/{petition_id}/auto-asignar")
async def auto_asignar_tramite(petition_id: str, current_user: dict = Depends(get_current_user)):
    """Atención al usuario, gestor, coordinador o admin se auto-asigna un trámite"""
    # Roles que pueden auto-asignarse
    if current_user['role'] not in [UserRole.ATENCION_USUARIO, UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para auto-asignarse")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    gestores_asignados = petition.get('gestores_asignados', [])
    
    if current_user['id'] in gestores_asignados:
        raise HTTPException(status_code=400, detail="Ya está asignado a esta petición")
    
    gestores_asignados.append(current_user['id'])
    
    update_data = {
        "gestores_asignados": gestores_asignados,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Si es el primer asignado, cambiar estado
    if petition['estado'] == PetitionStatus.RADICADO:
        update_data['estado'] = PetitionStatus.ASIGNADO
    
    # Historial
    rol_labels = {
        UserRole.ATENCION_USUARIO: "Atención al Usuario",
        UserRole.COORDINADOR: "Coordinador",
        UserRole.ADMINISTRADOR: "Administrador"
    }
    
    historial_entry = {
        "accion": f"Auto-asignación: {current_user['full_name']} ({rol_labels.get(current_user['role'], 'Staff')})",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "notas": "Auto-asignación de trámite",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    update_data['historial'] = current_historial
    
    await db.petitions.update_one({"id": petition_id}, {"$set": update_data})
    
    return {"message": "Se ha asignado exitosamente al trámite"}


@api_router.post("/petitions/{petition_id}/asignar/{gestor_id}")
async def asignar_gestor_a_tramite(petition_id: str, gestor_id: str, current_user: dict = Depends(get_current_user)):
    """Atención al usuario, coordinador o admin asigna un gestor a un trámite"""
    # Solo estos roles pueden asignar gestores
    if current_user['role'] not in [UserRole.ATENCION_USUARIO, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para asignar gestores")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    # Verificar que el gestor existe
    gestor = await db.users.find_one({"id": gestor_id}, {"_id": 0})
    if not gestor:
        raise HTTPException(status_code=404, detail="Gestor no encontrado")
    
    # Verificar que el gestor tenga rol adecuado
    if gestor['role'] not in [UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=400, detail="El usuario seleccionado no puede ser asignado como gestor")
    
    gestores_asignados = petition.get('gestores_asignados', [])
    
    if gestor_id in gestores_asignados:
        raise HTTPException(status_code=400, detail="El gestor ya está asignado a esta petición")
    
    gestores_asignados.append(gestor_id)
    
    update_data = {
        "gestores_asignados": gestores_asignados,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Si es el primer asignado, cambiar estado
    if petition['estado'] == PetitionStatus.RADICADO:
        update_data['estado'] = PetitionStatus.ASIGNADO
    
    # Historial
    historial_entry = {
        "accion": f"Asignación de gestor: {gestor['full_name']} asignado por {current_user['full_name']}",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "notas": f"Gestor {gestor['full_name']} asignado al trámite",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    update_data['historial'] = current_historial
    
    await db.petitions.update_one({"id": petition_id}, {"$set": update_data})
    
    # Enviar notificación al gestor asignado
    try:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": gestor_id,
            "tipo": "asignacion",
            "mensaje": f"Se te ha asignado el trámite {petition['radicado']} - {petition['tipo_tramite']}",
            "link": f"/dashboard/peticiones/{petition_id}",
            "leido": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    except:
        pass
    
    return {"message": f"Gestor {gestor['full_name']} asignado exitosamente al trámite"}


@api_router.post("/petitions/{petition_id}/marcar-completado")
async def marcar_trabajo_completado(petition_id: str, current_user: dict = Depends(get_current_user)):
    """El gestor marca su trabajo como completado en el trámite"""
    if current_user['role'] not in [UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo gestores pueden marcar su trabajo como completado")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    # Verificar que el usuario esté asignado
    if current_user['id'] not in petition.get('gestores_asignados', []):
        raise HTTPException(status_code=403, detail="No está asignado a este trámite")
    
    gestores_finalizados = petition.get('gestores_finalizados', [])
    
    if current_user['id'] in gestores_finalizados:
        raise HTTPException(status_code=400, detail="Ya ha marcado su trabajo como completado")
    
    gestores_finalizados.append(current_user['id'])
    
    # Agregar al historial
    historial_entry = {
        "accion": f"Trabajo completado por {current_user['full_name']}",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "notas": "El gestor marcó su trabajo como finalizado",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    
    update_data = {
        "gestores_finalizados": gestores_finalizados,
        "historial": current_historial,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Si todos los gestores han completado su trabajo, notificar
    gestores_asignados = petition.get('gestores_asignados', [])
    todos_completaron = all(g in gestores_finalizados for g in gestores_asignados)
    
    await db.petitions.update_one({"id": petition_id}, {"$set": update_data})
    
    return {
        "message": "Trabajo marcado como completado",
        "todos_completaron": todos_completaron,
        "gestores_finalizados": len(gestores_finalizados),
        "gestores_totales": len(gestores_asignados)
    }


@api_router.post("/petitions/{petition_id}/desmarcar-completado")
async def desmarcar_trabajo_completado(petition_id: str, current_user: dict = Depends(get_current_user)):
    """El gestor desmarca su trabajo como completado (por si necesita hacer más cambios)"""
    if current_user['role'] not in [UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo gestores pueden modificar el estado de su trabajo")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    gestores_finalizados = petition.get('gestores_finalizados', [])
    
    if current_user['id'] not in gestores_finalizados:
        raise HTTPException(status_code=400, detail="No ha marcado su trabajo como completado")
    
    gestores_finalizados.remove(current_user['id'])
    
    # Agregar al historial
    historial_entry = {
        "accion": f"{current_user['full_name']} retomó el trabajo",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "notas": "El gestor desmarcó su trabajo como finalizado para continuar",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    
    await db.petitions.update_one({"id": petition_id}, {"$set": {
        "gestores_finalizados": gestores_finalizados,
        "historial": current_historial,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }})
    
    return {"message": "Trabajo desmarcado como completado"}


@api_router.patch("/petitions/{petition_id}")
async def update_petition(petition_id: str, update_data: PetitionUpdate, current_user: dict = Depends(get_current_user)):
    # Citizens cannot update petitions
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para actualizar peticiones")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Determine what fields can be updated based on role
    update_dict = {}
    historial_entry = None
    
    if current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        # Coordinador and Admin can update all fields including approval
        update_dict = update_data.model_dump(exclude_none=True)
    elif current_user['role'] == UserRole.ATENCION_USUARIO:
        # Atención al usuario can update status, notes, and can finalize/reject
        if update_data.estado:
            update_dict['estado'] = update_data.estado
        if update_data.notas:
            update_dict['notas'] = update_data.notas
    elif current_user['role'] in [UserRole.GESTOR]:
        # Gestores can update notes and change to: EN_PROCESO, REVISION, RECHAZADO
        if update_data.notas:
            update_dict['notas'] = update_data.notas
        if update_data.estado in [PetitionStatus.EN_PROCESO, PetitionStatus.REVISION, PetitionStatus.RECHAZADO]:
            update_dict['estado'] = update_data.estado
    
    if update_dict:
        update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
        
        # Create historial entry if status changed
        if 'estado' in update_dict:
            estado_anterior = petition.get('estado')
            estado_nuevo = update_dict['estado']
            
            status_names = {
                PetitionStatus.RADICADO: "Radicado",
                PetitionStatus.ASIGNADO: "Asignado",
                PetitionStatus.EN_PROCESO: "En Proceso",
                PetitionStatus.REVISION: "En Revisión",
                PetitionStatus.APROBADO: "Aprobado",
                PetitionStatus.RECHAZADO: "Rechazado",
                PetitionStatus.DEVUELTO: "Devuelto",
                PetitionStatus.FINALIZADO: "Finalizado"
            }
            
            # Si se está devolviendo, guardar info del que devuelve y observaciones
            if estado_nuevo == PetitionStatus.DEVUELTO:
                update_dict['devuelto_por_id'] = current_user['id']
                update_dict['devuelto_por_nombre'] = current_user['full_name']
                if update_data.observaciones_devolucion:
                    update_dict['observaciones_devolucion'] = update_data.observaciones_devolucion
            
            # Si se está aprobando, guardar info del aprobador
            if estado_nuevo == PetitionStatus.APROBADO:
                update_dict['aprobado_por_id'] = current_user['id']
                update_dict['aprobado_por_nombre'] = current_user['full_name']
                update_dict['fecha_aprobacion'] = datetime.now(timezone.utc).isoformat()
                if hasattr(update_data, 'comentario_aprobacion') and update_data.comentario_aprobacion:
                    update_dict['comentario_aprobacion'] = update_data.comentario_aprobacion
            
            historial_entry = {
                "accion": f"Estado cambiado de {status_names.get(estado_anterior, estado_anterior)} a {status_names.get(estado_nuevo, estado_nuevo)}",
                "usuario": current_user['full_name'],
                "usuario_rol": current_user['role'],
                "estado_anterior": estado_anterior,
                "estado_nuevo": estado_nuevo,
                "notas": update_dict.get('notas', ''),
                "observaciones_devolucion": update_data.observaciones_devolucion if estado_nuevo == PetitionStatus.DEVUELTO else None,
                "fecha": datetime.now(timezone.utc).isoformat()
            }
            
            # Add to historial
            current_historial = petition.get('historial', [])
            current_historial.append(historial_entry)
            update_dict['historial'] = current_historial
            
            # AUTO-ASIGNACIÓN: Cuando pasa a "revisión", asignar automáticamente 
            # al coordinador y gestores con permiso de aprobar cambios
            if estado_nuevo == PetitionStatus.REVISION:
                # Buscar coordinadores y gestores con permiso de aprobar
                aprobadores = await db.users.find({
                    "$or": [
                        {"role": UserRole.COORDINADOR},
                        {"role": UserRole.ADMINISTRADOR},
                        {"permissions": Permission.APPROVE_CHANGES}
                    ]
                }, {"_id": 0, "id": 1, "full_name": 1, "role": 1}).to_list(100)
                
                gestores_asignados = petition.get('gestores_asignados', [])
                nuevos_asignados = []
                
                for aprobador in aprobadores:
                    if aprobador['id'] not in gestores_asignados:
                        gestores_asignados.append(aprobador['id'])
                        nuevos_asignados.append(aprobador)
                
                if nuevos_asignados:
                    update_dict['gestores_asignados'] = gestores_asignados
                    
                    # Agregar entrada al historial
                    nombres_asignados = [a['full_name'] for a in nuevos_asignados]
                    auto_historial = {
                        "accion": f"Auto-asignado a revisores: {', '.join(nombres_asignados)}",
                        "usuario": "Sistema",
                        "usuario_rol": "sistema",
                        "notas": "Asignación automática al pasar a revisión",
                        "fecha": datetime.now(timezone.utc).isoformat()
                    }
                    update_dict['historial'].append(auto_historial)
                    
                    # Notificar a los aprobadores
                    for aprobador in nuevos_asignados:
                        await crear_notificacion(
                            usuario_id=aprobador['id'],
                            tipo="info",
                            titulo="Trámite en Revisión",
                            mensaje=f"El trámite {petition['radicado']} está listo para su revisión y aprobación.",
                            enlace=f"/dashboard/peticion/{petition_id}",
                            enviar_email=False
                        )
        
        await db.petitions.update_one({"id": petition_id}, {"$set": update_dict})
        
        # Send email notification to citizen if status changed
        if 'estado' in update_dict:
            # Solo enviar email si hay un ciudadano asociado y correo electrónico
            citizen_email = petition.get('correo')
            nombre_solicitante = petition.get('nombre_completo', 'Usuario')
            
            # Intentar obtener más info del usuario si existe
            if petition.get('user_id'):
                citizen = await db.users.find_one({"id": petition['user_id']}, {"_id": 0})
                if citizen:
                    citizen_email = citizen_email or citizen.get('email')
                    nombre_solicitante = nombre_solicitante or citizen.get('full_name', 'Usuario')
            
            if citizen_email:
                estado_nuevo = update_dict['estado']
                
                # Usar plantilla especial para finalización
                if estado_nuevo == PetitionStatus.FINALIZADO:
                    # Verificar si hay archivos del staff para adjuntar
                    archivos_staff = petition.get('archivos_staff', [])
                    enviar_archivos = update_data.enviar_archivos_finalizacion and len(archivos_staff) > 0
                    
                    email_body = get_finalizacion_email(
                        radicado=petition['radicado'],
                        tipo_tramite=petition['tipo_tramite'],
                        nombre_solicitante=nombre_solicitante,
                        con_archivos=enviar_archivos
                    )
                    
                    # Si se deben enviar archivos, prepararlos
                    attachment_path = None
                    attachment_name = None
                    if enviar_archivos:
                        uploads_dir = Path("uploads") / petition_id
                        for archivo in archivos_staff:
                            archivo_path = uploads_dir / archivo.get('filename', '')
                            if archivo_path.exists():
                                # Solo adjuntar el primer archivo (limitación de send_email)
                                attachment_path = str(archivo_path)
                                attachment_name = archivo.get('filename', 'archivo_adjunto')
                                break
                    
                    await send_email(
                        citizen_email,
                        f"¡Trámite Finalizado! - {petition['radicado']}",
                        email_body,
                        attachment_path,
                        attachment_name
                    )
                    
                    # Notificar a TODOS los gestores asignados que el trámite fue finalizado
                    gestores_asignados = petition.get('gestores_asignados', [])
                    for gestor_id in gestores_asignados:
                        if gestor_id != current_user['id']:  # No notificar al que finalizó
                            await crear_notificacion(
                                usuario_id=gestor_id,
                                tipo="info",
                                titulo="Trámite Finalizado",
                                mensaje=f"El trámite {petition['radicado']} ha sido finalizado por {current_user['full_name']}.",
                                enlace=f"/dashboard/peticiones/{petition_id}",
                                enviar_email=False
                            )
                else:
                    email_body = get_actualizacion_email(
                        radicado=petition['radicado'],
                        estado_nuevo=estado_nuevo,
                        nombre_solicitante=nombre_solicitante,
                        observaciones=update_data.observaciones_devolucion if estado_nuevo == PetitionStatus.DEVUELTO else None
                    )
                    
                    await send_email(
                        citizen_email,
                        f"Actualización de Trámite - {petition['radicado']}",
                        email_body
                    )
    
    updated_petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    
    if isinstance(updated_petition['created_at'], str):
        updated_petition['created_at'] = datetime.fromisoformat(updated_petition['created_at'])
    if isinstance(updated_petition['updated_at'], str):
        updated_petition['updated_at'] = datetime.fromisoformat(updated_petition['updated_at'])
    
    return updated_petition

@api_router.get("/petitions/stats/dashboard")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    user_id = current_user['id']
    user_role = current_user['role']
    user_municipios = current_user.get('municipios', [])
    
    # Verificar si tiene permiso de aprobar cambios
    user_permisos = await db.user_permissions.find_one({"user_id": user_id})
    puede_aprobar = user_permisos and user_permisos.get('permissions', {}).get('aprobar_cambios')
    es_aprobador = user_role in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR] or puede_aprobar
    
    # ===== ESTADÍSTICAS DE PETICIONES/RADICADOS =====
    if user_role == UserRole.USUARIO:
        query_peticiones = {"user_id": user_id}
    elif user_role == UserRole.GESTOR or user_role == 'gestor_auxiliar':
        query_peticiones = {
            "$or": [
                {"gestores_asignados": user_id},
                {"user_id": user_id}
            ]
        }
    elif user_role == UserRole.EMPRESA:
        return {
            "total": 0, "radicado": 0, "asignado": 0, "rechazado": 0,
            "revision": 0, "devuelto": 0, "finalizado": 0,
            "predios_creados": 0, "predios_asignados": 0, "modificaciones_asignadas": 0,
            "predios_revision": 0, "modificaciones_pendientes": 0, "reapariciones_pendientes": 0,
            "mis_radicados": 0, "aprobados_mes": 0, "rechazados_mes": 0,
            "mensaje": "Sin acceso a estadísticas de peticiones"
        }
    else:
        query_peticiones = {}  # Coordinador, Admin, Atención ven todo
    
    # Contar peticiones
    total = await db.petitions.count_documents(query_peticiones)
    radicado = await db.petitions.count_documents({**query_peticiones, "estado": PetitionStatus.RADICADO})
    asignado = await db.petitions.count_documents({**query_peticiones, "estado": PetitionStatus.ASIGNADO})
    rechazado = await db.petitions.count_documents({**query_peticiones, "estado": PetitionStatus.RECHAZADO})
    revision = await db.petitions.count_documents({**query_peticiones, "estado": PetitionStatus.REVISION})
    devuelto = await db.petitions.count_documents({**query_peticiones, "estado": PetitionStatus.DEVUELTO})
    finalizado = await db.petitions.count_documents({**query_peticiones, "estado": PetitionStatus.FINALIZADO})
    
    # Mis radicados (los que yo creé)
    mis_radicados = await db.petitions.count_documents({"user_id": user_id})
    
    # ===== ESTADÍSTICAS DE PREDIOS NUEVOS =====
    # Predios que creé
    predios_creados = await db.predios_nuevos.count_documents({
        "gestor_creador_id": user_id,
        "estado_flujo": {"$in": ["creado", "digitalizacion", "devuelto", "revision"]}
    })
    
    # Predios asignados a mí (soy apoyo)
    predios_asignados = await db.predios_nuevos.count_documents({
        "gestor_apoyo_id": user_id,
        "estado_flujo": {"$in": ["creado", "digitalizacion", "devuelto"]}
    })
    
    # Modificaciones asignadas a mí
    modificaciones_asignadas = await db.cambios_pendientes.count_documents({
        "gestor_apoyo_id": user_id,
        "estado": "pendiente"
    })
    
    # ===== ESTADÍSTICAS PARA APROBADORES =====
    predios_revision = 0
    modificaciones_pendientes = 0
    reapariciones_pendientes = 0
    aprobados_mes = 0
    rechazados_mes = 0
    
    if es_aprobador:
        # Query base para municipios (si el usuario tiene municipios asignados)
        query_municipio = {}
        if user_municipios:
            query_municipio = {"municipio": {"$in": user_municipios}}
        
        # Predios en revisión (pendientes de aprobar)
        predios_revision = await db.predios_nuevos.count_documents({
            **query_municipio,
            "estado_flujo": "revision"
        })
        
        # Modificaciones pendientes
        modificaciones_pendientes = await db.cambios_pendientes.count_documents({
            **query_municipio,
            "estado": "pendiente"
        })
        
        # Reapariciones pendientes
        reapariciones_pendientes = await db.reapariciones.count_documents({
            **query_municipio,
            "estado": {"$in": ["pendiente", "revision"]}
        })
        
        # Estadísticas del mes actual
        from datetime import datetime, timezone
        inicio_mes = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        aprobados_mes = await db.predios_nuevos.count_documents({
            "estado_flujo": "aprobado",
            "fecha_aprobacion": {"$gte": inicio_mes.isoformat()}
        })
        
        rechazados_mes = await db.cambios_pendientes.count_documents({
            "estado": "rechazado",
            "fecha_resolucion": {"$gte": inicio_mes.isoformat()}
        })
    
    return {
        # Peticiones/Radicados
        "total": total,
        "radicado": radicado,
        "asignado": asignado,
        "rechazado": rechazado,
        "revision": revision,
        "devuelto": devuelto,
        "finalizado": finalizado,
        "mis_radicados": mis_radicados,
        # Predios/Asignaciones (para gestores)
        "predios_creados": predios_creados,
        "predios_asignados": predios_asignados,
        "modificaciones_asignadas": modificaciones_asignadas,
        # Pendientes de aprobación (para coordinadores)
        "predios_revision": predios_revision,
        "modificaciones_pendientes": modificaciones_pendientes,
        "reapariciones_pendientes": reapariciones_pendientes,
        # Estadísticas del mes
        "aprobados_mes": aprobados_mes,
        "rechazados_mes": rechazados_mes,
        # Info del rol
        "es_aprobador": es_aprobador
    }

@api_router.post("/petitions/{petition_id}/reenviar")
async def reenviar_petition(petition_id: str, current_user: dict = Depends(get_current_user)):
    """
    Permite al usuario reenviar una petición devuelta para revisión.
    Notifica al miembro del staff que la devolvió.
    """
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Solo el propietario puede reenviar (usar .get() para evitar KeyError)
    petition_user_id = petition.get('user_id')
    if not petition_user_id or petition_user_id != current_user['id']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para reenviar esta petición")
    
    # Solo se puede reenviar si está devuelta
    if petition.get('estado') != PetitionStatus.DEVUELTO:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo puede reenviar peticiones devueltas")
    
    # Cambiar estado a revisión
    historial_entry = {
        "accion": "Petición reenviada para revisión por el usuario",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": PetitionStatus.DEVUELTO,
        "estado_nuevo": PetitionStatus.REVISION,
        "notas": "El usuario ha realizado las correcciones solicitadas y reenvía el trámite para revisión.",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    
    await db.petitions.update_one(
        {"id": petition_id},
        {"$set": {
            "estado": PetitionStatus.REVISION,
            "historial": current_historial,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notificar al staff que devolvió el trámite (SOLO en plataforma, sin correo para evitar spam)
    devuelto_por_id = petition.get('devuelto_por_id')
    if devuelto_por_id:
        await crear_notificacion(
            usuario_id=devuelto_por_id,
            titulo="📋 Subsanación Recibida",
            mensaje=f"El usuario {current_user['full_name']} ha enviado la subsanación del trámite {petition['radicado']}. Requiere su revisión.",
            tipo="warning",
            enlace=f"/dashboard/peticion/{petition_id}",
            enviar_email=False  # Solo notificación en plataforma
        )
    
    # También notificar a los gestores asignados (si los hay y son diferentes)
    for gestor_id in petition.get('gestores_asignados', []):
        if gestor_id != devuelto_por_id:  # Evitar duplicados
            await crear_notificacion(
                usuario_id=gestor_id,
                titulo="📋 Subsanación Recibida",
                mensaje=f"El trámite {petition['radicado']} ha sido reenviado para revisión después de subsanación.",
                tipo="warning",
                enlace=f"/dashboard/peticion/{petition_id}",
                enviar_email=False  # Solo notificación en plataforma
            )
    
    return {"message": "Petición reenviada exitosamente para revisión"}

@api_router.get("/gestores")
async def get_gestores(current_user: dict = Depends(get_current_user)):
    # Get all users that can be assigned to petitions (gestores, auxiliares, atencion_usuario, coordinadores, admin)
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR, UserRole.ATENCION_USUARIO, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]}},
        {"_id": 0, "password": 0}
    ).sort("full_name", 1).to_list(1000)  # Ordenar alfabéticamente por nombre
    
    return gestores


# ===== PDF EXPORT AND DIGITAL SIGNATURE =====

def generate_petition_pdf(petition_data: dict, user_data: dict, signed_by: str = None) -> bytes:
    """Generate PDF report for a petition with optional digital signature"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#009846'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#064E3B'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=12
    )
    
    # Title
    story.append(Paragraph("ASOCIACIÓN DE MUNICIPIOS DEL CATATUMBO", title_style))
    story.append(Paragraph("Provincia de Ocaña y Sur del Cesar - ASOMUNICIPIOS", normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Radicado
    story.append(Paragraph(f"<b>Radicado:</b> {petition_data.get('radicado', 'N/A')}", heading_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Status
    status_names = {
        'radicado': 'Radicado',
        'asignado': 'Asignado',
        'rechazado': 'Rechazado',
        'revision': 'En Revisión',
        'devuelto': 'Devuelto',
        'finalizado': 'Finalizado'
    }
    status_label = status_names.get(petition_data.get('estado', ''), petition_data.get('estado', 'N/A'))
    story.append(Paragraph(f"<b>Estado:</b> {status_label}", normal_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Petition details table
    story.append(Paragraph("DATOS DEL SOLICITANTE", heading_style))
    
    data = [
        ['Campo', 'Información'],
        ['Nombre Completo', petition_data.get('nombre_completo', 'N/A')],
        ['Correo Electrónico', petition_data.get('correo', 'N/A')],
        ['Teléfono', petition_data.get('telefono', 'N/A')],
        ['Municipio', petition_data.get('municipio', 'N/A')],
        ['Tipo de Trámite', petition_data.get('tipo_tramite', 'N/A')],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#009846')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Notes if any
    if petition_data.get('notas'):
        story.append(Paragraph("NOTAS", heading_style))
        story.append(Paragraph(petition_data.get('notas', ''), normal_style))
        story.append(Spacer(1, 0.3*inch))
    
    # Dates
    story.append(Paragraph("INFORMACIÓN DE FECHAS", heading_style))
    created_date = petition_data.get('created_at', '')
    updated_date = petition_data.get('updated_at', '')
    
    if isinstance(created_date, str):
        created_date = datetime.fromisoformat(created_date).strftime('%d/%m/%Y %H:%M')
    else:
        created_date = created_date.strftime('%d/%m/%Y %H:%M') if created_date else 'N/A'
    
    if isinstance(updated_date, str):
        updated_date = datetime.fromisoformat(updated_date).strftime('%d/%m/%Y %H:%M')
    else:
        updated_date = updated_date.strftime('%d/%m/%Y %H:%M') if updated_date else 'N/A'
    
    story.append(Paragraph(f"<b>Fecha de Radicación:</b> {created_date}", normal_style))
    story.append(Paragraph(f"<b>Última Actualización:</b> {updated_date}", normal_style))
    story.append(Spacer(1, 0.5*inch))
    
    # Digital signature section
    if signed_by:
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph("___________________________", normal_style))
        story.append(Paragraph("<b>Firmado digitalmente por:</b>", normal_style))
        story.append(Paragraph(f"{signed_by}", normal_style))
        story.append(Paragraph(f"Fecha: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}", normal_style))
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        "Este documento ha sido generado por el Sistema de Gestión Catastral de ASOMUNICIPIOS",
        footer_style
    ))
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


@api_router.get("/documentacion/flujo-radicacion-pdf")
async def descargar_flujo_radicacion_pdf():
    """Descarga el PDF del flujo de radicación de trámites para socialización"""
    pdf_path = "/app/backend/static/flujo_radicacion_tramites.pdf"
    
    if not os.path.exists(pdf_path):
        # Regenerar el PDF si no existe
        try:
            import subprocess
            subprocess.run(["python", "/app/backend/generar_flujo_pdf.py"], check=True)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
    
    return FileResponse(
        path=pdf_path,
        filename="Flujo_Radicacion_Tramites_Asomunicipios.pdf",
        media_type="application/pdf"
    )


@api_router.get("/petitions/{petition_id}/export-pdf")
async def export_petition_pdf(petition_id: str, current_user: dict = Depends(get_current_user)):
    """Export single petition as PDF"""
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Check permissions
    if current_user['role'] == UserRole.USUARIO and petition['user_id'] != current_user['id']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get user data
    user = await db.users.find_one({"id": petition['user_id']}, {"_id": 0, "password": 0})
    
    # Generate PDF with digital signature if coordinator or admin
    signed_by = None
    if current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        signed_by = f"{current_user['full_name']} - {current_user['role'].replace('_', ' ').title()}"
    
    pdf_bytes = generate_petition_pdf(petition, user, signed_by)
    
    # Save to temp file
    temp_pdf_path = UPLOAD_DIR / f"petition_{petition_id}_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"{petition['radicado']}.pdf",
        media_type='application/pdf'
    )


@api_router.post("/petitions/export-multiple")
async def export_multiple_petitions(
    petition_ids: List[str],
    current_user: dict = Depends(get_current_user)
):
    """Export multiple petitions as PDF"""
    # Only staff can export multiple
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    from reportlab.platypus import PageBreak
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    
    for idx, petition_id in enumerate(petition_ids):
        petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
        if petition:
            # Generate petition content
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#009846'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            
            story.append(Paragraph(f"Petición {idx + 1} de {len(petition_ids)}", title_style))
            story.append(Paragraph(f"Radicado: {petition.get('radicado', 'N/A')}", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            
            # Add basic info
            info = [
                ['Solicitante', petition.get('nombre_completo', 'N/A')],
                ['Tipo de Trámite', petition.get('tipo_tramite', 'N/A')],
                ['Estado', petition.get('estado', 'N/A')],
                ['Municipio', petition.get('municipio', 'N/A')],
            ]
            
            table = Table(info, colWidths=[2*inch, 4*inch])
            table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            story.append(table)
            
            # Add page break between petitions except for the last one
            if idx < len(petition_ids) - 1:
                story.append(PageBreak())
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    # Save to temp file
    temp_pdf_path = UPLOAD_DIR / f"petitions_report_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"reporte_peticiones_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type='application/pdf'
    )


# ===== PRODUCTIVITY REPORTS =====

@api_router.get("/reports/gestor-productivity")
async def get_gestor_productivity(current_user: dict = Depends(get_current_user)):
    """Get productivity report for all gestores"""
    # Only admin, coordinador, and atencion_usuario can view reports
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get all gestores and auxiliares
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR]}},
        {"_id": 0, "password": 0}
    ).to_list(1000)
    
    productivity_data = []
    
    for gestor in gestores:
        gestor_id = gestor['id']
        
        # Total petitions assigned
        total_assigned = await db.petitions.count_documents({
            "gestores_asignados": gestor_id
        })
        
        # Completed petitions
        completed = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.FINALIZADO
        })
        
        # In process
        in_process = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": {"$in": [PetitionStatus.ASIGNADO, PetitionStatus.REVISION, PetitionStatus.DEVUELTO]}
        })
        
        # Rejected
        rejected = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.RECHAZADO
        })
        
        # Calculate average time to complete (for completed petitions)
        completed_petitions = await db.petitions.find({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.FINALIZADO
        }, {"_id": 0, "created_at": 1, "updated_at": 1}).to_list(1000)
        
        avg_days = 0
        if completed_petitions:
            total_days = 0
            for petition in completed_petitions:
                created = petition['created_at']
                updated = petition['updated_at']
                
                if isinstance(created, str):
                    created = datetime.fromisoformat(created)
                if isinstance(updated, str):
                    updated = datetime.fromisoformat(updated)
                
                delta = (updated - created).days
                total_days += delta
            
            avg_days = round(total_days / len(completed_petitions), 1) if completed_petitions else 0
        
        # Calculate completion rate
        completion_rate = round((completed / total_assigned * 100), 1) if total_assigned > 0 else 0
        
        productivity_data.append({
            "gestor_id": gestor_id,
            "gestor_name": gestor['full_name'],
            "gestor_email": gestor['email'],
            "gestor_role": gestor['role'],
            "total_assigned": total_assigned,
            "completed": completed,
            "in_process": in_process,
            "rejected": rejected,
            "avg_completion_days": avg_days,
            "completion_rate": completion_rate
        })
    
    # Sort by completion rate descending
    productivity_data.sort(key=lambda x: x['completion_rate'], reverse=True)
    
    return productivity_data


@api_router.get("/reports/gestor-productivity/export-pdf")
async def export_gestor_productivity_pdf(current_user: dict = Depends(get_current_user)):
    """Export gestor productivity report as PDF"""
    # Only admin, coordinador, and atencion_usuario can export reports
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get productivity data
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR]}},
        {"_id": 0, "password": 0}
    ).to_list(1000)
    
    productivity_data = []
    
    for gestor in gestores:
        gestor_id = gestor['id']
        
        total_assigned = await db.petitions.count_documents({"gestores_asignados": gestor_id})
        completed = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.FINALIZADO
        })
        in_process = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": {"$in": [PetitionStatus.ASIGNADO, PetitionStatus.REVISION, PetitionStatus.DEVUELTO]}
        })
        rejected = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.RECHAZADO
        })
        
        completion_rate = round((completed / total_assigned * 100), 1) if total_assigned > 0 else 0
        
        productivity_data.append({
            "name": gestor['full_name'],
            "role": "Gestor" if gestor['role'] == UserRole.GESTOR else "Gestor Auxiliar",
            "total": total_assigned,
            "completed": completed,
            "in_process": in_process,
            "rejected": rejected,
            "rate": completion_rate
        })
    
    productivity_data.sort(key=lambda x: x['rate'], reverse=True)
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#009846'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph("ASOCIACIÓN DE MUNICIPIOS DEL CATATUMBO", title_style))
    story.append(Paragraph("Reporte de Productividad de Gestores", styles['Heading2']))
    story.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 0.5*inch))
    
    # Table data
    table_data = [
        ['Gestor', 'Rol', 'Total', 'Finalizados', 'En Proceso', 'Rechazados', 'Tasa (%)']
    ]
    
    for data in productivity_data:
        table_data.append([
            data['name'],
            data['role'],
            str(data['total']),
            str(data['completed']),
            str(data['in_process']),
            str(data['rejected']),
            f"{data['rate']}%"
        ])
    
    # Create table
    col_widths = [1.8*inch, 1*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#009846')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.5*inch))
    
    # Summary
    story.append(Paragraph("Resumen General", styles['Heading3']))
    total_gestores = len(productivity_data)
    total_tramites = sum(d['total'] for d in productivity_data)
    total_finalizados = sum(d['completed'] for d in productivity_data)
    avg_rate = round(sum(d['rate'] for d in productivity_data) / total_gestores, 1) if total_gestores > 0 else 0
    
    summary_text = f"""
    <b>Total de Gestores:</b> {total_gestores}<br/>
    <b>Total de Trámites Asignados:</b> {total_tramites}<br/>
    <b>Total Finalizados:</b> {total_finalizados}<br/>
    <b>Tasa Promedio de Finalización:</b> {avg_rate}%
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        "Reporte generado por el Sistema de Gestión Catastral de ASOMUNICIPIOS",
        footer_style
    ))
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    # Save to temp file
    temp_pdf_path = UPLOAD_DIR / f"productivity_report_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"reporte_productividad_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type='application/pdf'
    )


@api_router.get("/reports/listado-tramites/export-pdf")
async def export_listado_tramites_pdf(
    municipio: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export petition list as PDF with professional statistics"""
    from reportlab.graphics.shapes import Drawing, String, Rect, Circle, Line
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.lib.colors import HexColor
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO, UserRole.GESTOR]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Build query
    query = {}
    if municipio:
        query["municipio"] = municipio
    if estado:
        query["estado"] = estado
    if fecha_inicio:
        query["created_at"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "created_at" in query:
            query["created_at"]["$lte"] = fecha_fin
        else:
            query["created_at"] = {"$lte": fecha_fin}
    
    # Get petitions
    petitions = await db.petitions.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    # Get user names for gestores
    users = await db.users.find({}, {"_id": 0, "id": 1, "full_name": 1}).to_list(1000)
    user_map = {u['id']: u['full_name'] for u in users}
    
    # Calculate statistics
    total = len(petitions)
    stats_estado = {}
    stats_municipio = {}
    stats_tipo = {}
    
    estado_labels = {
        'radicado': 'Radicado',
        'asignado': 'Asignado',
        'revision': 'En Revisión',
        'devuelto': 'Devuelto',
        'rechazado': 'Rechazado',
        'finalizado': 'Finalizado'
    }
    
    for pet in petitions:
        # Por estado
        est = pet.get('estado', 'radicado')
        stats_estado[est] = stats_estado.get(est, 0) + 1
        # Por municipio
        mun = pet.get('municipio', 'Sin municipio')
        stats_municipio[mun] = stats_municipio.get(mun, 0) + 1
        # Por tipo
        tipo = pet.get('tipo_tramite', 'Otro')
        stats_tipo[tipo] = stats_tipo.get(tipo, 0) + 1
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.4*inch, bottomMargin=0.4*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Define colors
    primary_color = HexColor('#009846')
    secondary_color = HexColor('#064e3b')
    light_bg = HexColor('#f0fdf4')
    gray_text = HexColor('#64748b')
    
    # Stats values
    finalizados = stats_estado.get('finalizado', 0)
    en_proceso = stats_estado.get('asignado', 0) + stats_estado.get('revision', 0) + stats_estado.get('devuelto', 0)
    radicados = stats_estado.get('radicado', 0)
    rechazados = stats_estado.get('rechazado', 0)
    
    # ===================== PROFESSIONAL HEADER =====================
    # Header bar with gradient effect using table
    header_bar_data = [['']]
    header_bar = Table(header_bar_data, colWidths=[9.5*inch], rowHeights=[0.08*inch])
    header_bar.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), primary_color),
    ]))
    story.append(header_bar)
    story.append(Spacer(1, 0.15*inch))
    
    # Logo and title row
    logo_path = Path("/app/frontend/public/logo-asomunicipios.png")
    header_content = []
    
    if logo_path.exists():
        logo_img = RLImage(str(logo_path), width=0.9*inch, height=0.4*inch)
        header_content.append([logo_img, '', ''])
    
    header_title = Paragraph(
        '<font size="18" color="#009846"><b>INFORME DE GESTIÓN</b></font><br/>'
        '<font size="10" color="#64748b">Trámites Catastrales</font>',
        ParagraphStyle('Title', alignment=TA_CENTER, leading=22)
    )
    
    date_info = Paragraph(
        f'<font size="8" color="#94a3b8">Generado el</font><br/>'
        f'<font size="10" color="#374151"><b>{datetime.now().strftime("%d de %B, %Y")}</b></font><br/>'
        f'<font size="8" color="#94a3b8">{datetime.now().strftime("%H:%M")} hrs</font>',
        ParagraphStyle('DateInfo', alignment=TA_CENTER, leading=12)
    )
    
    header_data = [['', header_title, date_info]]
    if logo_path.exists():
        header_data[0][0] = logo_img
    
    header_table = Table(header_data, colWidths=[1.5*inch, 5.5*inch, 2.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    story.append(header_table)
    
    # Subtitle line
    story.append(Spacer(1, 0.1*inch))
    subtitle_bar = Table([['']], colWidths=[9.5*inch], rowHeights=[0.02*inch])
    subtitle_bar.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), HexColor('#e2e8f0'))]))
    story.append(subtitle_bar)
    story.append(Spacer(1, 0.2*inch))
    
    # ===================== RESUMEN EJECUTIVO =====================
    resumen_title = Paragraph(
        '<font size="11" color="#009846"><b>RESUMEN EJECUTIVO</b></font>',
        ParagraphStyle('Section', spaceAfter=8)
    )
    story.append(resumen_title)
    
    # Filter info
    filter_parts = []
    if municipio:
        filter_parts.append(f"<b>Municipio:</b> {municipio}")
    if estado:
        filter_parts.append(f"<b>Estado:</b> {estado_labels.get(estado, estado)}")
    filter_text = " &nbsp;|&nbsp; ".join(filter_parts) if filter_parts else "<b>Filtro:</b> Todos los trámites"
    story.append(Paragraph(f'<font size="8" color="#64748b">{filter_text}</font>', styles['Normal']))
    story.append(Spacer(1, 0.15*inch))
    
    # ===================== KPI CARDS =====================
    def create_kpi_card(number, label, color, bg_color):
        return Paragraph(
            f'<font size="28" color="{color}"><b>{number}</b></font><br/>'
            f'<font size="8" color="#64748b">{label}</font>',
            ParagraphStyle('KPI', alignment=TA_CENTER, leading=32)
        )
    
    kpi_data = [[
        create_kpi_card(total, "TOTAL TRÁMITES", "#1e293b", "#f8fafc"),
        create_kpi_card(finalizados, "FINALIZADOS", "#16a34a", "#f0fdf4"),
        create_kpi_card(en_proceso, "EN PROCESO", "#d97706", "#fffbeb"),
        create_kpi_card(radicados, "RADICADOS", "#2563eb", "#eff6ff"),
        create_kpi_card(rechazados, "RECHAZADOS", "#dc2626", "#fef2f2"),
    ]]
    
    kpi_table = Table(kpi_data, colWidths=[1.9*inch]*5, rowHeights=[0.85*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), HexColor('#f8fafc')),
        ('BACKGROUND', (1, 0), (1, 0), HexColor('#f0fdf4')),
        ('BACKGROUND', (2, 0), (2, 0), HexColor('#fffbeb')),
        ('BACKGROUND', (3, 0), (3, 0), HexColor('#eff6ff')),
        ('BACKGROUND', (4, 0), (4, 0), HexColor('#fef2f2')),
        ('BOX', (0, 0), (0, 0), 1, HexColor('#e2e8f0')),
        ('BOX', (1, 0), (1, 0), 1, HexColor('#bbf7d0')),
        ('BOX', (2, 0), (2, 0), 1, HexColor('#fde68a')),
        ('BOX', (3, 0), (3, 0), 1, HexColor('#bfdbfe')),
        ('BOX', (4, 0), (4, 0), 1, HexColor('#fecaca')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.25*inch))
    
    # ===================== CHARTS SECTION =====================
    charts_title = Paragraph(
        '<font size="11" color="#009846"><b>DISTRIBUCIÓN DE TRÁMITES</b></font>',
        ParagraphStyle('Section', spaceAfter=10)
    )
    story.append(charts_title)
    
    # Create chart drawing
    chart_drawing = Drawing(650, 180)
    
    # Pie Chart (left side)
    if stats_estado:
        pie = Pie()
        pie.x = 60
        pie.y = 25
        pie.width = 130
        pie.height = 130
        
        pie_colors = {
            'radicado': HexColor('#3b82f6'),
            'asignado': HexColor('#8b5cf6'),
            'revision': HexColor('#f59e0b'),
            'devuelto': HexColor('#f97316'),
            'rechazado': HexColor('#ef4444'),
            'finalizado': HexColor('#22c55e'),
        }
        
        estado_order = ['finalizado', 'radicado', 'asignado', 'revision', 'devuelto', 'rechazado']
        pie_data = []
        active_estados = []
        
        for est in estado_order:
            if est in stats_estado and stats_estado[est] > 0:
                pie_data.append(stats_estado[est])
                active_estados.append(est)
        
        if pie_data:
            pie.data = pie_data
            pie.labels = None  # No labels on pie
            pie.slices.strokeWidth = 1
            pie.slices.strokeColor = colors.white
            
            for i, est in enumerate(active_estados):
                pie.slices[i].fillColor = pie_colors.get(est, HexColor('#94a3b8'))
            
            chart_drawing.add(pie)
            
            # Legend (right of pie)
            legend_x = 220
            legend_y = 150
            for i, est in enumerate(active_estados):
                # Color box
                rect = Rect(legend_x, legend_y - i*22, 14, 14, 
                           fillColor=pie_colors.get(est, HexColor('#94a3b8')), 
                           strokeWidth=0)
                chart_drawing.add(rect)
                # Label
                pct = (stats_estado[est] / total * 100) if total > 0 else 0
                label_text = f"{estado_labels.get(est, est)}: {stats_estado[est]} ({pct:.1f}%)"
                chart_drawing.add(String(legend_x + 20, legend_y - i*22 + 2, label_text, 
                                        fontSize=9, fillColor=HexColor('#374151')))
    
    # Bar chart for top municipalities (right side)
    if stats_municipio:
        # Sort and get top 5
        top_municipios = sorted(stats_municipio.items(), key=lambda x: x[1], reverse=True)[:5]
        
        if top_municipios:
            bar = VerticalBarChart()
            bar.x = 420
            bar.y = 30
            bar.width = 200
            bar.height = 120
            bar.data = [[m[1] for m in top_municipios]]
            bar.categoryAxis.categoryNames = [m[0][:10] for m in top_municipios]
            bar.categoryAxis.labels.fontSize = 7
            bar.categoryAxis.labels.angle = 0
            bar.valueAxis.valueMin = 0
            bar.valueAxis.labels.fontSize = 7
            bar.bars[0].fillColor = primary_color
            bar.bars[0].strokeWidth = 0
            bar.barWidth = 25
            
            chart_drawing.add(bar)
            chart_drawing.add(String(420, 160, "Top 5 Municipios", fontSize=9, fillColor=gray_text))
    
    story.append(chart_drawing)
    story.append(Spacer(1, 0.15*inch))
    
    # Footer for page 1
    footer_p1 = Paragraph(
        '<font size="7" color="#94a3b8">ASOMUNICIPIOS - Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar</font>',
        ParagraphStyle('FooterP1', alignment=TA_CENTER)
    )
    story.append(footer_p1)
    
    # ===================== PAGE BREAK - TABLE =====================
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    
    # Header for page 2
    story.append(header_bar)
    story.append(Spacer(1, 0.1*inch))
    
    table_title = Paragraph(
        '<font size="12" color="#009846"><b>LISTADO DETALLADO DE TRÁMITES</b></font>',
        ParagraphStyle('TableTitle', spaceAfter=5)
    )
    story.append(table_title)
    story.append(Paragraph(
        f'<font size="8" color="#64748b">Mostrando {min(len(petitions), 200)} de {len(petitions)} registros</font>',
        styles['Normal']
    ))
    story.append(Spacer(1, 0.1*inch))
    
    # Table data
    table_data = [['#', 'RADICADO', 'FECHA', 'SOLICITANTE', 'TIPO DE TRÁMITE', 'MUNICIPIO', 'ESTADO']]
    
    for idx, pet in enumerate(petitions[:200], 1):
        created_at = pet.get('created_at', '')
        if isinstance(created_at, str):
            try:
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                fecha_str = created_at.strftime('%d/%m/%y')
            except:
                fecha_str = str(created_at)[:8]
        else:
            fecha_str = created_at.strftime('%d/%m/%y') if created_at else ''
        
        table_data.append([
            str(idx),
            pet.get('radicado', '')[:24],
            fecha_str,
            pet.get('nombre_completo', '')[:28],
            pet.get('tipo_tramite', '')[:22],
            pet.get('municipio', '')[:12],
            estado_labels.get(pet.get('estado', ''), '')[:11]
        ])
    
    # Create table
    col_widths = [0.35*inch, 1.7*inch, 0.6*inch, 2.2*inch, 1.7*inch, 1*inch, 0.85*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('ALIGN', (6, 0), (6, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#f0fdf4')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    story.append(table)
    
    # Footer
    story.append(Spacer(1, 0.2*inch))
    footer_style = ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
    story.append(Paragraph(
        f"Documento generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} | ASOMUNICIPIOS - Sistema de Gestión Catastral",
        footer_style
    ))
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    temp_pdf_path = UPLOAD_DIR / f"listado_tramites_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"listado_tramites_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type='application/pdf'
    )


@api_router.get("/reports/tramites/export-excel")
async def export_tramites_excel(
    municipio: Optional[str] = None,
    estado: Optional[str] = None,
    gestor_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export petition history as Excel (for coordinators/admins)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo coordinadores y administradores pueden exportar el histórico")
    
    # Build query
    query = {}
    if municipio and municipio != 'todos':
        query["municipio"] = municipio
    if estado and estado != 'todos':
        query["estado"] = estado
    if gestor_id and gestor_id != 'todos':
        query["gestores_asignados"] = gestor_id
    
    # Date filters
    if fecha_desde or fecha_hasta:
        query["created_at"] = {}
        if fecha_desde:
            query["created_at"]["$gte"] = fecha_desde
        if fecha_hasta:
            query["created_at"]["$lte"] = fecha_hasta + "T23:59:59"
        if not query["created_at"]:
            del query["created_at"]
    
    # Get petitions
    petitions = await db.petitions.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Get user names for gestores
    users = await db.users.find({}, {"_id": 0, "id": 1, "full_name": 1, "email": 1}).to_list(1000)
    user_map = {u['id']: u for u in users}
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Trámites"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Headers
    headers = [
        "No.", "Radicado", "Fecha Creación", "Solicitante", "Correo", "Teléfono",
        "Tipo de Trámite", "Municipio", "Estado", "Gestor Asignado", "Descripción"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Estado labels
    estado_labels = {
        'radicado': 'Radicado',
        'asignado': 'Asignado',
        'revision': 'En Revisión',
        'devuelto': 'Devuelto',
        'rechazado': 'Rechazado',
        'finalizado': 'Finalizado'
    }
    
    # Data rows
    for idx, pet in enumerate(petitions, 1):
        # Parse date
        created_at = pet.get('created_at', '')
        if isinstance(created_at, str):
            try:
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                fecha_str = created_at.strftime('%d/%m/%Y %H:%M')
            except:
                fecha_str = str(created_at)[:16]
        else:
            fecha_str = created_at.strftime('%d/%m/%Y %H:%M') if created_at else ''
        
        # Get gestor names
        gestor_names = []
        for g_id in pet.get('gestores_asignados', []):
            if g_id in user_map:
                gestor_names.append(user_map[g_id]['full_name'])
        gestor_str = ', '.join(gestor_names) if gestor_names else 'Sin asignar'
        
        row_data = [
            idx,
            pet.get('radicado', pet.get('radicado_id', '')),
            fecha_str,
            pet.get('nombre_completo', pet.get('creator_name', '')),
            pet.get('correo', ''),
            pet.get('telefono', ''),
            pet.get('tipo_tramite', ''),
            pet.get('municipio', ''),
            estado_labels.get(pet.get('estado', ''), pet.get('estado', '')),
            gestor_str,
            (pet.get('descripcion', '') or '')[:200]  # Truncate long descriptions
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=idx + 1, column=col, value=value)
    
    # Adjust column widths
    column_widths = [6, 25, 18, 25, 25, 15, 30, 15, 12, 25, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Resumen")
    ws_summary.cell(row=1, column=1, value="RESUMEN DEL HISTÓRICO DE TRÁMITES").font = Font(bold=True, size=14)
    ws_summary.cell(row=2, column=1, value=f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws_summary.cell(row=3, column=1, value=f"Total de trámites: {len(petitions)}")
    
    # Count by status
    status_counts = {}
    for pet in petitions:
        est = pet.get('estado', 'unknown')
        status_counts[est] = status_counts.get(est, 0) + 1
    
    row = 5
    ws_summary.cell(row=row, column=1, value="Por Estado:").font = Font(bold=True)
    row += 1
    for est, count in status_counts.items():
        ws_summary.cell(row=row, column=1, value=estado_labels.get(est, est))
        ws_summary.cell(row=row, column=2, value=count)
        row += 1
    
    # Count by municipio
    muni_counts = {}
    for pet in petitions:
        muni = pet.get('municipio', 'Sin municipio')
        muni_counts[muni] = muni_counts.get(muni, 0) + 1
    
    row += 1
    ws_summary.cell(row=row, column=1, value="Por Municipio:").font = Font(bold=True)
    row += 1
    for muni, count in sorted(muni_counts.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=row, column=1, value=muni)
        ws_summary.cell(row=row, column=2, value=count)
        row += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Save to temp file
    temp_path = UPLOAD_DIR / f"historico_tramites_{uuid.uuid4()}.xlsx"
    with open(temp_path, 'wb') as f:
        f.write(output.getvalue())
    
    return FileResponse(
        path=temp_path,
        filename=f"Historico_Tramites_{datetime.now().strftime('%Y%m%d')}.xlsx",
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ===== ADVANCED STATISTICS =====

@api_router.get("/stats/by-municipality")
async def get_stats_by_municipality(current_user: dict = Depends(get_current_user)):
    """Get petition statistics grouped by municipality"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    pipeline = [
        {"$group": {
            "_id": "$municipio",
            "total": {"$sum": 1},
            "radicado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RADICADO]}, 1, 0]}},
            "asignado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.ASIGNADO]}, 1, 0]}},
            "revision": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.REVISION]}, 1, 0]}},
            "finalizado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.FINALIZADO]}, 1, 0]}},
            "rechazado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RECHAZADO]}, 1, 0]}},
            "devuelto": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.DEVUELTO]}, 1, 0]}}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await db.petitions.aggregate(pipeline).to_list(100)
    
    return [
        {
            "municipio": r["_id"] or "Sin especificar",
            "total": r["total"],
            "radicado": r["radicado"],
            "asignado": r["asignado"],
            "revision": r["revision"],
            "finalizado": r["finalizado"],
            "rechazado": r["rechazado"],
            "devuelto": r["devuelto"]
        }
        for r in results
    ]

@api_router.get("/stats/by-tramite")
async def get_stats_by_tramite(current_user: dict = Depends(get_current_user)):
    """Get petition statistics grouped by tramite type"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    pipeline = [
        {"$group": {
            "_id": "$tipo_tramite",
            "total": {"$sum": 1},
            "radicado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RADICADO]}, 1, 0]}},
            "asignado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.ASIGNADO]}, 1, 0]}},
            "revision": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.REVISION]}, 1, 0]}},
            "finalizado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.FINALIZADO]}, 1, 0]}},
            "rechazado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RECHAZADO]}, 1, 0]}},
            "devuelto": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.DEVUELTO]}, 1, 0]}}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await db.petitions.aggregate(pipeline).to_list(100)
    
    return [
        {
            "tipo_tramite": r["_id"] or "Sin especificar",
            "total": r["total"],
            "radicado": r["radicado"],
            "asignado": r["asignado"],
            "revision": r["revision"],
            "finalizado": r["finalizado"],
            "rechazado": r["rechazado"],
            "devuelto": r["devuelto"]
        }
        for r in results
    ]

@api_router.get("/stats/by-gestor")
async def get_stats_by_gestor(current_user: dict = Depends(get_current_user)):
    """Get petition statistics grouped by assigned gestor"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get all gestores
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR]}},
        {"_id": 0, "id": 1, "full_name": 1, "role": 1}
    ).to_list(100)
    
    gestor_stats = []
    
    for gestor in gestores:
        gestor_id = gestor['id']
        
        # Count by status for this gestor
        total = await db.petitions.count_documents({"gestores_asignados": gestor_id})
        radicado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.RADICADO})
        asignado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.ASIGNADO})
        revision = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.REVISION})
        finalizado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.FINALIZADO})
        rechazado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.RECHAZADO})
        devuelto = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.DEVUELTO})
        
        completion_rate = round((finalizado / total * 100), 1) if total > 0 else 0
        
        gestor_stats.append({
            "gestor_id": gestor_id,
            "gestor_name": gestor['full_name'],
            "gestor_role": "Gestor" if gestor['role'] == UserRole.GESTOR else "Gestor Auxiliar",
            "total": total,
            "radicado": radicado,
            "asignado": asignado,
            "revision": revision,
            "finalizado": finalizado,
            "rechazado": rechazado,
            "devuelto": devuelto,
            "completion_rate": completion_rate
        })
    
    # Sort by total descending
    gestor_stats.sort(key=lambda x: x['total'], reverse=True)
    
    return gestor_stats

@api_router.get("/stats/summary")
async def get_stats_summary(current_user: dict = Depends(get_current_user)):
    """Get overall statistics summary"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Total counts
    total_petitions = await db.petitions.count_documents({})
    total_users = await db.users.count_documents({})
    total_gestores = await db.users.count_documents({"role": UserRole.GESTOR})
    
    # Staff counts by role
    staff_counts = {
        "coordinadores": await db.users.count_documents({"role": UserRole.COORDINADOR}),
        "gestores": await db.users.count_documents({"role": UserRole.GESTOR}),
        "atencion_usuario": await db.users.count_documents({"role": UserRole.ATENCION_USUARIO}),
        "administradores": await db.users.count_documents({"role": UserRole.ADMINISTRADOR}),
        "ciudadanos": await db.users.count_documents({"role": UserRole.USUARIO})
    }
    
    # Status counts
    status_counts = {
        "radicado": await db.petitions.count_documents({"estado": PetitionStatus.RADICADO}),
        "asignado": await db.petitions.count_documents({"estado": PetitionStatus.ASIGNADO}),
        "revision": await db.petitions.count_documents({"estado": PetitionStatus.REVISION}),
        "finalizado": await db.petitions.count_documents({"estado": PetitionStatus.FINALIZADO}),
        "rechazado": await db.petitions.count_documents({"estado": PetitionStatus.RECHAZADO}),
        "devuelto": await db.petitions.count_documents({"estado": PetitionStatus.DEVUELTO})
    }
    
    # Completion rate
    completion_rate = round((status_counts["finalizado"] / total_petitions * 100), 1) if total_petitions > 0 else 0
    
    # Recent petitions (last 30 days)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    recent_petitions = await db.petitions.count_documents({
        "created_at": {"$gte": thirty_days_ago}
    })
    
    return {
        "total_petitions": total_petitions,
        "total_users": total_users,
        "total_gestores": total_gestores,
        "staff_counts": staff_counts,
        "status_counts": status_counts,
        "completion_rate": completion_rate,
        "recent_petitions_30_days": recent_petitions
    }


# ===== PREDIOS ROUTES (Código Nacional Catastral) =====

async def generate_codigo_predial(municipio: str, zona: str, sector: str, manzana_vereda: str, 
                                   terreno: str, condicion: str, ph: str) -> str:
    """Genera el código predial nacional de 30 dígitos"""
    if municipio not in MUNICIPIOS_DIVIPOLA:
        raise HTTPException(status_code=400, detail=f"Municipio '{municipio}' no válido")
    
    divipola = MUNICIPIOS_DIVIPOLA[municipio]
    
    # Construir código de 30 dígitos
    codigo = (
        divipola["departamento"].zfill(2) +  # 2 dígitos
        divipola["municipio"].zfill(3) +     # 3 dígitos
        zona.zfill(2) +                       # 2 dígitos
        sector.zfill(2) +                     # 2 dígitos
        manzana_vereda.zfill(4) +            # 4 dígitos
        terreno.zfill(4) +                    # 4 dígitos
        condicion.zfill(4) +                  # 4 dígitos
        ph.zfill(4) +                         # 4 dígitos
        "00000"                               # 5 dígitos (unidad predial)
    )
    
    return codigo

async def generate_codigo_homologado(municipio: str) -> str:
    """Genera un código homologado único de 11 caracteres"""
    import string
    import random
    
    # Obtener último código para este municipio - buscar solo número_predio que sea un entero razonable
    last_predio = await db.predios.find_one(
        {
            "municipio": municipio, 
            "deleted": {"$ne": True},
            "numero_predio": {"$type": "int"}  # Solo buscar donde sea entero
        },
        sort=[("numero_predio", -1)]
    )
    
    if last_predio:
        num_predio = last_predio.get("numero_predio", 0)
        # Verificar que sea un número razonable (menos de 100 millones)
        if isinstance(num_predio, (int, float)) and num_predio < 100000000:
            next_num = int(num_predio) + 1
        else:
            # Si no hay un número razonable, buscar el máximo de predios_nuevos
            last_nuevo = await db.predios_nuevos.find_one(
                {"municipio": municipio},
                sort=[("numero_predio", -1)]
            )
            if last_nuevo and isinstance(last_nuevo.get("numero_predio"), int):
                next_num = last_nuevo["numero_predio"] + 1
            else:
                # Contar predios como fallback
                count = await db.predios.count_documents({"municipio": municipio})
                next_num = count + 1
    else:
        # No hay predios con numero_predio entero, verificar predios_nuevos
        last_nuevo = await db.predios_nuevos.find_one(
            {"municipio": municipio},
            sort=[("numero_predio", -1)]
        )
        if last_nuevo and isinstance(last_nuevo.get("numero_predio"), int):
            next_num = last_nuevo["numero_predio"] + 1
        else:
            next_num = 1
    
    # Generar código: BPP + número + letras aleatorias
    letters = ''.join(random.choices(string.ascii_uppercase, k=4))
    codigo = f"BPP{str(next_num).zfill(4)}{letters}"
    
    return codigo, next_num

async def get_next_terreno_number(municipio: str, zona: str, sector: str, manzana_vereda: str) -> str:
    """Obtiene el siguiente número de terreno disponible (incluyendo eliminados)"""
    # Buscar el máximo terreno usado (incluyendo eliminados para no reutilizar)
    pipeline = [
        {"$match": {
            "municipio": municipio,
            "zona": zona,
            "sector": sector,
            "manzana_vereda": manzana_vereda
        }},
        {"$group": {
            "_id": None,
            "max_terreno": {"$max": "$terreno_num"}
        }}
    ]
    
    result = await db.predios.aggregate(pipeline).to_list(1)
    
    if result and result[0].get("max_terreno"):
        next_num = result[0]["max_terreno"] + 1
    else:
        next_num = 1
    
    return str(next_num).zfill(4), next_num

@api_router.get("/predios/catalogos")
async def get_predios_catalogos(current_user: dict = Depends(get_current_user)):
    """Obtiene los catálogos para el formulario de predios"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    return {
        "municipios": list(MUNICIPIOS_DIVIPOLA.keys()),
        "destino_economico": DESTINO_ECONOMICO,
        "tipo_documento": TIPO_DOCUMENTO_PREDIO,
        "estado_civil": ESTADO_CIVIL_PREDIO,
        "divipola": MUNICIPIOS_DIVIPOLA
    }

# ===== SISTEMA DE CÓDIGOS HOMOLOGADOS =====

@api_router.post("/codigos-homologados/cargar")
async def cargar_codigos_homologados(
    file: UploadFile = File(...),
    municipio: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Cargar códigos homologados desde un archivo Excel.
    
    Compara cada código contra los predios existentes para determinar si está usado o disponible.
    Un código está "usado" si ya existe un predio con ese codigo_homologado asignado.
    
    El archivo puede tener:
    1. Dos columnas: 'Municipio' y 'Codigo_Homologado' (o similar)
    2. Una sola columna con códigos - en este caso se requiere el parámetro 'municipio'
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden cargar códigos")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")
    
    try:
        import pandas as pd
        from io import BytesIO
        
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        # Normalizar nombres de columnas
        df.columns = df.columns.str.strip().str.lower()
        
        # Buscar columnas de municipio y código
        municipio_col = None
        codigo_col = None
        
        for col in df.columns:
            if 'municipio' in col:
                municipio_col = col
            if 'codigo' in col or 'homologado' in col:
                codigo_col = col
        
        # Si no hay columna de código identificada, asumir que la primera columna son los códigos
        if not codigo_col and len(df.columns) >= 1:
            codigo_col = df.columns[0]
        
        # Si no hay columna de municipio, debe proporcionarse el parámetro
        # IMPORTANTE: Si el usuario seleccionó un municipio, SIEMPRE usarlo (tiene prioridad)
        if municipio:
            # El usuario seleccionó explícitamente un municipio - usarlo siempre
            municipio_fijo = municipio.strip()
            logging.info(f"Usando municipio seleccionado por usuario: {municipio_fijo}")
        elif not municipio_col:
            raise HTTPException(
                status_code=400, 
                detail="El archivo no tiene columna 'Municipio'. Por favor seleccione el municipio."
            )
        else:
            municipio_fijo = None
        
        # === OPTIMIZACIÓN: Obtener todos los datos necesarios de una sola vez ===
        
        # 1. Identificar municipios en el archivo
        municipios_en_archivo = set()
        if municipio_fijo:
            municipios_en_archivo.add(municipio_fijo)
        else:
            for _, row in df.iterrows():
                muni = str(row[municipio_col]).strip() if municipio_col else municipio_fijo
                if muni and muni != 'nan':
                    municipios_en_archivo.add(muni)
        
        # 2. Obtener códigos ya cargados (para evitar duplicados) - búsqueda case-insensitive
        codigos_existentes = set()
        for muni in municipios_en_archivo:
            # Usar regex case-insensitive para encontrar códigos existentes
            existing_docs = await db.codigos_homologados.find(
                {'municipio': {'$regex': f'^{muni}$', '$options': 'i'}},
                {'codigo': 1, 'municipio': 1}
            ).to_list(100000)
            for doc in existing_docs:
                # Normalizar: usar el nombre del archivo en la clave
                codigos_existentes.add(f"{muni}:{doc['codigo'].upper()}")
        
        # 3. Obtener predios con codigo_homologado asignado - UNA sola consulta por municipio
        # Esto nos dice qué códigos ya están EN USO
        # IMPORTANTE: Usamos búsqueda case-insensitive para evitar problemas de mayúsculas/minúsculas
        codigos_en_uso = {}  # codigo -> {predio_id, codigo_predial, propietario, municipio_real}
        municipio_normalizado = {}  # muni_lower -> muni_original_en_db
        
        for muni in municipios_en_archivo:
            # Búsqueda case-insensitive usando regex
            predios_con_codigo = await db.predios.find(
                {
                    'municipio': {'$regex': f'^{muni}$', '$options': 'i'},
                    'codigo_homologado': {'$exists': True, '$ne': None, '$ne': ''},
                    'deleted': {'$ne': True}
                },
                {'_id': 0, 'id': 1, 'codigo_homologado': 1, 'codigo_predial_nacional': 1, 'nombre_propietario': 1, 'municipio': 1}
            ).to_list(100000)
            
            for predio in predios_con_codigo:
                codigo = predio.get('codigo_homologado', '').strip().upper()
                municipio_db = predio.get('municipio', muni)
                if codigo:
                    # Usar el nombre del municipio del archivo para la clave
                    clave = f"{muni}:{codigo}"
                    codigos_en_uso[clave] = {
                        'predio_id': predio['id'],
                        'codigo_predial': predio.get('codigo_predial_nacional'),
                        'propietario': predio.get('nombre_propietario')
                    }
                    # Guardar mapping de municipio para referencia
                    if muni.lower() not in municipio_normalizado:
                        municipio_normalizado[muni.lower()] = municipio_db
        
        # Log de debug para verificar cuántos códigos en uso se encontraron
        logging.info(f"Códigos homologados: municipios={list(municipios_en_archivo)}, predios_con_codigo_encontrados={len(codigos_en_uso)}")
        
        # 4. Preparar documentos para inserción bulk
        documentos_a_insertar = []
        codigos_duplicados = 0
        codigos_usados = 0
        codigos_disponibles = 0
        codigos_por_municipio = {}
        fecha_carga = datetime.utcnow().isoformat()
        
        for _, row in df.iterrows():
            muni = municipio_fijo if municipio_fijo else str(row[municipio_col]).strip()
            codigo = str(row[codigo_col]).strip().upper() if codigo_col else None
            
            # Validar datos
            if not muni or not codigo or muni == 'nan' or codigo == 'NAN':
                continue
            
            clave = f"{muni}:{codigo}"
            
            # Verificar duplicado en la colección
            if clave in codigos_existentes:
                codigos_duplicados += 1
                continue
            
            # Agregar a existentes para evitar duplicados dentro del mismo archivo
            codigos_existentes.add(clave)
            
            # Verificar si está en uso (predio ya tiene este código)
            info_uso = codigos_en_uso.get(clave)
            
            if info_uso:
                # Código YA ESTÁ ASIGNADO a un predio
                documentos_a_insertar.append({
                    'municipio': muni,
                    'codigo': codigo,
                    'usado': True,
                    'predio_id': info_uso['predio_id'],
                    'codigo_predial': info_uso['codigo_predial'],
                    'propietario': info_uso['propietario'],
                    'fecha_asignacion': None,
                    'cargado_por': current_user['id'],
                    'cargado_por_nombre': current_user['full_name'],
                    'fecha_carga': fecha_carga
                })
                codigos_usados += 1
            else:
                # Código DISPONIBLE
                documentos_a_insertar.append({
                    'municipio': muni,
                    'codigo': codigo,
                    'usado': False,
                    'predio_id': None,
                    'codigo_predial': None,
                    'propietario': None,
                    'fecha_asignacion': None,
                    'cargado_por': current_user['id'],
                    'cargado_por_nombre': current_user['full_name'],
                    'fecha_carga': fecha_carga
                })
                codigos_disponibles += 1
            
            codigos_por_municipio[muni] = codigos_por_municipio.get(muni, 0) + 1
        
        # 5. Inserción bulk (mucho más rápido que inserciones individuales)
        codigos_insertados = 0
        if documentos_a_insertar:
            batch_size = 1000
            for i in range(0, len(documentos_a_insertar), batch_size):
                batch = documentos_a_insertar[i:i + batch_size]
                result = await db.codigos_homologados.insert_many(batch)
                codigos_insertados += len(result.inserted_ids)
        
        return {
            "success": True,
            "message": f"Códigos cargados exitosamente",
            "codigos_insertados": codigos_insertados,
            "codigos_duplicados": codigos_duplicados,
            "codigos_usados": codigos_usados,
            "codigos_disponibles": codigos_disponibles,
            "por_municipio": codigos_por_municipio,
            "debug_info": {
                "predios_encontrados_con_codigo": len(codigos_en_uso),
                "codigos_existentes_en_bd": len(codigos_existentes) - len(documentos_a_insertar),
                "municipios_procesados": list(municipios_en_archivo)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error procesando archivo de códigos homologados: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")


@api_router.get("/codigos-homologados/stats")
async def get_codigos_homologados_stats(
    current_user: dict = Depends(get_current_user)
):
    """Obtener estadísticas de códigos homologados por municipio"""
    pipeline = [
        {
            "$group": {
                "_id": "$municipio",
                "total": {"$sum": 1},
                "usados": {"$sum": {"$cond": ["$usado", 1, 0]}},
                "disponibles": {"$sum": {"$cond": ["$usado", 0, 1]}}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    results = await db.codigos_homologados.aggregate(pipeline).to_list(100)
    
    stats = []
    for r in results:
        stats.append({
            "municipio": r["_id"],
            "total": r["total"],
            "usados": r["usados"],
            "disponibles": r["disponibles"]
        })
    
    return {
        "stats": stats,
        "total_municipios": len(stats)
    }


@api_router.get("/codigos-homologados/usados/{municipio}")
async def get_codigos_usados(
    municipio: str,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Obtener códigos homologados usados con información del predio asociado"""
    # Obtener códigos usados
    codigos = await db.codigos_homologados.find(
        {'municipio': municipio, 'usado': True},
        {'_id': 0}
    ).sort('codigo', 1).skip(skip).limit(limit).to_list(limit)
    
    # Para cada código, obtener info del predio si no está en el registro
    codigos_con_info = []
    for c in codigos:
        codigo_info = {
            'codigo': c['codigo'],
            'predio_id': c.get('predio_id'),
            'codigo_predial': c.get('codigo_predial'),
            'propietario': c.get('propietario'),
            'fecha_asignacion': c.get('fecha_asignacion')
        }
        
        # Si no tiene info del predio, buscarla
        if c.get('predio_id') and not c.get('codigo_predial'):
            predio = await db.predios.find_one(
                {'id': c['predio_id'], 'deleted': {'$ne': True}},
                {'_id': 0, 'codigo_predial_nacional': 1, 'nombre_propietario': 1}
            )
            if predio:
                codigo_info['codigo_predial'] = predio.get('codigo_predial_nacional')
                codigo_info['propietario'] = predio.get('nombre_propietario')
        
        codigos_con_info.append(codigo_info)
    
    total_usados = await db.codigos_homologados.count_documents({
        'municipio': municipio,
        'usado': True
    })
    
    return {
        "municipio": municipio,
        "codigos": codigos_con_info,
        "total_usados": total_usados,
        "skip": skip,
        "limit": limit
    }


@api_router.get("/codigos-homologados/disponibles/{municipio}")
async def get_codigos_disponibles(
    municipio: str,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Obtener códigos homologados disponibles para un municipio"""
    codigos = await db.codigos_homologados.find(
        {'municipio': municipio, 'usado': False},
        {'_id': 0, 'codigo': 1}
    ).sort('codigo', 1).limit(limit).to_list(limit)
    
    total_disponibles = await db.codigos_homologados.count_documents({
        'municipio': municipio,
        'usado': False
    })
    
    return {
        "municipio": municipio,
        "codigos": [c['codigo'] for c in codigos],
        "total_disponibles": total_disponibles
    }


@api_router.get("/codigos-homologados/siguiente/{municipio}")
async def get_siguiente_codigo_homologado(
    municipio: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener el siguiente código homologado disponible para un municipio"""
    # Buscar el primer código disponible (ordenado alfabéticamente)
    codigo_doc = await db.codigos_homologados.find_one(
        {'municipio': municipio, 'usado': False},
        sort=[('codigo', 1)]
    )
    
    if not codigo_doc:
        return {
            "municipio": municipio,
            "codigo": None,
            "disponibles": 0,
            "mensaje": "No hay códigos homologados disponibles para este municipio"
        }
    
    total_disponibles = await db.codigos_homologados.count_documents({
        'municipio': municipio,
        'usado': False
    })
    
    return {
        "municipio": municipio,
        "codigo": codigo_doc['codigo'],
        "disponibles": total_disponibles
    }


async def asignar_codigo_homologado(municipio: str, predio_id: str) -> str:
    """Asigna el siguiente código homologado disponible a un predio"""
    # Buscar y marcar como usado en una operación atómica
    result = await db.codigos_homologados.find_one_and_update(
        {'municipio': municipio, 'usado': False},
        {
            '$set': {
                'usado': True,
                'predio_id': predio_id,
                'fecha_asignacion': datetime.utcnow().isoformat()
            }
        },
        sort=[('codigo', 1)],
        return_document=True
    )
    
    if result:
        return result['codigo']
    return None


@api_router.delete("/codigos-homologados/{municipio}")
async def eliminar_codigos_municipio(
    municipio: str,
    current_user: dict = Depends(get_current_user)
):
    """Eliminar todos los códigos no usados de un municipio"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar códigos")
    
    result = await db.codigos_homologados.delete_many({
        'municipio': municipio,
        'usado': False
    })
    
    return {
        "success": True,
        "eliminados": result.deleted_count,
        "municipio": municipio
    }


@api_router.get("/codigos-homologados/diagnostico/{municipio}")
async def diagnostico_codigos_municipio(
    municipio: str,
    current_user: dict = Depends(get_current_user)
):
    """Diagnóstico detallado de códigos homologados para un municipio.
    Compara los códigos en la colección con los predios que realmente los tienen asignados.
    OPTIMIZADO: Usa búsquedas en memoria para evitar N+1 queries."""
    
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # 1. Contar códigos en la colección codigos_homologados (case-insensitive)
    total_en_coleccion = await db.codigos_homologados.count_documents({
        'municipio': {'$regex': f'^{municipio}$', '$options': 'i'}
    })
    usados_en_coleccion = await db.codigos_homologados.count_documents({
        'municipio': {'$regex': f'^{municipio}$', '$options': 'i'}, 
        'usado': True
    })
    disponibles_en_coleccion = await db.codigos_homologados.count_documents({
        'municipio': {'$regex': f'^{municipio}$', '$options': 'i'}, 
        'usado': False
    })
    
    # 2. Contar predios que REALMENTE tienen codigo_homologado asignado
    predios_con_codigo = await db.predios.count_documents({
        'municipio': {'$regex': f'^{municipio}$', '$options': 'i'},
        'codigo_homologado': {'$exists': True, '$ne': None, '$ne': ''},
        'deleted': {'$ne': True}
    })
    
    # 3. Obtener los códigos que están en predios (una sola consulta)
    codigos_en_predios = await db.predios.distinct('codigo_homologado', {
        'municipio': {'$regex': f'^{municipio}$', '$options': 'i'},
        'codigo_homologado': {'$exists': True, '$ne': None, '$ne': ''},
        'deleted': {'$ne': True}
    })
    # Normalizar a mayúsculas para comparación
    codigos_en_predios_set = set(c.strip().upper() for c in codigos_en_predios if c)
    
    # 4. Obtener códigos de la colección (una sola consulta)
    codigos_en_coleccion = await db.codigos_homologados.distinct('codigo', {
        'municipio': {'$regex': f'^{municipio}$', '$options': 'i'}
    })
    codigos_en_coleccion_set = set(c.strip().upper() for c in codigos_en_coleccion if c)
    
    # Códigos en predios que NO están en la colección
    codigos_solo_en_predios = codigos_en_predios_set - codigos_en_coleccion_set
    
    # 5. OPTIMIZADO: Detectar códigos huérfanos sin N+1 queries
    # Obtener todos los códigos marcados como "usados" de una sola vez
    codigos_usados_docs = await db.codigos_homologados.find(
        {'municipio': {'$regex': f'^{municipio}$', '$options': 'i'}, 'usado': True},
        {'_id': 0, 'codigo': 1}
    ).to_list(100000)
    
    # Códigos huérfanos = marcados como usados pero NO están en ningún predio
    codigos_huerfanos = []
    for doc in codigos_usados_docs:
        codigo = doc.get('codigo', '').strip().upper()
        if codigo and codigo not in codigos_en_predios_set:
            codigos_huerfanos.append(codigo)
    
    return {
        "municipio": municipio,
        "en_coleccion": {
            "total": total_en_coleccion,
            "usados": usados_en_coleccion,
            "disponibles": disponibles_en_coleccion
        },
        "en_predios": {
            "predios_con_codigo_homologado": predios_con_codigo,
            "codigos_unicos": len(codigos_en_predios_set)
        },
        "inconsistencias": {
            "codigos_solo_en_predios": len(codigos_solo_en_predios),
            "codigos_huerfanos_en_coleccion": len(codigos_huerfanos),
            "ejemplos_solo_en_predios": list(codigos_solo_en_predios)[:10],
            "ejemplos_huerfanos": codigos_huerfanos[:10]
        },
        "recomendacion": "Si hay muchos códigos huérfanos, use el endpoint /codigos-homologados/recalcular/{municipio} para corregir"
    }


@api_router.post("/codigos-homologados/recalcular/{municipio}")
async def recalcular_codigos_municipio(
    municipio: str,
    current_user: dict = Depends(get_current_user)
):
    """Recalcula el estado de los códigos homologados basándose en los predios REALMENTE asignados.
    
    Esto corrige situaciones donde:
    - Un código está marcado como 'usado' pero el predio ya no lo tiene
    - Un código está marcado como 'disponible' pero un predio lo tiene asignado
    """
    
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden recalcular códigos")
    
    codigos_liberados = 0
    codigos_marcados_usados = 0
    
    # 1. Obtener todos los códigos del municipio de una sola vez
    codigos_docs = await db.codigos_homologados.find(
        {'municipio': municipio},
        {'_id': 0, 'codigo': 1, 'usado': 1}
    ).to_list(100000)
    
    if not codigos_docs:
        return {
            "success": True,
            "municipio": municipio,
            "total_revisados": 0,
            "codigos_corregidos": 0,
            "codigos_liberados": 0,
            "codigos_marcados_usados": 0,
            "mensaje": "No hay códigos para este municipio"
        }
    
    # 2. Obtener todos los predios del municipio que tienen codigo_homologado de una sola vez
    # Usar búsqueda case-insensitive para municipio
    predios_con_codigo = await db.predios.find(
        {
            'municipio': {'$regex': f'^{municipio}$', '$options': 'i'},
            'codigo_homologado': {'$exists': True, '$ne': None, '$ne': ''},
            'deleted': {'$ne': True}
        },
        {'_id': 0, 'id': 1, 'codigo_homologado': 1, 'codigo_predial_nacional': 1, 'nombre_propietario': 1}
    ).to_list(100000)
    
    # 3. Crear diccionario de códigos -> predio para búsqueda O(1)
    codigos_en_predios = {}
    for predio in predios_con_codigo:
        codigo = predio.get('codigo_homologado', '').strip().upper()
        if codigo:
            codigos_en_predios[codigo] = predio
    
    # 4. Preparar operaciones bulk
    from pymongo import UpdateOne
    operaciones = []
    fecha_ahora = datetime.utcnow().isoformat()
    
    for codigo_doc in codigos_docs:
        codigo = codigo_doc['codigo']
        esta_marcado_usado = codigo_doc.get('usado', False)
        predio_real = codigos_en_predios.get(codigo)
        
        if predio_real and not esta_marcado_usado:
            # El predio tiene el código pero está marcado como disponible -> marcar como usado
            operaciones.append(UpdateOne(
                {'municipio': municipio, 'codigo': codigo},
                {'$set': {
                    'usado': True,
                    'predio_id': predio_real['id'],
                    'codigo_predial': predio_real.get('codigo_predial_nacional'),
                    'propietario': predio_real.get('nombre_propietario'),
                    'corregido_en': fecha_ahora
                }}
            ))
            codigos_marcados_usados += 1
            
        elif not predio_real and esta_marcado_usado:
            # Está marcado como usado pero ningún predio lo tiene -> liberar
            operaciones.append(UpdateOne(
                {'municipio': municipio, 'codigo': codigo},
                {'$set': {
                    'usado': False,
                    'predio_id': None,
                    'codigo_predial': None,
                    'propietario': None,
                    'liberado_en': fecha_ahora
                }}
            ))
            codigos_liberados += 1
    
    # 5. Ejecutar operaciones bulk
    if operaciones:
        await db.codigos_homologados.bulk_write(operaciones)
    
    return {
        "success": True,
        "municipio": municipio,
        "total_revisados": len(codigos_docs),
        "codigos_corregidos": codigos_liberados + codigos_marcados_usados,
        "codigos_liberados": codigos_liberados,
        "codigos_marcados_usados": codigos_marcados_usados,
        "mensaje": f"Se liberaron {codigos_liberados} códigos y se marcaron {codigos_marcados_usados} como usados"
    }


# ===== FIN SISTEMA DE CÓDIGOS HOMOLOGADOS =====

@api_router.get("/predios")
async def get_predios(
    municipio: Optional[str] = None,
    vigencia: Optional[int] = None,
    destino_economico: Optional[str] = None,
    zona: Optional[str] = None,  # '00' = rural, '01' = urbano, '02-99' = corregimientos
    tiene_geometria: Optional[str] = None,  # Filtro para predios con/sin geometría GDB ('true' o 'false')
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50000,  # Aumentado para soportar municipios grandes en modo offline
    current_user: dict = Depends(get_current_user)
):
    """Lista todos los predios (solo staff)"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {"deleted": {"$ne": True}}
    
    # Si es usuario empresa, filtrar por municipios asignados
    if current_user['role'] == UserRole.EMPRESA:
        municipios_asignados = current_user.get('municipios_asignados', [])
        if not municipios_asignados:
            return {"total": 0, "predios": [], "mensaje": "No tiene municipios asignados. Contacte al coordinador."}
        
        # Si hay búsqueda activa, primero buscar sin restricción de municipio para determinar si existe
        if search:
            # Buscar el predio sin restricción de municipio
            search_query_general = {"deleted": {"$ne": True}}
            is_matricula_search = bool(re.match(r'^\d{3}-\d+$', search.strip()))
            
            if is_matricula_search:
                search_query_general["r2_registros.matricula_inmobiliaria"] = search.strip()
            else:
                search_query_general["$or"] = [
                    {"codigo_predial_nacional": {"$regex": search, "$options": "i"}},
                    {"codigo_homologado": {"$regex": search, "$options": "i"}},
                    {"propietarios.nombre_propietario": {"$regex": search, "$options": "i"}},
                    {"propietarios.numero_documento": {"$regex": search, "$options": "i"}},
                    {"direccion": {"$regex": search, "$options": "i"}},
                    {"r2_registros.matricula_inmobiliaria": {"$regex": search, "$options": "i"}}
                ]
            
            predio_encontrado = await db.predios.find_one(search_query_general, {"_id": 0, "municipio": 1, "codigo_predial_nacional": 1})
            
            if predio_encontrado:
                # Verificar si el municipio del predio está en los asignados
                municipio_predio = predio_encontrado.get('municipio')
                if municipio_predio and municipio_predio not in municipios_asignados:
                    # El predio existe pero el usuario no tiene acceso
                    return {
                        "total": 0, 
                        "predios": [], 
                        "sin_acceso": True,
                        "mensaje": f"No tiene acceso al municipio {municipio_predio}"
                    }
        
        # Si se especifica un municipio, verificar que esté en los asignados
        if municipio:
            if municipio not in municipios_asignados:
                raise HTTPException(status_code=403, detail=f"No tiene acceso al municipio {municipio}")
            query["municipio"] = {"$regex": f"^{municipio}$", "$options": "i"}
        else:
            # Si no se especifica municipio, filtrar por todos los asignados
            query["municipio"] = {"$in": municipios_asignados}
    elif municipio:
        # Búsqueda case-insensitive para soportar variaciones de acentos
        query["municipio"] = {"$regex": f"^{municipio}$", "$options": "i"}
    
    if vigencia:
        query["vigencia"] = vigencia
    if destino_economico:
        query["destino_economico"] = destino_economico
    if zona:
        if zona == 'rural':
            query["zona"] = "00"
        elif zona == 'urbano':
            query["zona"] = "01"
        elif zona == 'corregimiento':
            query["zona"] = {"$gte": "02", "$lte": "99"}
    
    # Filtro de geometría - convertir string a booleano
    geometria_filter = None
    if tiene_geometria is not None:
        if tiene_geometria.lower() == 'true':
            query["tiene_geometria"] = True
        elif tiene_geometria.lower() == 'false':
            geometria_filter = {
                "$or": [
                    {"tiene_geometria": False},
                    {"tiene_geometria": {"$exists": False}}
                ]
            }
    
    # Filtro de búsqueda
    search_filter = None
    if search:
        # Detectar si es una búsqueda de matrícula (formato XXX-XXXXX)
        is_matricula_search = bool(re.match(r'^\d{3}-\d+$', search.strip()))
        
        if is_matricula_search:
            # Búsqueda EXACTA para matrículas
            search_filter = {
                "r2_registros.matricula_inmobiliaria": search.strip()
            }
        else:
            # Búsqueda parcial para otros campos
            search_filter = {
                "$or": [
                    {"codigo_predial_nacional": {"$regex": search, "$options": "i"}},
                    {"codigo_homologado": {"$regex": search, "$options": "i"}},
                    {"propietarios.nombre_propietario": {"$regex": search, "$options": "i"}},
                    {"propietarios.numero_documento": {"$regex": search, "$options": "i"}},
                    {"direccion": {"$regex": search, "$options": "i"}},
                    {"r2_registros.matricula_inmobiliaria": {"$regex": search, "$options": "i"}}
                ]
            }
    
    # Combinar filtros usando $and si ambos existen
    if geometria_filter and search_filter:
        query["$and"] = [geometria_filter, search_filter]
    elif geometria_filter:
        query.update(geometria_filter)
    elif search_filter:
        query.update(search_filter)
    
    total = await db.predios.count_documents(query)
    
    # Usar agregación para extraer zona del código predial y ordenar
    # La zona está en las posiciones 6-7 (índice 5-7) del codigo_predial_nacional
    pipeline = [
        {"$match": query},
        {"$addFields": {
            "zona_orden": {"$substr": ["$codigo_predial_nacional", 5, 2]}
        }},
        {"$sort": {"zona_orden": 1, "codigo_predial_nacional": 1}},  # Ordenar por zona ascendente, luego por código
        {"$skip": skip},
        {"$limit": limit},
        {"$project": {"_id": 0, "zona_orden": 0}}  # Excluir campos auxiliares
    ]
    
    predios = await db.predios.aggregate(pipeline).to_list(limit)
    
    return {
        "total": total,
        "predios": predios
    }

@api_router.get("/predios/stats/summary")
async def get_predios_stats(current_user: dict = Depends(get_current_user)):
    """Obtiene estadísticas de predios - SOLO la vigencia más alta GLOBAL del sistema"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Función para extraer el año de una vigencia
    def get_year(vig):
        vig_str = str(vig)
        if len(vig_str) >= 7:
            return int(vig_str[-4:])
        return int(vig_str)
    
    # Obtener TODAS las vigencias disponibles en el sistema
    all_vigencias = await db.predios.distinct("vigencia", {"deleted": {"$ne": True}})
    
    if not all_vigencias:
        return {
            "total_predios": 0,
            "total_avaluo": 0,
            "total_area_terreno": 0,
            "by_municipio": [],
            "by_destino": [],
            "vigencia_actual": None
        }
    
    # Encontrar la vigencia más alta (año más reciente) GLOBALMENTE
    vigencia_mas_alta = max(all_vigencias, key=lambda x: get_year(x))
    vigencia_year = get_year(vigencia_mas_alta)
    
    # Filtrar solo predios de la vigencia más alta
    pipeline_municipios = [
        {"$match": {"vigencia": vigencia_mas_alta, "deleted": {"$ne": True}}},
        {"$group": {"_id": "$municipio", "count": {"$sum": 1}, "avaluo": {"$sum": "$avaluo"}, "area": {"$sum": "$area_terreno"}}},
        {"$sort": {"count": -1}}
    ]
    
    municipios_result = await db.predios.aggregate(pipeline_municipios).to_list(100)
    
    by_municipio = []
    total_predios = 0
    total_avaluo = 0
    total_area = 0
    
    for r in municipios_result:
        by_municipio.append({
            "municipio": r['_id'], 
            "count": r['count'],
            "vigencia": vigencia_mas_alta,
            "vigencia_display": str(vigencia_year)
        })
        total_predios += r['count']
        total_avaluo += r.get('avaluo', 0) or 0
        total_area += r.get('area', 0) or 0
    
    # Por destino económico (solo vigencia más alta)
    pipeline_destino = [
        {"$match": {"vigencia": vigencia_mas_alta, "deleted": {"$ne": True}}},
        {"$group": {"_id": "$destino_economico", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    destinos_result = await db.predios.aggregate(pipeline_destino).to_list(20)
    
    by_destino = [
        {"destino": d["_id"], "nombre": DESTINO_ECONOMICO.get(d["_id"], "Desconocido"), "count": d["count"]}
        for d in destinos_result
    ]
    
    # Conteo de registros R2 y área de GDB
    pipeline_r2_gdb = [
        {"$match": {"vigencia": vigencia_mas_alta, "deleted": {"$ne": True}}},
        {"$project": {
            "has_r2": {"$cond": [{"$gt": [{"$size": {"$ifNull": ["$r2_registros", []]}}, 0]}, 1, 0]},
            "r2_count": {"$size": {"$ifNull": ["$r2_registros", []]}},
            "area_gdb": {"$ifNull": ["$area_gdb", 0]},
            "tiene_geometria": 1
        }},
        {"$group": {
            "_id": None,
            "total_con_r2": {"$sum": "$has_r2"},
            "total_registros_r2": {"$sum": "$r2_count"},
            "total_area_gdb": {"$sum": "$area_gdb"},
            "total_con_geometria": {"$sum": {"$cond": ["$tiene_geometria", 1, 0]}}
        }}
    ]
    r2_gdb_result = await db.predios.aggregate(pipeline_r2_gdb).to_list(1)
    r2_gdb_stats = r2_gdb_result[0] if r2_gdb_result else {}
    
    return {
        "total_predios": total_predios,
        "total_avaluo": total_avaluo,
        "total_area_terreno": total_area,
        "total_registros_r2": r2_gdb_stats.get("total_registros_r2", 0),
        "total_con_r2": r2_gdb_stats.get("total_con_r2", 0),
        "total_area_gdb": r2_gdb_stats.get("total_area_gdb", 0),
        "total_con_geometria": r2_gdb_stats.get("total_con_geometria", 0),
        "by_municipio": by_municipio,
        "by_destino": by_destino[:20],
        "vigencia_actual": vigencia_year
    }

@api_router.get("/predios/eliminados")
async def get_predios_eliminados(
    municipio: Optional[str] = None,
    vigencia: Optional[int] = None,
    skip: int = 0,
    limit: int = 10000,  # Aumentado de 50 a 10000
    current_user: dict = Depends(get_current_user)
):
    """Lista predios eliminados - filtrable por municipio y vigencia"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {}
    if municipio:
        query["municipio"] = municipio
    if vigencia:
        query["vigencia_eliminacion"] = vigencia
    
    total = await db.predios_eliminados.count_documents(query)
    predios = await db.predios_eliminados.find(query, {"_id": 0}).sort("eliminado_en", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "total": total,
        "predios": predios
    }

@api_router.get("/predios/eliminados/stats")
async def get_predios_eliminados_stats(current_user: dict = Depends(get_current_user)):
    """Obtiene estadísticas de predios eliminados por municipio"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    pipeline = [
        {"$group": {"_id": {"municipio": "$municipio", "vigencia": "$vigencia_eliminacion"}, "count": {"$sum": 1}}},
        {"$sort": {"_id.municipio": 1}}
    ]
    
    result = await db.predios_eliminados.aggregate(pipeline).to_list(100)
    
    return {
        "by_municipio": [
            {"municipio": r["_id"]["municipio"], "vigencia": r["_id"]["vigencia"], "count": r["count"]}
            for r in result
        ],
        "total": sum(r["count"] for r in result)
    }


@api_router.post("/predios/analisis-historico")
async def analisis_historico_predios(
    current_user: dict = Depends(get_current_user)
):
    """Analiza todas las vigencias desde 2022 para detectar predios eliminados y reapariciones"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden ejecutar análisis histórico")
    
    # Obtener todas las vigencias ordenadas
    vigencias = await db.predios.distinct("vigencia")
    vigencias = sorted([v for v in vigencias if v >= 2022])
    
    if len(vigencias) < 2:
        return {"message": "Se necesitan al menos 2 vigencias para comparar", "vigencias": vigencias}
    
    # Obtener todos los municipios
    municipios = await db.predios.distinct("municipio")
    
    resultados = {
        "total_eliminados": 0,
        "total_reapariciones": 0,
        "por_municipio": [],
        "reapariciones": [],
        "vigencias_analizadas": vigencias
    }
    
    for municipio in municipios:
        eliminados_mun = 0
        reapariciones_mun = []
        
        # Obtener todos los códigos por vigencia para este municipio
        codigos_por_vigencia = {}
        for vig in vigencias:
            predios = await db.predios.find(
                {"municipio": municipio, "vigencia": vig},
                {"_id": 0, "codigo_predial_nacional": 1}
            ).to_list(50000)
            codigos_por_vigencia[vig] = {p['codigo_predial_nacional'] for p in predios}
        
        # Track de códigos eliminados históricamente
        codigos_eliminados_historico = set()
        
        # Comparar vigencias consecutivas
        for i in range(len(vigencias) - 1):
            vig_anterior = vigencias[i]
            vig_siguiente = vigencias[i + 1]
            
            codigos_anterior = codigos_por_vigencia[vig_anterior]
            codigos_siguiente = codigos_por_vigencia[vig_siguiente]
            
            # Predios eliminados (estaban antes, no están ahora)
            eliminados = codigos_anterior - codigos_siguiente
            
            # Verificar si ya están registrados como eliminados
            for codigo in eliminados:
                existe = await db.predios_eliminados.find_one({"codigo_predial_nacional": codigo, "municipio": municipio})
                if not existe:
                    # Obtener datos del predio
                    predio = await db.predios.find_one(
                        {"municipio": municipio, "vigencia": vig_anterior, "codigo_predial_nacional": codigo},
                        {"_id": 0}
                    )
                    if predio:
                        await db.predios_eliminados.insert_one({
                            **predio,
                            "id": str(uuid.uuid4()),
                            "eliminado_en": datetime.now(timezone.utc).isoformat(),
                            "vigencia_eliminacion": vig_siguiente,
                            "vigencia_origen": vig_anterior,
                            "motivo": f"No incluido en vigencia {vig_siguiente}",
                            "detectado_por": "análisis histórico"
                        })
                        eliminados_mun += 1
                        codigos_eliminados_historico.add(codigo)
            
            # Detectar reapariciones (predios eliminados que vuelven a aparecer)
            for codigo in codigos_siguiente:
                if codigo in codigos_eliminados_historico:
                    reapariciones_mun.append({
                        "codigo_predial_nacional": codigo,
                        "municipio": municipio,
                        "vigencia_reaparicion": vig_siguiente,
                        "mensaje": f"ALERTA: Predio eliminado reaparece en vigencia {vig_siguiente}"
                    })
        
        resultados["por_municipio"].append({
            "municipio": municipio,
            "eliminados_detectados": eliminados_mun,
            "reapariciones": len(reapariciones_mun)
        })
        resultados["total_eliminados"] += eliminados_mun
        resultados["reapariciones"].extend(reapariciones_mun)
        resultados["total_reapariciones"] += len(reapariciones_mun)
    
    return resultados


@api_router.get("/predios/reapariciones/conteo-por-municipio")
async def get_conteo_reapariciones_por_municipio(
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el conteo de reapariciones pendientes por municipio"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Obtener todos los códigos eliminados
    eliminados = await db.predios_eliminados.find(
        {}, 
        {"_id": 0, "codigo_predial_nacional": 1, "municipio": 1, "vigencia_eliminacion": 1}
    ).to_list(50000)
    
    reapariciones_por_municipio = {}
    
    for elim in eliminados:
        codigo = elim["codigo_predial_nacional"]
        mun = elim["municipio"]
        vig_elim = elim.get("vigencia_eliminacion", 0)
        
        # Verificar si ya fue aprobado/rechazado
        decision = await db.predios_reapariciones_aprobadas.find_one({
            "codigo_predial_nacional": codigo,
            "municipio": mun
        })
        
        if decision:
            continue
        
        # Buscar si existe en vigencias posteriores (reaparición)
        existe = await db.predios.find_one({
            "codigo_predial_nacional": codigo,
            "municipio": mun,
            "vigencia": {"$gt": vig_elim}
        })
        
        if existe:
            if mun not in reapariciones_por_municipio:
                reapariciones_por_municipio[mun] = 0
            reapariciones_por_municipio[mun] += 1
    
    return {
        "conteo": reapariciones_por_municipio,
        "total": sum(reapariciones_por_municipio.values())
    }


@api_router.post("/predios/reapariciones/solicitar")
async def solicitar_reaparicion(
    codigo_predial: str = Form(...),
    municipio: str = Form(...),
    justificacion: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Permite al gestor solicitar la reaparición de un predio eliminado para aprobación del coordinador"""
    if current_user['role'] not in [UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para solicitar reapariciones")
    
    # Verificar que el predio esté en eliminados
    eliminado = await db.predios_eliminados.find_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    })
    
    if not eliminado:
        raise HTTPException(status_code=404, detail="Predio no encontrado en la lista de eliminados")
    
    # Verificar si ya existe una solicitud pendiente
    solicitud_existente = await db.predios_reapariciones_solicitudes.find_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio,
        "estado": "pendiente"
    })
    
    if solicitud_existente:
        raise HTTPException(status_code=400, detail="Ya existe una solicitud pendiente para este predio")
    
    # Crear solicitud
    solicitud = {
        "id": str(uuid.uuid4()),
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio,
        "vigencia_eliminacion": eliminado.get("vigencia_eliminacion"),
        "vigencia_origen": eliminado.get("vigencia_origen"),
        "estado": "pendiente",
        "justificacion_gestor": justificacion,
        "solicitado_por": current_user['id'],
        "solicitado_por_nombre": current_user['full_name'],
        "fecha_solicitud": datetime.now(timezone.utc).isoformat(),
        "datos_predio_eliminado": {
            "propietarios": eliminado.get("propietarios", []),
            "direccion": eliminado.get("direccion"),
            "avaluo": eliminado.get("avaluo"),
            "area_terreno": eliminado.get("area_terreno")
        }
    }
    
    await db.predios_reapariciones_solicitudes.insert_one(solicitud)
    
    # Notificar a coordinadores
    coordinadores = await db.users.find(
        {"role": {"$in": [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]}},
        {"_id": 0, "id": 1, "full_name": 1}
    ).to_list(20)
    
    for coord in coordinadores:
        await crear_notificacion(
            usuario_id=coord['id'],
            titulo=f"Solicitud de Reaparición - {municipio}",
            mensaje=f"{current_user['full_name']} solicita aprobar la reaparición del predio {codigo_predial}. Justificación: {justificacion[:100]}...",
            tipo="warning",
            enviar_email=True
        )
    
    return {
        "message": "Solicitud de reaparición enviada al coordinador",
        "solicitud_id": solicitud["id"],
        "estado": "pendiente"
    }


@api_router.get("/predios/reapariciones/solicitudes-pendientes")
async def get_solicitudes_reaparicion_pendientes(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las solicitudes de reaparición pendientes de aprobación por el coordinador"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden ver solicitudes pendientes")
    
    query = {"estado": "pendiente"}
    if municipio:
        query["municipio"] = municipio
    
    solicitudes = await db.predios_reapariciones_solicitudes.find(query, {"_id": 0}).sort("fecha_solicitud", -1).to_list(500)
    
    return {
        "total": len(solicitudes),
        "solicitudes": solicitudes
    }


@api_router.get("/predios/verificar-codigo-eliminado/{codigo}")
async def verificar_codigo_eliminado(
    codigo: str,
    current_user: dict = Depends(get_current_user)
):
    """Verifica si un código predial está en la lista de eliminados"""
    eliminado = await db.predios_eliminados.find_one(
        {"codigo_predial_nacional": codigo},
        {"_id": 0}
    )
    
    # Verificar si está aprobado para reaparecer
    aprobacion = await db.predios_reapariciones_aprobadas.find_one(
        {"codigo_predial_nacional": codigo, "estado": "aprobado"},
        {"_id": 0}
    )
    
    # Verificar si hay solicitud pendiente
    solicitud_pendiente = await db.predios_reapariciones_solicitudes.find_one(
        {"codigo_predial_nacional": codigo, "estado": "pendiente"},
        {"_id": 0}
    )
    
    if eliminado and not aprobacion:
        return {
            "eliminado": True,
            "tiene_solicitud_pendiente": solicitud_pendiente is not None,
            "puede_solicitar_reaparicion": solicitud_pendiente is None,
            "mensaje": "Este código predial fue eliminado y NO puede ser reutilizado sin aprobación",
            "detalles": {
                "municipio": eliminado.get("municipio"),
                "vigencia_origen": eliminado.get("vigencia_origen"),
                "vigencia_eliminacion": eliminado.get("vigencia_eliminacion"),
                "fecha_eliminacion": eliminado.get("eliminado_en"),
                "motivo": eliminado.get("motivo")
            }
        }
    
    if eliminado and aprobacion:
        return {
            "eliminado": False,
            "mensaje": "Este código predial fue eliminado pero su reaparición fue APROBADA",
            "aprobacion": {
                "aprobado_por": aprobacion.get("aprobado_por_nombre"),
                "fecha_aprobacion": aprobacion.get("fecha_aprobacion"),
                "justificacion": aprobacion.get("justificacion")
            }
        }
    
    return {
        "eliminado": False,
        "mensaje": "Este código predial NO está en la lista de eliminados"
    }


@api_router.get("/predios/estructura-codigo/{municipio}")
async def get_estructura_codigo_predial(
    municipio: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Retorna la estructura del código predial nacional para un municipio.
    Incluye los primeros 5 dígitos fijos (departamento + municipio).
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    divipola = MUNICIPIOS_DIVIPOLA.get(municipio)
    if not divipola:
        raise HTTPException(status_code=404, detail=f"Municipio {municipio} no encontrado")
    
    # Estructura del código predial nacional (30 dígitos)
    estructura = {
        "departamento": {"posicion": "1-2", "valor": divipola["departamento"], "editable": False, "descripcion": "Departamento"},
        "municipio": {"posicion": "3-5", "valor": divipola["municipio"], "editable": False, "descripcion": "Municipio"},
        "zona": {"posicion": "6-7", "valor": "", "editable": True, "descripcion": "Zona (00=Rural, 01=Urbano, 02-99=Corregimientos)"},
        "sector": {"posicion": "8-9", "valor": "", "editable": True, "descripcion": "Sector"},
        "comuna": {"posicion": "10-11", "valor": "", "editable": True, "descripcion": "Comuna"},
        "barrio": {"posicion": "12-13", "valor": "", "editable": True, "descripcion": "Barrio"},
        "manzana_vereda": {"posicion": "14-17", "valor": "", "editable": True, "descripcion": "Vereda o Manzana"},
        "terreno": {"posicion": "18-21", "valor": "", "editable": True, "descripcion": "Condición del Predio (Terreno)"},
        "edificio": {"posicion": "22", "valor": "0", "editable": True, "descripcion": "No. Edificio o Torre"},
        "piso": {"posicion": "23-26", "valor": "0000", "editable": True, "descripcion": "No. del Piso"},
        "unidad": {"posicion": "27-30", "valor": "0000", "editable": True, "descripcion": "No. Unidad PH/Mejora"}
    }
    
    prefijo = divipola["departamento"] + divipola["municipio"]
    
    return {
        "municipio": municipio,
        "prefijo_fijo": prefijo,
        "estructura": estructura,
        "total_digitos": 30
    }


@api_router.get("/predios/ultima-manzana/{municipio}")
async def get_ultima_manzana_sector(
    municipio: str,
    zona: str = "00",
    sector: str = "00",
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene la última manzana (número más alto) registrada para un sector específico.
    Útil para guiar al usuario al crear nuevos predios.
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    divipola = MUNICIPIOS_DIVIPOLA.get(municipio)
    if not divipola:
        raise HTTPException(status_code=404, detail=f"Municipio {municipio} no encontrado")
    
    # Construir prefijo para buscar: departamento + municipio + zona + sector
    prefijo_sector = f"{divipola['departamento']}{divipola['municipio']}{zona}{sector}"
    
    # Buscar todos los predios en este sector
    regex_pattern = f"^{prefijo_sector}"
    predios_en_sector = await db.predios.find(
        {"codigo_predial_nacional": {"$regex": regex_pattern}},
        {"_id": 0, "codigo_predial_nacional": 1}
    ).to_list(50000)
    
    if not predios_en_sector:
        return {
            "municipio": municipio,
            "zona": zona,
            "sector": sector,
            "ultima_manzana": None,
            "total_predios_sector": 0,
            "mensaje": f"No hay predios registrados en la zona {zona}, sector {sector}"
        }
    
    # Extraer las manzanas (posiciones 14-17, es decir índices 13:17)
    manzanas = set()
    for p in predios_en_sector:
        codigo = p["codigo_predial_nacional"]
        if len(codigo) >= 17:
            # Las posiciones van: 
            # 0-1: depto, 2-4: mpio, 5-6: zona, 7-8: sector, 9-10: comuna, 11-12: barrio, 13-16: manzana
            manzana = codigo[13:17]
            manzanas.add(manzana)
    
    # Encontrar la manzana más alta (máximo numérico)
    manzanas_numericas = []
    for m in manzanas:
        try:
            manzanas_numericas.append(int(m))
        except ValueError:
            continue
    
    if not manzanas_numericas:
        return {
            "municipio": municipio,
            "zona": zona,
            "sector": sector,
            "ultima_manzana": None,
            "total_predios_sector": len(predios_en_sector),
            "mensaje": "No se encontraron manzanas válidas"
        }
    
    ultima_manzana = str(max(manzanas_numericas)).zfill(4)
    
    return {
        "municipio": municipio,
        "zona": zona,
        "sector": sector,
        "ultima_manzana": ultima_manzana,
        "total_manzanas": len(manzanas_numericas),
        "total_predios_sector": len(predios_en_sector),
        "mensaje": f"Última manzana registrada: {ultima_manzana}"
    }


@api_router.get("/predios/por-manzana/{municipio}")
async def get_predios_por_manzana(
    municipio: str,
    zona: str = "00",
    sector: str = "00",
    comuna: str = "00",
    barrio: str = "00",
    manzana_vereda: str = "0000",
    limit: int = 5,
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene los últimos terrenos únicos en una manzana específica.
    Agrupa por número de terreno y cuenta registros por cada uno.
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    divipola = MUNICIPIOS_DIVIPOLA.get(municipio)
    if not divipola:
        raise HTTPException(status_code=404, detail=f"Municipio {municipio} no encontrado")
    
    # Construir prefijo para buscar: departamento + municipio + zona + sector + comuna + barrio + manzana
    prefijo_manzana = f"{divipola['departamento']}{divipola['municipio']}{zona}{sector}{comuna}{barrio}{manzana_vereda}"
    
    # Usar aggregation para agrupar por terreno (posiciones 18-21 del código)
    pipeline = [
        {"$match": {"codigo_predial_nacional": {"$regex": f"^{prefijo_manzana}"}}},
        {"$addFields": {
            "terreno": {"$substr": ["$codigo_predial_nacional", 17, 4]}
        }},
        {"$group": {
            "_id": "$terreno",
            "count": {"$sum": 1},
            "direccion": {"$first": "$direccion"},
            "area_terreno": {"$first": "$area_terreno"},
            "codigo_ejemplo": {"$first": "$codigo_predial_nacional"}
        }},
        {"$sort": {"_id": -1}},  # Ordenar descendente para obtener los últimos
        {"$limit": limit}
    ]
    
    terrenos = await db.predios.aggregate(pipeline).to_list(limit)
    
    # Reordenar de menor a mayor para mostrar
    terrenos.sort(key=lambda x: x["_id"])
    
    # Calcular siguiente terreno sugerido
    siguiente_terreno = "0001"
    if terrenos:
        ultimo_terreno = max(t["_id"] for t in terrenos)
        try:
            siguiente_num = int(ultimo_terreno) + 1
            siguiente_terreno = str(siguiente_num).zfill(4)
        except:
            pass
    
    # Formatear respuesta
    predios_formateados = []
    for t in terrenos:
        predios_formateados.append({
            "terreno": t["_id"],
            "direccion": t.get("direccion", "Sin dirección") or "Sin dirección",
            "area_terreno": t.get("area_terreno"),
            "registros": t["count"]
        })
    
    return {
        "municipio": municipio,
        "manzana": manzana_vereda,
        "total_terrenos": len(terrenos),
        "siguiente_terreno": siguiente_terreno,
        "predios": predios_formateados
    }


@api_router.get("/predios/sugerir-codigo/{municipio}")
async def sugerir_codigo_disponible(
    municipio: str,
    zona: str = "00",
    sector: str = "00",
    comuna: str = "00",
    barrio: str = "00",
    manzana_vereda: str = "0000",
    current_user: dict = Depends(get_current_user)
):
    """
    Sugiere el próximo código de terreno disponible para una manzana/vereda específica.
    También verifica si hay geometría GDB disponible.
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    divipola = MUNICIPIOS_DIVIPOLA.get(municipio)
    if not divipola:
        raise HTTPException(status_code=404, detail=f"Municipio {municipio} no encontrado")
    
    # Construir prefijo base (primeros 17 dígitos)
    prefijo_base = f"{divipola['departamento']}{divipola['municipio']}{zona}{sector}{comuna}{barrio}{manzana_vereda}"
    
    # Buscar predios existentes en esta manzana
    regex_pattern = f"^{prefijo_base}"
    predios_existentes = await db.predios.find(
        {"codigo_predial_nacional": {"$regex": regex_pattern}},
        {"_id": 0, "codigo_predial_nacional": 1}
    ).to_list(10000)
    
    # Buscar predios eliminados en esta manzana
    predios_eliminados = await db.predios_eliminados.find(
        {"codigo_predial_nacional": {"$regex": regex_pattern}},
        {"_id": 0, "codigo_predial_nacional": 1, "vigencia_eliminacion": 1}
    ).to_list(10000)
    
    # Extraer los números de terreno usados (posiciones 18-21)
    terrenos_usados = set()
    for p in predios_existentes:
        codigo = p["codigo_predial_nacional"]
        if len(codigo) >= 21:
            terreno = codigo[17:21]
            terrenos_usados.add(terreno)
    
    terrenos_eliminados = []
    for p in predios_eliminados:
        codigo = p["codigo_predial_nacional"]
        if len(codigo) >= 21:
            terreno = codigo[17:21]
            terrenos_eliminados.append({
                "numero": terreno,
                "codigo_completo": codigo,
                "vigencia_eliminacion": p.get("vigencia_eliminacion")
            })
            terrenos_usados.add(terreno)  # También marcar como usados
    
    # Encontrar el siguiente terreno disponible
    siguiente_terreno = "0001"
    for i in range(1, 10000):
        candidato = str(i).zfill(4)
        if candidato not in terrenos_usados:
            siguiente_terreno = candidato
            break
    
    # Construir código sugerido completo
    codigo_sugerido = f"{prefijo_base}{siguiente_terreno}000000000"
    
    # Verificar si hay geometría GDB disponible para el código sugerido específico
    # Buscar geometría que coincida con el código sugerido (21 primeros dígitos: prefijo + terreno)
    codigo_terreno = f"{prefijo_base}{siguiente_terreno}"  # 21 dígitos
    geometria_codigo_exacto = await db.gdb_geometrias.find_one(
        {"municipio": municipio, "codigo": {"$regex": f"^{codigo_terreno}"}},
        {"_id": 0, "codigo": 1, "area_gdb": 1}
    )
    
    # También verificar si hay alguna geometría en la manzana (para información)
    geometrias_en_manzana = await db.gdb_geometrias.count_documents(
        {"municipio": municipio, "codigo": {"$regex": f"^{prefijo_base}"}}
    )
    
    # Construir mensaje según disponibilidad
    if geometria_codigo_exacto:
        area_info = geometria_codigo_exacto.get('area_gdb')
        if area_info:
            mensaje_geo = f"✅ Geometría GDB disponible (área: {area_info:,.2f} m²)"
        else:
            mensaje_geo = "✅ Geometría GDB disponible"
    else:
        if geometrias_en_manzana > 0:
            mensaje_geo = f"⚠️ Sin geometría para este código. La manzana tiene {geometrias_en_manzana} geometrías registradas."
        else:
            mensaje_geo = "⚠️ No hay información gráfica (GDB) para esta zona. Se relacionará cuando se cargue el GDB."
    
    return {
        "prefijo_base": prefijo_base,
        "total_activos": len(predios_existentes),
        "terrenos_usados": list(terrenos_usados)[:20],  # Limitar para la respuesta
        "terrenos_eliminados": terrenos_eliminados,
        "siguiente_terreno": siguiente_terreno,
        "codigo_sugerido": codigo_sugerido,
        "tiene_geometria_gdb": geometria_codigo_exacto is not None,
        "geometrias_en_manzana": geometrias_en_manzana,
        "mensaje_geometria": mensaje_geo
    }


@api_router.get("/predios/verificar-codigo-completo/{codigo}")
async def verificar_codigo_completo(
    codigo: str,
    municipio: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Verifica un código predial completo de 30 dígitos:
    - Si ya existe activo
    - Si está eliminado (y ofrece reactivar)
    - Si está disponible
    - Si tiene geometría GDB
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    if len(codigo) != 30:
        raise HTTPException(status_code=400, detail="El código predial debe tener exactamente 30 dígitos")
    
    # Verificar si ya existe activo
    predio_existente = await db.predios.find_one(
        {"codigo_predial_nacional": codigo},
        {"_id": 0, "id": 1, "municipio": 1, "nombre_propietario": 1, "estado": 1}
    )
    
    if predio_existente:
        return {
            "estado": "existente",
            "disponible": False,
            "mensaje": "Este código predial ya está registrado en la base de datos",
            "predio": predio_existente
        }
    
    # Verificar si está eliminado
    eliminado = await db.predios_eliminados.find_one(
        {"codigo_predial_nacional": codigo},
        {"_id": 0}
    )
    
    # Verificar si tiene geometría GDB
    geometria = await db.gdb_geometrias.find_one(
        {"codigo": codigo[:21], "municipio": municipio},  # Los primeros 21 caracteres para terreno
        {"_id": 0, "area_m2": 1}
    )
    
    if eliminado:
        # Verificar si ya tiene aprobación de reaparición
        aprobacion = await db.predios_reapariciones_aprobadas.find_one(
            {"codigo_predial_nacional": codigo, "estado": "aprobado"}
        )
        
        return {
            "estado": "eliminado",
            "disponible": aprobacion is not None,
            "puede_reactivar": True,
            "mensaje": "⚠️ Este código pertenece a un predio ELIMINADO. ¿Desea reactivarlo?",
            "detalles_eliminacion": {
                "municipio": eliminado.get("municipio"),
                "vigencia_origen": eliminado.get("vigencia_origen"),
                "vigencia_eliminacion": eliminado.get("vigencia_eliminacion"),
                "motivo": eliminado.get("motivo", "Mutación catastral")
            },
            "aprobacion_existente": aprobacion is not None,
            "tiene_geometria": geometria is not None,
            "area_gdb": geometria.get("area_m2") if geometria else None
        }
    
    # Código disponible
    return {
        "estado": "disponible",
        "disponible": True,
        "mensaje": "✅ Este código predial está disponible para usar",
        "tiene_geometria": geometria is not None,
        "area_gdb": geometria.get("area_m2") if geometria else None,
        "mensaje_geometria": "Tiene información gráfica (GDB)" if geometria else "⚠️ Sin información gráfica. Se relacionará cuando se cargue el GDB."
    }


@api_router.post("/predios/crear-con-workflow")
async def crear_predio_con_workflow(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Crea un nuevo predio con el flujo de aprobación completo.
    Solo staff puede crear predios.
    El predio queda pendiente de aprobación del coordinador.
    """
    # Solo staff puede crear predios
    if current_user['role'] not in [UserRole.ATENCION_USUARIO, UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo usuarios staff pueden crear predios")
    
    codigo_predial = request.get("codigo_predial_nacional")
    municipio = request.get("municipio")
    es_reactivacion = request.get("es_reactivacion", False)
    justificacion = request.get("justificacion", "Creación de nuevo predio")
    
    if not codigo_predial or len(codigo_predial) != 30:
        raise HTTPException(status_code=400, detail="El código predial debe tener exactamente 30 dígitos")
    
    # Verificar que el municipio coincida con el código
    divipola = MUNICIPIOS_DIVIPOLA.get(municipio)
    if not divipola:
        raise HTTPException(status_code=400, detail=f"Municipio {municipio} no válido")
    
    prefijo_esperado = divipola["departamento"] + divipola["municipio"]
    if not codigo_predial.startswith(prefijo_esperado):
        raise HTTPException(status_code=400, detail=f"El código no corresponde al municipio {municipio}")
    
    # Verificar si ya existe
    existente = await db.predios.find_one({"codigo_predial_nacional": codigo_predial})
    if existente:
        raise HTTPException(status_code=400, detail="Este código predial ya existe")
    
    # Si es reactivación, verificar que esté eliminado
    if es_reactivacion:
        eliminado = await db.predios_eliminados.find_one({"codigo_predial_nacional": codigo_predial})
        if not eliminado:
            raise HTTPException(status_code=400, detail="Este código no está en la lista de eliminados")
    
    # Verificar geometría GDB
    geometria = await db.gdb_geometrias.find_one(
        {"codigo": codigo_predial[:21], "municipio": municipio},
        {"_id": 0, "area_m2": 1, "geometry": 1}
    )
    
    # Crear propuesta de cambio
    cambio_id = str(uuid.uuid4())
    propuesta = {
        "id": cambio_id,
        "tipo_cambio": "creacion",
        "municipio": municipio,
        "codigo_predial_nacional": codigo_predial,
        "es_reactivacion": es_reactivacion,
        "datos_propuestos": {
            "codigo_predial_nacional": codigo_predial,
            "municipio": municipio,
            "nombre_propietario": request.get("nombre_propietario", ""),
            "tipo_documento": request.get("tipo_documento", "C"),
            "numero_documento": request.get("numero_documento", ""),
            "direccion": request.get("direccion", ""),
            "destino_economico": request.get("destino_economico", "D"),
            "area_terreno": request.get("area_terreno", 0),
            "area_construida": request.get("area_construida", 0),
            "avaluo": request.get("avaluo", 0),
            "zona": codigo_predial[5:7],
            "sector": codigo_predial[7:9],
            "comuna": codigo_predial[9:11],
            "barrio": codigo_predial[11:13],
            "manzana_vereda": codigo_predial[13:17],
            "condicion_predio": codigo_predial[17:21],
            "predio_horizontal": codigo_predial[21:30],
            "tiene_geometria_gdb": geometria is not None,
            "area_gdb": geometria.get("area_m2") if geometria else None,
            "propietarios": request.get("propietarios", []),
            # Soporte para nuevo formato R2 separado (zonas + construcciones)
            "zonas": request.get("zonas", []),
            "construcciones": request.get("construcciones", []),
            # Campo legacy para compatibilidad
            "zonas_fisicas": request.get("zonas_fisicas", []),
            "matricula_inmobiliaria": request.get("matricula_inmobiliaria", "")
        },
        "justificacion": justificacion,
        "estado": "pendiente_aprobacion",
        "creado_por": current_user["id"],
        "creado_por_nombre": current_user["full_name"],
        "creado_por_rol": current_user["role"],
        "gestor_asignado": None,
        "gestor_asignado_nombre": None,
        "gestor_continuar": None,
        "gestor_continuar_nombre": None,
        "historial": [{
            "accion": "Creación de propuesta",
            "usuario": current_user["full_name"],
            "usuario_id": current_user["id"],
            "rol": current_user["role"],
            "fecha": datetime.now(timezone.utc).isoformat(),
            "notas": justificacion
        }],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Si se asignó a otro gestor para continuar
    gestor_asignado_id = request.get("gestor_asignado_id")
    if gestor_asignado_id:
        gestor_info = await db.users.find_one({"id": gestor_asignado_id}, {"_id": 0, "full_name": 1})
        if gestor_info:
            propuesta["gestor_continuar"] = gestor_asignado_id
            propuesta["gestor_continuar_nombre"] = gestor_info["full_name"]
            propuesta["estado"] = "en_proceso"
            propuesta["historial"].append({
                "accion": "Asignado a otro gestor para continuar",
                "usuario": current_user["full_name"],
                "usuario_id": current_user["id"],
                "rol": current_user["role"],
                "fecha": datetime.now(timezone.utc).isoformat(),
                "notas": f"Asignado a {gestor_info['full_name']} para continuar el diligenciamiento"
            })
    
    await db.predios_cambios_propuestos.insert_one(propuesta)
    
    # Si es coordinador o admin, puede aprobar directamente
    requiere_aprobacion = current_user["role"] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]
    
    # Actualizar contador de trámites del gestor
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$inc": {"tramites_creados": 1}}
    )
    
    return {
        "cambio_id": cambio_id,
        "requiere_aprobacion": requiere_aprobacion,
        "mensaje": "Predio propuesto. Pendiente de aprobación del coordinador." if requiere_aprobacion else "Predio listo para aprobar.",
        "tiene_geometria": geometria is not None
    }


@api_router.post("/predios/cambios/{cambio_id}/asignar-gestor")
async def asignar_gestor_a_cambio(
    cambio_id: str,
    gestor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Coordinador asigna un gestor para revisar el cambio propuesto.
    """
    if current_user["role"] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden asignar gestores")
    
    cambio = await db.predios_cambios_propuestos.find_one({"id": cambio_id})
    if not cambio:
        raise HTTPException(status_code=404, detail="Cambio no encontrado")
    
    gestor = await db.users.find_one({"id": gestor_id}, {"_id": 0, "full_name": 1, "role": 1})
    if not gestor:
        raise HTTPException(status_code=404, detail="Gestor no encontrado")
    
    # Actualizar cambio
    historial_entry = {
        "accion": "Asignación de gestor para revisión",
        "usuario": current_user["full_name"],
        "usuario_id": current_user["id"],
        "rol": current_user["role"],
        "fecha": datetime.now(timezone.utc).isoformat(),
        "notas": f"Gestor asignado: {gestor['full_name']}"
    }
    
    await db.predios_cambios_propuestos.update_one(
        {"id": cambio_id},
        {
            "$set": {
                "gestor_asignado": gestor_id,
                "gestor_asignado_nombre": gestor["full_name"],
                "estado": "en_revision_gestor",
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$push": {"historial": historial_entry}
        }
    )
    
    return {"mensaje": f"Gestor {gestor['full_name']} asignado para revisar el cambio"}


@api_router.post("/predios/cambios/{cambio_id}/revision-gestor")
async def revision_gestor_cambio(
    cambio_id: str,
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Gestor asignado realiza la revisión y puede aplicar ajustes.
    El cambio queda pendiente de revisión final del coordinador.
    """
    if current_user["role"] not in [UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para revisar cambios")
    
    cambio = await db.predios_cambios_propuestos.find_one({"id": cambio_id})
    if not cambio:
        raise HTTPException(status_code=404, detail="Cambio no encontrado")
    
    # Verificar que sea el gestor asignado (o coordinador/admin)
    if current_user["role"] in [UserRole.GESTOR]:
        if cambio.get("gestor_asignado") != current_user["id"]:
            raise HTTPException(status_code=403, detail="No está asignado a este cambio")
    
    observaciones = request.get("observaciones", "")
    datos_revisados = request.get("datos_revisados", {})
    
    historial_entry = {
        "accion": "Revisión completada por gestor",
        "usuario": current_user["full_name"],
        "usuario_id": current_user["id"],
        "rol": current_user["role"],
        "fecha": datetime.now(timezone.utc).isoformat(),
        "notas": observaciones
    }
    
    update_data = {
        "estado": "pendiente_revision_final",
        "observaciones_gestor": observaciones,
        "fecha_revision_gestor": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if datos_revisados:
        update_data["datos_revisados"] = datos_revisados
    
    await db.predios_cambios_propuestos.update_one(
        {"id": cambio_id},
        {
            "$set": update_data,
            "$push": {"historial": historial_entry}
        }
    )
    
    # Actualizar contador de trámites del gestor
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$inc": {"tramites_revisados": 1}}
    )
    
    return {"mensaje": "Revisión completada. Pendiente de aprobación final del coordinador."}


@api_router.get("/predios/cambios/estadisticas-gestores")
async def get_estadisticas_gestores(
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene estadísticas de trámites por gestor para informes.
    """
    if current_user["role"] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden ver estadísticas")
    
    # Obtener todos los gestores con sus contadores
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR, UserRole.ATENCION_USUARIO]}},
        {"_id": 0, "id": 1, "full_name": 1, "role": 1, "tramites_creados": 1, "tramites_revisados": 1, "tramites_aprobados": 1}
    ).to_list(100)
    
    # Obtener conteo de cambios por gestor desde la colección de cambios
    pipeline = [
        {"$match": {"gestor_asignado": {"$ne": None}}},
        {"$group": {
            "_id": "$gestor_asignado",
            "total_asignados": {"$sum": 1},
            "pendientes": {"$sum": {"$cond": [{"$eq": ["$estado", "pendiente_revision_final"]}, 1, 0]}},
            "aprobados": {"$sum": {"$cond": [{"$eq": ["$estado", "aprobado"]}, 1, 0]}}
        }}
    ]
    
    stats_cambios = await db.predios_cambios_propuestos.aggregate(pipeline).to_list(100)
    stats_dict = {s["_id"]: s for s in stats_cambios}
    
    resultado = []
    for gestor in gestores:
        stats = stats_dict.get(gestor["id"], {})
        resultado.append({
            "id": gestor["id"],
            "nombre": gestor["full_name"],
            "rol": gestor["role"],
            "tramites_creados": gestor.get("tramites_creados", 0),
            "tramites_revisados": gestor.get("tramites_revisados", 0),
            "cambios_asignados": stats.get("total_asignados", 0),
            "cambios_pendientes": stats.get("pendientes", 0),
            "cambios_aprobados": stats.get("aprobados", 0)
        })
    
    return {"gestores": resultado}

@api_router.get("/predios/reapariciones/pendientes")
async def get_reapariciones_pendientes(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene la lista de predios reaparecidos pendientes de aprobación"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden ver reapariciones pendientes")
    
    # Obtener todos los códigos eliminados
    query_elim = {}
    if municipio:
        query_elim["municipio"] = municipio
    
    eliminados = await db.predios_eliminados.find(
        query_elim, 
        {"_id": 0, "codigo_predial_nacional": 1, "municipio": 1, "vigencia_eliminacion": 1, "vigencia_origen": 1}
    ).to_list(50000)
    
    reapariciones_pendientes = []
    
    for elim in eliminados:
        codigo = elim["codigo_predial_nacional"]
        mun = elim["municipio"]
        vig_elim = elim.get("vigencia_eliminacion", 0)
        
        # Verificar si ya fue aprobado/rechazado
        decision = await db.predios_reapariciones_aprobadas.find_one({
            "codigo_predial_nacional": codigo,
            "municipio": mun
        })
        
        if decision:
            continue  # Ya tiene decisión, no está pendiente
        
        # Buscar si este código existe en vigencias posteriores (reaparición)
        predio_actual = await db.predios.find_one({
            "codigo_predial_nacional": codigo,
            "municipio": mun,
            "vigencia": {"$gt": vig_elim}
        }, {"_id": 0})
        
        if predio_actual:
            # Obtener datos del predio eliminado original
            predio_eliminado = await db.predios_eliminados.find_one(
                {"codigo_predial_nacional": codigo, "municipio": mun},
                {"_id": 0}
            )
            
            reapariciones_pendientes.append({
                "codigo_predial_nacional": codigo,
                "municipio": mun,
                "vigencia_eliminacion": vig_elim,
                "vigencia_origen": elim.get("vigencia_origen"),
                "vigencia_reaparicion": predio_actual.get("vigencia"),
                "propietario_anterior": predio_eliminado.get("propietarios", [{}])[0].get("nombre_propietario", "N/A") if predio_eliminado else "N/A",
                "propietario_actual": predio_actual.get("propietarios", [{}])[0].get("nombre_propietario", "N/A") if predio_actual.get("propietarios") else "N/A",
                "direccion": predio_actual.get("direccion", ""),
                "avaluo_anterior": predio_eliminado.get("avaluo", 0) if predio_eliminado else 0,
                "avaluo_actual": predio_actual.get("avaluo", 0),
                "estado": "pendiente"
            })
    
    return {
        "total_pendientes": len(reapariciones_pendientes),
        "reapariciones": reapariciones_pendientes,
        "mensaje": f"Hay {len(reapariciones_pendientes)} reapariciones pendientes de revisión"
    }


@api_router.post("/predios/reapariciones/aprobar")
async def aprobar_reaparicion(
    codigo_predial: str = Form(...),
    municipio: str = Form(...),
    justificacion: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Aprueba la reaparición de un predio eliminado después del análisis técnico"""
    has_permission = await check_permission(current_user, Permission.APPROVE_CHANGES)
    if not has_permission:
        raise HTTPException(status_code=403, detail="No tiene permiso para aprobar reapariciones")
    
    # Verificar que el predio esté en eliminados
    eliminado = await db.predios_eliminados.find_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    })
    
    if not eliminado:
        raise HTTPException(status_code=404, detail="Predio no encontrado en la lista de eliminados")
    
    # Verificar que exista la reaparición
    predio_actual = await db.predios.find_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    })
    
    if not predio_actual:
        raise HTTPException(status_code=404, detail="No se encontró el predio reaparecido en la vigencia actual")
    
    # Registrar aprobación
    aprobacion = {
        "id": str(uuid.uuid4()),
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio,
        "vigencia_eliminacion": eliminado.get("vigencia_eliminacion"),
        "vigencia_reaparicion": predio_actual.get("vigencia"),
        "estado": "aprobado",
        "justificacion": justificacion,
        "aprobado_por": current_user['id'],
        "aprobado_por_nombre": current_user['full_name'],
        "fecha_aprobacion": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios_reapariciones_aprobadas.insert_one(aprobacion)
    
    # Remover _id antes de retornar
    aprobacion.pop("_id", None)
    
    # Opcionalmente, remover de eliminados ya que fue aprobada la reaparición
    await db.predios_eliminados.delete_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    })
    
    return {
        "message": f"Reaparición del predio {codigo_predial} APROBADA",
        "aprobacion": aprobacion
    }


@api_router.post("/predios/reapariciones/rechazar")
async def rechazar_reaparicion(
    codigo_predial: str = Form(...),
    municipio: str = Form(...),
    justificacion: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Rechaza la reaparición - Envía a subsanación al gestor que la solicitó"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden rechazar reapariciones")
    
    # Verificar que el predio esté en eliminados
    eliminado = await db.predios_eliminados.find_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    })
    
    if not eliminado:
        raise HTTPException(status_code=404, detail="Predio no encontrado en la lista de eliminados")
    
    # Obtener el predio actual
    predio_actual = await db.predios.find_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    }, {"_id": 0})
    
    if not predio_actual:
        raise HTTPException(status_code=404, detail="No se encontró el predio reaparecido")
    
    # Buscar si existe una solicitud previa de este predio para saber quién la hizo
    solicitud_previa = await db.predios_reapariciones_solicitudes.find_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    }, sort=[("fecha_solicitud", -1)])
    
    gestor_id = solicitud_previa.get("solicitado_por") if solicitud_previa else None
    gestor_nombre = solicitud_previa.get("solicitado_por_nombre") if solicitud_previa else "Sistema"
    gestor_email = None
    
    # Obtener email del gestor si existe
    if gestor_id:
        gestor = await db.users.find_one({"id": gestor_id})
        if gestor:
            gestor_email = gestor.get("email")
    
    # Crear registro de subsanación pendiente
    subsanacion_id = str(uuid.uuid4())
    subsanacion = {
        "id": subsanacion_id,
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio,
        "vigencia_eliminacion": eliminado.get("vigencia_eliminacion"),
        "vigencia_reaparicion": predio_actual.get("vigencia"),
        "estado": "pendiente_subsanacion",
        "motivo_rechazo": justificacion,
        "rechazado_por": current_user['id'],
        "rechazado_por_nombre": current_user['full_name'],
        "fecha_rechazo": datetime.now(timezone.utc).isoformat(),
        "gestor_asignado": gestor_id,
        "gestor_nombre": gestor_nombre,
        "intentos": 1,
        "historial": [{
            "accion": "rechazado",
            "fecha": datetime.now(timezone.utc).isoformat(),
            "usuario": current_user['full_name'],
            "motivo": justificacion
        }],
        "datos_predio": {
            "direccion": predio_actual.get("direccion", ""),
            "propietarios": predio_actual.get("propietarios", []),
            "avaluo": predio_actual.get("avaluo", 0),
            "area_terreno": predio_actual.get("area_terreno", 0),
            "area_construida": predio_actual.get("area_construida", 0)
        }
    }
    
    await db.reapariciones_subsanacion.insert_one(subsanacion)
    
    # Registrar también en historial de decisiones
    decision = {
        "id": str(uuid.uuid4()),
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio,
        "vigencia_eliminacion": eliminado.get("vigencia_eliminacion"),
        "vigencia_reaparicion": predio_actual.get("vigencia"),
        "estado": "enviado_subsanacion",
        "justificacion": justificacion,
        "rechazado_por": current_user['id'],
        "rechazado_por_nombre": current_user['full_name'],
        "fecha_rechazo": datetime.now(timezone.utc).isoformat(),
        "subsanacion_id": subsanacion_id
    }
    await db.predios_reapariciones_aprobadas.insert_one(decision)
    
    # Enviar notificación por correo al gestor
    if gestor_email:
        try:
            asunto = f"[Asomunicipios] Reaparición Rechazada - Requiere Subsanación"
            cuerpo = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <h2 style="color: #dc2626;">Reaparición Rechazada - Requiere Subsanación</h2>
                <p>Estimado(a) {gestor_nombre},</p>
                <p>La solicitud de reaparición del siguiente predio ha sido rechazada y requiere subsanación:</p>
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 15px 0;">
                    <p><strong>Código Predial:</strong> {codigo_predial}</p>
                    <p><strong>Municipio:</strong> {municipio}</p>
                    <p><strong>Dirección:</strong> {predio_actual.get('direccion', 'N/A')}</p>
                </div>
                <div style="background: #fef2f2; padding: 15px; border-radius: 8px; border-left: 4px solid #dc2626;">
                    <p><strong>Motivo del Rechazo:</strong></p>
                    <p>{justificacion}</p>
                    <p><strong>Rechazado por:</strong> {current_user['full_name']}</p>
                </div>
                <p style="margin-top: 20px;">Por favor, ingrese al sistema para revisar y subsanar la información del predio.</p>
                <p>Una vez corregida la información, podrá reenviar la solicitud para su aprobación.</p>
                <hr style="margin: 20px 0;">
                <p style="color: #666; font-size: 12px;">Este es un mensaje automático del Sistema de Gestión Catastral de Asomunicipios.</p>
            </body>
            </html>
            """
            await enviar_email(gestor_email, asunto, cuerpo)
        except Exception as e:
            print(f"Error enviando notificación de subsanación: {e}")
    
    # Remover _id antes de retornar
    subsanacion.pop("_id", None)
    
    return {
        "message": f"Reaparición del predio {codigo_predial} enviada a SUBSANACIÓN",
        "subsanacion": subsanacion,
        "notificacion_enviada": gestor_email is not None
    }


@api_router.get("/predios/reapariciones/subsanaciones-pendientes")
async def get_subsanaciones_pendientes(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las reapariciones pendientes de subsanación para el gestor actual"""
    query = {"estado": "pendiente_subsanacion"}
    
    # Si es gestor, solo ver las asignadas a él
    if current_user['role'] == UserRole.GESTOR:
        query["gestor_asignado"] = current_user['id']
    elif current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver subsanaciones")
    
    if municipio:
        query["municipio"] = municipio
    
    subsanaciones = await db.reapariciones_subsanacion.find(query, {"_id": 0}).sort("fecha_rechazo", -1).to_list(500)
    
    return {
        "total": len(subsanaciones),
        "subsanaciones": subsanaciones
    }


@api_router.post("/predios/reapariciones/subsanar")
async def subsanar_reaparicion(
    subsanacion_id: str = Form(...),
    justificacion_subsanacion: str = Form(...),
    direccion: Optional[str] = Form(None),
    avaluo: Optional[float] = Form(None),
    area_terreno: Optional[float] = Form(None),
    area_construida: Optional[float] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """El gestor subsana y reenvía la reaparición para aprobación"""
    if current_user['role'] not in [UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para subsanar reapariciones")
    
    # Buscar la subsanación
    subsanacion = await db.reapariciones_subsanacion.find_one({"id": subsanacion_id})
    if not subsanacion:
        raise HTTPException(status_code=404, detail="Subsanación no encontrada")
    
    # Verificar que esté pendiente
    if subsanacion.get("estado") != "pendiente_subsanacion":
        raise HTTPException(status_code=400, detail="Esta subsanación ya fue procesada")
    
    # Si es gestor, verificar que esté asignada a él
    if current_user['role'] == UserRole.GESTOR and subsanacion.get("gestor_asignado") != current_user['id']:
        raise HTTPException(status_code=403, detail="Esta subsanación no está asignada a usted")
    
    codigo_predial = subsanacion["codigo_predial_nacional"]
    municipio = subsanacion["municipio"]
    
    # Actualizar datos del predio si se proporcionaron
    update_predio = {}
    if direccion:
        update_predio["direccion"] = direccion
    if avaluo is not None:
        update_predio["avaluo"] = avaluo
    if area_terreno is not None:
        update_predio["area_terreno"] = area_terreno
    if area_construida is not None:
        update_predio["area_construida"] = area_construida
    
    if update_predio:
        await db.predios.update_one(
            {"codigo_predial_nacional": codigo_predial, "municipio": municipio},
            {"$set": update_predio}
        )
    
    # Agregar al historial
    historial_entry = {
        "accion": "subsanado_reenviado",
        "fecha": datetime.now(timezone.utc).isoformat(),
        "usuario": current_user['full_name'],
        "justificacion": justificacion_subsanacion,
        "cambios_realizados": update_predio if update_predio else "Sin cambios en datos"
    }
    
    # Actualizar subsanación a "reenviado"
    await db.reapariciones_subsanacion.update_one(
        {"id": subsanacion_id},
        {
            "$set": {
                "estado": "reenviado",
                "fecha_subsanacion": datetime.now(timezone.utc).isoformat(),
                "subsanado_por": current_user['id'],
                "subsanado_por_nombre": current_user['full_name'],
                "justificacion_subsanacion": justificacion_subsanacion
            },
            "$inc": {"intentos": 1},
            "$push": {"historial": historial_entry}
        }
    )
    
    # Notificar al coordinador que rechazó
    coordinador = await db.users.find_one({"id": subsanacion.get("rechazado_por")})
    if coordinador and coordinador.get("email"):
        try:
            asunto = f"[Asomunicipios] Reaparición Subsanada - Pendiente de Revisión"
            cuerpo = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <h2 style="color: #059669;">Reaparición Subsanada - Requiere Nueva Revisión</h2>
                <p>El gestor {current_user['full_name']} ha subsanado la siguiente reaparición:</p>
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 15px 0;">
                    <p><strong>Código Predial:</strong> {codigo_predial}</p>
                    <p><strong>Municipio:</strong> {municipio}</p>
                    <p><strong>Intento #:</strong> {subsanacion.get('intentos', 1) + 1}</p>
                </div>
                <div style="background: #ecfdf5; padding: 15px; border-radius: 8px; border-left: 4px solid #059669;">
                    <p><strong>Justificación de Subsanación:</strong></p>
                    <p>{justificacion_subsanacion}</p>
                </div>
                <p style="margin-top: 20px;">Por favor, ingrese al sistema para revisar y aprobar/rechazar nuevamente.</p>
            </body>
            </html>
            """
            await enviar_email(coordinador["email"], asunto, cuerpo)
        except Exception as e:
            print(f"Error enviando notificación al coordinador: {e}")
    
    return {
        "message": f"Reaparición subsanada y reenviada para aprobación",
        "codigo_predial": codigo_predial,
        "intentos": subsanacion.get("intentos", 1) + 1
    }


@api_router.get("/predios/reapariciones/reenviadas")
async def get_reapariciones_reenviadas(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las reapariciones reenviadas pendientes de revisión (para coordinadores)"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden ver reapariciones reenviadas")
    
    query = {"estado": "reenviado"}
    if municipio:
        query["municipio"] = municipio
    
    reenviadas = await db.reapariciones_subsanacion.find(query, {"_id": 0}).sort("fecha_subsanacion", -1).to_list(500)
    
    return {
        "total": len(reenviadas),
        "reenviadas": reenviadas
    }


@api_router.post("/predios/reapariciones/aprobar-subsanacion")
async def aprobar_subsanacion(
    subsanacion_id: str = Form(...),
    justificacion: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Aprueba una reaparición subsanada - El predio permanece en la vigencia actual"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden aprobar subsanaciones")
    
    subsanacion = await db.reapariciones_subsanacion.find_one({"id": subsanacion_id})
    if not subsanacion:
        raise HTTPException(status_code=404, detail="Subsanación no encontrada")
    
    if subsanacion.get("estado") != "reenviado":
        raise HTTPException(status_code=400, detail="Esta subsanación no está pendiente de revisión")
    
    codigo_predial = subsanacion["codigo_predial_nacional"]
    municipio = subsanacion["municipio"]
    
    # Actualizar subsanación a aprobada
    historial_entry = {
        "accion": "aprobado_final",
        "fecha": datetime.now(timezone.utc).isoformat(),
        "usuario": current_user['full_name'],
        "justificacion": justificacion
    }
    
    await db.reapariciones_subsanacion.update_one(
        {"id": subsanacion_id},
        {
            "$set": {
                "estado": "aprobado",
                "fecha_aprobacion_final": datetime.now(timezone.utc).isoformat(),
                "aprobado_por": current_user['id'],
                "aprobado_por_nombre": current_user['full_name'],
                "justificacion_aprobacion": justificacion
            },
            "$push": {"historial": historial_entry}
        }
    )
    
    # Registrar aprobación definitiva
    aprobacion = {
        "id": str(uuid.uuid4()),
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio,
        "vigencia_eliminacion": subsanacion.get("vigencia_eliminacion"),
        "vigencia_reaparicion": subsanacion.get("vigencia_reaparicion"),
        "estado": "aprobado",
        "justificacion": justificacion,
        "aprobado_por": current_user['id'],
        "aprobado_por_nombre": current_user['full_name'],
        "fecha_aprobacion": datetime.now(timezone.utc).isoformat(),
        "subsanacion_id": subsanacion_id,
        "intentos_totales": subsanacion.get("intentos", 1) + 1
    }
    await db.predios_reapariciones_aprobadas.insert_one(aprobacion)
    
    # Eliminar de predios_eliminados
    await db.predios_eliminados.delete_one({
        "codigo_predial_nacional": codigo_predial,
        "municipio": municipio
    })
    
    # Notificar al gestor
    if subsanacion.get("gestor_asignado"):
        gestor = await db.users.find_one({"id": subsanacion["gestor_asignado"]})
        if gestor and gestor.get("email"):
            try:
                asunto = f"[Asomunicipios] Reaparición APROBADA"
                cuerpo = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2 style="color: #059669;">✓ Reaparición Aprobada</h2>
                    <p>La reaparición del predio ha sido APROBADA:</p>
                    <div style="background: #ecfdf5; padding: 15px; border-radius: 8px;">
                        <p><strong>Código Predial:</strong> {codigo_predial}</p>
                        <p><strong>Municipio:</strong> {municipio}</p>
                        <p><strong>Aprobado por:</strong> {current_user['full_name']}</p>
                    </div>
                </body>
                </html>
                """
                await enviar_email(gestor["email"], asunto, cuerpo)
            except Exception as e:
                print(f"Error enviando notificación: {e}")
    
    return {
        "message": f"Reaparición del predio {codigo_predial} APROBADA definitivamente",
        "codigo_predial": codigo_predial
    }


@api_router.post("/predios/reapariciones/rechazar-subsanacion")
async def rechazar_subsanacion_nuevamente(
    subsanacion_id: str = Form(...),
    justificacion: str = Form(...),
    rechazo_definitivo: bool = Form(False),
    current_user: dict = Depends(get_current_user)
):
    """Rechaza nuevamente una subsanación - Puede ser definitivo o permitir otro intento"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden rechazar subsanaciones")
    
    subsanacion = await db.reapariciones_subsanacion.find_one({"id": subsanacion_id})
    if not subsanacion:
        raise HTTPException(status_code=404, detail="Subsanación no encontrada")
    
    if subsanacion.get("estado") != "reenviado":
        raise HTTPException(status_code=400, detail="Esta subsanación no está pendiente de revisión")
    
    codigo_predial = subsanacion["codigo_predial_nacional"]
    municipio = subsanacion["municipio"]
    intentos = subsanacion.get("intentos", 1)
    
    if rechazo_definitivo or intentos >= 3:
        # Rechazo definitivo - eliminar el predio
        historial_entry = {
            "accion": "rechazado_definitivo",
            "fecha": datetime.now(timezone.utc).isoformat(),
            "usuario": current_user['full_name'],
            "justificacion": justificacion
        }
        
        await db.reapariciones_subsanacion.update_one(
            {"id": subsanacion_id},
            {
                "$set": {
                    "estado": "rechazado_definitivo",
                    "fecha_rechazo_final": datetime.now(timezone.utc).isoformat(),
                    "rechazado_final_por": current_user['id'],
                    "rechazado_final_por_nombre": current_user['full_name'],
                    "justificacion_rechazo_final": justificacion
                },
                "$push": {"historial": historial_entry}
            }
        )
        
        # Registrar rechazo definitivo
        rechazo = {
            "id": str(uuid.uuid4()),
            "codigo_predial_nacional": codigo_predial,
            "municipio": municipio,
            "estado": "rechazado_definitivo",
            "justificacion": justificacion,
            "rechazado_por": current_user['id'],
            "rechazado_por_nombre": current_user['full_name'],
            "fecha_rechazo": datetime.now(timezone.utc).isoformat(),
            "intentos_totales": intentos + 1
        }
        await db.predios_reapariciones_aprobadas.insert_one(rechazo)
        
        # Eliminar el predio de la vigencia actual
        predio = await db.predios.find_one({
            "codigo_predial_nacional": codigo_predial,
            "municipio": municipio
        })
        if predio:
            await db.predios.delete_one({
                "codigo_predial_nacional": codigo_predial,
                "municipio": municipio,
                "vigencia": predio.get("vigencia")
            })
        
        mensaje = f"Reaparición RECHAZADA DEFINITIVAMENTE - Predio eliminado"
    else:
        # Permitir otro intento de subsanación
        historial_entry = {
            "accion": "rechazado_nueva_subsanacion",
            "fecha": datetime.now(timezone.utc).isoformat(),
            "usuario": current_user['full_name'],
            "justificacion": justificacion
        }
        
        await db.reapariciones_subsanacion.update_one(
            {"id": subsanacion_id},
            {
                "$set": {
                    "estado": "pendiente_subsanacion",
                    "motivo_rechazo": justificacion,
                    "fecha_rechazo": datetime.now(timezone.utc).isoformat(),
                    "rechazado_por": current_user['id'],
                    "rechazado_por_nombre": current_user['full_name']
                },
                "$push": {"historial": historial_entry}
            }
        )
        
        mensaje = f"Reaparición rechazada - Enviada nuevamente a subsanación (Intento {intentos + 1}/3)"
    
    # Notificar al gestor
    if subsanacion.get("gestor_asignado"):
        gestor = await db.users.find_one({"id": subsanacion["gestor_asignado"]})
        if gestor and gestor.get("email"):
            try:
                estado_msg = "RECHAZADA DEFINITIVAMENTE" if rechazo_definitivo or intentos >= 3 else "requiere nueva subsanación"
                asunto = f"[Asomunicipios] Reaparición {estado_msg}"
                cuerpo = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2 style="color: #dc2626;">Reaparición {estado_msg}</h2>
                    <div style="background: #fef2f2; padding: 15px; border-radius: 8px;">
                        <p><strong>Código Predial:</strong> {codigo_predial}</p>
                        <p><strong>Municipio:</strong> {municipio}</p>
                        <p><strong>Motivo:</strong> {justificacion}</p>
                        <p><strong>Intento:</strong> {intentos + 1}/3</p>
                    </div>
                    {"<p>El predio ha sido eliminado del sistema.</p>" if rechazo_definitivo or intentos >= 3 else "<p>Por favor, revise y subsane nuevamente.</p>"}
                </body>
                </html>
                """
                await enviar_email(gestor["email"], asunto, cuerpo)
            except Exception as e:
                print(f"Error enviando notificación: {e}")
    
    return {
        "message": mensaje,
        "codigo_predial": codigo_predial,
        "rechazo_definitivo": rechazo_definitivo or intentos >= 3
    }


class SolicitudRespuestaRequest(BaseModel):
    solicitud_id: str
    aprobado: bool
    comentario: str


@api_router.post("/predios/reapariciones/solicitud-responder")
async def responder_solicitud_reaparicion(
    request: SolicitudRespuestaRequest,
    current_user: dict = Depends(get_current_user)
):
    """Responde a una solicitud de reaparición hecha por un gestor"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden responder solicitudes")
    
    # Buscar la solicitud
    solicitud = await db.predios_reapariciones_solicitudes.find_one(
        {"id": request.solicitud_id},
        {"_id": 0}
    )
    
    if not solicitud:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    
    if solicitud.get("estado") != "pendiente":
        raise HTTPException(status_code=400, detail="Esta solicitud ya fue procesada")
    
    nuevo_estado = "aprobado" if request.aprobado else "rechazado"
    
    # Actualizar la solicitud
    await db.predios_reapariciones_solicitudes.update_one(
        {"id": request.solicitud_id},
        {"$set": {
            "estado": nuevo_estado,
            "respondido_por": current_user['id'],
            "respondido_por_nombre": current_user['full_name'],
            "comentario_respuesta": request.comentario,
            "fecha_respuesta": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Si fue aprobado, registrar en reapariciones aprobadas para permitir la creación del predio
    if request.aprobado:
        aprobacion = {
            "id": str(uuid.uuid4()),
            "codigo_predial_nacional": solicitud["codigo_predial_nacional"],
            "municipio": solicitud["municipio"],
            "vigencia_eliminacion": solicitud.get("vigencia_eliminacion"),
            "estado": "aprobado",
            "justificacion": f"Solicitud de gestor aprobada: {request.comentario}",
            "justificacion_gestor": solicitud.get("justificacion_gestor"),
            "aprobado_por": current_user['id'],
            "aprobado_por_nombre": current_user['full_name'],
            "fecha_aprobacion": datetime.now(timezone.utc).isoformat(),
            "origen": "solicitud_gestor"
        }
        await db.predios_reapariciones_aprobadas.insert_one(aprobacion)
        
        # Eliminar de la lista de predios eliminados
        await db.predios_eliminados.delete_one({
            "codigo_predial_nacional": solicitud["codigo_predial_nacional"],
            "municipio": solicitud["municipio"]
        })
    
    # Notificar al gestor que hizo la solicitud
    await crear_notificacion(
        usuario_id=solicitud.get("solicitado_por"),
        titulo=f"Solicitud de Reaparición {'Aprobada' if request.aprobado else 'Rechazada'}",
        mensaje=f"Su solicitud para el predio {solicitud['codigo_predial_nacional']} ha sido {'aprobada' if request.aprobado else 'rechazada'}. {request.comentario}",
        tipo="success" if request.aprobado else "error",
        enviar_email=True
    )
    
    return {
        "message": f"Solicitud {'aprobada' if request.aprobado else 'rechazada'} correctamente",
        "estado": nuevo_estado,
        "puede_crear_predio": request.aprobado
    }


@api_router.get("/predios/reapariciones/historial")
async def get_historial_reapariciones(
    municipio: Optional[str] = None,
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el historial de decisiones sobre reapariciones"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {}
    if municipio:
        query["municipio"] = municipio
    if estado:
        query["estado"] = estado
    
    historial = await db.predios_reapariciones_aprobadas.find(query, {"_id": 0}).sort("fecha_aprobacion", -1).to_list(500)
    
    return {
        "total": len(historial),
        "historial": historial
    }


@api_router.get("/predios/reapariciones")
async def get_reapariciones(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene la lista de predios que fueron eliminados y reaparecieron"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Obtener todos los códigos eliminados
    query_elim = {}
    if municipio:
        query_elim["municipio"] = municipio
    
    eliminados = await db.predios_eliminados.find(query_elim, {"_id": 0, "codigo_predial_nacional": 1, "municipio": 1, "vigencia_eliminacion": 1}).to_list(50000)
    
    reapariciones = []
    for elim in eliminados:
        codigo = elim["codigo_predial_nacional"]
        mun = elim["municipio"]
        vig_elim = elim.get("vigencia_eliminacion", 0)
        
        # Buscar si este código existe en vigencias posteriores
        existe = await db.predios.find_one({
            "codigo_predial_nacional": codigo,
            "municipio": mun,
            "vigencia": {"$gt": vig_elim}
        })
        
        if existe:
            reapariciones.append({
                "codigo_predial_nacional": codigo,
                "municipio": mun,
                "vigencia_eliminacion": vig_elim,
                "vigencia_reaparicion": existe.get("vigencia"),
                "alerta": "INCONSISTENCIA: Predio eliminado que reaparece"
            })
    
    return {
        "total_reapariciones": len(reapariciones),
        "reapariciones": reapariciones,
        "mensaje": f"Se encontraron {len(reapariciones)} predios eliminados que reaparecieron" if reapariciones else "No se encontraron reapariciones"
    }


@api_router.post("/predios/comparar-vigencias")
async def comparar_vigencias_predios(
    municipio: str = Form(...),
    vigencia_anterior: int = Form(...),
    vigencia_nueva: int = Form(...),
    radicado: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Compara dos vigencias para detectar predios eliminados entre ellas"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Obtener predios de vigencia anterior
    predios_anterior = await db.predios.find(
        {"municipio": municipio, "vigencia": vigencia_anterior},
        {"_id": 0, "codigo_predial_nacional": 1}
    ).to_list(50000)
    codigos_anterior = {p['codigo_predial_nacional'] for p in predios_anterior}
    
    # Obtener predios de vigencia nueva
    predios_nueva = await db.predios.find(
        {"municipio": municipio, "vigencia": vigencia_nueva},
        {"_id": 0, "codigo_predial_nacional": 1}
    ).to_list(50000)
    codigos_nueva = {p['codigo_predial_nacional'] for p in predios_nueva}
    
    # Predios que estaban en anterior pero no en nueva = eliminados
    codigos_eliminados = codigos_anterior - codigos_nueva
    
    if not codigos_eliminados:
        return {
            "message": "No se encontraron predios eliminados entre las vigencias",
            "municipio": municipio,
            "vigencia_anterior": vigencia_anterior,
            "vigencia_nueva": vigencia_nueva,
            "predios_eliminados": 0
        }
    
    # Obtener datos completos de predios eliminados
    predios_eliminados = await db.predios.find(
        {"municipio": municipio, "vigencia": vigencia_anterior, "codigo_predial_nacional": {"$in": list(codigos_eliminados)}},
        {"_id": 0}
    ).to_list(50000)
    
    # Guardar en colección de eliminados
    eliminados_docs = []
    for p in predios_eliminados:
        eliminados_docs.append({
            **p,
            "eliminado_en": datetime.now(timezone.utc).isoformat(),
            "vigencia_eliminacion": vigencia_nueva,
            "vigencia_origen": vigencia_anterior,
            "radicado_eliminacion": radicado,
            "motivo": f"No incluido en vigencia {vigencia_nueva}",
            "eliminado_por": current_user['id'],
            "eliminado_por_nombre": current_user['full_name']
        })
    
    if eliminados_docs:
        await db.predios_eliminados.insert_many(eliminados_docs)
    
    return {
        "message": f"Se detectaron {len(eliminados_docs)} predios eliminados",
        "municipio": municipio,
        "vigencia_anterior": vigencia_anterior,
        "vigencia_nueva": vigencia_nueva,
        "predios_eliminados": len(eliminados_docs),
        "radicado": radicado
    }


@api_router.get("/predios/eliminados/exportar-excel")
async def exportar_predios_eliminados_excel(
    municipio: Optional[str] = None,
    vigencia: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exporta predios eliminados a Excel con radicado"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    query = {}
    if municipio:
        query["municipio"] = municipio
    if vigencia:
        query["vigencia_eliminacion"] = vigencia
    
    predios = await db.predios_eliminados.find(query, {"_id": 0}).sort("eliminado_en", -1).to_list(50000)
    
    if not predios:
        raise HTTPException(status_code=404, detail="No hay predios eliminados para exportar")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predios Eliminados"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Código Predial Nacional",
        "Municipio",
        "Dirección",
        "Propietario",
        "Área Terreno (m²)",
        "Área Construida (m²)",
        "Avalúo",
        "Vigencia Origen",
        "Vigencia Eliminación",
        "Radicado Eliminación",
        "Fecha Eliminación",
        "Motivo",
        "Eliminado Por"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row_idx, predio in enumerate(predios, 2):
        propietario = ""
        if predio.get("propietarios"):
            propietario = predio["propietarios"][0].get("nombre_propietario", "")
        
        data = [
            predio.get("codigo_predial_nacional", ""),
            predio.get("municipio", ""),
            predio.get("direccion", ""),
            propietario,
            predio.get("area_terreno", 0),
            predio.get("area_construida", 0),
            predio.get("avaluo", 0),
            predio.get("vigencia_origen", predio.get("vigencia", "")),
            predio.get("vigencia_eliminacion", ""),
            predio.get("radicado_eliminacion", ""),
            predio.get("eliminado_en", "")[:10] if predio.get("eliminado_en") else "",
            predio.get("motivo", ""),
            predio.get("eliminado_por_nombre", "")
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in [5, 6, 7]:  # Números
                cell.alignment = Alignment(horizontal="right")
    
    # Ajustar anchos de columna
    column_widths = [35, 15, 30, 30, 15, 15, 15, 12, 12, 20, 12, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"predios_eliminados_{municipio or 'todos'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.patch("/predios/eliminados/{predio_id}/radicado")
async def actualizar_radicado_eliminado(
    predio_id: str,
    radicado: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Actualiza el radicado de un predio eliminado"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    result = await db.predios_eliminados.update_one(
        {"id": predio_id},
        {"$set": {
            "radicado_eliminacion": radicado,
            "radicado_actualizado_por": current_user['full_name'],
            "radicado_actualizado_en": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Predio eliminado no encontrado")
    
    return {"message": "Radicado actualizado correctamente", "radicado": radicado}


@api_router.post("/predios/import-excel")
async def import_predios_excel(
    file: UploadFile = File(...),
    vigencia: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """Importa predios desde archivo Excel R1-R2 con soporte de vigencia"""
    import openpyxl
    
    # Procesar vigencia - puede venir como "2023", "2025", "01012023", "1012023", etc.
    vigencia_int = 2025  # valor por defecto
    if vigencia:
        vigencia_str = str(vigencia).strip()
        # Si tiene más de 4 dígitos, extraer los últimos 4 (el año)
        if len(vigencia_str) > 4:
            vigencia_int = int(vigencia_str[-4:])
        else:
            vigencia_int = int(vigencia_str)
    
    logger.info(f"Importando con vigencia: {vigencia} -> {vigencia_int}")
    
    # Helper para convertir números con formato de coma decimal
    def parse_number(value, default=0):
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        try:
            # Convertir string, reemplazando coma por punto
            s = str(value).strip().replace('$', '').replace(' ', '')
            # Si tiene punto y coma, el punto es separador de miles
            if '.' in s and ',' in s:
                s = s.replace('.', '').replace(',', '.')
            elif ',' in s:
                s = s.replace(',', '.')
            return float(s) if s else default
        except:
            return default
    
    # Solo coordinador o admin pueden importar, o usuarios con permiso explícito
    has_permission = await check_permission(current_user, Permission.IMPORT_R1R2)
    if not has_permission:
        raise HTTPException(status_code=403, detail="No tiene permiso para importar datos R1/R2")
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")
    
    try:
        # Guardar archivo temporalmente
        temp_path = UPLOAD_DIR / f"temp_import_{uuid.uuid4()}.xlsx"
        content = await file.read()
        with open(temp_path, 'wb') as f:
            f.write(content)
        
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
        
        # Buscar hoja R1 con nombres alternativos (normalizando espacios)
        r1_sheet_names = ['REGISTRO_R1', 'REGISTRO R1', 'R1', 'Registro_R1', 'Registro R1', 'registro_r1', 'Hoja1', 'Sheet1']
        ws_r1 = None
        # Primero intentar coincidencia exacta
        for name in r1_sheet_names:
            if name in wb.sheetnames:
                ws_r1 = wb[name]
                break
        # Si no encontró, intentar con nombres que contengan espacios
        if ws_r1 is None:
            for sheet_name in wb.sheetnames:
                sheet_name_clean = sheet_name.strip().upper()
                for name in r1_sheet_names:
                    if sheet_name_clean == name.upper():
                        ws_r1 = wb[sheet_name]
                        break
                if ws_r1:
                    break
        
        if ws_r1 is None:
            wb.close()
            temp_path.unlink()
            raise HTTPException(
                status_code=400, 
                detail=f"No se encontró hoja R1. Hojas disponibles: {', '.join(wb.sheetnames)}. Se esperaba: REGISTRO_R1, R1, o similar."
            )
        
        # Leer R1 (propietarios)
        r1_data = {}
        rows_read = 0
        
        # Detectar formato del archivo leyendo headers
        headers = []
        for row in ws_r1.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).upper().strip() if h else '' for h in row]
            break
        
        # Mapear columnas según headers o usar posiciones por defecto
        # Formato 1: Sin headers explícitos (col[0]=depto, col[1]=mun, col[3]=codigo_predial)
        # Formato 2: Con headers (NUMERO_DEL_PREDIO, Codigo_Predial_Nacional, etc.)
        
        col_map = {
            'departamento': 0,
            'municipio': 1,
            'numero_predio': 2,
            'codigo_predial': 3,
            'codigo_homologado': 4,
            'nombre': 8,
            'estado_civil': 9,
            'tipo_documento': 10,
            'numero_documento': 11,
            'direccion': 12,
            'comuna': 13,
            'destino_economico': 14,
            'area_terreno': 15,
            'area_construida': 16,
            'avaluo': 17,
            'vigencia_excel': 18,
            'tipo_mutacion': 19,
            'numero_resolucion': 20,
            'fecha_resolucion': 21,
        }
        
        # Detectar si tiene headers conocidos
        tiene_headers = any('CODIGO' in h or 'PREDIO' in h or 'NUMERO' in h for h in headers)
        
        if tiene_headers:
            logger.info(f"Detectado archivo con headers: {headers[:10]}")
            # Buscar posiciones de columnas por nombre de header
            for i, h in enumerate(headers):
                h_upper = h.upper().replace('_', ' ').replace('-', ' ')
                # Detectar código predial nacional con múltiples variantes
                if ('CODIGO' in h_upper and 'PREDIAL' in h_upper and 'NACIONAL' in h_upper) or \
                   ('NUMERO' in h_upper and 'PREDIAL' in h_upper and 'NACIONAL' in h_upper) or \
                   (h_upper == 'N PREDIAL') or \
                   ('PREDIAL' in h_upper and 'NACIONAL' in h_upper):
                    col_map['codigo_predial'] = i
                elif 'CODIGO' in h_upper and 'HOMOLOG' in h_upper:
                    col_map['codigo_homologado'] = i
                elif ('NUMERO' in h_upper and 'PREDIO' in h_upper and 'NACIONAL' not in h_upper) or h_upper == 'NUMERO DEL PREDIO':
                    col_map['numero_predio'] = i
                elif h_upper in ['NOMBRE', 'PROPIETARIO', 'NOMBRE PROPIETARIO']:
                    col_map['nombre'] = i
                elif 'ESTADO' in h_upper and 'CIVIL' in h_upper:
                    col_map['estado_civil'] = i
                elif 'TIPO' in h_upper and 'DOCUMENTO' in h_upper:
                    col_map['tipo_documento'] = i
                elif 'NUMERO' in h_upper and 'DOCUMENTO' in h_upper:
                    col_map['numero_documento'] = i
                elif h_upper in ['DIRECCION', 'DIRECCIÓN']:
                    col_map['direccion'] = i
                elif h_upper == 'COMUNA':
                    col_map['comuna'] = i
                elif 'DESTINO' in h_upper and 'ECONOMICO' in h_upper:
                    col_map['destino_economico'] = i
                elif 'AREA' in h_upper and 'TERRENO' in h_upper:
                    col_map['area_terreno'] = i
                elif 'AREA' in h_upper and 'CONSTRUIDA' in h_upper:
                    col_map['area_construida'] = i
                elif h_upper in ['AVALUO', 'AVALÚO']:
                    col_map['avaluo'] = i
                elif h_upper == 'VIGENCIA':
                    col_map['vigencia_excel'] = i
                elif 'TIPO' in h_upper and 'MUTACION' in h_upper:
                    col_map['tipo_mutacion'] = i
                elif 'RESOLUCION' in h_upper or 'RESOLUCIÓN' in h_upper:
                    if 'FECHA' in h_upper:
                        col_map['fecha_resolucion'] = i
                    elif 'NO' in h_upper or 'NUMERO' in h_upper:
                        col_map['numero_resolucion'] = i
            
            logger.info(f"Mapa de columnas detectado: codigo_predial={col_map['codigo_predial']}, nombre={col_map['nombre']}")
            start_row = 2  # Datos empiezan en fila 2
        else:
            logger.info("Archivo sin headers explícitos, usando formato estándar")
            start_row = 2  # Datos empiezan en fila 2
        
        # Función helper para obtener valor de columna de forma segura
        def get_col(row, key, default=''):
            idx = col_map.get(key, -1)
            if idx >= 0 and idx < len(row):
                return row[idx]
            return default
        
        for row in ws_r1.iter_rows(min_row=start_row, values_only=True):
            # Verificar fila no vacía
            first_col = get_col(row, 'numero_predio') or get_col(row, 'codigo_predial')
            if not first_col:
                continue
            
            rows_read += 1
            codigo_predial = str(get_col(row, 'codigo_predial') or '').strip()
            if not codigo_predial or len(codigo_predial) < 10:
                continue
            
            if codigo_predial not in r1_data:
                r1_data[codigo_predial] = {
                    'departamento': str(get_col(row, 'departamento') or '').strip()[:10],
                    'municipio': str(get_col(row, 'municipio') or '').strip(),
                    'numero_predio': str(get_col(row, 'numero_predio') or '').strip(),
                    'codigo_predial_nacional': codigo_predial,
                    'codigo_homologado': str(get_col(row, 'codigo_homologado') or '').strip(),
                    'direccion': str(get_col(row, 'direccion') or '').strip(),
                    'comuna': str(get_col(row, 'comuna') or '').strip(),
                    'destino_economico': str(get_col(row, 'destino_economico') or '').strip(),
                    'area_terreno': parse_number(get_col(row, 'area_terreno')),
                    'area_construida': parse_number(get_col(row, 'area_construida')),
                    'avaluo': parse_number(get_col(row, 'avaluo')),
                    'vigencia': vigencia_int,
                    'tipo_mutacion': str(get_col(row, 'tipo_mutacion') or '').strip(),
                    'numero_resolucion': str(get_col(row, 'numero_resolucion') or '').strip(),
                    'fecha_resolucion': str(get_col(row, 'fecha_resolucion') or '').strip(),
                    'propietarios': [],
                    'r2_registros': [],
                    'id': str(uuid.uuid4()),
                    'created_at': datetime.now(timezone.utc).isoformat(),
                    'status': 'aprobado'
                }
            
            # Agregar propietario
            nombre = str(get_col(row, 'nombre') or '').strip()
            if nombre:
                r1_data[codigo_predial]['propietarios'].append({
                    'nombre_propietario': nombre,
                    'estado_civil': str(get_col(row, 'estado_civil') or '').strip(),
                    'tipo_documento': str(get_col(row, 'tipo_documento') or '').strip(),
                    'numero_documento': str(get_col(row, 'numero_documento') or '').strip()
                })
        
        # Buscar hoja R2 con nombres alternativos (normalizando espacios)
        r2_sheet_names = ['REGISTRO_R2', 'REGISTRO R2', 'R2', 'Registro_R2', 'Registro R2', 'registro_r2', 'Hoja2', 'Sheet2']
        ws_r2 = None
        # Primero intentar coincidencia exacta
        for name in r2_sheet_names:
            if name in wb.sheetnames:
                ws_r2 = wb[name]
                break
        # Si no encontró, intentar con nombres que contengan espacios
        if ws_r2 is None:
            for sheet_name in wb.sheetnames:
                sheet_name_clean = sheet_name.strip().upper()
                for name in r2_sheet_names:
                    if sheet_name_clean == name.upper():
                        ws_r2 = wb[sheet_name]
                        break
                if ws_r2:
                    break
        
        if ws_r2 is None:
            wb.close()
            temp_path.unlink()
            raise HTTPException(
                status_code=400, 
                detail=f"No se encontró hoja R2. Hojas disponibles: {', '.join(wb.sheetnames)}. Se esperaba: REGISTRO_R2, R2, o similar."
            )
        
        # Leer R2 (físico)
        for row in ws_r2.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            codigo_predial = str(row[3] or '').strip()
            if codigo_predial not in r1_data:
                continue
            
            matricula = str(row[7] or '').strip() if len(row) > 7 else ''
            
            # Buscar si ya existe este registro R2 (por matrícula o agregar nuevo)
            r2_exists = False
            for r2 in r1_data[codigo_predial]['r2_registros']:
                if r2.get('matricula_inmobiliaria') == matricula:
                    r2_exists = True
                    break
            
            if not r2_exists:
                zonas = []
                
                # Zona 1
                if len(row) > 10 and row[10]:
                    area_t = parse_number(row[10])
                    zonas.append({
                        'zona_fisica': str(row[8] or '').strip() if len(row) > 8 else '',
                        'zona_economica': str(row[9] or '').strip() if len(row) > 9 else '',
                        'area_terreno': area_t,
                        'habitaciones': int(parse_number(row[14])) if len(row) > 14 else 0,
                        'banos': int(parse_number(row[15])) if len(row) > 15 else 0,
                        'locales': int(parse_number(row[16])) if len(row) > 16 else 0,
                        'pisos': int(parse_number(row[17])) if len(row) > 17 else 0,
                        'tipificacion': str(row[18] or '').strip() if len(row) > 18 else '',
                        'uso': str(row[19] or '').strip() if len(row) > 19 else '',
                        'puntaje': int(parse_number(row[20])) if len(row) > 20 else 0,
                        'area_construida': parse_number(row[21]) if len(row) > 21 else 0
                    })
                
                # Zona 2
                if len(row) > 13 and row[13]:
                    area_t2 = parse_number(row[13])
                    if area_t2 > 0:
                        zonas.append({
                            'zona_fisica': str(row[11] or '').strip() if len(row) > 11 else '',
                            'zona_economica': str(row[12] or '').strip() if len(row) > 12 else '',
                            'area_terreno': area_t2,
                            'habitaciones': int(parse_number(row[22])) if len(row) > 22 else 0,
                            'banos': int(parse_number(row[23])) if len(row) > 23 else 0,
                            'locales': int(parse_number(row[24])) if len(row) > 24 else 0,
                            'pisos': int(parse_number(row[25])) if len(row) > 25 else 0,
                            'tipificacion': str(row[26] or '').strip() if len(row) > 26 else '',
                            'uso': str(row[27] or '').strip() if len(row) > 27 else '',
                            'puntaje': int(parse_number(row[28])) if len(row) > 28 else 0,
                            'area_construida': parse_number(row[29]) if len(row) > 29 else 0
                        })
                
                r1_data[codigo_predial]['r2_registros'].append({
                    'matricula_inmobiliaria': matricula,
                    'zonas': zonas
                })
        
        wb.close()
        temp_path.unlink()
        
        # Mapeo de códigos de municipio a nombres
        MUNICIPIO_CODIGOS = {
            '003': 'Ábrego', '54003': 'Ábrego', '3': 'Ábrego',
            '109': 'Bucarasica', '54109': 'Bucarasica',
            '128': 'Cáchira', '54128': 'Cáchira',
            '206': 'Convención', '54206': 'Convención',
            '245': 'El Carmen', '54245': 'El Carmen',
            '250': 'El Tarra', '54250': 'El Tarra',
            '344': 'Hacarí', '54344': 'Hacarí',
            '398': 'La Playa', '54398': 'La Playa',
            '670': 'San Calixto', '54670': 'San Calixto',
            '720': 'Sardinata', '54720': 'Sardinata',
            '800': 'Teorama', '54800': 'Teorama',
            '614': 'Río de Oro', '20614': 'Río de Oro',
        }
        
        # Función para extraer municipio del código predial nacional
        def get_municipio_from_codigo(codigo_predial):
            """Extrae el código de municipio del código predial nacional.
            Formato: DDMMMXXXX... donde DD=depto, MMM=municipio
            """
            if not codigo_predial or len(codigo_predial) < 5:
                return None
            # Los primeros 2 dígitos son departamento, los siguientes 3 son municipio
            codigo_mun = codigo_predial[2:5]
            # Intentar con el código completo depto+mun
            codigo_completo = codigo_predial[:5]
            if codigo_completo in MUNICIPIO_CODIGOS:
                return MUNICIPIO_CODIGOS[codigo_completo]
            if codigo_mun in MUNICIPIO_CODIGOS:
                return MUNICIPIO_CODIGOS[codigo_mun]
            # Intentar sin ceros a la izquierda
            codigo_mun_int = codigo_mun.lstrip('0')
            if codigo_mun_int in MUNICIPIO_CODIGOS:
                return MUNICIPIO_CODIGOS[codigo_mun_int]
            return None
        
        # Guardar en historial antes de actualizar
        municipio_raw = list(r1_data.values())[0]['municipio'] if r1_data else 'Desconocido'
        primer_codigo = list(r1_data.keys())[0] if r1_data else ''
        
        # Primero intentar extraer del código predial nacional (más confiable)
        municipio = get_municipio_from_codigo(primer_codigo)
        
        # Si no funciona, intentar con el valor raw
        if not municipio:
            municipio = MUNICIPIO_CODIGOS.get(municipio_raw, None)
        
        # Si el municipio_raw es un código predial largo, intentar extraerlo también
        if not municipio and len(str(municipio_raw)) >= 5:
            municipio = get_municipio_from_codigo(municipio_raw)
        
        if not municipio:
            # Usar el valor raw como último recurso
            municipio = municipio_raw
            logger.warning(f"No se pudo determinar el municipio. raw={municipio_raw}, codigo={primer_codigo[:30]}")
        
        # Actualizar el municipio en todos los predios
        for predio in r1_data.values():
            predio['municipio'] = municipio
        
        # Obtener los códigos prediales de los nuevos predios (CNP)
        new_codigos = set(r1_data.keys())
        
        # === LÓGICA CORREGIDA: Comparar con TODAS las vigencias anteriores ===
        
        # 1. Obtener predios de la vigencia actual (para historial)
        existing_predios_vigencia_actual = await db.predios.find(
            {
                "municipio": {"$regex": f"^{municipio}$", "$options": "i"}, 
                "vigencia": vigencia_int
            }, 
            {"_id": 0}
        ).to_list(50000)
        
        # 2. Obtener TODOS los CNP de TODAS las vigencias anteriores
        all_previous_predios = await db.predios.find(
            {
                "municipio": {"$regex": f"^{municipio}$", "$options": "i"},
                "vigencia": {"$lt": vigencia_int}
            },
            {"_id": 0, "codigo_predial_nacional": 1, "vigencia": 1}
        ).to_list(100000)
        
        all_previous_cnp = {p['codigo_predial_nacional'] for p in all_previous_predios if p.get('codigo_predial_nacional')}
        
        # 3. CNP eliminados = existían en alguna vigencia anterior pero NO vienen en la nueva
        cnp_eliminados = all_previous_cnp - new_codigos
        
        # Log para debug
        logger.info(f"R1 Import - Municipio: {municipio}, Vigencia: {vigencia_int}")
        logger.info(f"  - Predios vigencia actual: {len(existing_predios_vigencia_actual)}")
        logger.info(f"  - CNP históricos (todas vigencias anteriores): {len(all_previous_cnp)}")
        logger.info(f"  - CNP en nueva importación: {len(new_codigos)}")
        logger.info(f"  - CNP eliminados detectados: {len(cnp_eliminados)}")
        
        # 4. Guardar predios eliminados (evitando duplicados)
        predios_eliminados_count = 0
        predios_eliminados_nuevos = []
        
        for cnp in cnp_eliminados:
            # Verificar si ya está marcado como eliminado
            ya_eliminado = await db.predios_eliminados.find_one({"codigo_predial_nacional": cnp})
            if not ya_eliminado:
                # Obtener datos del predio de la última vigencia donde existió
                predio_original = await db.predios.find_one(
                    {
                        "codigo_predial_nacional": cnp, 
                        "municipio": {"$regex": f"^{municipio}$", "$options": "i"}
                    },
                    {"_id": 0},
                    sort=[("vigencia", -1)]
                )
                if predio_original:
                    predios_eliminados_nuevos.append({
                        **predio_original,
                        "eliminado_en": datetime.now(timezone.utc).isoformat(),
                        "vigencia_eliminacion": vigencia_int,
                        "motivo": f"No incluido en importación R1-R2 vigencia {vigencia_int}"
                    })
                    predios_eliminados_count += 1
        
        # Insertar todos los nuevos eliminados de una vez
        if predios_eliminados_nuevos:
            await db.predios_eliminados.insert_many(predios_eliminados_nuevos)
            logger.info(f"  - Predios marcados como eliminados: {len(predios_eliminados_nuevos)}")
        
        # Guardar historial de predios existentes de esta vigencia
        if existing_predios_vigencia_actual:
            await db.predios_historico.insert_many([
                {**p, "archivado_en": datetime.now(timezone.utc).isoformat(), "vigencia_archivo": vigencia_int}
                for p in existing_predios_vigencia_actual
            ])
        
        # Eliminar predios actuales de este municipio SOLO PARA ESTA VIGENCIA
        await db.predios.delete_many({"municipio": municipio, "vigencia": vigencia_int})
        
        # Insertar nuevos predios
        predios_list = list(r1_data.values())
        if predios_list:
            await db.predios.insert_many(predios_list)
        
        # Calcular predios nuevos (CNP que no estaban en ninguna vigencia anterior)
        predios_nuevos_count = len(new_codigos - all_previous_cnp)
        
        # Registrar importación
        logger.info(f"Import stats: rows_read={rows_read}, unique_predios={len(r1_data)}, municipio={municipio}")
        await db.importaciones.insert_one({
            "id": str(uuid.uuid4()),
            "municipio": municipio,
            "vigencia": vigencia_int,
            "total_predios": len(predios_list),
            "predios_anteriores": len(existing_predios_vigencia_actual),
            "predios_eliminados": predios_eliminados_count,
            "predios_nuevos": predios_nuevos_count,
            "archivo": file.filename,
            "importado_por": current_user['id'],
            "importado_por_nombre": current_user['full_name'],
            "fecha": datetime.now(timezone.utc).isoformat()
        })
        
        # Broadcast WebSocket para sincronización en tiempo real
        await ws_manager.broadcast({
            "type": "predio_actualizado",
            "action": "import",
            "municipio": municipio,
            "vigencia": vigencia_int,
            "predios_importados": len(predios_list),
            "predios_nuevos": predios_nuevos_count,
            "imported_by": current_user['full_name'],
            "timestamp": datetime.now(timezone.utc).isoformat()
        }, exclude_user=current_user['id'])
        
        return {
            "message": f"Importación exitosa para {municipio}",
            "vigencia": vigencia_int,
            "predios_importados": len(predios_list),
            "predios_anteriores": len(existing_predios),
            "predios_eliminados": predios_eliminados_count,
            "predios_nuevos": predios_nuevos_count,
            "municipio": municipio
        }
        
    except Exception as e:
        logger.error(f"Error importing Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error al importar: {str(e)}")


@api_router.get("/predios/vigencias")
async def get_vigencias_disponibles(current_user: dict = Depends(get_current_user)):
    """Obtiene las vigencias disponibles por municipio, ordenadas de más reciente a más antigua"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Vigencias actuales - solo predios con vigencia definida
    pipeline = [
        {"$match": {"vigencia": {"$ne": None}}},
        {"$group": {"_id": {"municipio": "$municipio", "vigencia": "$vigencia"}, "count": {"$sum": 1}}},
        {"$sort": {"_id.municipio": 1}}
    ]
    result = await db.predios.aggregate(pipeline).to_list(1000)
    
    vigencias = {}
    for r in result:
        mun = r['_id'].get('municipio')
        vig = r['_id'].get('vigencia')
        if mun and vig:
            if mun not in vigencias:
                vigencias[mun] = []
            vigencias[mun].append({"vigencia": vig, "predios": r['count']})
    
    # Vigencias históricas
    historico_pipeline = [
        {"$match": {"vigencia": {"$ne": None}}},
        {"$group": {"_id": {"municipio": "$municipio", "vigencia": "$vigencia"}, "count": {"$sum": 1}}},
        {"$sort": {"_id.municipio": 1}}
    ]
    historico = await db.predios_historico.aggregate(historico_pipeline).to_list(1000)
    
    for h in historico:
        mun = h['_id'].get('municipio')
        vig = h['_id'].get('vigencia')
        if mun and vig:
            if mun not in vigencias:
                vigencias[mun] = []
            # Agregar si no existe
            if not any(v['vigencia'] == vig for v in vigencias[mun]):
                vigencias[mun].append({"vigencia": vig, "predios": h['count'], "historico": True})
    
    # Función para extraer el año de una vigencia (puede ser 2025, 01012025, 1012025)
    def get_year(vig):
        vig_str = str(vig)
        if len(vig_str) >= 7:
            return int(vig_str[-4:])
        return int(vig_str)
    
    # Ordenar vigencias: primero datos actuales (no históricos) por año descendente, luego históricos
    for mun in vigencias:
        # Separar actuales de históricos
        actuales = [v for v in vigencias[mun] if not v.get('historico')]
        historicos = [v for v in vigencias[mun] if v.get('historico')]
        
        # Ordenar cada grupo por año descendente
        actuales.sort(key=lambda x: get_year(x['vigencia']), reverse=True)
        historicos.sort(key=lambda x: get_year(x['vigencia']), reverse=True)
        
        # Combinar: actuales primero, luego históricos
        vigencias[mun] = actuales + historicos
    
    return vigencias


@api_router.get("/predios/export-excel")
async def export_predios_excel(
    municipio: Optional[str] = None,
    vigencia: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exporta predios a Excel en formato EXACTO al archivo original R1-R2"""
    # Usuario externo y Empresa no pueden descargar Excel
    if current_user['role'] in [UserRole.USUARIO, UserRole.EMPRESA]:
        raise HTTPException(status_code=403, detail="No tiene permiso para exportar datos")
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Query
    query = {"deleted": {"$ne": True}}
    if municipio:
        query["municipio"] = municipio
    
    # Aplicar filtro de vigencia
    vigencia_exportada = vigencia
    if vigencia:
        query["vigencia"] = vigencia
    else:
        # Si no se especifica vigencia, usar la más alta disponible
        all_vigencias = await db.predios.distinct("vigencia", {"deleted": {"$ne": True}})
        if all_vigencias:
            # Filtrar solo valores numéricos válidos
            numeric_vigencias = [v for v in all_vigencias if isinstance(v, (int, float)) and v is not None]
            if numeric_vigencias:
                vigencia_exportada = max(numeric_vigencias)
                query["vigencia"] = vigencia_exportada
    
    predios = await db.predios.find(query, {"_id": 0}).to_list(50000)
    
    # Crear workbook
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === HOJA REGISTRO_R1 (Propietarios) ===
    ws_r1 = wb.active
    ws_r1.title = "REGISTRO_R1"
    
    # Headers R1 - Formato actualizado con campos XTF
    headers_r1 = [
        "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL", 
        "CODIGO_HOMOLOGADO", "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS",
        "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "PRIMER_NOMBRE", "SEGUNDO_NOMBRE",
        "NOMBRE", "ESTADO", "TIPO_DOCUMENTO", "NUMERO_DOCUMENTO", "DIRECCION",
        "COMUNA", "DESTINO_ECONOMICO", "AREA_TERRENO", "AREA_CONSTRUIDA", "AVALUO",
        "VIGENCIA", "TIPO_MUTACION", "NO_RESOLUCION", "FECHA_RESOLUCION"
    ]
    
    for col, header in enumerate(headers_r1, 1):
        cell = ws_r1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Función para formatear número de documento con padding de 0s (12 dígitos)
    def formatear_documento(numero):
        if not numero:
            return ''
        # Eliminar caracteres no numéricos
        solo_numeros = ''.join(filter(str.isdigit, str(numero)))
        # Rellenar con ceros a la izquierda hasta 12 dígitos
        return solo_numeros.zfill(12) if solo_numeros else ''
    
    # Función para generar nombre completo desde campos separados
    def generar_nombre_completo(prop):
        partes = [
            prop.get('primer_apellido', ''),
            prop.get('segundo_apellido', ''),
            prop.get('primer_nombre', ''),
            prop.get('segundo_nombre', '')
        ]
        nombre = ' '.join(p for p in partes if p and p.strip())
        # Si no hay campos separados, usar nombre_propietario
        return nombre if nombre else prop.get('nombre_propietario', '')
    
    # Escribir datos R1
    row = 2
    for predio in predios:
        propietarios = predio.get('propietarios', [])
        if not propietarios:
            propietarios = [{
                'primer_apellido': predio.get('primer_apellido', ''),
                'segundo_apellido': predio.get('segundo_apellido', ''),
                'primer_nombre': predio.get('primer_nombre', ''),
                'segundo_nombre': predio.get('segundo_nombre', ''),
                'nombre_propietario': predio.get('nombre_propietario', ''),
                'estado': predio.get('estado', predio.get('estado_civil', '')),
                'tipo_documento': predio.get('tipo_documento', ''),
                'numero_documento': predio.get('numero_documento', '')
            }]
        
        total_props = len(propietarios)
        for idx, prop in enumerate(propietarios, 1):
            ws_r1.cell(row=row, column=1, value=predio.get('departamento', ''))
            ws_r1.cell(row=row, column=2, value=predio.get('municipio', ''))
            ws_r1.cell(row=row, column=3, value=predio.get('numero_predio', ''))
            ws_r1.cell(row=row, column=4, value=predio.get('codigo_predial_nacional', ''))
            ws_r1.cell(row=row, column=5, value=predio.get('codigo_homologado', ''))
            ws_r1.cell(row=row, column=6, value='1')  # Tipo de registro R1
            ws_r1.cell(row=row, column=7, value=str(idx).zfill(2))  # Número de orden
            ws_r1.cell(row=row, column=8, value=str(total_props).zfill(2))  # Total registros
            # Campos de nombre (formato XTF)
            ws_r1.cell(row=row, column=9, value=prop.get('primer_apellido', ''))
            ws_r1.cell(row=row, column=10, value=prop.get('segundo_apellido', ''))
            ws_r1.cell(row=row, column=11, value=prop.get('primer_nombre', ''))
            ws_r1.cell(row=row, column=12, value=prop.get('segundo_nombre', ''))
            ws_r1.cell(row=row, column=13, value=generar_nombre_completo(prop))  # Nombre completo
            ws_r1.cell(row=row, column=14, value=prop.get('estado', prop.get('estado_civil', '')))  # Estado
            ws_r1.cell(row=row, column=15, value=prop.get('tipo_documento', ''))
            ws_r1.cell(row=row, column=16, value=formatear_documento(prop.get('numero_documento', '')))  # Documento con padding
            ws_r1.cell(row=row, column=17, value=predio.get('direccion', ''))
            ws_r1.cell(row=row, column=18, value=predio.get('comuna', ''))
            ws_r1.cell(row=row, column=19, value=predio.get('destino_economico', ''))
            ws_r1.cell(row=row, column=20, value=predio.get('area_terreno', 0))
            ws_r1.cell(row=row, column=21, value=predio.get('area_construida', 0))
            ws_r1.cell(row=row, column=22, value=predio.get('avaluo', 0))
            # Vigencia: extraer solo el año si viene en formato largo
            vigencia_raw = predio.get('vigencia', datetime.now().year)
            if isinstance(vigencia_raw, str) and len(vigencia_raw) >= 4:
                vigencia = vigencia_raw[-4:]  # Tomar últimos 4 dígitos (año)
            else:
                vigencia = vigencia_raw
            ws_r1.cell(row=row, column=23, value=vigencia)
            ws_r1.cell(row=row, column=24, value=predio.get('tipo_mutacion', ''))
            ws_r1.cell(row=row, column=25, value=predio.get('numero_resolucion', ''))
            ws_r1.cell(row=row, column=26, value=predio.get('fecha_resolucion', ''))
            row += 1
    
    # === HOJA REGISTRO_R2 (Físico - Formato original con columnas horizontales) ===
    ws_r2 = wb.create_sheet(title="REGISTRO_R2")
    
    # Headers R2 - Formato original con zonas en columnas horizontales
    headers_r2 = [
        "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL",
        "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS", "MATRICULA_INMOBILIARIA",
        # Zona 1
        "ZONA_FISICA_1", "ZONA_ECONOMICA_1", "AREA_TERRENO_1",
        # Zona 2
        "ZONA_FISICA_2", "ZONA_ECONOMICA_2", "AREA_TERRENO_2",
        # Construcción 1
        "HABITACIONES_1", "BANOS_1", "LOCALES_1", "PISOS_1", "TIPIFICACION_1", "USO_1", "PUNTAJE_1", "AREA_CONSTRUIDA_1",
        # Construcción 2
        "HABITACIONES_2", "BANOS_2", "LOCALES_2", "PISOS_2", "TIPIFICACION_2", "USO_2", "PUNTAJE_2", "AREA_CONSTRUIDA_2",
        # Construcción 3
        "HABITACIONES_3", "BANOS_3", "LOCALES_3", "PISOS_3", "TIPIFICACION_3", "USO_3", "PUNTAJE_3", "AREA_CONSTRUIDA_3",
        "VIGENCIA"
    ]
    
    for col, header in enumerate(headers_r2, 1):
        cell = ws_r2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos R2 - Una fila por registro R2 con zonas en columnas
    row = 2
    for predio in predios:
        r2_registros = predio.get('r2_registros', [])
        total_r2 = len(r2_registros) if r2_registros else 0
        
        # Si el predio tiene el nuevo formato (zonas y construcciones separadas), convertir a r2_registros
        if not r2_registros and (predio.get('zonas') or predio.get('construcciones')):
            zonas = predio.get('zonas', [])
            construcciones = predio.get('construcciones', [])
            matricula = predio.get('matricula_inmobiliaria', '')
            
            # Crear un r2_registro combinando zonas y construcciones
            zonas_combinadas = []
            max_items = max(len(zonas), len(construcciones))
            for i in range(max_items):
                zona_data = zonas[i] if i < len(zonas) else {}
                const_data = construcciones[i] if i < len(construcciones) else {}
                zonas_combinadas.append({
                    'zona_fisica': zona_data.get('zona_fisica', 0),
                    'zona_economica': zona_data.get('zona_economica', 0),
                    'area_terreno': zona_data.get('area_terreno', 0),
                    'habitaciones': const_data.get('habitaciones', 0),
                    'banos': const_data.get('banos', 0),
                    'locales': const_data.get('locales', 0),
                    'pisos': const_data.get('piso', const_data.get('pisos', 0)),
                    'tipificacion': const_data.get('tipificacion', ''),
                    'uso': const_data.get('uso', ''),
                    'puntaje': const_data.get('puntaje', 0),
                    'area_construida': const_data.get('area_construida', 0)
                })
            
            r2_registros = [{
                'matricula_inmobiliaria': matricula,
                'zonas': zonas_combinadas
            }]
            total_r2 = 1
        
        for r2_idx, r2 in enumerate(r2_registros, 1):
            ws_r2.cell(row=row, column=1, value=predio.get('departamento', ''))
            ws_r2.cell(row=row, column=2, value=predio.get('municipio', ''))
            ws_r2.cell(row=row, column=3, value=predio.get('numero_predio', ''))
            ws_r2.cell(row=row, column=4, value=predio.get('codigo_predial_nacional', ''))
            ws_r2.cell(row=row, column=5, value='2')
            ws_r2.cell(row=row, column=6, value=str(r2_idx).zfill(2))
            ws_r2.cell(row=row, column=7, value=str(total_r2).zfill(2))
            ws_r2.cell(row=row, column=8, value=r2.get('matricula_inmobiliaria', ''))
            
            zonas = r2.get('zonas', [])
            
            # Zona 1 (columnas 9-11) - siempre llenar con 0 si vacío
            z1 = zonas[0] if len(zonas) >= 1 else {}
            ws_r2.cell(row=row, column=9, value=z1.get('zona_fisica', 0) or 0)
            ws_r2.cell(row=row, column=10, value=z1.get('zona_economica', 0) or 0)
            ws_r2.cell(row=row, column=11, value=z1.get('area_terreno', 0) or 0)
            
            # Zona 2 (columnas 12-14)
            z2 = zonas[1] if len(zonas) >= 2 else {}
            ws_r2.cell(row=row, column=12, value=z2.get('zona_fisica', 0) or 0)
            ws_r2.cell(row=row, column=13, value=z2.get('zona_economica', 0) or 0)
            ws_r2.cell(row=row, column=14, value=z2.get('area_terreno', 0) or 0)
            
            # Construcción 1 (columnas 15-22)
            ws_r2.cell(row=row, column=15, value=z1.get('habitaciones', 0) or 0)
            ws_r2.cell(row=row, column=16, value=z1.get('banos', 0) or 0)
            ws_r2.cell(row=row, column=17, value=z1.get('locales', 0) or 0)
            ws_r2.cell(row=row, column=18, value=z1.get('pisos', 0) or 0)
            ws_r2.cell(row=row, column=19, value=z1.get('tipificacion', 0) or 0)
            ws_r2.cell(row=row, column=20, value=z1.get('uso', 0) or 0)
            ws_r2.cell(row=row, column=21, value=z1.get('puntaje', 0) or 0)
            ws_r2.cell(row=row, column=22, value=z1.get('area_construida', 0) or 0)
            
            # Construcción 2 (columnas 23-30)
            ws_r2.cell(row=row, column=23, value=z2.get('habitaciones', 0) or 0)
            ws_r2.cell(row=row, column=24, value=z2.get('banos', 0) or 0)
            ws_r2.cell(row=row, column=25, value=z2.get('locales', 0) or 0)
            ws_r2.cell(row=row, column=26, value=z2.get('pisos', 0) or 0)
            ws_r2.cell(row=row, column=27, value=z2.get('tipificacion', 0) or 0)
            ws_r2.cell(row=row, column=28, value=z2.get('uso', 0) or 0)
            ws_r2.cell(row=row, column=29, value=z2.get('puntaje', 0) or 0)
            ws_r2.cell(row=row, column=30, value=z2.get('area_construida', 0) or 0)
            
            # Construcción 3 (columnas 31-38)
            z3 = zonas[2] if len(zonas) >= 3 else {}
            ws_r2.cell(row=row, column=31, value=z3.get('habitaciones', 0) or 0)
            ws_r2.cell(row=row, column=32, value=z3.get('banos', 0) or 0)
            ws_r2.cell(row=row, column=33, value=z3.get('locales', 0) or 0)
            ws_r2.cell(row=row, column=34, value=z3.get('pisos', 0) or 0)
            ws_r2.cell(row=row, column=35, value=z3.get('tipificacion', 0) or 0)
            ws_r2.cell(row=row, column=36, value=z3.get('uso', 0) or 0)
            ws_r2.cell(row=row, column=37, value=z3.get('puntaje', 0) or 0)
            ws_r2.cell(row=row, column=38, value=z3.get('area_construida', 0) or 0)
            
            # Vigencia
            ws_r2.cell(row=row, column=39, value=predio.get('vigencia', datetime.now().year))
            row += 1
    
    # Ajustar anchos de columna
    for ws in [ws_r1, ws_r2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generar nombre de archivo con vigencia incluida
    fecha = datetime.now().strftime('%Y%m%d')
    vigencia_str = f"_Vigencia{vigencia_exportada}" if vigencia_exportada else ""
    filename = f"Predios_{municipio or 'Todos'}{vigencia_str}_{fecha}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# ===== CERTIFICADO CATASTRAL CON VERIFICACIÓN =====

def generar_codigo_verificacion():
    """Genera un código único de verificación para el certificado"""
    año = datetime.now().year
    aleatorio = uuid.uuid4().hex[:8].upper()
    return f"ASM-{año}-CC-{aleatorio}"

def generar_hash_documento(contenido: str) -> str:
    """Genera hash SHA256 del contenido para verificar integridad"""
    return hashlib.sha256(contenido.encode()).hexdigest()[:16].upper()

def generar_qr_verificacion(codigo_verificacion: str) -> bytes:
    """Genera código QR para verificación"""
    # La URL incluye /api porque el endpoint está en el router de API
    url = f"{VERIFICACION_BASE_URL}/api/verificar/{codigo_verificacion}"
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="#009846", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer.getvalue()

def generate_certificado_catastral(predio: dict, firmante: dict, proyectado_por: str, codigo_verificacion: str, radicado: str = None) -> bytes:
    """
    Genera un certificado catastral en PDF con soporte multi-página.
    Incluye QR de verificación y código de seguridad.
    
    Si se proporciona radicado, se usa en el campo correspondiente (viene de una petición).
    Si no hay radicado, el campo queda editable.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import simpleSplit
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    # Registrar fuente Carlito (reemplazo métrico de Calibri Light)
    try:
        pdfmetrics.registerFont(TTFont('Carlito', '/usr/share/fonts/truetype/crosextra/Carlito-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('Carlito-Bold', '/usr/share/fonts/truetype/crosextra/Carlito-Bold.ttf'))
        fuente_normal = 'Carlito'
        fuente_bold = 'Carlito-Bold'
    except:
        fuente_normal = 'Helvetica'
        fuente_bold = 'Helvetica-Bold'
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Colores - Verde institucional (#009846)
    verde_institucional = colors.HexColor('#009846')
    negro = colors.HexColor('#000000')
    gris_texto = colors.HexColor('#333333')
    gris_claro = colors.HexColor('#666666')
    linea_gris = colors.HexColor('#cccccc')
    blanco = colors.HexColor('#FFFFFF')
    
    # Márgenes hoja carta estándar
    left_margin = 2.0 * cm
    right_margin = width - 2.0 * cm
    content_width = right_margin - left_margin
    
    # Límite inferior para contenido (antes del pie de página)
    footer_limit = 2.5 * cm
    
    fecha_actual = datetime.now()
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 
             'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    
    # Importar imágenes embebidas en Base64 (funcionan en cualquier servidor)
    from certificado_images import get_encabezado_image, get_pie_pagina_image, get_firma_dalgie_image
    from reportlab.lib.utils import ImageReader
    
    # Obtener las imágenes desde Base64 (siempre disponibles)
    try:
        encabezado_img = ImageReader(get_encabezado_image())
        pie_pagina_img = ImageReader(get_pie_pagina_image())
        firma_dalgie_img = ImageReader(get_firma_dalgie_image())
        imagenes_embebidas_ok = True
        print(f"📋 Generando certificado - Usando imágenes embebidas (Base64)")
    except Exception as e:
        print(f"⚠️ Error cargando imágenes embebidas: {e}")
        imagenes_embebidas_ok = False
    
    print(f"   - URL Verificación QR: {VERIFICACION_BASE_URL}/api/verificar/...")
    
    # Variables para manejo de páginas
    current_page = 1
    total_pages = 1  # Se calculará después
    
    def draw_header(is_first_page=True):
        """Dibuja el encabezado en cada página usando imágenes embebidas"""
        if imagenes_embebidas_ok:
            encabezado_width = content_width + 1 * cm
            encabezado_height = 2.0 * cm
            encabezado_x = left_margin - 0.5 * cm
            encabezado_y = height - 2.2 * cm
            c.drawImage(encabezado_img, encabezado_x, encabezado_y, 
                        width=encabezado_width, height=encabezado_height, 
                        preserveAspectRatio=True, mask='auto')
        else:
            # Encabezado alternativo si las imágenes embebidas fallaron
            c.setFillColor(verde_institucional)
            c.setFont(fuente_bold, 14)
            c.drawCentredString(width/2, height - 1.5 * cm, "ASOMUNICIPIOS - Gestor Catastral")
        return height - 2.8 * cm
    
    def draw_footer():
        """Dibuja el pie de página en cada página usando imágenes embebidas"""
        if imagenes_embebidas_ok:
            footer_height = 2.0 * cm
            c.drawImage(pie_pagina_img, 0, 0, 
                        width=width, height=footer_height, 
                        preserveAspectRatio=False, mask='auto')
        else:
            c.setFillColor(verde_institucional)
            c.rect(0, 0, width, 28, fill=1, stroke=0)
            c.setFillColor(blanco)
            c.setFont(fuente_normal, 8)
            c.drawCentredString(width/2, 10, "comunicaciones@asomunicipios.gov.co")
    
    def new_page():
        """Crea una nueva página con encabezado y pie de página"""
        nonlocal current_page
        draw_footer()
        c.showPage()
        current_page += 1
        y = draw_header(is_first_page=False)
        
        # Título de continuación
        c.setFillColor(negro)
        c.setFont(fuente_bold, 12)
        c.drawCentredString(width/2, y, "CERTIFICADO CATASTRAL SENCILLO (Continuación)")
        y -= 14
        
        # Número de certificado y predio para referencia
        c.setFont(fuente_normal, 9)
        c.setFillColor(gris_claro)
        codigo_predial = predio.get('codigo_predial_nacional', '')
        c.drawCentredString(width/2, y, f"Predio: {codigo_predial}")
        y -= 20
        
        return y
    
    def check_page_break(y, needed_space=20):
        """Verifica si hay espacio suficiente, si no, crea nueva página"""
        if y < footer_limit + needed_space:
            return new_page()
        return y
    
    # ===============================================
    # === PÁGINA 1: CONTENIDO PRINCIPAL ===
    # ===============================================
    
    y = draw_header(is_first_page=True)
    
    # Fecha (izquierda) y Certificado N° + Radicado N° (derecha)
    fecha_str = f"{fecha_actual.day} de {meses[fecha_actual.month-1]} del {fecha_actual.year}"
    c.setFont(fuente_normal, 11)
    c.setFillColor(negro)
    c.drawString(left_margin, y, fecha_str)
    
    # Certificado N°: (campo editable)
    c.drawRightString(right_margin - 140, y, "Certificado N°:")
    c.acroForm.textfield(
        name='certificado_numero',
        x=right_margin - 135,
        y=y - 4,
        width=135,
        height=16,
        fontSize=11,
        fontName='Helvetica',
        borderWidth=0,
        fillColor=colors.white,
        textColor=negro,
        value=''
    )
    
    # Radicado N°: debajo de Certificado N°:
    y -= 18
    c.setFillColor(negro)
    c.setFont(fuente_normal, 11)
    c.drawRightString(right_margin - 100, y, "Radicado N°:")
    
    if radicado:
        # Radicado fijo (viene de una petición)
        c.setFont(fuente_bold, 11)
        c.setFillColor(verde_institucional)
        c.drawString(right_margin - 95, y, radicado)
    else:
        # Campo editable (certificado manual)
        c.acroForm.textfield(
            name='radicado',
            x=right_margin - 95,
            y=y - 4,
            width=95,
            height=16,
            fontSize=11,
            fontName='Helvetica',
            borderWidth=0,
            fillColor=colors.white,
            textColor=negro,
            value=''
        )
    
    # Título (con espacio adicional después del Radicado)
    y -= 1.2 * cm
    c.setFillColor(negro)
    c.setFont(fuente_bold, 14)
    c.drawCentredString(width/2, y, "CERTIFICADO CATASTRAL SENCILLO")
    y -= 14
    
    # Base legal
    c.setFillColor(gris_claro)
    c.setFont(fuente_normal, 8)
    c.drawCentredString(width/2, y, "ESTE CERTIFICADO TIENE VALIDEZ DE ACUERDO CON LA LEY 527 DE 1999 (AGOSTO 18)")
    y -= 10
    c.drawCentredString(width/2, y, "Directiva Presidencial No. 02 del 2000, Ley 962 de 2005 (Anti trámites), Artículo 6, Parágrafo 3.")
    y -= 14
    
    # Texto certificador
    c.setFillColor(negro)
    c.setFont(fuente_normal, 11)
    c.drawCentredString(width/2, y, "La Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar - Asomunicipios,")
    y -= 14
    c.drawCentredString(width/2, y, "certifica que el siguiente predio se encuentra inscrito en la base de datos catastral con la siguiente información:")
    y -= 18
    
    # Variables para el cuadro
    cuadro_x = left_margin
    cuadro_width = content_width
    cuadro_start_y = y
    
    def draw_section_header(title, y_pos):
        """Dibuja encabezado de sección (barra verde) - CENTRADO"""
        y_pos = check_page_break(y_pos, 30)
        c.setFillColor(verde_institucional)
        c.rect(cuadro_x, y_pos - 14, cuadro_width, 16, fill=1, stroke=0)
        c.setFillColor(blanco)
        c.setFont(fuente_bold, 11)
        c.drawCentredString(cuadro_x + cuadro_width/2, y_pos - 10, title)
        return y_pos - 18
    
    def draw_field(label, value, y_pos, label_width=180):
        """Dibuja una fila de campo"""
        y_pos = check_page_break(y_pos, 18)
        c.setFillColor(negro)
        c.setFont(fuente_bold, 10)
        c.drawString(cuadro_x + 5, y_pos - 10, label)
        c.setFont(fuente_normal, 10)
        value_str = str(value) if value else ""
        c.drawString(cuadro_x + label_width, y_pos - 10, value_str)
        c.setStrokeColor(linea_gris)
        c.setLineWidth(0.5)
        c.line(cuadro_x, y_pos - 14, cuadro_x + cuadro_width, y_pos - 14)
        return y_pos - 16
    
    # === SECCIÓN 1: INFORMACIÓN CATASTRAL DEL PREDIO ===
    y = draw_section_header("INFORMACIÓN CATASTRAL DEL PREDIO", y)
    
    # === SECCIÓN 2: INFORMACIÓN JURÍDICA ===
    y = draw_section_header("INFORMACIÓN JURÍDICA", y)
    
    # Obtener propietarios del R1/R2
    propietarios = predio.get('propietarios', [])
    r2_registros = predio.get('r2_registros', [])
    
    # Mostrar TODOS los propietarios (con salto de página si es necesario)
    if propietarios:
        for i, prop in enumerate(propietarios, 1):
            nombre = prop.get('nombre_propietario', '')
            tipo_doc = prop.get('tipo_documento', 'N')
            num_doc = prop.get('numero_documento', '')
            derecho = prop.get('tipo_derecho', '')
            
            y = draw_field(f"Propietario {i}", nombre, y)
            y = draw_field("Tipo documento", tipo_doc, y)
            y = draw_field("Número documento", num_doc, y)
            if derecho:
                y = draw_field("Tipo derecho", derecho, y)
    else:
        y = draw_field("Propietario", predio.get('nombre_propietario', 'N/A'), y)
        y = draw_field("Tipo documento", predio.get('tipo_documento', 'N'), y)
        y = draw_field("Número documento", predio.get('numero_documento', ''), y)
    
    # Matrícula inmobiliaria del R2
    matricula = ''
    if r2_registros:
        matricula = r2_registros[0].get('matricula_inmobiliaria', '')
    y = draw_field("Matrícula inmobiliaria", matricula or 'Sin matrícula', y)
    
    # === SECCIÓN 3: INFORMACIÓN FÍSICA ===
    y = draw_section_header("INFORMACIÓN FÍSICA", y)
    
    municipio = predio.get('municipio', '')
    if municipio in ['Río de Oro', 'Rio de Oro']:
        depto_cod = "20 - CESAR"
        muni_cod = "614 - RIO DE ORO"
    else:
        depto_cod = "54 - NORTE DE SANTANDER"
        muni_mapping = {
            'Ábrego': '003 - ÁBREGO', 'Bucarasica': '109 - BUCARASICA',
            'Convención': '206 - CONVENCIÓN', 'Cáchira': '128 - CÁCHIRA',
            'El Carmen': '245 - EL CARMEN', 'El Tarra': '250 - EL TARRA',
            'Hacarí': '344 - HACARÍ', 'La Playa': '398 - LA PLAYA',
            'San Calixto': '670 - SAN CALIXTO', 'Sardinata': '720 - SARDINATA',
            'Teorama': '800 - TEORAMA'
        }
        muni_cod = muni_mapping.get(municipio, municipio)
    
    y = draw_field("Departamento", depto_cod, y)
    y = draw_field("Municipio", muni_cod, y)
    y = draw_field("Número predial", predio.get('codigo_predial_nacional', ''), y)
    y = draw_field("Código homologado", predio.get('codigo_homologado', ''), y)
    y = draw_field("Dirección", predio.get('direccion', ''), y)
    
    # Área terreno - Formato: "39 ha 3750 m²"
    area_terreno = predio.get('area_terreno', 0)
    if area_terreno and area_terreno >= 10000:
        ha = int(area_terreno // 10000)
        m2_restantes = int(area_terreno % 10000)
        area_str = f"{ha} ha {m2_restantes} m²"
    elif area_terreno:
        area_str = f"{int(area_terreno)} m²"
    else:
        area_str = "N/A"
    y = draw_field("Área terreno", area_str, y)
    
    area_construida = predio.get('area_construida', 0)
    area_const_str = f"{int(area_construida)} m²" if area_construida else "0 m²"
    y = draw_field("Área construida", area_const_str, y)
    
    # === SECCIÓN 4: INFORMACIÓN ECONÓMICA ===
    y = draw_section_header("INFORMACIÓN ECONÓMICA", y)
    
    avaluo = predio.get('avaluo', 0)
    avaluo_str = f"$ {int(avaluo):,}".replace(',', '.') if avaluo else "N/A"
    y = draw_field("Avalúo catastral", avaluo_str, y)
    
    # Verificar espacio para el texto de expedición y firma
    y = check_page_break(y, 120)
    
    # === TEXTO DE EXPEDICIÓN ===
    y -= 10
    c.setFillColor(negro)
    c.setFont(fuente_normal, 11)
    fecha_exp = f"{fecha_actual.day} de {meses[fecha_actual.month-1]} del {fecha_actual.year}"
    texto_exp = f"El presente certificado se expide a favor del interesado el {fecha_exp}."
    c.drawString(left_margin, y, texto_exp)
    y -= 20
    
    # === SECCIÓN DE FIRMA Y VERIFICACIÓN (lado a lado, centrados) ===
    y = check_page_break(y, 110)
    
    # Generar QR de verificación
    from reportlab.lib.utils import ImageReader
    qr_bytes = generar_qr_verificacion(codigo_verificacion)
    qr_buffer = io.BytesIO(qr_bytes)
    qr_image = ImageReader(qr_buffer)
    
    # Calcular fecha/hora de generación
    fecha_hora_gen = fecha_actual.strftime("%d/%m/%Y %H:%M")
    
    # Hash del documento (simplificado)
    hash_doc = generar_hash_documento(f"{predio.get('codigo_predial_nacional', '')}-{codigo_verificacion}-{fecha_hora_gen}")
    
    # Calcular posiciones centradas para ambos bloques
    # Ancho total disponible
    total_width = content_width
    firma_block_width = 180  # Ancho del bloque de firma
    verif_block_width = 185  # Ancho del bloque de verificación
    gap_between = 30  # Espacio entre ambos bloques
    
    # Calcular inicio para centrar ambos bloques
    total_blocks_width = firma_block_width + gap_between + verif_block_width
    start_x = left_margin + (total_width - total_blocks_width) / 2
    
    # === BLOQUE IZQUIERDO: FIRMA ===
    firma_block_y = y - 65
    
    if imagenes_embebidas_ok:
        # Dibujar la firma usando imagen embebida
        firma_width = 90
        firma_height = 54
        firma_x = start_x + (firma_block_width - firma_width) / 2
        firma_y = firma_block_y
        
        # Dibujar la imagen de la firma embebida
        c.drawImage(firma_dalgie_img, firma_x, firma_y, width=firma_width, height=firma_height, 
                    preserveAspectRatio=True, mask='auto')
    else:
        # Si las imágenes embebidas fallaron, mostrar línea para firma manual
        firma_x = start_x + 20
        firma_y = firma_block_y + 30
        c.setStrokeColor(negro)
        c.setLineWidth(0.5)
        c.line(firma_x, firma_y, firma_x + 140, firma_y)
        c.setFont(fuente_normal, 7)
        c.setFillColor(gris_claro)
        c.drawCentredString(firma_x + 70, firma_y + 5, "(Firma)")
    
    # Nombre y cargo debajo de la firma (centrados en el bloque)
    nombre_y = firma_block_y - 10
    c.setFillColor(negro)
    c.setFont(fuente_bold, 9)
    c.drawString(start_x, nombre_y, firmante.get('full_name', 'DALGIE ESPERANZA TORRADO RIZO'))
    
    cargo_y = nombre_y - 10
    c.setFont(fuente_normal, 8)
    c.drawString(start_x, cargo_y, firmante.get('cargo', 'Subdirectora Financiera y Administrativa'))
    
    # === BLOQUE DERECHO: CUADRO DE VERIFICACIÓN ===
    marco_width = verif_block_width
    marco_height = 58
    marco_x = start_x + firma_block_width + gap_between
    marco_y = y - 68
    
    # Fondo del marco
    c.setFillColor(colors.HexColor('#f0fdf4'))  # Verde muy claro
    c.roundRect(marco_x, marco_y, marco_width, marco_height, 5, fill=1, stroke=0)
    
    # Borde verde
    c.setStrokeColor(verde_institucional)
    c.setLineWidth(1.5)
    c.roundRect(marco_x, marco_y, marco_width, marco_height, 5, fill=0, stroke=1)
    
    # QR dentro del marco (izquierda del bloque)
    qr_size = 46
    c.drawImage(qr_image, marco_x + 5, marco_y + 6, width=qr_size, height=qr_size, mask='auto')
    
    # Información de verificación (derecha del QR)
    info_x = marco_x + 55
    
    # Título
    c.setFillColor(verde_institucional)
    c.setFont(fuente_bold, 8)
    c.drawString(info_x, marco_y + marco_height - 11, "CERTIFICADO VERIFICABLE")
    
    # Código
    c.setFillColor(negro)
    c.setFont(fuente_bold, 7)
    c.drawString(info_x, marco_y + 36, f"Código: {codigo_verificacion}")
    
    # Detalles
    c.setFont(fuente_normal, 6)
    c.setFillColor(gris_claro)
    c.drawString(info_x, marco_y + 26, f"Generado: {fecha_hora_gen}")
    c.drawString(info_x, marco_y + 17, f"Hash: SHA256:{hash_doc[:12]}...")
    
    # URL de verificación
    c.setFillColor(verde_institucional)
    c.setFont(fuente_normal, 6)
    c.drawString(info_x, marco_y + 8, "Escanear QR para verificar")
    
    # Calcular posición Y para "Elaboró" (debajo de ambos bloques)
    y = min(cargo_y - 15, marco_y - 10)
    
    # === ELABORÓ ===
    c.setFont(fuente_normal, 8)
    c.setFillColor(negro)
    c.drawString(left_margin, y, f"Elaboró: {proyectado_por}")
    y -= 14
    
    # === NOTAS ===
    c.setFillColor(negro)
    c.setFont(fuente_bold, 8)
    c.drawString(left_margin, y, "NOTA:")
    y -= 10
    
    c.setFont(fuente_normal, 7)
    notas = [
        "• La presente información no sirve como prueba para establecer actos constitutivos de posesión.",
        "• De conformidad con el artículo 2.2.2.2.8 del Decreto 148 de 2020, Inscripción o incorporación catastral.",
        "• Adicionalmente de conformidad con el artículo 29 de la resolución No. 1149 de 2021 del IGAC.",
        "• La base catastral de Asomunicipios sólo incluye información de los municipios habilitados.",
        "• Ante cualquier inquietud: comunicaciones@asomunicipios.gov.co",
    ]
    
    for nota in notas:
        y = check_page_break(y, 10)
        c.drawString(left_margin, y, nota)
        y -= 8
    
    # === PIE DE PÁGINA FINAL ===
    draw_footer()
    
    c.save()
    return buffer.getvalue()


@api_router.get("/predios/{predio_id}/certificado")
async def generar_certificado_catastral_endpoint(
    predio_id: str, 
    current_user: dict = Depends(get_current_user)
):
    """
    Genera un certificado catastral PDF para un predio específico.
    Incluye QR de verificación y código de seguridad verificable en línea.
    """
    # Solo coordinador, administrador y atencion_usuario pueden generar certificados
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=403, detail="No tiene permiso para generar certificados")
    
    # Obtener predio
    predio = await db.predios.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Verificar que el predio tenga información de propietarios (R1/R2)
    propietarios = predio.get('propietarios', [])
    if not propietarios or len(propietarios) == 0:
        raise HTTPException(
            status_code=400,
            detail="Certificado no disponible. El predio no tiene información de propietarios (R1/R2) registrada en la base de datos catastral."
        )
    
    # Generar código de verificación único
    codigo_verificacion = generar_codigo_verificacion()
    
    # Extraer datos críticos del predio para verificación
    datos_criticos = {
        "codigo_predial": predio.get('codigo_predial_nacional', ''),
        "propietarios": [p.get('nombre_propietario', '') for p in propietarios],
        "area_terreno": predio.get('area_terreno', ''),
        "area_construida": predio.get('area_construida', ''),
        "avaluo_catastral": predio.get('avaluo_catastral', ''),
        "direccion": predio.get('direccion', ''),
        "municipio": predio.get('municipio', ''),
        "matricula_inmobiliaria": next((r.get('matricula_inmobiliaria', '') for r in predio.get('r2_registros', [])), '')
    }
    
    # Formatear área para mostrar en verificación
    area_raw = predio.get('area_terreno', 0)
    if area_raw and area_raw >= 10000:
        ha = int(area_raw // 10000)
        m2_restantes = int(area_raw % 10000)
        area_str = f"{ha} ha {m2_restantes} m²"
    elif area_raw:
        area_str = f"{int(area_raw)} m²"
    else:
        area_str = "N/A"
    
    # Formatear avalúo
    avaluo_raw = predio.get('avaluo', 0)
    avaluo_str = f"$ {int(avaluo_raw):,}".replace(',', '.') if avaluo_raw else "N/A"
    
    # Generar hash de datos críticos (para comparación rápida)
    datos_string = "|".join([
        str(datos_criticos['codigo_predial']),
        ",".join(datos_criticos['propietarios']),
        str(datos_criticos['area_terreno']),
        str(datos_criticos['avaluo_catastral'])
    ])
    hash_datos = hashlib.sha256(datos_string.encode()).hexdigest()
    
    # Registrar certificado en la base de datos
    certificado_record = {
        "id": str(uuid.uuid4()),
        "codigo_verificacion": codigo_verificacion,
        "hash_datos": hash_datos,  # Hash de datos críticos
        "hash_pdf": "",  # Se actualizará después de generar el PDF
        "predio_id": predio_id,
        "codigo_predial": predio.get('codigo_predial_nacional', ''),
        "municipio": predio.get('municipio', ''),
        "direccion": predio.get('direccion', ''),
        "propietarios": [p.get('nombre_propietario', '') for p in propietarios],
        "datos_criticos": datos_criticos,  # Guardar datos originales
        # Datos formateados para verificación
        "area_terreno": area_str,
        "avaluo_catastral": avaluo_str,
        "generado_por": current_user['id'],
        "generado_por_nombre": current_user['full_name'],
        "generado_por_email": current_user['email'],
        "fecha_generacion": datetime.now(timezone.utc).isoformat(),
        "estado": "activo",  # activo, anulado
        "motivo_anulacion": None,
        "fecha_anulacion": None,
        "anulado_por": None
    }
    await db.certificados_verificables.insert_one(certificado_record)
    
    # Firmante siempre es Dalgie Esperanza Torrado Rizo
    firmante = {
        "full_name": "DALGIE ESPERANZA TORRADO RIZO",
        "cargo": "Subdirectora Financiera y Administrativa"
    }
    
    # Quien proyecta es el usuario actual
    proyectado_por = current_user['full_name']
    
    # Generar PDF con verificación
    pdf_bytes = generate_certificado_catastral(predio, firmante, proyectado_por, codigo_verificacion)
    
    # Calcular hash del PDF generado
    hash_pdf = hashlib.sha256(pdf_bytes).hexdigest()
    
    # Actualizar el certificado con el hash del PDF
    await db.certificados_verificables.update_one(
        {"codigo_verificacion": codigo_verificacion},
        {"$set": {"hash_pdf": hash_pdf}}
    )
    
    # Guardar temporalmente
    temp_path = UPLOAD_DIR / f"certificado_{predio_id}_{uuid.uuid4()}.pdf"
    with open(temp_path, 'wb') as f:
        f.write(pdf_bytes)
    
    # Nombre del archivo
    codigo = predio.get('codigo_predial_nacional', predio_id)
    filename = f"Certificado_Catastral_{codigo}.pdf"
    
    return FileResponse(
        path=temp_path,
        filename=filename,
        media_type='application/pdf'
    )


@api_router.get("/petitions/{petition_id}/certificado")
async def generar_certificado_desde_peticion(
    petition_id: str,
    enviar_correo: bool = Query(False, description="Si es True, envía el certificado por correo y finaliza el trámite"),
    current_user: dict = Depends(get_current_user)
):
    """
    Genera un certificado catastral PDF desde una petición.
    El radicado de la petición se usa automáticamente en el PDF.
    Si enviar_correo=True, envía al peticionario y finaliza el trámite.
    """
    # Solo coordinador, administrador y atencion_usuario pueden generar certificados
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=403, detail="No tiene permiso para generar certificados")
    
    # Buscar la petición
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    # Verificar que sea un tipo de certificado
    tipo = petition.get('tipo_tramite', '').lower()
    if 'certificado' not in tipo:
        raise HTTPException(status_code=400, detail="Esta petición no es de tipo certificado catastral")
    
    # Buscar el predio relacionado
    predio = None
    predio_relacionado = petition.get('predio_relacionado')
    
    if predio_relacionado and predio_relacionado.get('predio_id'):
        predio = await db.predios.find_one({"id": predio_relacionado['predio_id']}, {"_id": 0})
    
    # Si no hay predio relacionado, buscar por código o matrícula guardada
    if not predio:
        codigo_buscado = petition.get('codigo_predial_buscado')
        matricula_buscada = petition.get('matricula_buscada')
        
        if codigo_buscado:
            predio = await db.predios.find_one({"codigo_predial_nacional": codigo_buscado}, {"_id": 0})
        elif matricula_buscada:
            # Buscar matrícula en r2_registros.matricula_inmobiliaria (fuente R1/R2)
            predio = await db.predios.find_one(
                {"r2_registros.matricula_inmobiliaria": matricula_buscada}, 
                {"_id": 0}
            )
    
    if not predio:
        raise HTTPException(
            status_code=404, 
            detail="No se encontró el predio asociado a esta petición. La petición debe incluir un código predial nacional o matrícula inmobiliaria válidos. Si la petición es antigua, solicite al peticionario radicar una nueva con los datos correctos."
        )
    
    # Verificar que el predio tenga información de propietarios (R1/R2)
    propietarios = predio.get('propietarios', [])
    if not propietarios or len(propietarios) == 0:
        raise HTTPException(
            status_code=400,
            detail="Certificado no disponible. El predio no tiene información de propietarios (R1/R2) registrada en la base de datos catastral."
        )
    
    # Generar código de verificación
    codigo_verificacion = generar_codigo_verificacion()
    
    # Registrar en la base de datos de certificados verificables
    propietarios_nombres = [p.get('nombre_propietario', 'N/A') for p in propietarios] if propietarios else ['N/A']
    
    # Obtener área terreno y avalúo del predio
    area_terreno = predio.get('area_terreno', 0)
    avaluo = predio.get('avaluo', 0)
    
    # Formatear área para mostrar
    if area_terreno and area_terreno >= 10000:
        ha = int(area_terreno // 10000)
        m2_restantes = int(area_terreno % 10000)
        area_str = f"{ha} ha {m2_restantes} m²"
    elif area_terreno:
        area_str = f"{int(area_terreno)} m²"
    else:
        area_str = "N/A"
    
    # Formatear avalúo
    avaluo_str = f"$ {int(avaluo):,}".replace(',', '.') if avaluo else "N/A"
    
    certificado_record = {
        "id": str(uuid.uuid4()),
        "codigo_verificacion": codigo_verificacion,
        "predio_id": predio.get('id'),
        "codigo_predial": predio.get('codigo_predial_nacional'),
        "municipio": predio.get('municipio'),
        "direccion": predio.get('direccion'),
        "propietarios": propietarios_nombres,
        "fecha_generacion": datetime.now(timezone.utc).isoformat(),
        "generado_por": current_user['id'],
        "generado_por_nombre": current_user['full_name'],
        "estado": "activo",
        "hash_documento": "",
        # Datos adicionales para verificación
        "area_terreno": area_str,
        "avaluo_catastral": avaluo_str,
        # Vinculación con la petición
        "petition_id": petition_id,
        "radicado": petition.get('radicado'),
        "generado_desde_peticion": True
    }
    
    # Calcular hash después de generar el PDF
    hash_doc = generar_hash_documento(f"{predio.get('codigo_predial_nacional', '')}-{codigo_verificacion}-{datetime.now(timezone.utc).isoformat()}")
    certificado_record["hash_documento"] = hash_doc
    
    await db.certificados_verificables.insert_one(certificado_record)
    
    # Firmante siempre es Dalgie Esperanza Torrado Rizo
    firmante = {
        "full_name": "DALGIE ESPERANZA TORRADO RIZO",
        "cargo": "Subdirectora Financiera y Administrativa"
    }
    
    # Quien proyecta es el usuario actual
    proyectado_por = current_user['full_name']
    
    # Generar PDF con verificación y radicado de la petición
    pdf_bytes = generate_certificado_catastral(
        predio, 
        firmante, 
        proyectado_por, 
        codigo_verificacion,
        radicado=petition.get('radicado')  # Pasar el radicado de la petición
    )
    
    # Guardar PDF permanentemente para que el usuario pueda descargarlo después
    cert_filename = f"certificado_{petition_id}_{codigo_verificacion}.pdf"
    cert_path = UPLOAD_DIR / cert_filename
    with open(cert_path, 'wb') as f:
        f.write(pdf_bytes)
    
    # Extraer matrícula del R2 si existe
    matricula_r2 = None
    r2_registros = predio.get("r2_registros", [])
    if r2_registros:
        matricula_r2 = r2_registros[0].get("matricula_inmobiliaria")
    
    # Preparar información del predio relacionado
    predio_relacionado_info = {
        "predio_id": predio.get("id"),
        "codigo_predial": predio.get("codigo_predial_nacional"),
        "matricula": matricula_r2,
        "direccion": predio.get("direccion"),
        "municipio": predio.get("municipio")
    }
    
    # Actualizar la petición para indicar que se generó el certificado
    update_data = {
        "certificado_generado": True,
        "certificado_codigo": codigo_verificacion,
        "certificado_fecha": datetime.now(timezone.utc).isoformat(),
        "certificado_archivo": str(cert_path),
        "predio_relacionado": predio_relacionado_info,  # Guardar la info del predio
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Si enviar_correo es True, también finalizar el trámite
    if enviar_correo:
        update_data["estado"] = PetitionStatus.FINALIZADO
        
        # Agregar entrada al historial
        historial_entry = {
            "accion": "Certificado generado y enviado",
            "usuario": current_user['full_name'],
            "usuario_rol": current_user['role'],
            "estado_anterior": petition.get('estado'),
            "estado_nuevo": PetitionStatus.FINALIZADO,
            "notas": f"Certificado catastral generado con código {codigo_verificacion}. Enviado al correo del peticionario.",
            "fecha": datetime.now(timezone.utc).isoformat()
        }
        
        await db.petitions.update_one(
            {"id": petition_id},
            {
                "$set": update_data,
                "$push": {"historial": historial_entry}
            }
        )
        
        # Enviar correo al peticionario con el certificado adjunto
        try:
            radicado_pet = petition.get('radicado', petition_id)
            correo_destino = petition.get('correo')
            nombre_peticionario = petition.get('nombre_completo', 'Estimado usuario')
            codigo_predial = predio.get('codigo_predial_nacional', 'N/A')
            verificacion_url = f"{VERIFICACION_BASE_URL}/verificar/{codigo_verificacion}"
            
            email_html = f"""
            <html>
            <body style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="background: linear-gradient(135deg, #009846, #059669); padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">✅ Certificado Catastral Listo</h1>
                </div>
                <div style="background: #f8fafc; padding: 30px; border: 1px solid #e2e8f0; border-top: none; border-radius: 0 0 10px 10px;">
                    <p style="color: #334155; font-size: 16px;">Estimado/a <strong>{nombre_peticionario}</strong>,</p>
                    
                    <p style="color: #334155;">Su trámite de <strong>Certificado Catastral Sencillo</strong> ha sido aprobado y el certificado está listo.</p>
                    
                    <div style="background: white; border: 2px solid #009846; border-radius: 8px; padding: 20px; margin: 20px 0;">
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="padding: 8px 0; color: #64748b;">Radicado:</td>
                                <td style="padding: 8px 0; font-weight: bold; color: #009846;">{radicado_pet}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b;">Código de Verificación:</td>
                                <td style="padding: 8px 0; font-family: monospace; font-weight: bold;">{codigo_verificacion}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b;">Código Predial:</td>
                                <td style="padding: 8px 0; font-family: monospace;">{codigo_predial}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <p style="color: #334155;"><strong>📎 El certificado está adjunto a este correo.</strong></p>
                    
                    <p style="color: #334155;">También puede descargarlo ingresando a su cuenta en la plataforma de Asomunicipios.</p>
                    
                    <div style="text-align: center; margin: 25px 0;">
                        <a href="{verificacion_url}" style="background: #009846; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                            🔗 Verificar Certificado en Línea
                        </a>
                    </div>
                    
                    <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 25px 0;">
                    
                    <p style="color: #64748b; font-size: 12px; text-align: center;">
                        Asomunicipios - Gestor Catastral<br>
                        Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar<br>
                        comunicaciones@asomunicipios.gov.co | +57 3102327647
                    </p>
                </div>
            </body>
            </html>
            """
            
            # Enviar correo con adjunto
            await enviar_correo_con_adjunto(
                destinatario=correo_destino,
                asunto=f"✅ Certificado Catastral Listo - {radicado_pet}",
                contenido_html=email_html,
                adjunto_path=str(cert_path),
                adjunto_nombre=f"Certificado_Catastral_{radicado_pet}.pdf"
            )
            
        except Exception as e:
            print(f"Error enviando correo: {e}")
            # No fallar si el correo no se envía
        
        # Crear notificación en plataforma para el peticionario
        try:
            await crear_notificacion(
                usuario_id=petition.get('user_id'),
                tipo="success",
                titulo="✅ Certificado Catastral Listo",
                mensaje=f"Su certificado catastral ({radicado_pet}) ha sido generado y enviado a su correo. También puede descargarlo desde la plataforma.",
                enlace=f"/dashboard/peticiones/{petition_id}"
            )
        except Exception as e:
            print(f"Error creando notificación: {e}")
    else:
        # Solo actualizar sin finalizar
        await db.petitions.update_one(
            {"id": petition_id},
            {"$set": update_data}
        )
    
    # Nombre del archivo para descarga
    radicado_file = petition.get('radicado', petition_id)
    codigo_file = predio.get('codigo_predial_nacional', '')
    filename = f"Certificado_Catastral_{radicado_file}_{codigo_file}.pdf"
    
    return FileResponse(
        path=cert_path,
        filename=filename,
        media_type='application/pdf'
    )


@api_router.post("/petitions/{petition_id}/regenerar-certificado")
async def regenerar_certificado_desde_peticion(
    petition_id: str,
    enviar_correo: bool = Query(False, description="Si es True, envía el certificado por correo"),
    current_user: dict = Depends(get_current_user)
):
    """
    Regenera un certificado catastral para una petición que ya tiene un certificado previo.
    Útil cuando el certificado original ha vencido (vigencia de 1 mes) y se necesita uno actualizado
    con fecha reciente pero manteniendo el mismo radicado.
    
    - Genera un NUEVO código de verificación
    - Mantiene el mismo radicado de la petición original
    - Actualiza la fecha de generación
    - El certificado anterior queda inactivo pero se mantiene en el historial
    """
    # Solo coordinador, administrador y atencion_usuario pueden regenerar certificados
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=403, detail="No tiene permiso para regenerar certificados")
    
    # Buscar la petición
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    # Verificar que sea un tipo de certificado
    tipo = petition.get('tipo_tramite', '').lower()
    if 'certificado' not in tipo:
        raise HTTPException(status_code=400, detail="Esta petición no es de tipo certificado catastral")
    
    # Verificar que ya tenga un certificado generado previamente
    if not petition.get('certificado_generado'):
        raise HTTPException(
            status_code=400, 
            detail="Esta petición no tiene un certificado generado previamente. Use el endpoint normal de generación."
        )
    
    # Buscar el predio relacionado
    predio = None
    predio_relacionado = petition.get('predio_relacionado')
    
    if predio_relacionado and predio_relacionado.get('predio_id'):
        predio = await db.predios.find_one({"id": predio_relacionado['predio_id']}, {"_id": 0})
    
    # Si no hay predio relacionado, buscar por código o matrícula guardada
    if not predio:
        codigo_buscado = petition.get('codigo_predial_buscado')
        matricula_buscada = petition.get('matricula_buscada')
        
        if codigo_buscado:
            predio = await db.predios.find_one({"codigo_predial_nacional": codigo_buscado}, {"_id": 0})
        elif matricula_buscada:
            predio = await db.predios.find_one(
                {"r2_registros.matricula_inmobiliaria": matricula_buscada}, 
                {"_id": 0}
            )
    
    if not predio:
        raise HTTPException(
            status_code=404, 
            detail="No se encontró el predio asociado a esta petición."
        )
    
    # Verificar que el predio tenga información de propietarios
    propietarios = predio.get('propietarios', [])
    if not propietarios or len(propietarios) == 0:
        raise HTTPException(
            status_code=400,
            detail="Certificado no disponible. El predio no tiene información de propietarios (R1/R2) registrada."
        )
    
    # Marcar el certificado anterior como inactivo
    codigo_anterior = petition.get('certificado_codigo')
    if codigo_anterior:
        await db.certificados_verificables.update_one(
            {"codigo_verificacion": codigo_anterior},
            {"$set": {
                "estado": "reemplazado",
                "reemplazado_fecha": datetime.now(timezone.utc).isoformat(),
                "reemplazado_por": current_user['id']
            }}
        )
    
    # Generar NUEVO código de verificación
    codigo_verificacion = generar_codigo_verificacion()
    
    # Registrar el nuevo certificado en la base de datos
    propietarios_nombres = [p.get('nombre_propietario', 'N/A') for p in propietarios]
    
    # Obtener área terreno y avalúo del predio
    area_terreno = predio.get('area_terreno', 0)
    avaluo = predio.get('avaluo', 0)
    
    # Formatear área para mostrar
    if area_terreno and area_terreno >= 10000:
        ha = int(area_terreno // 10000)
        m2_restantes = int(area_terreno % 10000)
        area_str = f"{ha} ha {m2_restantes} m²"
    elif area_terreno:
        area_str = f"{int(area_terreno)} m²"
    else:
        area_str = "N/A"
    
    # Formatear avalúo
    avaluo_str = f"$ {int(avaluo):,}".replace(',', '.') if avaluo else "N/A"
    
    certificado_record = {
        "id": str(uuid.uuid4()),
        "codigo_verificacion": codigo_verificacion,
        "predio_id": predio.get('id'),
        "codigo_predial": predio.get('codigo_predial_nacional'),
        "municipio": predio.get('municipio'),
        "direccion": predio.get('direccion'),
        "propietarios": propietarios_nombres,
        "fecha_generacion": datetime.now(timezone.utc).isoformat(),
        "generado_por": current_user['id'],
        "generado_por_nombre": current_user['full_name'],
        "estado": "activo",
        "hash_documento": "",
        # Datos adicionales para verificación
        "area_terreno": area_str,
        "avaluo_catastral": avaluo_str,
        "petition_id": petition_id,
        "radicado": petition.get('radicado'),
        "generado_desde_peticion": True,
        "es_regeneracion": True,
        "certificado_anterior": codigo_anterior,
        "numero_regeneracion": petition.get('regeneraciones_count', 0) + 1
    }
    
    hash_doc = generar_hash_documento(f"{predio.get('codigo_predial_nacional', '')}-{codigo_verificacion}-{datetime.now(timezone.utc).isoformat()}")
    certificado_record["hash_documento"] = hash_doc
    
    await db.certificados_verificables.insert_one(certificado_record)
    
    # Firmante
    firmante = {
        "full_name": "DALGIE ESPERANZA TORRADO RIZO",
        "cargo": "Subdirectora Financiera y Administrativa"
    }
    
    proyectado_por = current_user['full_name']
    
    # Generar PDF con el mismo radicado de la petición original
    pdf_bytes = generate_certificado_catastral(
        predio, 
        firmante, 
        proyectado_por, 
        codigo_verificacion,
        radicado=petition.get('radicado')
    )
    
    # Guardar PDF
    cert_filename = f"certificado_{petition_id}_{codigo_verificacion}.pdf"
    cert_path = UPLOAD_DIR / cert_filename
    with open(cert_path, 'wb') as f:
        f.write(pdf_bytes)
    
    # Actualizar la petición con el nuevo certificado
    update_data = {
        "certificado_codigo": codigo_verificacion,
        "certificado_fecha": datetime.now(timezone.utc).isoformat(),
        "certificado_archivo": str(cert_path),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Historial de regeneración
    historial_entry = {
        "accion": "Certificado regenerado",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": petition.get('estado'),
        "estado_nuevo": petition.get('estado'),
        "notas": f"Certificado regenerado con nuevo código {codigo_verificacion}. Código anterior: {codigo_anterior}.",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    await db.petitions.update_one(
        {"id": petition_id},
        {
            "$set": update_data,
            "$push": {"historial": historial_entry},
            "$inc": {"regeneraciones_count": 1}
        }
    )
    
    # Si enviar_correo es True, enviar al peticionario
    if enviar_correo:
        try:
            radicado_pet = petition.get('radicado', petition_id)
            correo_destino = petition.get('correo')
            nombre_peticionario = petition.get('nombre_completo', 'Estimado usuario')
            codigo_predial = predio.get('codigo_predial_nacional', 'N/A')
            
            if correo_destino:
                subject = f"Certificado Catastral Actualizado - {radicado_pet}"
                
                html_body = f"""
                <div style="font-family: 'Segoe UI', sans-serif; max-width: 600px; margin: 0 auto;">
                    <div style="background-color: #009846; padding: 20px; text-align: center;">
                        <h1 style="color: white; margin: 0;">Asomunicipios</h1>
                        <p style="color: #e0e0e0; margin: 5px 0 0 0;">Gestión Catastral</p>
                    </div>
                    <div style="padding: 30px; background: #f8fafc;">
                        <h2 style="color: #009846;">Certificado Catastral Actualizado</h2>
                        <p style="color: #334155;">Estimado/a <strong>{nombre_peticionario}</strong>,</p>
                        <p style="color: #334155;">Se ha generado una versión actualizada de su certificado catastral.</p>
                        
                        <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #009846;">
                            <p style="margin: 5px 0;"><strong>Radicado:</strong> {radicado_pet}</p>
                            <p style="margin: 5px 0;"><strong>Código Predial:</strong> {codigo_predial}</p>
                            <p style="margin: 5px 0;"><strong>Nuevo Código de Verificación:</strong> {codigo_verificacion}</p>
                            <p style="margin: 5px 0;"><strong>Fecha de Regeneración:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                        </div>
                        
                        <p style="color: #334155;"><strong>📎 El certificado actualizado está adjunto a este correo.</strong></p>
                        
                        <p style="color: #64748b; font-size: 12px; margin-top: 30px;">
                            Este certificado tiene vigencia de un (1) mes a partir de su fecha de emisión.
                        </p>
                    </div>
                    <div style="background: #1e293b; color: #94a3b8; padding: 15px; text-align: center; font-size: 12px;">
                        <p style="margin: 0;">Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar</p>
                        <p style="margin: 5px 0 0 0;">comunicaciones@asomunicipios.gov.co</p>
                    </div>
                </div>
                """
                
                await send_email_with_attachment(
                    to_email=correo_destino,
                    subject=subject,
                    html_body=html_body,
                    attachment_path=str(cert_path),
                    attachment_name=f"Certificado_Catastral_{radicado_pet}.pdf"
                )
                
                # Agregar notificación de envío al historial
                await db.petitions.update_one(
                    {"id": petition_id},
                    {"$push": {"historial": {
                        "accion": "Certificado regenerado enviado por correo",
                        "usuario": current_user['full_name'],
                        "usuario_rol": current_user['role'],
                        "notas": f"Certificado actualizado enviado a {correo_destino}",
                        "fecha": datetime.now(timezone.utc).isoformat()
                    }}}
                )
        except Exception as e:
            print(f"Error enviando correo de certificado regenerado: {e}")
    
    # Nombre del archivo para descarga
    filename = f"Certificado_Catastral_{petition.get('radicado', petition_id)}.pdf"
    
    return FileResponse(
        path=cert_path,
        filename=filename,
        media_type='application/pdf'
    )


# === VERIFICACIÓN PÚBLICA DE CERTIFICADOS ===

@api_router.get("/petitions/{petition_id}/descargar-certificado")
async def descargar_certificado_peticion(
    petition_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Permite al usuario (peticionario) o staff descargar el certificado de su petición.
    """
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=404, detail="Petición no encontrada")
    
    # Verificar permisos: el peticionario o staff pueden descargar
    is_owner = petition.get('user_id') == current_user['id']
    is_staff = current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.ATENCION_USUARIO]
    
    if not is_owner and not is_staff:
        raise HTTPException(status_code=403, detail="No tiene permiso para descargar este certificado")
    
    # Verificar que el certificado existe
    if not petition.get('certificado_generado'):
        raise HTTPException(status_code=400, detail="El certificado aún no ha sido generado")
    
    cert_path = petition.get('certificado_archivo')
    if not cert_path or not os.path.exists(cert_path):
        raise HTTPException(status_code=404, detail="El archivo del certificado no se encuentra")
    
    radicado = petition.get('radicado', petition_id)
    filename = f"Certificado_Catastral_{radicado}.pdf"
    
    return FileResponse(
        path=cert_path,
        filename=filename,
        media_type='application/pdf'
    )


# === VERIFICACIÓN PÚBLICA DE CERTIFICADOS ===

@api_router.get("/verificar/{codigo_verificacion}", response_class=HTMLResponse)
async def verificar_certificado_publico(codigo_verificacion: str):
    """
    Endpoint PÚBLICO para verificar un certificado catastral.
    Devuelve una página HTML con la información del certificado.
    No requiere autenticación.
    """
    frontend_url = os.environ.get('FRONTEND_URL', 'https://land-registry-17.preview.emergentagent.com')
    logo_url = f"{frontend_url}/logo-asomunicipios.png"
    
    # Buscar certificado
    certificado = await db.certificados_verificables.find_one(
        {"codigo_verificacion": codigo_verificacion},
        {"_id": 0}
    )
    
    if not certificado:
        # Certificado no encontrado
        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Verificación - Asomunicipios</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
                .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
                .header {{ background: #dc2626; color: white; padding: 20px; text-align: center; }}
                .header img {{ max-width: 180px; margin-bottom: 15px; background: white; padding: 8px; border-radius: 8px; }}
                .content {{ padding: 30px; text-align: center; }}
                .icon {{ font-size: 48px; margin-bottom: 20px; }}
                .code {{ background: #fee2e2; padding: 10px; border-radius: 5px; font-family: monospace; color: #dc2626; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <img src="{logo_url}" alt="Asomunicipios">
                    <h1>⚠️ Certificado NO Encontrado</h1>
                </div>
                <div class="content">
                    <div class="icon">❌</div>
                    <p>El código de verificación no corresponde a ningún certificado emitido por Asomunicipios.</p>
                    <p class="code">{codigo_verificacion}</p>
                    <p style="margin-top: 20px; color: #666;">
                        Si cree que esto es un error, contacte a:<br>
                        <strong>comunicaciones@asomunicipios.gov.co</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        return HTMLResponse(content=html_content, status_code=404)
    
    # Verificar estado
    estado = certificado.get('estado', 'activo')
    es_valido = estado == 'activo'
    
    # Formatear fecha
    fecha_gen = certificado.get('fecha_generacion', '')
    if fecha_gen:
        try:
            dt = datetime.fromisoformat(fecha_gen.replace('Z', '+00:00'))
            fecha_formateada = dt.strftime("%d/%m/%Y a las %H:%M")
        except:
            fecha_formateada = fecha_gen
    else:
        fecha_formateada = "No disponible"
    
    # Propietarios
    propietarios = certificado.get('propietarios', [])
    propietarios_html = "<br>".join([f"• {p}" for p in propietarios]) if propietarios else "No disponible"
    
    # Obtener área terreno y avalúo directamente del certificado (ya vienen formateados)
    area_terreno = certificado.get('area_terreno', 'N/A')
    avaluo = certificado.get('avaluo_catastral', 'N/A')
    
    # Verificar si tiene hash de PDF (sistema nuevo)
    tiene_verificacion_integridad = bool(certificado.get('hash_pdf'))
    
    if es_valido:
        # Certificado VÁLIDO
        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>✅ Certificado Válido - Asomunicipios</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: #f0fdf4; }}
                .container {{ max-width: 650px; margin: 0 auto; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
                .header {{ background: #009846; color: white; padding: 25px; text-align: center; }}
                .header img {{ max-width: 180px; margin-bottom: 15px; background: white; padding: 8px; border-radius: 8px; }}
                .header h1 {{ margin: 10px 0 0 0; font-size: 22px; }}
                .badge {{ display: inline-block; background: #10b981; padding: 5px 15px; border-radius: 20px; margin-top: 10px; font-size: 14px; }}
                .content {{ padding: 30px; }}
                .info-row {{ display: flex; border-bottom: 1px solid #e5e7eb; padding: 12px 0; }}
                .info-label {{ font-weight: bold; color: #374151; width: 150px; flex-shrink: 0; }}
                .info-value {{ color: #6b7280; }}
                .code {{ background: #f0fdf4; padding: 15px; border-radius: 8px; font-family: monospace; color: #009846; text-align: center; margin: 20px 0; border: 2px solid #009846; font-size: 16px; }}
                .footer {{ background: #009846; padding: 20px; text-align: center; font-size: 12px; color: white; }}
                .footer a {{ color: #a7f3d0; }}
                .verify-section {{ background: #fffbeb; border: 2px solid #f59e0b; border-radius: 8px; padding: 20px; margin-top: 25px; }}
                .verify-section h3 {{ margin-top: 0; color: #d97706; }}
                .upload-form {{ margin-top: 15px; }}
                .upload-form input[type="file"] {{ margin-bottom: 10px; }}
                .upload-form button {{ background: #009846; color: white; border: none; padding: 10px 20px; border-radius: 5px; cursor: pointer; }}
                .upload-form button:hover {{ background: #007a38; }}
                #resultado {{ margin-top: 15px; padding: 15px; border-radius: 8px; display: none; }}
                .resultado-ok {{ background: #d1fae5; border: 2px solid #10b981; color: #065f46; }}
                .resultado-error {{ background: #fee2e2; border: 2px solid #dc2626; color: #991b1b; }}
                .datos-originales {{ background: #f3f4f6; padding: 15px; border-radius: 8px; margin-top: 20px; }}
                .datos-originales h4 {{ margin-top: 0; color: #374151; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <img src="{logo_url}" alt="Asomunicipios">
                    <h1>✅ CERTIFICADO VÁLIDO</h1>
                    <span class="badge">Verificado por Asomunicipios</span>
                </div>
                <div class="content">
                    <div class="code">
                        <strong>Código de Verificación:</strong><br>
                        {codigo_verificacion}
                    </div>
                    
                    <h3 style="color: #009846; border-bottom: 2px solid #009846; padding-bottom: 10px;">📋 Datos Originales del Certificado</h3>
                    <p style="font-size: 13px; color: #666; margin-bottom: 15px;">Compare estos datos con su documento físico o digital</p>
                    
                    <div class="info-row">
                        <span class="info-label">📋 Código Predial:</span>
                        <span class="info-value"><strong>{certificado.get('codigo_predial', 'N/A')}</strong></span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">🏛️ Municipio:</span>
                        <span class="info-value">{certificado.get('municipio', 'N/A')}</span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">📍 Dirección:</span>
                        <span class="info-value">{certificado.get('direccion', 'N/A')}</span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">👥 Propietarios:</span>
                        <span class="info-value"><strong>{propietarios_html}</strong></span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">📐 Área Terreno:</span>
                        <span class="info-value"><strong>{area_terreno}</strong></span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">💰 Avalúo:</span>
                        <span class="info-value"><strong>{avaluo}</strong></span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">📅 Generado:</span>
                        <span class="info-value">{fecha_formateada}</span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">👤 Elaborado por:</span>
                        <span class="info-value">{certificado.get('generado_por_nombre', 'N/A')}</span>
                    </div>
                    
                    <div class="verify-section">
                        <h3>🔍 Verificar Integridad del Documento</h3>
                        <p style="font-size: 14px; color: #666;">
                            ¿Desea verificar si su PDF ha sido modificado? Suba el archivo para compararlo con el original.
                        </p>
                        <form class="upload-form" id="verifyForm" enctype="multipart/form-data">
                            <input type="file" id="pdfFile" accept=".pdf" required>
                            <input type="hidden" id="codigoVerif" value="{codigo_verificacion}">
                            <br>
                            <button type="submit">🔐 Verificar Integridad</button>
                        </form>
                        <div id="resultado"></div>
                    </div>
                </div>
                <div class="footer">
                    <strong>Asomunicipios - Gestor Catastral</strong><br>
                    Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar<br>
                    <a href="mailto:comunicaciones@asomunicipios.gov.co">comunicaciones@asomunicipios.gov.co</a> | +57 3102327647
                </div>
            </div>
            
            <script>
                document.getElementById('verifyForm').addEventListener('submit', async function(e) {{
                    e.preventDefault();
                    const fileInput = document.getElementById('pdfFile');
                    const codigo = document.getElementById('codigoVerif').value;
                    const resultado = document.getElementById('resultado');
                    
                    if (!fileInput.files[0]) {{
                        alert('Por favor seleccione un archivo PDF');
                        return;
                    }}
                    
                    resultado.style.display = 'block';
                    resultado.className = '';
                    resultado.innerHTML = '⏳ Verificando integridad del documento...';
                    
                    const formData = new FormData();
                    formData.append('pdf_file', fileInput.files[0]);
                    formData.append('codigo_verificacion', codigo);
                    
                    try {{
                        const response = await fetch('/api/certificados/verificar-integridad', {{
                            method: 'POST',
                            body: formData
                        }});
                        
                        const data = await response.json();
                        
                        if (data.pdf_integro === true) {{
                            resultado.className = 'resultado-ok';
                            resultado.innerHTML = '<strong>✅ DOCUMENTO ÍNTEGRO</strong><br><br>El PDF no ha sido modificado desde su generación. Este es el documento original emitido por Asomunicipios.';
                        }} else if (data.pdf_integro === false) {{
                            resultado.className = 'resultado-error';
                            resultado.innerHTML = '<strong>🚨 ¡ALERTA! DOCUMENTO ALTERADO</strong><br><br>El PDF ha sido modificado después de su generación.<br><br><strong>IMPORTANTE:</strong> Los datos mostrados arriba son los ORIGINALES del sistema. Si el documento que tiene muestra datos diferentes, ha sido manipulado.';
                        }} else {{
                            resultado.className = 'resultado-ok';
                            resultado.innerHTML = '<strong>ℹ️ Verificación Parcial</strong><br><br>Este certificado fue generado antes del sistema de verificación de integridad. Compare manualmente los datos mostrados arriba con su documento.';
                        }}
                    }} catch (error) {{
                        resultado.className = 'resultado-error';
                        resultado.innerHTML = '<strong>❌ Error</strong><br>No se pudo verificar el documento. Intente nuevamente.';
                    }}
                }});
            </script>
        </body>
        </html>
        """
    else:
        # Certificado ANULADO
        fecha_anulacion = certificado.get('fecha_anulacion', '')
        if fecha_anulacion:
            try:
                dt = datetime.fromisoformat(fecha_anulacion.replace('Z', '+00:00'))
                fecha_anulacion_fmt = dt.strftime("%d/%m/%Y a las %H:%M")
            except:
                fecha_anulacion_fmt = fecha_anulacion
        else:
            fecha_anulacion_fmt = "No disponible"
        motivo = certificado.get('motivo_anulacion', 'No especificado')
        
        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>⚠️ Certificado Anulado - Asomunicipios</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: #fef2f2; }}
                .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
                .header {{ background: #dc2626; color: white; padding: 25px; text-align: center; }}
                .header img {{ max-width: 180px; margin-bottom: 15px; background: white; padding: 8px; border-radius: 8px; }}
                .header h1 {{ margin: 10px 0 0 0; font-size: 22px; }}
                .content {{ padding: 30px; }}
                .warning {{ background: #fee2e2; border: 2px solid #dc2626; border-radius: 8px; padding: 20px; margin-bottom: 20px; text-align: center; }}
                .info-row {{ display: flex; border-bottom: 1px solid #e5e7eb; padding: 12px 0; }}
                .info-label {{ font-weight: bold; color: #374151; width: 150px; flex-shrink: 0; }}
                .info-value {{ color: #6b7280; }}
                .footer {{ background: #009846; padding: 20px; text-align: center; font-size: 12px; color: white; }}
                .footer a {{ color: #a7f3d0; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <img src="{logo_url}" alt="Asomunicipios">
                    <h1>⚠️ CERTIFICADO ANULADO</h1>
                </div>
                <div class="content">
                    <div class="warning">
                        <strong style="font-size: 16px;">⚠️ Este certificado ha sido ANULADO</strong><br><br>
                        Este documento ya no tiene validez legal.<br><br>
                        <strong>Motivo:</strong> {motivo}
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">🔑 Código:</span>
                        <span class="info-value" style="font-family: monospace;">{codigo_verificacion}</span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">📋 Código Predial:</span>
                        <span class="info-value">{certificado.get('codigo_predial', 'N/A')}</span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">📅 Fecha Anulación:</span>
                        <span class="info-value">{fecha_anulacion_fmt}</span>
                    </div>
                    
                    <div class="info-row">
                        <span class="info-label">👤 Anulado por:</span>
                        <span class="info-value">{certificado.get('anulado_por_nombre', 'N/A')}</span>
                    </div>
                </div>
                <div class="footer">
                    <strong>Asomunicipios - Gestor Catastral</strong><br>
                    <a href="mailto:comunicaciones@asomunicipios.gov.co">comunicaciones@asomunicipios.gov.co</a>
                </div>
            </div>
        </body>
        </html>
        """
    
    return HTMLResponse(content=html_content)


@api_router.post("/certificados/{codigo_verificacion}/anular")
async def anular_certificado(
    codigo_verificacion: str,
    motivo: str = Form(..., description="Motivo de la anulación"),
    current_user: dict = Depends(get_current_user)
):
    """
    Anula un certificado existente.
    Solo coordinadores y administradores pueden anular certificados.
    """
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores y administradores pueden anular certificados")
    
    # Buscar certificado
    certificado = await db.certificados_verificables.find_one({"codigo_verificacion": codigo_verificacion})
    if not certificado:
        raise HTTPException(status_code=404, detail="Certificado no encontrado")
    
    if certificado.get('estado') == 'anulado':
        raise HTTPException(status_code=400, detail="El certificado ya está anulado")
    
    # Anular certificado
    await db.certificados_verificables.update_one(
        {"codigo_verificacion": codigo_verificacion},
        {
            "$set": {
                "estado": "anulado",
                "motivo_anulacion": motivo,
                "fecha_anulacion": datetime.now(timezone.utc).isoformat(),
                "anulado_por": current_user['id'],
                "anulado_por_nombre": current_user['full_name']
            }
        }
    )
    
    return {
        "success": True,
        "message": f"Certificado {codigo_verificacion} anulado exitosamente",
        "codigo_verificacion": codigo_verificacion
    }


@api_router.post("/certificados/verificar-integridad")
async def verificar_integridad_pdf(
    pdf_file: UploadFile = File(..., description="PDF del certificado a verificar"),
    codigo_verificacion: str = Form(..., description="Código de verificación del certificado")
):
    """
    Verifica la integridad de un PDF comparando su hash con el original.
    Devuelve si el documento ha sido alterado o no.
    """
    # Buscar certificado en la base de datos
    certificado = await db.certificados_verificables.find_one(
        {"codigo_verificacion": codigo_verificacion},
        {"_id": 0}
    )
    
    if not certificado:
        return {
            "verificado": False,
            "estado": "no_encontrado",
            "mensaje": "Código de verificación no encontrado en el sistema",
            "detalles": None
        }
    
    # Leer el PDF subido
    pdf_content = await pdf_file.read()
    
    # Calcular hash del PDF subido
    hash_pdf_subido = hashlib.sha256(pdf_content).hexdigest()
    
    # Obtener hash original
    hash_pdf_original = certificado.get('hash_pdf', '')
    
    # Comparar hashes
    pdf_integro = hash_pdf_subido == hash_pdf_original if hash_pdf_original else None
    
    # Preparar respuesta con datos originales
    datos_originales = certificado.get('datos_criticos', {})
    
    if certificado.get('estado') == 'anulado':
        return {
            "verificado": True,
            "estado": "anulado",
            "mensaje": "⚠️ CERTIFICADO ANULADO - Este documento ya no tiene validez legal",
            "motivo_anulacion": certificado.get('motivo_anulacion'),
            "fecha_anulacion": certificado.get('fecha_anulacion'),
            "pdf_integro": pdf_integro,
            "datos_originales": datos_originales
        }
    
    if pdf_integro is None:
        return {
            "verificado": True,
            "estado": "activo_sin_hash",
            "mensaje": "✅ Certificado válido (generado antes del sistema de integridad)",
            "pdf_integro": None,
            "advertencia": "Este certificado fue generado antes de implementar la verificación de integridad. Compare manualmente los datos.",
            "datos_originales": datos_originales,
            "fecha_generacion": certificado.get('fecha_generacion'),
            "generado_por": certificado.get('generado_por_nombre')
        }
    
    if pdf_integro:
        return {
            "verificado": True,
            "estado": "activo_integro",
            "mensaje": "✅ CERTIFICADO VÁLIDO - El documento NO ha sido modificado",
            "pdf_integro": True,
            "datos_originales": datos_originales,
            "fecha_generacion": certificado.get('fecha_generacion'),
            "generado_por": certificado.get('generado_por_nombre'),
            "hash_verificado": hash_pdf_original[:16].upper()
        }
    else:
        return {
            "verificado": True,
            "estado": "activo_alterado",
            "mensaje": "🚨 DOCUMENTO ALTERADO - El PDF ha sido modificado después de su generación",
            "pdf_integro": False,
            "advertencia": "Los datos del documento pueden haber sido manipulados. Use los datos originales mostrados abajo.",
            "datos_originales": datos_originales,
            "fecha_generacion": certificado.get('fecha_generacion'),
            "generado_por": certificado.get('generado_por_nombre'),
            "hash_original": hash_pdf_original[:16].upper(),
            "hash_subido": hash_pdf_subido[:16].upper()
        }


@api_router.get("/certificados/verificables")
async def listar_certificados_verificables(
    skip: int = 0,
    limit: int = 50,
    estado: str = Query(None, description="Filtrar por estado: activo, anulado"),
    current_user: dict = Depends(get_current_user)
):
    """Lista todos los certificados verificables generados"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {}
    if estado:
        query["estado"] = estado
    
    certificados = await db.certificados_verificables.find(query, {"_id": 0}).sort("fecha_generacion", -1).skip(skip).limit(limit).to_list(length=limit)
    total = await db.certificados_verificables.count_documents(query)
    
    return {
        "certificados": certificados,
        "total": total,
        "skip": skip,
        "limit": limit
    }


@api_router.get("/certificados/historial")
async def get_certificados_historial(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el historial de certificados generados"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    certificados = await db.certificados.find({}, {"_id": 0}).sort("fecha_generacion", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.certificados.count_documents({})
    
    return {
        "certificados": certificados,
        "total": total
    }


@api_router.get("/predios/terreno-info/{municipio}")
async def get_terreno_info(
    municipio: str,
    zona: str = "00",
    sector: str = "01", 
    manzana_vereda: str = "0000",
    current_user: dict = Depends(get_current_user)
):
    """Obtiene información sobre el siguiente terreno disponible en una manzana"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Buscar todos los terrenos en esta manzana (incluyendo eliminados)
    query = {
        "municipio": municipio,
        "zona": zona,
        "sector": sector,
        "manzana_vereda": manzana_vereda
    }
    
    predios = await db.predios.find(query, {"_id": 0, "terreno": 1, "terreno_num": 1, "deleted": 1, "codigo_homologado": 1}).to_list(10000)
    
    # Clasificar terrenos
    terrenos_activos = []
    terrenos_eliminados = []
    
    for p in predios:
        terreno_num = p.get('terreno_num', 0)
        if p.get('deleted'):
            terrenos_eliminados.append({
                "numero": p.get('terreno'),
                "codigo": p.get('codigo_homologado')
            })
        else:
            terrenos_activos.append(terreno_num)
    
    # Encontrar el máximo terreno usado
    max_terreno = max(terrenos_activos + [t.get('terreno_num', 0) for t in predios], default=0)
    siguiente_terreno = max_terreno + 1
    
    return {
        "municipio": municipio,
        "zona": zona,
        "sector": sector,
        "manzana_vereda": manzana_vereda,
        "total_activos": len(terrenos_activos),
        "ultimo_terreno": str(max_terreno).zfill(4) if max_terreno > 0 else "N/A",
        "siguiente_terreno": str(siguiente_terreno).zfill(4),
        "terrenos_eliminados": terrenos_eliminados,
        "terrenos_no_reutilizables": len(terrenos_eliminados)
    }

@api_router.get("/predios/{predio_id}")
async def get_predio(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene un predio por ID"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    predio = await db.predios.find_one({"id": predio_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    return predio


@api_router.get("/predios/{predio_id}/construcciones")
async def get_construcciones_predio(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene las construcciones asociadas a un predio"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Obtener el predio para conocer su código
    predio = await db.predios.find_one({"id": predio_id, "deleted": {"$ne": True}}, {"_id": 0, "codigo_predial_nacional": 1, "codigo_gdb": 1})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    codigo = predio.get("codigo_gdb") or predio.get("codigo_predial_nacional")
    if not codigo:
        return {"construcciones": [], "total": 0}
    
    # Buscar construcciones con match flexible
    construcciones = await db.gdb_construcciones.find({
        "$or": [
            {"codigo_predio": codigo},
            {"codigo_predio": {"$regex": f"^{codigo[:20]}"}}  # Match por prefijo (20 dígitos)
        ]
    }, {"_id": 0}).to_list(100)
    
    return {
        "construcciones": construcciones,
        "total": len(construcciones)
    }


@api_router.get("/gdb/construcciones/{codigo_predio}")
async def get_construcciones_by_codigo(codigo_predio: str, current_user: dict = Depends(get_current_user)):
    """Obtiene las construcciones por código de predio - match EXACTO"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Buscar con match EXACTO solamente
    # El código de la construcción debe coincidir exactamente con el código del predio
    construcciones = await db.gdb_construcciones.find({
        "codigo_predio": codigo_predio
    }, {"_id": 0}).to_list(100)
    
    return {
        "construcciones": construcciones,
        "total": len(construcciones)
    }


@api_router.post("/predios")
async def create_predio(predio_data: PredioCreate, current_user: dict = Depends(get_current_user)):
    """Crea un nuevo predio"""
    # Usuario, Comunicaciones y Empresa no pueden crear predios
    if current_user['role'] in [UserRole.USUARIO, UserRole.COMUNICACIONES, UserRole.EMPRESA]:
        raise HTTPException(status_code=403, detail="No tiene permiso para crear predios")
    
    r1 = predio_data.r1
    r2 = predio_data.r2
    
    # Obtener siguiente número de terreno
    terreno, terreno_num = await get_next_terreno_number(
        r1.municipio, r1.zona, r1.sector, r1.manzana_vereda
    )
    
    # Generar código predial nacional
    codigo_predial = await generate_codigo_predial(
        r1.municipio, r1.zona, r1.sector, r1.manzana_vereda,
        terreno, r1.condicion_predio, r1.predio_horizontal
    )
    
    # Verificar que no exista
    existing = await db.predios.find_one({"codigo_predial_nacional": codigo_predial})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un predio con este código predial")
    
    # Verificar que no sea un código eliminado
    eliminado = await db.predios_eliminados.find_one({"codigo_predial_nacional": codigo_predial})
    if eliminado:
        # Verificar si tiene aprobación de reaparición
        aprobacion = await db.predios_reapariciones_aprobadas.find_one({
            "codigo_predial_nacional": codigo_predial,
            "estado": "aprobado"
        })
        
        if not aprobacion:
            # Verificar si ya hay solicitud pendiente
            solicitud_pendiente = await db.predios_reapariciones_solicitudes.find_one({
                "codigo_predial_nacional": codigo_predial,
                "estado": "pendiente"
            })
            
            error_detail = {
                "error": "PREDIO_ELIMINADO",
                "mensaje": f"Este código predial fue ELIMINADO en vigencia {eliminado.get('vigencia_eliminacion')}",
                "codigo_predial": codigo_predial,
                "municipio": eliminado.get("municipio"),
                "vigencia_eliminacion": eliminado.get("vigencia_eliminacion"),
                "puede_solicitar_reaparicion": solicitud_pendiente is None,
                "tiene_solicitud_pendiente": solicitud_pendiente is not None,
                "instrucciones": "Use el endpoint /api/predios/reapariciones/solicitar para solicitar la reaparición con justificación técnica" if not solicitud_pendiente else "Ya existe una solicitud pendiente de aprobación"
            }
            raise HTTPException(status_code=400, detail=error_detail)
    
    # Generar código homologado - Primero intentar de la colección de códigos cargados
    codigo_homologado = await asignar_codigo_homologado(r1.municipio, None)  # Se actualizará con el ID real después
    
    if codigo_homologado:
        # Obtener el último número de predio para mantener secuencia
        last_predio = await db.predios.find_one(
            {"municipio": r1.municipio, "deleted": {"$ne": True}},
            sort=[("numero_predio", -1)]
        )
        if last_predio:
            num_predio = last_predio.get("numero_predio", 0)
            if isinstance(num_predio, str):
                try:
                    num_predio = int(num_predio)
                except (ValueError, TypeError):
                    num_predio = 0
            numero_predio = num_predio + 1
        else:
            numero_predio = 1
    else:
        # Fallback: generar código aleatorio si no hay códigos cargados
        codigo_homologado, numero_predio = await generate_codigo_homologado(r1.municipio)
    
    # Obtener código Código Nacional Catastral
    divipola = MUNICIPIOS_DIVIPOLA[r1.municipio]
    
    # Crear el predio
    predio = {
        "id": str(uuid.uuid4()),
        "departamento": divipola["departamento"],
        "municipio": r1.municipio,
        "municipio_codigo": divipola["municipio"],
        "numero_predio": numero_predio,
        "codigo_predial_nacional": codigo_predial,
        "codigo_homologado": codigo_homologado,
        "zona": r1.zona,
        "sector": r1.sector,
        "manzana_vereda": r1.manzana_vereda,
        "terreno": terreno,
        "terreno_num": terreno_num,
        "condicion_predio": r1.condicion_predio,
        "predio_horizontal": r1.predio_horizontal,
        
        # R1 - Información jurídica
        "tipo_registro": 1,
        "nombre_propietario": r1.nombre_propietario,
        "tipo_documento": r1.tipo_documento,
        "numero_documento": r1.numero_documento,
        "estado_civil": r1.estado_civil,
        "direccion": r1.direccion,
        "comuna": r1.comuna,
        "destino_economico": r1.destino_economico,
        "area_terreno": r1.area_terreno,
        "area_construida": r1.area_construida,
        "avaluo": r1.avaluo,
        "tipo_mutacion": r1.tipo_mutacion,
        "numero_resolucion": r1.numero_resolucion,
        "fecha_resolucion": r1.fecha_resolucion,
        
        # R2 - Información física (si se proporciona)
        "r2": r2.model_dump() if r2 else None,
        
        # Metadata
        "vigencia": datetime.now().strftime("%m%d%Y"),
        "created_by": current_user['id'],
        "created_by_name": current_user['full_name'],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted": False,
        
        # Historial
        "historial": [{
            "accion": "Predio creado",
            "usuario": current_user['full_name'],
            "usuario_id": current_user['id'],
            "fecha": datetime.now(timezone.utc).isoformat()
        }]
    }
    
    await db.predios.insert_one(predio)
    
    # Actualizar la asignación del código homologado con el ID real del predio
    if codigo_homologado:
        await db.codigos_homologados.update_one(
            {'codigo': codigo_homologado, 'municipio': r1.municipio},
            {'$set': {'predio_id': predio['id']}}
        )
    
    # Remover _id antes de retornar
    predio.pop("_id", None)
    
    # Broadcast WebSocket para sincronización en tiempo real
    await ws_manager.broadcast({
        "type": "predio_actualizado",
        "action": "create",
        "predio_id": predio['id'],
        "municipio": predio.get('municipio', ''),
        "codigo_predial": predio.get('codigo_predial_nacional', ''),
        "created_by": current_user['full_name'],
        "timestamp": datetime.now(timezone.utc).isoformat()
    }, exclude_user=current_user['id'])
    
    return predio

@api_router.patch("/predios/{predio_id}")
async def update_predio(predio_id: str, update_data: PredioUpdate, current_user: dict = Depends(get_current_user)):
    """Actualiza un predio existente"""
    # Usuario, Comunicaciones y Empresa no pueden modificar predios directamente
    if current_user['role'] in [UserRole.USUARIO, UserRole.COMUNICACIONES, UserRole.EMPRESA]:
        raise HTTPException(status_code=403, detail="No tiene permiso para modificar predios")
    
    predio = await db.predios.find_one({"id": predio_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Filtrar campos no nulos
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if not update_dict:
        return predio
    
    # Agregar metadata de actualización
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Agregar al historial
    historial_entry = {
        "accion": "Predio modificado",
        "usuario": current_user['full_name'],
        "usuario_id": current_user['id'],
        "campos_modificados": list(update_dict.keys()),
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios.update_one(
        {"id": predio_id},
        {
            "$set": update_dict,
            "$push": {"historial": historial_entry}
        }
    )
    
    # Retornar predio actualizado
    updated_predio = await db.predios.find_one({"id": predio_id}, {"_id": 0})
    
    # Broadcast WebSocket para sincronización en tiempo real
    await ws_manager.broadcast({
        "type": "predio_actualizado",
        "action": "update",
        "predio_id": predio_id,
        "municipio": updated_predio.get('municipio', ''),
        "codigo_predial": updated_predio.get('codigo_predial_nacional', ''),
        "campos_modificados": list(update_dict.keys()),
        "updated_by": current_user['full_name'],
        "timestamp": datetime.now(timezone.utc).isoformat()
    }, exclude_user=current_user['id'])
    
    return updated_predio

@api_router.delete("/predios/{predio_id}")
async def delete_predio(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Elimina un predio (soft delete)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores y administradores pueden eliminar predios")
    
    predio = await db.predios.find_one({"id": predio_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Soft delete - NO eliminamos físicamente para evitar reutilizar códigos
    historial_entry = {
        "accion": "Predio eliminado",
        "usuario": current_user['full_name'],
        "usuario_id": current_user['id'],
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios.update_one(
        {"id": predio_id},
        {
            "$set": {
                "deleted": True,
                "deleted_at": datetime.now(timezone.utc).isoformat(),
                "deleted_by": current_user['id'],
                "deleted_by_name": current_user['full_name']
            },
            "$push": {"historial": historial_entry}
        }
    )
    
    # Broadcast WebSocket para sincronización en tiempo real
    await ws_manager.broadcast({
        "type": "predio_actualizado",
        "action": "delete",
        "predio_id": predio_id,
        "municipio": predio.get('municipio', ''),
        "codigo_predial": predio.get('codigo_predial_nacional', ''),
        "deleted_by": current_user['full_name'],
        "timestamp": datetime.now(timezone.utc).isoformat()
    }, exclude_user=current_user['id'])
    
    return {"message": "Predio eliminado exitosamente"}


# ===== FLUJO DE TRABAJO PARA PREDIOS NUEVOS (CONSERVACIÓN) =====

@api_router.post("/predios-nuevos")
async def crear_predio_nuevo(
    predio_data: PredioNuevoCreate,
    current_user: dict = Depends(get_current_user)
):
    """
    Crea un nuevo predio con flujo de trabajo de digitalización.
    El predio queda en estado 'creado' y se asigna a un gestor de apoyo.
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso para crear predios")
    
    r1 = predio_data.r1
    r2 = predio_data.r2
    
    # Verificar que el gestor de apoyo existe y es gestor
    gestor_apoyo = await db.users.find_one({"id": predio_data.gestor_apoyo_id}, {"_id": 0})
    if not gestor_apoyo:
        raise HTTPException(status_code=400, detail="El gestor de apoyo no existe")
    if gestor_apoyo.get('role') not in ['gestor', 'coordinador', 'administrador']:
        raise HTTPException(status_code=400, detail="El usuario asignado no tiene rol de gestor")
    
    # Construir código predial nacional usando los datos que el usuario especificó
    divipola = MUNICIPIOS_DIVIPOLA.get(r1.municipio)
    if not divipola:
        raise HTTPException(status_code=400, detail=f"Municipio {r1.municipio} no válido")
    
    prefijo = divipola["departamento"] + divipola["municipio"]
    
    # Construir el código completo de 30 dígitos con los valores del usuario
    # Estructura: prefijo(5) + zona(2) + sector(2) + comuna(2) + barrio(2) + manzana_vereda(4) + terreno(4) + condicion_predio(4) + predio_horizontal(5)
    codigo_predial = (
        f"{prefijo}"
        f"{r1.zona.zfill(2)}"
        f"{r1.sector.zfill(2)}"
        f"{r1.comuna.zfill(2)}"
        f"{r1.barrio.zfill(2)}"
        f"{r1.manzana_vereda.zfill(4)}"
        f"{r1.terreno.zfill(4)}"
        f"{r1.condicion_predio.zfill(4)}"
        f"{r1.predio_horizontal.zfill(5)}"
    )
    
    # Asegurar que el código tenga exactamente 30 dígitos
    codigo_predial = codigo_predial[:30].ljust(30, '0')
    
    # Extraer el número de terreno para guardarlo
    terreno = r1.terreno.zfill(4)
    terreno_num = int(terreno) if terreno.isdigit() else 1
    
    # Verificar que no exista
    existing = await db.predios.find_one({"codigo_predial_nacional": codigo_predial})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un predio con este código predial")
    
    # Verificar que no sea un código eliminado sin aprobación
    eliminado = await db.predios_eliminados.find_one({"codigo_predial_nacional": codigo_predial})
    if eliminado:
        aprobacion = await db.predios_reapariciones_aprobadas.find_one({
            "codigo_predial_nacional": codigo_predial,
            "estado": "aprobado"
        })
        if not aprobacion:
            raise HTTPException(status_code=400, detail={
                "error": "PREDIO_ELIMINADO",
                "mensaje": f"Este código predial fue ELIMINADO. Debe solicitar reaparición."
            })
    
    # Generar código homologado - Primero intentar de la colección de códigos cargados del Excel
    codigo_homologado = await asignar_codigo_homologado(r1.municipio, None)
    
    if codigo_homologado:
        # Obtener el último número de predio para mantener secuencia
        last_predio = await db.predios.find_one(
            {"municipio": r1.municipio, "deleted": {"$ne": True}},
            sort=[("numero_predio", -1)]
        )
        if last_predio:
            num_predio = last_predio.get("numero_predio", 0)
            if isinstance(num_predio, str):
                try:
                    num_predio = int(num_predio)
                except (ValueError, TypeError):
                    num_predio = 0
            numero_predio = num_predio + 1
        else:
            numero_predio = 1
    else:
        # Fallback: generar código aleatorio si no hay códigos cargados en el Excel
        codigo_homologado, numero_predio = await generate_codigo_homologado(r1.municipio)
    
    # Obtener código Divipola
    divipola = MUNICIPIOS_DIVIPOLA[r1.municipio]
    
    # Construir radicado completo si se proporcionó número
    radicado_completo = None
    radicado_fecha = None
    if predio_data.radicado_numero:
        # Buscar la petición con este radicado para obtener la fecha
        peticion = await db.petitions.find_one({
            "radicado": {"$regex": f"-{predio_data.radicado_numero}-", "$options": "i"}
        }, {"_id": 0})
        if peticion:
            fecha_rad = peticion.get('created_at') or peticion.get('fecha_radicacion')
            if fecha_rad:
                if isinstance(fecha_rad, str):
                    fecha_obj = datetime.fromisoformat(fecha_rad.replace('Z', '+00:00'))
                else:
                    fecha_obj = fecha_rad
                radicado_fecha = fecha_obj.strftime("%d-%m-%Y")
        radicado_completo = f"RASMGC-{predio_data.radicado_numero}-{radicado_fecha or datetime.now().strftime('%d-%m-%Y')}"
    
    # Crear el predio nuevo
    predio_nuevo = {
        "id": str(uuid.uuid4()),
        "departamento": divipola["departamento"],
        "municipio": r1.municipio,
        "municipio_codigo": divipola["municipio"],
        "numero_predio": numero_predio,
        "codigo_predial_nacional": codigo_predial,
        "codigo_homologado": codigo_homologado,
        "zona": r1.zona,
        "sector": r1.sector,
        "manzana_vereda": r1.manzana_vereda,
        "terreno": terreno,
        "terreno_num": terreno_num,
        "condicion_predio": r1.condicion_predio,
        "predio_horizontal": r1.predio_horizontal,
        
        # R1 - Información jurídica
        "tipo_registro": 1,
        "nombre_propietario": r1.nombre_propietario,
        "tipo_documento": r1.tipo_documento,
        "numero_documento": r1.numero_documento,
        "estado_civil": r1.estado_civil,
        "direccion": r1.direccion,
        "comuna": r1.comuna,
        "destino_economico": r1.destino_economico,
        "area_terreno": r1.area_terreno,
        "area_construida": r1.area_construida,
        "avaluo": r1.avaluo,
        "tipo_mutacion": r1.tipo_mutacion,
        "numero_resolucion": r1.numero_resolucion,
        "fecha_resolucion": r1.fecha_resolucion,
        
        # R2 - Información física
        "r2": r2.model_dump() if r2 else None,
        
        # === FLUJO DE TRABAJO ===
        "es_predio_nuevo": True,
        "estado_flujo": PredioNuevoEstado.CREADO,
        
        # Gestor creador
        "gestor_creador_id": current_user['id'],
        "gestor_creador_nombre": current_user['full_name'],
        "gestor_creador_email": current_user.get('email'),
        
        # Gestor de apoyo (digitalización)
        "gestor_apoyo_id": predio_data.gestor_apoyo_id,
        "gestor_apoyo_nombre": gestor_apoyo['full_name'],
        "gestor_apoyo_email": gestor_apoyo.get('email'),
        
        # Relaciones
        "radicado_relacionado": radicado_completo,
        "radicado_numero": predio_data.radicado_numero,
        "peticiones_relacionadas": predio_data.peticiones_ids or [],
        
        # Observaciones
        "observaciones": predio_data.observaciones,
        
        # Metadata
        "vigencia": datetime.now().year,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted": False,
        
        # Historial de trazabilidad
        "historial_flujo": [{
            "fecha": datetime.now(timezone.utc).isoformat(),
            "accion": "Predio creado",
            "estado_anterior": None,
            "estado_nuevo": PredioNuevoEstado.CREADO,
            "usuario_id": current_user['id'],
            "usuario_nombre": current_user['full_name'],
            "observaciones": f"Predio creado y asignado a {gestor_apoyo['full_name']} para digitalización"
        }]
    }
    
    # Guardar en colección de predios en proceso
    await db.predios_nuevos.insert_one(predio_nuevo)
    
    # Crear notificación para el gestor de apoyo
    notificacion = {
        "id": str(uuid.uuid4()),
        "user_id": predio_data.gestor_apoyo_id,
        "tipo": "predio_asignado",
        "titulo": "Nuevo predio asignado para digitalización",
        "mensaje": f"Se te ha asignado el predio {codigo_predial} para digitalización de base gráfica.",
        "predio_id": predio_nuevo['id'],
        "codigo_predial": codigo_predial,
        "leida": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notificacion)
    
    predio_nuevo.pop("_id", None)
    return predio_nuevo


@api_router.get("/predios-nuevos")
async def listar_predios_nuevos(
    estado: Optional[str] = None,
    municipio: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """
    Lista predios nuevos en proceso según el rol del usuario.
    - Gestores: ven predios que crearon o que les asignaron
    - Coordinadores: ven todos los predios en revisión
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {}
    
    # Filtrar por estado
    if estado:
        query["estado_flujo"] = estado
    
    # Filtrar por municipio
    if municipio:
        query["municipio"] = municipio
    
    # Filtrar según rol
    if current_user['role'] == UserRole.GESTOR:
        # Gestores ven predios que crearon, que les asignaron, o devueltos a ellos
        query["$or"] = [
            {"gestor_creador_id": current_user['id']},
            {"gestor_apoyo_id": current_user['id']}
        ]
    elif current_user['role'] == UserRole.COORDINADOR:
        # Coordinadores ven TODOS los predios (pueden aprobar por defecto)
        # Solo filtramos por municipio si el coordinador tiene municipios asignados
        user_municipios = current_user.get('municipios', [])
        if user_municipios:
            query["municipio"] = {"$in": user_municipios}
        # Si no tiene municipios asignados, ve todos
    # Administradores ven todo
    
    # Obtener predios
    cursor = db.predios_nuevos.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
    predios = await cursor.to_list(length=limit)
    
    # Contar total
    total = await db.predios_nuevos.count_documents(query)
    
    return {
        "predios": predios,
        "total": total,
        "skip": skip,
        "limit": limit
    }


@api_router.get("/predios-nuevos/pendientes")
async def predios_nuevos_pendientes(current_user: dict = Depends(get_current_user)):
    """
    Lista predios pendientes de acción para el usuario actual.
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    pendientes = {
        "digitalizacion": [],  # Pendientes para gestor de apoyo
        "revision": [],        # Pendientes para coordinador
        "devueltos": []        # Devueltos al gestor/apoyo
    }
    
    user_id = current_user['id']
    
    # Predios en digitalización asignados a este usuario
    cursor = db.predios_nuevos.find({
        "gestor_apoyo_id": user_id,
        "estado_flujo": {"$in": [PredioNuevoEstado.CREADO, PredioNuevoEstado.DIGITALIZACION]}
    }, {"_id": 0})
    pendientes["digitalizacion"] = await cursor.to_list(length=100)
    
    # Predios en revisión (para coordinadores con permiso)
    user_permisos = await db.user_permissions.find_one({"user_id": user_id})
    puede_aprobar = (
        current_user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR] or
        (user_permisos and user_permisos.get('permissions', {}).get('aprobar_cambios'))
    )
    
    if puede_aprobar:
        cursor = db.predios_nuevos.find({
            "estado_flujo": PredioNuevoEstado.REVISION
        }, {"_id": 0})
        pendientes["revision"] = await cursor.to_list(length=100)
    
    # Predios devueltos a este usuario
    cursor = db.predios_nuevos.find({
        "estado_flujo": PredioNuevoEstado.DEVUELTO,
        "$or": [
            {"gestor_creador_id": user_id},
            {"gestor_apoyo_id": user_id}
        ]
    }, {"_id": 0})
    pendientes["devueltos"] = await cursor.to_list(length=100)
    
    return pendientes


@api_router.get("/predios-nuevos/{predio_id}")
async def obtener_predio_nuevo(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene un predio nuevo por ID"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    predio = await db.predios_nuevos.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    return predio


@api_router.post("/predios-nuevos/{predio_id}/accion")
async def ejecutar_accion_predio_nuevo(
    predio_id: str,
    accion_data: PredioNuevoAccion,
    current_user: dict = Depends(get_current_user)
):
    """
    Ejecuta una acción en el flujo de trabajo del predio.
    Acciones: enviar_revision, aprobar, devolver, rechazar
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    predio = await db.predios_nuevos.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    estado_actual = predio['estado_flujo']
    accion = accion_data.accion
    user_id = current_user['id']
    
    # Validar permisos según acción
    if accion == "enviar_revision":
        # Solo el gestor de apoyo puede enviar a revisión
        if predio['gestor_apoyo_id'] != user_id and current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
            raise HTTPException(status_code=403, detail="Solo el gestor de apoyo asignado puede enviar a revisión")
        
        if estado_actual not in [PredioNuevoEstado.CREADO, PredioNuevoEstado.DIGITALIZACION, PredioNuevoEstado.DEVUELTO]:
            raise HTTPException(status_code=400, detail=f"No se puede enviar a revisión desde estado '{estado_actual}'")
        
        nuevo_estado = PredioNuevoEstado.REVISION
        mensaje_notif = f"El predio {predio['codigo_predial_nacional']} ha sido enviado para revisión"
    
    elif accion == "rechazar_asignacion":
        # Solo el gestor de apoyo asignado puede rechazar su asignación
        if predio['gestor_apoyo_id'] != user_id:
            raise HTTPException(status_code=403, detail="Solo el gestor de apoyo asignado puede rechazar esta asignación")
        
        if estado_actual not in [PredioNuevoEstado.CREADO, PredioNuevoEstado.DIGITALIZACION, PredioNuevoEstado.DEVUELTO]:
            raise HTTPException(status_code=400, detail=f"No se puede rechazar la asignación desde estado '{estado_actual}'")
        
        nuevo_estado = PredioNuevoEstado.CREADO  # Vuelve a estado creado sin asignación
        mensaje_notif = f"La asignación del predio {predio['codigo_predial_nacional']} ha sido rechazada por {current_user['full_name']}"
        
    elif accion in ["aprobar", "devolver", "rechazar"]:
        # Verificar permiso de aprobar cambios
        user_permisos = await db.user_permissions.find_one({"user_id": user_id})
        puede_aprobar = (
            current_user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR] or
            (user_permisos and user_permisos.get('permissions', {}).get('aprobar_cambios'))
        )
        
        if not puede_aprobar:
            raise HTTPException(status_code=403, detail="No tiene permiso para aprobar/devolver/rechazar predios")
        
        if estado_actual != PredioNuevoEstado.REVISION:
            raise HTTPException(status_code=400, detail="El predio debe estar en revisión para esta acción")
        
        if accion == "aprobar":
            nuevo_estado = PredioNuevoEstado.APROBADO
            mensaje_notif = f"El predio {predio['codigo_predial_nacional']} ha sido APROBADO"
        elif accion == "devolver":
            nuevo_estado = PredioNuevoEstado.DEVUELTO
            mensaje_notif = f"El predio {predio['codigo_predial_nacional']} ha sido DEVUELTO para corrección"
        else:  # rechazar
            nuevo_estado = PredioNuevoEstado.RECHAZADO
            mensaje_notif = f"El predio {predio['codigo_predial_nacional']} ha sido RECHAZADO"
    else:
        raise HTTPException(status_code=400, detail=f"Acción no válida: {accion}")
    
    # Crear entrada de historial
    historial_entry = {
        "fecha": datetime.now(timezone.utc).isoformat(),
        "accion": accion,
        "estado_anterior": estado_actual,
        "estado_nuevo": nuevo_estado,
        "usuario_id": user_id,
        "usuario_nombre": current_user['full_name'],
        "observaciones": accion_data.observaciones
    }
    
    # Actualizar predio
    update_data = {
        "estado_flujo": nuevo_estado,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Si se aprueba, agregar datos de aprobación
    if accion == "aprobar":
        update_data["aprobado_por_id"] = user_id
        update_data["aprobado_por_nombre"] = current_user['full_name']
        update_data["fecha_aprobacion"] = datetime.now(timezone.utc).isoformat()
    
    # Si se rechaza la asignación, quitar el gestor de apoyo asignado
    if accion == "rechazar_asignacion":
        update_data["gestor_apoyo_id"] = None
        update_data["gestor_apoyo_nombre"] = None
        update_data["comentario_devolucion"] = accion_data.observaciones
    
    await db.predios_nuevos.update_one(
        {"id": predio_id},
        {
            "$set": update_data,
            "$push": {"historial_flujo": historial_entry}
        }
    )
    
    # Si se aprueba, mover a la colección principal de predios
    if accion == "aprobar":
        predio_actualizado = await db.predios_nuevos.find_one({"id": predio_id}, {"_id": 0})
        
        # Preparar datos para colección principal
        predio_aprobado = {**predio_actualizado}
        predio_aprobado["historial"] = predio_aprobado.pop("historial_flujo", [])
        predio_aprobado["historial"].append({
            "accion": "Predio aprobado e integrado",
            "usuario": current_user['full_name'],
            "usuario_id": user_id,
            "fecha": datetime.now(timezone.utc).isoformat()
        })
        
        # Normalizar vigencia al año actual (formato consistente con colección principal)
        vigencia_actual = datetime.now(timezone.utc).year
        predio_aprobado["vigencia"] = vigencia_actual
        
        # Asegurar que tenga los campos necesarios para R1/R2
        if "r1" not in predio_aprobado or not predio_aprobado.get("r1"):
            predio_aprobado["r1"] = {
                "numero_orden": "1",
                "calificacion_no_certificada": "0",
                "tipo_predio": predio_aprobado.get("tipo_predio", ""),
                "numero_predial_anterior": "",
                "complemento_nom_predio": "",
                "area_total_terreno": predio_aprobado.get("area_terreno", 0),
                "valor_referencia": predio_aprobado.get("avaluo", 0),
                "tipo_avaluo_catastral": "1"
            }
        
        # Preparar propietarios en formato R1
        if predio_aprobado.get("propietarios"):
            predio_aprobado["r1_registros"] = []
            for i, prop in enumerate(predio_aprobado["propietarios"]):
                r1_registro = {
                    "numero_orden": str(i + 1),
                    "calificacion_no_certificada": "0",
                    "tipo_documento": prop.get("tipo_documento", "C"),
                    "numero_documento": prop.get("numero_documento", ""),
                    "nombre_propietario": prop.get("nombre_propietario", ""),
                    "estado_civil": prop.get("estado_civil", ""),
                    "direccion": predio_aprobado.get("direccion", ""),
                    "destino_economico": predio_aprobado.get("destino_economico", ""),
                    "area_terreno": predio_aprobado.get("area_terreno", 0),
                    "area_construida": predio_aprobado.get("area_construida", 0),
                    "avaluo": predio_aprobado.get("avaluo", 0)
                }
                predio_aprobado["r1_registros"].append(r1_registro)
        
        # Preparar zonas físicas en formato R2
        if predio_aprobado.get("zonas_fisicas"):
            predio_aprobado["r2_registros"] = []
            # Obtener matrícula del R2 o del nivel superior
            matricula = predio_aprobado.get("matricula_inmobiliaria", "")
            if not matricula and predio_aprobado.get("r2"):
                matricula = predio_aprobado["r2"].get("matricula_inmobiliaria", "")
            
            for i, zona in enumerate(predio_aprobado["zonas_fisicas"]):
                r2_registro = {
                    "numero_orden": str(i + 1),
                    "tipo_construccion": zona.get("tipo_construccion", "C"),
                    "numero_pisos": zona.get("pisos", 0),
                    "numero_habitaciones": zona.get("habitaciones", 0),
                    "numero_banios": zona.get("banos", 0),
                    "numero_locales": zona.get("locales", 0),
                    "anio_construccion": zona.get("anio_construccion", ""),
                    "area_construida": zona.get("area_construida", 0),
                    "area_terreno": zona.get("area_terreno", 0),
                    "puntaje": zona.get("puntaje", 0),
                    "zona_fisica": zona.get("zona_fisica", ""),
                    "zona_economica": zona.get("zona_economica", ""),
                    "matricula_inmobiliaria": matricula
                }
                predio_aprobado["r2_registros"].append(r2_registro)
            
            # Asegurar que la matrícula también esté a nivel superior del predio
            if matricula:
                predio_aprobado["matricula_inmobiliaria"] = matricula
                
        elif predio_aprobado.get("r2"):
            # Si tiene r2 pero no zonas_fisicas, crear r2_registros desde r2
            r2 = predio_aprobado["r2"]
            matricula = r2.get("matricula_inmobiliaria", "")
            
            predio_aprobado["r2_registros"] = [{
                "numero_orden": r2.get("numero_orden", "1"),
                "tipo_construccion": r2.get("tipo_construccion", "C"),
                "numero_pisos": r2.get("pisos_1", 0),
                "numero_habitaciones": r2.get("habitaciones_1", 0),
                "numero_banios": r2.get("banos_1", 0),
                "numero_locales": r2.get("locales_1", 0),
                "anio_construccion": r2.get("anio_construccion", ""),
                "area_construida": r2.get("area_construida_1", 0),
                "area_terreno": r2.get("area_terreno_1", 0),
                "puntaje": r2.get("puntaje_1", 0),
                "zona_fisica": r2.get("zona_fisica_1", ""),
                "zona_economica": r2.get("zona_economica_1", ""),
                "matricula_inmobiliaria": matricula
            }]
            
            # Asegurar que la matrícula también esté a nivel superior del predio
            if matricula:
                predio_aprobado["matricula_inmobiliaria"] = matricula
        
        # Insertar en colección principal
        await db.predios.insert_one(predio_aprobado)
        
        # Eliminar de predios_nuevos
        await db.predios_nuevos.delete_one({"id": predio_id})
    
    # Crear notificaciones
    destinatarios = set()
    if predio['gestor_creador_id'] and predio['gestor_creador_id'] != user_id:
        destinatarios.add(predio['gestor_creador_id'])
    if predio.get('gestor_apoyo_id') and predio['gestor_apoyo_id'] != user_id:
        destinatarios.add(predio['gestor_apoyo_id'])
    
    # Si se envía a revisión, notificar también a coordinadores y usuarios con permiso de aprobar
    if accion == "enviar_revision":
        municipio_predio = predio.get('municipio')
        
        # Obtener coordinadores y administradores
        coordinadores_cursor = db.users.find({
            "role": {"$in": [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]},
            "deleted": {"$ne": True}
        }, {"id": 1, "municipios": 1})
        
        async for coord in coordinadores_cursor:
            # Si el coordinador tiene municipios asignados, verificar que incluya el del predio
            coord_municipios = coord.get('municipios', [])
            if not coord_municipios or municipio_predio in coord_municipios:
                if coord['id'] != user_id:
                    destinatarios.add(coord['id'])
        
        # Obtener gestores con permiso de aprobar cambios
        permisos_cursor = db.user_permissions.find({
            "permissions.aprobar_cambios": True
        }, {"user_id": 1})
        
        async for perm in permisos_cursor:
            if perm['user_id'] != user_id:
                # Verificar que el usuario esté activo
                usuario = await db.users.find_one({"id": perm['user_id'], "deleted": {"$ne": True}})
                if usuario:
                    # Si tiene municipios asignados, verificar que incluya el del predio
                    user_municipios = usuario.get('municipios', [])
                    if not user_municipios or municipio_predio in user_municipios:
                        destinatarios.add(perm['user_id'])
        
        logger.info(f"Predio {predio['codigo_predial_nacional']} enviado a revisión. Notificando a {len(destinatarios)} usuarios aprobadores.")
    
    for dest_id in destinatarios:
        if dest_id:  # Verificar que el ID no sea None
            notif = {
                "id": str(uuid.uuid4()),
                "user_id": dest_id,
                "tipo": f"predio_{accion}",
                "titulo": mensaje_notif,
                "mensaje": accion_data.observaciones or mensaje_notif,
                "predio_id": predio_id,
                "codigo_predial": predio['codigo_predial_nacional'],
                "leida": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notif)
    
    # Retornar predio actualizado
    if accion == "aprobar":
        predio_final = await db.predios.find_one({"id": predio_id}, {"_id": 0})
    else:
        predio_final = await db.predios_nuevos.find_one({"id": predio_id}, {"_id": 0})
    
    return {
        "success": True,
        "mensaje": mensaje_notif,
        "predio": predio_final
    }


@api_router.get("/predios-nuevos/buscar-radicado/{numero}")
async def buscar_radicado(numero: str, current_user: dict = Depends(get_current_user)):
    """
    Busca un radicado por su número y retorna la fecha.
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Buscar petición con este número de radicado
    peticion = await db.petitions.find_one({
        "radicado": {"$regex": f"-{numero}-", "$options": "i"}
    }, {"_id": 0, "radicado": 1, "created_at": 1, "fecha_radicacion": 1, "solicitante": 1, "tipo_tramite": 1})
    
    if not peticion:
        return {
            "encontrado": False,
            "mensaje": "No se encontró petición con este número de radicado"
        }
    
    fecha_rad = peticion.get('created_at') or peticion.get('fecha_radicacion')
    if isinstance(fecha_rad, str):
        fecha_obj = datetime.fromisoformat(fecha_rad.replace('Z', '+00:00'))
    else:
        fecha_obj = fecha_rad
    
    return {
        "encontrado": True,
        "radicado_completo": peticion.get('radicado'),
        "fecha": fecha_obj.strftime("%d-%m-%Y") if fecha_obj else None,
        "solicitante": peticion.get('solicitante'),
        "tipo_tramite": peticion.get('tipo_tramite')
    }


@api_router.patch("/predios-nuevos/{predio_id}")
async def actualizar_predio_nuevo(
    predio_id: str,
    update_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Actualiza un predio nuevo (solo en estados editables).
    """
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    predio = await db.predios_nuevos.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Solo se puede editar en estados: creado, digitalizacion, devuelto
    if predio['estado_flujo'] not in [PredioNuevoEstado.CREADO, PredioNuevoEstado.DIGITALIZACION, PredioNuevoEstado.DEVUELTO]:
        raise HTTPException(status_code=400, detail="No se puede editar el predio en este estado")
    
    # Verificar que sea el gestor creador o de apoyo
    if predio['gestor_creador_id'] != current_user['id'] and predio['gestor_apoyo_id'] != current_user['id']:
        if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
            raise HTTPException(status_code=403, detail="Solo el gestor creador o de apoyo puede editar")
    
    # Campos protegidos que no se pueden cambiar
    campos_protegidos = ['id', 'codigo_predial_nacional', 'gestor_creador_id', 'created_at', 'historial_flujo']
    for campo in campos_protegidos:
        update_data.pop(campo, None)
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Agregar entrada al historial
    historial_entry = {
        "fecha": datetime.now(timezone.utc).isoformat(),
        "accion": "Predio actualizado",
        "estado_anterior": predio['estado_flujo'],
        "estado_nuevo": predio['estado_flujo'],
        "usuario_id": current_user['id'],
        "usuario_nombre": current_user['full_name'],
        "observaciones": f"Campos modificados: {', '.join(update_data.keys())}"
    }
    
    await db.predios_nuevos.update_one(
        {"id": predio_id},
        {
            "$set": update_data,
            "$push": {"historial_flujo": historial_entry}
        }
    )
    
    predio_actualizado = await db.predios_nuevos.find_one({"id": predio_id}, {"_id": 0})
    return predio_actualizado


@api_router.delete("/predios-nuevos/{predio_id}")
async def eliminar_predio_nuevo(
    predio_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """
    Elimina una solicitud de predio nuevo.
    Solo el gestor creador puede eliminar su propia solicitud.
    Solo se puede eliminar en estados: creado, digitalizacion, devuelto.
    """
    predio = await db.predios_nuevos.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Solo el gestor creador puede eliminar
    if predio['gestor_creador_id'] != current_user['id']:
        raise HTTPException(status_code=403, detail="Solo el gestor creador puede eliminar esta solicitud")
    
    # Solo se puede eliminar en estados editables
    if predio['estado_flujo'] not in [PredioNuevoEstado.CREADO, PredioNuevoEstado.DIGITALIZACION, PredioNuevoEstado.DEVUELTO]:
        raise HTTPException(status_code=400, detail="No se puede eliminar el predio en este estado. Solo se puede eliminar en estados: creado, digitalizacion o devuelto")
    
    # Obtener motivo del body
    try:
        body = await request.json()
        motivo = body.get('motivo', '')
    except:
        motivo = ''
    
    if not motivo:
        raise HTTPException(status_code=400, detail="Debe proporcionar un motivo para eliminar la solicitud")
    
    # Guardar registro de eliminación para auditoría
    registro_eliminacion = {
        "id": str(uuid.uuid4()),
        "predio_id": predio_id,
        "codigo_predial_nacional": predio.get('codigo_predial_nacional'),
        "municipio": predio.get('municipio'),
        "direccion": predio.get('direccion'),
        "gestor_creador_id": predio.get('gestor_creador_id'),
        "gestor_creador_nombre": predio.get('gestor_creador_nombre'),
        "eliminado_por_id": current_user['id'],
        "eliminado_por_nombre": current_user['full_name'],
        "motivo": motivo,
        "datos_predio": predio,
        "fecha_eliminacion": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios_nuevos_eliminados.insert_one(registro_eliminacion)
    
    # Eliminar el predio
    await db.predios_nuevos.delete_one({"id": predio_id})
    
    logger.info(f"Predio nuevo {predio_id} eliminado por {current_user['full_name']}. Motivo: {motivo}")
    
    return {"message": "Solicitud eliminada correctamente", "motivo": motivo}


# ===== SISTEMA DE APROBACIÓN DE PREDIOS =====

@api_router.post("/predios/cambios/proponer")
async def proponer_cambio_predio(
    cambio: CambioPendienteCreate,
    current_user: dict = Depends(get_current_user)
):
    """
    Propone un cambio en un predio (crear, modificar, eliminar).
    Solo gestores y atención pueden proponer. Coordinadores aprueban directamente.
    Opcionalmente puede asignar un gestor de apoyo para modificaciones.
    """
    # Verificar permisos - Usuarios, Comunicaciones y Empresa no pueden proponer cambios
    if current_user['role'] in [UserRole.USUARIO, UserRole.COMUNICACIONES, UserRole.EMPRESA]:
        raise HTTPException(status_code=403, detail="No tiene permiso para proponer cambios en predios")
    
    # Coordinadores y administradores aprueban directamente (a menos que asignen gestor de apoyo)
    aprueba_directo = current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]
    
    # Si se asigna gestor de apoyo, NO aprueba directo, va al flujo de apoyo
    usa_gestor_apoyo = cambio.gestor_apoyo_id is not None and cambio.tipo_cambio == "modificacion"
    if usa_gestor_apoyo:
        aprueba_directo = False
        # Verificar que el gestor de apoyo existe y tiene rol válido
        gestor_apoyo = await db.users.find_one({"id": cambio.gestor_apoyo_id}, {"_id": 0})
        if not gestor_apoyo:
            raise HTTPException(status_code=400, detail="El gestor de apoyo no existe")
        if gestor_apoyo.get('role') not in ['gestor', 'coordinador', 'administrador', 'atencion_usuario']:
            raise HTTPException(status_code=400, detail="El usuario asignado no tiene rol de gestor")
    
    # Obtener información del radicado si se proporcionó
    radicado_info = None
    if cambio.radicado_id:
        radicado_info = await db.petitions.find_one({"id": cambio.radicado_id}, {"_id": 0, "radicado": 1, "tipo_tramite": 1})
    
    # Obtener datos actuales del predio para modificaciones
    predio_actual_data = None
    if cambio.tipo_cambio == "modificacion" and cambio.predio_id:
        predio_existente = await db.predios.find_one({"id": cambio.predio_id}, {"_id": 0})
        if predio_existente:
            predio_actual_data = {
                "codigo_predial_nacional": predio_existente.get("codigo_predial_nacional"),
                "municipio": predio_existente.get("municipio"),
                "direccion": predio_existente.get("direccion"),
                "nombre_propietario": predio_existente.get("nombre_propietario"),
                "propietarios": predio_existente.get("propietarios", []),
                "area_terreno": predio_existente.get("area_terreno"),
                "area_construida": predio_existente.get("area_construida"),
                "avaluo": predio_existente.get("avaluo"),
                "destino_economico": predio_existente.get("destino_economico"),
                "zona": predio_existente.get("zona"),
            }
    
    # Determinar estado inicial según el flujo
    if aprueba_directo:
        estado_inicial = PredioEstadoAprobacion.APROBADO
    elif usa_gestor_apoyo:
        estado_inicial = "en_digitalizacion"  # Nuevo estado para flujo de apoyo
    else:
        estado_inicial = f"pendiente_{cambio.tipo_cambio}"
    
    cambio_doc = {
        "id": str(uuid.uuid4()),
        "predio_id": cambio.predio_id,
        "tipo_cambio": cambio.tipo_cambio,
        "datos_propuestos": cambio.datos_propuestos,
        "predio_actual": predio_actual_data,  # Guardar datos actuales para comparación
        "justificacion": cambio.justificacion,
        "radicado_id": cambio.radicado_id,
        "radicado_numero": radicado_info['radicado'] if radicado_info else cambio.radicado_numero,
        "estado": estado_inicial,
        "propuesto_por": current_user['id'],
        "propuesto_por_nombre": current_user['full_name'],
        "propuesto_por_rol": current_user['role'],
        "fecha_propuesta": datetime.now(timezone.utc).isoformat(),
        "aprobado_por": current_user['id'] if aprueba_directo else None,
        "aprobado_por_nombre": current_user['full_name'] if aprueba_directo else None,
        "fecha_aprobacion": datetime.now(timezone.utc).isoformat() if aprueba_directo else None,
        "comentario_aprobacion": "Aprobación directa por coordinador/administrador" if aprueba_directo else None
    }
    
    # Agregar datos del gestor de apoyo si se usa ese flujo
    if usa_gestor_apoyo:
        cambio_doc["gestor_apoyo_id"] = cambio.gestor_apoyo_id
        cambio_doc["gestor_apoyo_nombre"] = gestor_apoyo.get('full_name')
        cambio_doc["observaciones_apoyo"] = cambio.observaciones_apoyo
        cambio_doc["fecha_asignacion_apoyo"] = datetime.now(timezone.utc).isoformat()
    
    # Si aprueba directo, guardar también los datos_anteriores
    if aprueba_directo and predio_actual_data:
        cambio_doc["datos_anteriores"] = predio_actual_data
    
    # Si aprueba directo, aplicar el cambio inmediatamente
    if aprueba_directo:
        resultado = await aplicar_cambio_predio(cambio_doc, current_user)
        cambio_doc["resultado"] = resultado
    
    # Guardar el cambio en la colección de cambios
    await db.predios_cambios.insert_one(cambio_doc)
    
    # Notificar al gestor de apoyo si fue asignado
    if usa_gestor_apoyo:
        predio_codigo = predio_actual_data.get('codigo_predial_nacional', 'N/A') if predio_actual_data else 'N/A'
        await crear_notificacion(
            user_id=cambio.gestor_apoyo_id,
            tipo="modificacion_asignada",
            titulo="Modificación de predio asignada",
            mensaje=f"{current_user['full_name']} te ha asignado la modificación del predio {predio_codigo}",
            datos={"cambio_id": cambio_doc["id"], "predio_id": cambio.predio_id}
        )
    
    # Construir mensaje de respuesta
    if aprueba_directo:
        mensaje = "Cambio aplicado directamente"
    elif usa_gestor_apoyo:
        mensaje = f"Modificación asignada a {gestor_apoyo.get('full_name')} para completar"
    else:
        mensaje = "Cambio propuesto, pendiente de aprobación"
    
    return {
        "id": cambio_doc["id"],
        "estado": cambio_doc["estado"],
        "mensaje": mensaje,
        "requiere_aprobacion": not aprueba_directo,
        "asignado_gestor_apoyo": usa_gestor_apoyo
    }


@api_router.get("/predios/cambios/pendientes")
async def get_cambios_pendientes(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista todos los cambios pendientes de aprobación (coordinadores/admin o usuarios con permiso)"""
    # Verificar si es coordinador/admin o tiene permiso de aprobar cambios
    user_permissions = current_user.get('permissions', [])
    has_approve_permission = 'approve_changes' in user_permissions
    
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR] and not has_approve_permission:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver cambios pendientes")
    
    query = {
        "estado": {"$in": [
            PredioEstadoAprobacion.PENDIENTE_CREACION,
            PredioEstadoAprobacion.PENDIENTE_MODIFICACION,
            PredioEstadoAprobacion.PENDIENTE_ELIMINACION
        ]}
    }
    
    total = await db.predios_cambios.count_documents(query)
    cambios = await db.predios_cambios.find(query, {"_id": 0}).sort("fecha_propuesta", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enriquecer con datos del predio actual si existe (para comparación)
    for cambio in cambios:
        if cambio.get("predio_id"):
            # Obtener TODOS los campos del predio actual para permitir comparación
            predio = await db.predios.find_one(
                {"id": cambio["predio_id"]}, 
                {"_id": 0, "historial": 0}  # Excluir historial para reducir tamaño
            )
            cambio["predio_actual"] = predio
    
    return {
        "total": total,
        "cambios": cambios
    }


@api_router.post("/predios/cambios/aprobar")
async def aprobar_rechazar_cambio(
    request: CambioAprobacionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Aprueba o rechaza un cambio pendiente (solo coordinadores/admin o con permiso)"""
    has_permission = await check_permission(current_user, Permission.APPROVE_CHANGES)
    if not has_permission:
        raise HTTPException(status_code=403, detail="No tiene permiso para aprobar cambios")
    
    # Buscar el cambio
    cambio = await db.predios_cambios.find_one({"id": request.cambio_id}, {"_id": 0})
    
    if not cambio:
        raise HTTPException(status_code=404, detail="Cambio no encontrado")
    
    if cambio["estado"] not in [
        PredioEstadoAprobacion.PENDIENTE_CREACION,
        PredioEstadoAprobacion.PENDIENTE_MODIFICACION,
        PredioEstadoAprobacion.PENDIENTE_ELIMINACION
    ]:
        raise HTTPException(status_code=400, detail="Este cambio ya fue procesado")
    
    nuevo_estado = PredioEstadoAprobacion.APROBADO if request.aprobado else PredioEstadoAprobacion.RECHAZADO
    
    update_data = {
        "estado": nuevo_estado,
        "aprobado_por": current_user['id'],
        "aprobado_por_nombre": current_user['full_name'],
        "fecha_aprobacion": datetime.now(timezone.utc).isoformat(),
        "comentario_aprobacion": request.comentario
    }
    
    # Si es una modificación, guardar los datos anteriores del predio ANTES de aplicar el cambio
    if cambio["tipo_cambio"] == "modificacion" and cambio.get("predio_id"):
        predio_actual = await db.predios.find_one({"id": cambio["predio_id"]}, {"_id": 0})
        if predio_actual:
            # Guardar solo los campos relevantes como datos_anteriores
            datos_anteriores = {
                "codigo_predial_nacional": predio_actual.get("codigo_predial_nacional"),
                "municipio": predio_actual.get("municipio"),
                "direccion": predio_actual.get("direccion"),
                "nombre_propietario": predio_actual.get("nombre_propietario"),
                "propietarios": predio_actual.get("propietarios", []),
                "area_terreno": predio_actual.get("area_terreno"),
                "area_construida": predio_actual.get("area_construida"),
                "avaluo": predio_actual.get("avaluo"),
                "destino_economico": predio_actual.get("destino_economico"),
                "zona": predio_actual.get("zona"),
            }
            update_data["datos_anteriores"] = datos_anteriores
    
    # Si se aprueba, aplicar el cambio
    if request.aprobado:
        resultado = await aplicar_cambio_predio(cambio, current_user)
        update_data["resultado"] = resultado
    
    await db.predios_cambios.update_one(
        {"id": request.cambio_id},
        {"$set": update_data}
    )
    
    # Notificar al usuario que propuso el cambio
    if cambio.get("propuesto_por"):
        codigo_predio = cambio.get("datos_propuestos", {}).get("codigo_predial_nacional", 
                        cambio.get("datos_propuestos", {}).get("codigo_homologado", "N/A"))
        
        if request.aprobado:
            # Notificar aprobación - incluir instrucción para sincronizar
            await crear_notificacion(
                usuario_id=cambio["propuesto_por"],
                titulo="✅ Cambio de Predio Aprobado",
                mensaje=f"El cambio propuesto para el predio {codigo_predio} ha sido aprobado por {current_user['full_name']}. Por favor, sincronice los datos del municipio para ver los cambios actualizados.",
                tipo="success",
                enlace=f"/dashboard/conservacion?municipio={cambio.get('datos_propuestos', {}).get('municipio', '')}",
                enviar_email=False
            )
        else:
            # Notificar rechazo
            await crear_notificacion(
                usuario_id=cambio["propuesto_por"],
                titulo="❌ Cambio de Predio Rechazado",
                mensaje=f"El cambio propuesto para el predio {codigo_predio} ha sido rechazado por {current_user['full_name']}. Motivo: {request.comentario or 'Sin comentario'}",
                tipo="warning",
                enlace="/dashboard/pendientes",
                enviar_email=False
            )
    
    # === WEBSOCKET NOTIFICATION ===
    # Broadcast real-time notification to all connected users
    ws_notification = {
        "type": "cambio_predio",
        "action": "aprobado" if request.aprobado else "rechazado",
        "cambio_id": request.cambio_id,
        "predio_id": cambio.get("predio_id"),
        "codigo_predio": cambio.get("datos_propuestos", {}).get("codigo_homologado", "N/A"),
        "municipio": cambio.get("datos_propuestos", {}).get("municipio", ""),
        "decidido_por": current_user['full_name'],
        "propuesto_por": cambio.get("propuesto_por"),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await ws_manager.broadcast(ws_notification, exclude_user=current_user['id'])
    
    return {
        "mensaje": "Cambio aprobado y aplicado" if request.aprobado else "Cambio rechazado",
        "estado": nuevo_estado
    }


# ===== ENDPOINTS PARA FLUJO DE GESTOR DE APOYO EN MODIFICACIONES =====

@api_router.get("/predios/cambios/mis-asignaciones")
async def get_mis_asignaciones_apoyo(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las modificaciones asignadas al gestor de apoyo actual"""
    query = {
        "gestor_apoyo_id": current_user['id'],
        "estado": "en_digitalizacion"
    }
    
    total = await db.predios_cambios.count_documents(query)
    cambios = await db.predios_cambios.find(query, {"_id": 0}).sort("fecha_asignacion_apoyo", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enriquecer con datos del predio actual
    for cambio in cambios:
        if cambio.get("predio_id"):
            predio = await db.predios.find_one(
                {"id": cambio["predio_id"]}, 
                {"_id": 0, "historial": 0}
            )
            cambio["predio_actual"] = predio
    
    return {
        "total": total,
        "cambios": cambios
    }


@api_router.get("/predios/cambios/en-digitalizacion")
async def get_cambios_en_digitalizacion(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas las modificaciones en estado 'en_digitalizacion' (para coordinadores)"""
    # Verificar permisos
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        has_permission = 'approve_changes' in current_user.get('permissions', [])
        if not has_permission:
            raise HTTPException(status_code=403, detail="No tiene permiso para ver esta información")
    
    query = {"estado": "en_digitalizacion"}
    
    total = await db.predios_cambios.count_documents(query)
    cambios = await db.predios_cambios.find(query, {"_id": 0}).sort("fecha_asignacion_apoyo", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enriquecer con datos del predio actual
    for cambio in cambios:
        if cambio.get("predio_id"):
            predio = await db.predios.find_one(
                {"id": cambio["predio_id"]}, 
                {"_id": 0, "historial": 0}
            )
            cambio["predio_actual"] = predio
    
    return {
        "total": total,
        "cambios": cambios
    }


class CompletarApoyoRequest(BaseModel):
    """Request para completar una asignación de apoyo"""
    datos_actualizados: Optional[dict] = None  # Datos modificados por el gestor de apoyo
    observaciones: Optional[str] = None


@api_router.post("/predios/cambios/{cambio_id}/completar-apoyo")
async def completar_asignacion_apoyo(
    cambio_id: str,
    request: CompletarApoyoRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    El gestor de apoyo completa la modificación y la envía a revisión del coordinador.
    """
    # Buscar el cambio
    cambio = await db.predios_cambios.find_one({"id": cambio_id}, {"_id": 0})
    
    if not cambio:
        raise HTTPException(status_code=404, detail="Cambio no encontrado")
    
    if cambio["estado"] != "en_digitalizacion":
        raise HTTPException(status_code=400, detail="Este cambio no está en estado de digitalización")
    
    # Verificar que el usuario actual es el gestor de apoyo asignado o es coordinador/admin
    es_gestor_asignado = cambio.get("gestor_apoyo_id") == current_user['id']
    es_supervisor = current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]
    
    if not es_gestor_asignado and not es_supervisor:
        raise HTTPException(status_code=403, detail="Solo el gestor de apoyo asignado puede completar esta modificación")
    
    # Actualizar los datos propuestos si se proporcionaron actualizaciones
    datos_finales = cambio.get("datos_propuestos", {})
    if request.datos_actualizados:
        datos_finales.update(request.datos_actualizados)
    
    update_data = {
        "estado": PredioEstadoAprobacion.PENDIENTE_MODIFICACION,
        "datos_propuestos": datos_finales,
        "completado_por_apoyo": current_user['id'],
        "completado_por_apoyo_nombre": current_user['full_name'],
        "fecha_completado_apoyo": datetime.now(timezone.utc).isoformat(),
        "observaciones_completado": request.observaciones
    }
    
    await db.predios_cambios.update_one(
        {"id": cambio_id},
        {"$set": update_data}
    )
    
    # Notificar al coordinador/aprobadores que hay un cambio listo para revisión
    predio_codigo = cambio.get("predio_actual", {}).get("codigo_predial_nacional", "N/A")
    
    # Notificar al creador original del cambio
    if cambio.get("propuesto_por"):
        await crear_notificacion(
            user_id=cambio["propuesto_por"],
            tipo="modificacion_completada",
            titulo="Modificación completada por gestor de apoyo",
            mensaje=f"{current_user['full_name']} completó la modificación del predio {predio_codigo}. Pendiente de aprobación.",
            datos={"cambio_id": cambio_id, "predio_id": cambio.get("predio_id")}
        )
    
    return {
        "mensaje": "Modificación completada y enviada a revisión",
        "cambio_id": cambio_id,
        "nuevo_estado": PredioEstadoAprobacion.PENDIENTE_MODIFICACION
    }


@api_router.get("/predios/cambios/stats-apoyo")
async def get_stats_apoyo(
    current_user: dict = Depends(get_current_user)
):
    """Obtiene estadísticas del flujo de gestor de apoyo"""
    
    # Mis asignaciones pendientes
    mis_asignaciones = await db.predios_cambios.count_documents({
        "gestor_apoyo_id": current_user['id'],
        "estado": "en_digitalizacion"
    })
    
    # Total en digitalización (para coordinadores)
    total_en_digitalizacion = 0
    if current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        total_en_digitalizacion = await db.predios_cambios.count_documents({
            "estado": "en_digitalizacion"
        })
    
    return {
        "mis_asignaciones_pendientes": mis_asignaciones,
        "total_en_digitalizacion": total_en_digitalizacion
    }


class VincularRadicadoRequest(BaseModel):
    radicado_id: str
    radicado_numero: Optional[str] = None

@api_router.patch("/predios/cambios/{cambio_id}/vincular-radicado")
async def vincular_radicado_a_cambio(
    cambio_id: str,
    request: VincularRadicadoRequest,
    current_user: dict = Depends(get_current_user)
):
    """Vincula un radicado/petición a un cambio pendiente existente"""
    # Permitir a coordinadores, admins y gestores
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para vincular radicados")
    
    # Buscar el cambio
    cambio = await db.predios_cambios.find_one({"id": cambio_id}, {"_id": 0})
    if not cambio:
        raise HTTPException(status_code=404, detail="Cambio no encontrado")
    
    # Actualizar con el radicado
    update_data = {
        "radicado_id": request.radicado_id,
        "radicado_numero": request.radicado_numero,
        "vinculado_por": current_user['id'],
        "vinculado_por_nombre": current_user['full_name'],
        "fecha_vinculacion": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios_cambios.update_one(
        {"id": cambio_id},
        {"$set": update_data}
    )
    
    return {
        "mensaje": f"Radicado {request.radicado_numero} vinculado exitosamente",
        "cambio_id": cambio_id,
        "radicado_numero": request.radicado_numero
    }


async def aplicar_cambio_predio(cambio: dict, aprobador: dict) -> dict:
    """Aplica un cambio aprobado al predio"""
    tipo = cambio["tipo_cambio"]
    datos = cambio["datos_propuestos"]
    
    historial_entry = {
        "accion": f"Cambio {tipo} aprobado",
        "usuario": aprobador['full_name'],
        "usuario_rol": aprobador['role'],
        "propuesto_por": cambio.get("propuesto_por_nombre"),
        "fecha": datetime.now(timezone.utc).isoformat(),
        "comentario": cambio.get("comentario_aprobacion")
    }
    
    if tipo == "creacion":
        # Crear nuevo predio
        predio_doc = datos.copy()
        predio_doc["id"] = str(uuid.uuid4())
        predio_doc["estado_aprobacion"] = PredioEstadoAprobacion.APROBADO
        predio_doc["deleted"] = False
        predio_doc["created_at"] = datetime.now(timezone.utc).isoformat()
        predio_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
        predio_doc["historial"] = [historial_entry]
        
        await db.predios.insert_one(predio_doc)
        return {"predio_id": predio_doc["id"], "accion": "creado"}
    
    elif tipo == "modificacion":
        predio_id = cambio["predio_id"]
        
        # Actualizar predio
        datos["updated_at"] = datetime.now(timezone.utc).isoformat()
        datos["estado_aprobacion"] = PredioEstadoAprobacion.APROBADO
        
        # Convertir r2 a formato r2_registros para mantener consistencia
        if "r2" in datos:
            r2_data = datos["r2"]
            datos["r2_registros"] = [{
                "matricula_inmobiliaria": r2_data.get("matricula_inmobiliaria"),
                "zonas": r2_data.get("zonas", [])
            }]
            del datos["r2"]  # Eliminar campo r2 para evitar duplicados
        
        await db.predios.update_one(
            {"id": predio_id},
            {
                "$set": datos,
                "$push": {"historial": historial_entry}
            }
        )
        return {"predio_id": predio_id, "accion": "modificado"}
    
    elif tipo == "eliminacion":
        predio_id = cambio["predio_id"]
        
        # Soft delete
        await db.predios.update_one(
            {"id": predio_id},
            {
                "$set": {
                    "deleted": True,
                    "deleted_at": datetime.now(timezone.utc).isoformat(),
                    "deleted_by": aprobador['id'],
                    "deleted_by_name": aprobador['full_name'],
                    "estado_aprobacion": PredioEstadoAprobacion.APROBADO
                },
                "$push": {"historial": historial_entry}
            }
        )
        return {"predio_id": predio_id, "accion": "eliminado"}
    
    return {"error": "Tipo de cambio no reconocido"}


@api_router.get("/predios/cambios/mis-propuestas")
async def get_mis_propuestas(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista las propuestas de cambio del usuario actual"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {"propuesto_por": current_user['id']}
    
    total = await db.predios_cambios.count_documents(query)
    cambios = await db.predios_cambios.find(query, {"_id": 0}).sort("fecha_propuesta", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "total": total,
        "cambios": cambios
    }


@api_router.get("/predios/cambios/stats")
async def get_cambios_stats(current_user: dict = Depends(get_current_user)):
    """Obtiene estadísticas de cambios pendientes e historial"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    pendientes_creacion = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.PENDIENTE_CREACION})
    pendientes_modificacion = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.PENDIENTE_MODIFICACION})
    pendientes_eliminacion = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.PENDIENTE_ELIMINACION})
    
    # Contar modificaciones en digitalización (flujo de gestor de apoyo)
    en_digitalizacion = await db.predios_cambios.count_documents({"estado": "en_digitalizacion"})
    
    # Mis asignaciones de apoyo pendientes
    mis_asignaciones_apoyo = await db.predios_cambios.count_documents({
        "gestor_apoyo_id": current_user['id'],
        "estado": "en_digitalizacion"
    })
    
    # Contar historial de cambios procesados
    historial_aprobados = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.APROBADO})
    historial_rechazados = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.RECHAZADO})
    
    return {
        "pendientes_creacion": pendientes_creacion,
        "pendientes_modificacion": pendientes_modificacion,
        "pendientes_eliminacion": pendientes_eliminacion,
        "en_digitalizacion": en_digitalizacion,
        "mis_asignaciones_apoyo": mis_asignaciones_apoyo,
        "total_pendientes": pendientes_creacion + pendientes_modificacion + pendientes_eliminacion,
        "historial_aprobados": historial_aprobados,
        "historial_rechazados": historial_rechazados,
        "total_historial": historial_aprobados + historial_rechazados
    }


@api_router.get("/predios/cambios/historial")
async def get_cambios_historial(
    limit: int = Query(50, ge=1, le=200),
    estado: Optional[str] = Query(None, description="Filtrar por estado: 'aprobado' o 'rechazado'"),
    tipo_cambio: Optional[str] = Query(None, description="Filtrar por tipo: 'creacion', 'modificacion', 'eliminacion'"),
    municipio: Optional[str] = Query(None, description="Filtrar por municipio"),
    fecha_desde: Optional[str] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el historial de cambios aprobados y rechazados con filtros opcionales"""
    # Permitir a coordinadores, admins, gestores y atención ver el historial
    allowed_roles = [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.GESTOR, UserRole.ATENCION_USUARIO]
    if current_user['role'] not in allowed_roles:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver el historial de cambios")
    
    # Buscar cambios que ya fueron procesados (aprobados o rechazados)
    estados_procesados = [
        PredioEstadoAprobacion.APROBADO,
        PredioEstadoAprobacion.RECHAZADO
    ]
    
    # Construir query con filtros
    query = {"estado": {"$in": estados_procesados}}
    
    # Filtro por estado específico
    if estado:
        if estado == 'aprobado':
            query["estado"] = PredioEstadoAprobacion.APROBADO
        elif estado == 'rechazado':
            query["estado"] = PredioEstadoAprobacion.RECHAZADO
    
    # Filtro por tipo de cambio
    if tipo_cambio:
        query["tipo_cambio"] = tipo_cambio
    
    # Filtro por municipio
    if municipio:
        query["datos_propuestos.municipio"] = municipio
    
    # Filtro por fecha
    if fecha_desde or fecha_hasta:
        fecha_query = {}
        if fecha_desde:
            fecha_query["$gte"] = fecha_desde + "T00:00:00"
        if fecha_hasta:
            fecha_query["$lte"] = fecha_hasta + "T23:59:59"
        query["fecha_aprobacion"] = fecha_query
    
    cambios = await db.predios_cambios.find(
        query,
        {"_id": 0}
    ).sort("fecha_aprobacion", -1).limit(limit).to_list(limit)
    
    # Enriquecer con datos del predio original
    for cambio in cambios:
        # Solo enriquecer si NO tiene datos_anteriores guardados (para compatibilidad con registros antiguos)
        if cambio.get('predio_id') and not cambio.get('datos_anteriores'):
            predio = await db.predios.find_one({"id": cambio['predio_id']}, {"_id": 0, "codigo_predial_nacional": 1, "nombre_propietario": 1, "codigo_homologado": 1, "direccion": 1, "municipio": 1, "area_terreno": 1, "area_construida": 1, "avaluo": 1, "destino_economico": 1, "zona": 1, "propietarios": 1})
            if predio:
                # Usar como predio_actual para mostrar datos de referencia (actuales)
                # Nota: Este es el estado ACTUAL, no el estado en el momento del cambio
                cambio['predio_actual'] = predio
                cambio['predio_actual_es_referencia'] = True  # Flag para indicar que no son datos históricos
    
    # Obtener lista de municipios únicos para el selector de filtros
    municipios_cursor = await db.predios_cambios.distinct("datos_propuestos.municipio", {"estado": {"$in": estados_procesados}})
    municipios = sorted([m for m in municipios_cursor if m])
    
    return {
        "cambios": cambios,
        "total": len(cambios),
        "municipios": municipios
    }


# ===== GEOGRAPHIC DATABASE (GDB) INTEGRATION =====

GDB_PATH = Path("/app/gdb_data/54003.gdb")

async def get_gdb_geometry_async(codigo_predial: str) -> Optional[dict]:
    """Get geometry for a property from MongoDB gdb_geometrias collection first, then fallback to GDB files"""
    
    try:
        # Buscar en la colección gdb_geometrias de MongoDB por código exacto
        geometria = await db.gdb_geometrias.find_one(
            {"codigo": codigo_predial},
            {"_id": 0}
        )
        
        if geometria:
            return {
                "type": "Feature",
                "geometry": geometria.get("geometry"),
                "properties": {
                    "codigo": codigo_predial,
                    "tipo": geometria.get("tipo", "Rural"),
                    "municipio": geometria.get("municipio", ""),
                    "area_m2": geometria.get("area_m2", 0)
                }
            }
        
        # Si no se encontró con código exacto, buscar el terreno base para mejoras
        # Una mejora tiene los últimos 8 dígitos != "00000000"
        # El terreno base es: primeros 21 dígitos + "000000000"
        if len(codigo_predial) == 30:
            ultimos_8 = codigo_predial[-8:]
            es_mejora = ultimos_8 != "00000000"
            
            if es_mejora:
                # Buscar el terreno base (mismo predio, sin edificio/piso/unidad/mejora)
                codigo_terreno_base = codigo_predial[:21] + "000000000"
                geometria = await db.gdb_geometrias.find_one(
                    {"codigo": codigo_terreno_base},
                    {"_id": 0}
                )
                
                if geometria:
                    return {
                        "type": "Feature",
                        "geometry": geometria.get("geometry"),
                        "properties": {
                            "codigo": codigo_predial,  # Mantener código original de la mejora
                            "codigo_terreno": codigo_terreno_base,  # Indicar que es del terreno
                            "tipo": geometria.get("tipo", "Rural"),
                            "municipio": geometria.get("municipio", ""),
                            "area_m2": geometria.get("area_m2", 0),
                            "es_mejora": True
                        }
                    }
                
                # Buscar por predio (primeros 21 dígitos) con cualquier terminación
                geometria = await db.gdb_geometrias.find_one(
                    {"codigo": {"$regex": f"^{codigo_predial[:21]}"}},
                    {"_id": 0}
                )
                
                if geometria:
                    return {
                        "type": "Feature",
                        "geometry": geometria.get("geometry"),
                        "properties": {
                            "codigo": codigo_predial,
                            "codigo_terreno": geometria.get("codigo"),
                            "tipo": geometria.get("tipo", "Rural"),
                            "municipio": geometria.get("municipio", ""),
                            "area_m2": geometria.get("area_m2", 0),
                            "es_mejora": True
                        }
                    }
        
        # Si no está en MongoDB, NO hacer búsqueda parcial
        # El predio simplemente no tiene geometría disponible
        return None
        
    except Exception as e:
        logger.error(f"Error getting geometry from MongoDB: {e}")
        return None


def get_gdb_geometry(codigo_predial: str) -> Optional[dict]:
    """Get geometry for a property from multiple GDB files, transformed to WGS84 for web mapping"""
    import geopandas as gpd
    from shapely.geometry import mapping
    
    # Mapeo de códigos de municipio a archivos GDB
    GDB_FILES = {
        '54003': '/app/gdb_data/54003.gdb',  # Ábrego
        '54109': '/app/gdb_data/54109.gdb',  # Bucarasica
        '54128': '/app/gdb_data/54128.gdb',  # Cáchira
    }
    
    try:
        # Extraer código de municipio del código predial (posiciones 0-5)
        municipio_code = codigo_predial[:5]
        
        # Determinar si es rural o urbano
        sector = codigo_predial[5:8]
        is_urban = sector != "000"
        
        # Buscar el archivo GDB correcto
        gdb_path = None
        for code, path in GDB_FILES.items():
            if municipio_code == code:
                gdb_path = Path(path)
                break
        
        if not gdb_path or not gdb_path.exists():
            # Intentar con el GDB por defecto
            if GDB_PATH.exists():
                gdb_path = GDB_PATH
            else:
                return None
        
        # Determinar nombre de capa (diferentes GDBs usan diferentes nombres)
        # 54003 (Ábrego): R_TERRENO_1, U_TERRENO_1
        # Otros: R_TERRENO, U_TERRENO
        if '54003' in str(gdb_path):
            layer = "U_TERRENO_1" if is_urban else "R_TERRENO_1"
        else:
            layer = "U_TERRENO" if is_urban else "R_TERRENO"
        
        # Leer capa y buscar el código
        gdf = gpd.read_file(str(gdb_path), layer=layer)
        match = gdf[gdf['codigo'] == codigo_predial]
        
        if len(match) == 0:
            return None
        
        # Obtener área y perímetro originales
        area_m2 = float(match.iloc[0]['shape_Area']) if 'shape_Area' in match.columns else None
        perimetro_m = float(match.iloc[0]['shape_Length']) if 'shape_Length' in match.columns else None
        
        # Transformar a WGS84
        match_wgs84 = match.to_crs(epsg=4326)
        
        geom = match_wgs84.iloc[0]['geometry']
        if geom is None:
            return None
            
        geojson = mapping(geom)
        
        return {
            "type": "Feature",
            "geometry": geojson,
            "properties": {
                "codigo": codigo_predial,
                "area_m2": area_m2,
                "perimetro_m": perimetro_m,
                "tipo": "Urbano" if is_urban else "Rural"
            }
        }
    except Exception as e:
        logger.error(f"Error reading GDB geometry: {e}")
        return None


@api_router.get("/predios/{predio_id}/geometria")
async def get_predio_geometry(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Get geographic geometry for a property"""
    # Only staff can access geometry
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Get predio from database
    predio = await db.predios.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    codigo = predio.get("codigo_predial_nacional")
    if not codigo:
        raise HTTPException(status_code=404, detail="Predio sin código catastral")
    
    # Primero intentar con la función async que busca en MongoDB
    geometry = await get_gdb_geometry_async(codigo)
    if not geometry:
        # Fallback a la función que lee archivos GDB directamente
        geometry = get_gdb_geometry(codigo)
    if not geometry:
        raise HTTPException(status_code=404, detail="Geometría no disponible para este predio")
    
    return geometry


@api_router.get("/predios/codigo/{codigo_predial}/geometria")
async def get_geometry_by_code(codigo_predial: str, current_user: dict = Depends(get_current_user)):
    """Get geographic geometry directly by cadastral code"""
    # Only staff can access geometry
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Primero intentar con la función async que busca en MongoDB
    geometry = await get_gdb_geometry_async(codigo_predial)
    if not geometry:
        # Fallback a la función que lee archivos GDB directamente
        geometry = get_gdb_geometry(codigo_predial)
    if not geometry:
        raise HTTPException(status_code=404, detail="Geometría no disponible para este código")
    
    return geometry


@api_router.get("/gdb/geometrias")
async def get_geometrias_filtradas(
    municipio: Optional[str] = None,
    zona: Optional[str] = None,  # 'urbano' o 'rural'
    limit: int = 500,
    current_user: dict = Depends(get_current_user)
):
    """Get all geometries for a municipality/zone from MongoDB collection"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    if not municipio:
        raise HTTPException(status_code=400, detail="Debe especificar un municipio")
    
    # Si es usuario empresa, verificar que el municipio esté en sus asignados
    if current_user['role'] == UserRole.EMPRESA:
        municipios_asignados = current_user.get('municipios_asignados', [])
        if not municipios_asignados:
            raise HTTPException(status_code=403, detail="No tiene municipios asignados. Contacte al coordinador.")
        if municipio not in municipios_asignados:
            raise HTTPException(status_code=403, detail=f"No tiene acceso al municipio {municipio}")
    
    # Construir query
    query = {"municipio": municipio}
    if zona == 'urbano':
        query["tipo"] = "urbano"
    elif zona == 'rural':
        query["tipo"] = "rural"
    
    try:
        # Buscar en la colección gdb_geometrias
        geometrias = await db.gdb_geometrias.find(
            query,
            {"_id": 0, "codigo": 1, "tipo": 1, "geometry": 1}
        ).limit(limit).to_list(limit)
        
        if not geometrias:
            # Verificar si hay datos para este municipio
            total_municipio = await db.gdb_geometrias.count_documents({"municipio": municipio})
            if total_municipio == 0:
                raise HTTPException(status_code=404, detail=f"No hay datos geográficos para {municipio}. Cargue un archivo GDB primero.")
        
        # Convertir a GeoJSON FeatureCollection
        features = []
        for geom in geometrias:
            feature = {
                "type": "Feature",
                "geometry": geom.get("geometry"),
                "properties": {
                    "codigo": geom.get("codigo"),
                    "tipo": geom.get("tipo", "").capitalize()
                }
            }
            features.append(feature)
        
        return {
            "type": "FeatureCollection",
            "municipio": municipio,
            "zona_filter": zona,
            "total": len(features),
            "features": features
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting geometries: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener geometrías: {str(e)}")


@api_router.get("/gdb/limites-municipios")
async def get_limites_municipios(
    fuente: str = "gdb",  # "gdb" para calculados con líneas internas, "oficial" para DANE/IGAC
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene los límites de todos los municipios.
    - fuente="gdb": Límites calculados desde geometrías GDB (muestra líneas internas para revisar errores)
    - fuente="oficial": Límites oficiales DANE/IGAC (limpios, sin líneas internas)
    """
    from shapely.geometry import shape, mapping, box, Polygon
    from shapely.ops import unary_union
    
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    try:
        features = []
        municipios_con_limite = set()
        
        # Si se piden límites oficiales, usar la colección limites_municipales
        if fuente == "oficial":
            async for doc in db.limites_municipales.find({}, {"_id": 0}):
                municipio = doc.get("municipio")
                geometry = doc.get("geometry")
                sin_gdb = doc.get("sin_gdb", False)
                
                if municipio and geometry:
                    municipios_con_limite.add(municipio)
                    # Obtener stats de predios
                    stats = await db.gdb_geometrias.aggregate([
                        {"$match": {"municipio": municipio}},
                        {"$group": {"_id": "$tipo", "count": {"$sum": 1}}}
                    ]).to_list(10)
                    rural_count = 0
                    urbano_count = 0
                    for s in stats:
                        if s["_id"] == "rural":
                            rural_count = s["count"]
                        else:
                            urbano_count = s["count"]
                    
                    try:
                        geom = shape(geometry)
                        geom_simplified = geom.simplify(0.0005, preserve_topology=True)
                        centroid = geom_simplified.centroid
                        
                        features.append({
                            "type": "Feature",
                            "geometry": mapping(geom_simplified),
                            "properties": {
                                "municipio": municipio,
                                "total_predios": rural_count + urbano_count,
                                "rurales": rural_count,
                                "urbanos": urbano_count,
                                "centroid": [centroid.x, centroid.y],
                                "fuente": doc.get("fuente", "dane_igac"),
                                "sin_gdb": sin_gdb
                            }
                        })
                    except Exception as e:
                        logger.warning(f"Error procesando límite oficial de {municipio}: {e}")
            
            # Ordenar y retornar
            features.sort(key=lambda x: x["properties"]["municipio"])
            return {
                "type": "FeatureCollection",
                "total_municipios": len(features),
                "fuente": "oficial",
                "features": features
            }
        
        # Si se piden límites GDB (calculados), usar geometrías para mostrar líneas internas
        # Primero agregar municipios sin GDB desde límites oficiales
        async for doc in db.limites_municipales.find({"sin_gdb": True}, {"_id": 0}):
            municipio = doc.get("municipio")
            geometry = doc.get("geometry")
            if municipio and geometry:
                municipios_con_limite.add(municipio)
                try:
                    geom = shape(geometry)
                    centroid = geom.centroid
                    features.append({
                        "type": "Feature",
                        "geometry": geometry,
                        "properties": {
                            "municipio": municipio,
                            "total_predios": 0,
                            "rurales": 0,
                            "urbanos": 0,
                            "centroid": [centroid.x, centroid.y],
                            "fuente": "dane_igac",
                            "sin_gdb": True
                        }
                    })
                except Exception as e:
                    logger.warning(f"Error procesando límite de {municipio}: {e}")
        
        # Calcular límites desde geometrías GDB (muestra líneas internas)
        municipios_gdb = await db.gdb_geometrias.distinct("municipio")
        
        for municipio in municipios_gdb:
            if municipio in municipios_con_limite:
                continue
            
            try:
                # PRIMERO: Obtener el conteo REAL de TODAS las geometrías del municipio
                stats = await db.gdb_geometrias.aggregate([
                    {"$match": {"municipio": municipio}},
                    {"$group": {"_id": "$tipo", "count": {"$sum": 1}}}
                ]).to_list(10)
                rural_count = 0
                urbano_count = 0
                for s in stats:
                    if s["_id"] == "rural":
                        rural_count = s["count"]
                    else:
                        urbano_count = s["count"]
                total_real = rural_count + urbano_count
                
                # SEGUNDO: Obtener geometrías para calcular el límite visual (limitado para rendimiento)
                geometrias_cursor = db.gdb_geometrias.find(
                    {"municipio": municipio},
                    {"_id": 0, "geometry": 1}
                ).limit(5000)  # Limitar solo para cálculo visual, NO para conteo
                
                shapes = []
                
                async for doc in geometrias_cursor:
                    geom_dict = doc.get("geometry")
                    if geom_dict:
                        try:
                            geom = shape(geom_dict)
                            if geom.is_valid:
                                shapes.append(geom)
                        except:
                            continue
                
                if shapes:
                    # Crear el límite REAL del municipio usando unary_union
                    union_geom = unary_union(shapes)
                    # Simplificar para rendimiento pero mantener forma real
                    limite = union_geom.simplify(0.0008, preserve_topology=True)
                    centroid = limite.centroid
                    
                    features.append({
                        "type": "Feature",
                        "geometry": mapping(limite),
                        "properties": {
                            "municipio": municipio,
                            "total_predios": total_real,  # Usar conteo REAL, no len(shapes)
                            "rurales": rural_count,
                            "urbanos": urbano_count,
                            "centroid": [centroid.x, centroid.y],
                            "fuente": "calculado"
                        }
                    })
                    municipios_con_limite.add(municipio)
            except Exception as e:
                logger.warning(f"Error procesando {municipio}: {e}")
        
        # Ordenar por nombre de municipio
        features.sort(key=lambda x: x["properties"]["municipio"])
        
        return {
            "type": "FeatureCollection",
            "total_municipios": len(features),
            "fuente": "gdb",
            "features": features
        }
        
    except Exception as e:
        logger.error(f"Error getting municipality limits: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener límites: {str(e)}")


@api_router.get("/gdb/stats")
async def get_gdb_stats(current_user: dict = Depends(get_current_user)):
    """Get statistics about available geographic data from MongoDB collection"""
    # Only staff can access stats
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    try:
        # Agregar por municipio y tipo
        pipeline = [
            {"$group": {
                "_id": {"municipio": "$municipio", "tipo": "$tipo"},
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id.municipio": 1}}
        ]
        
        resultados = await db.gdb_geometrias.aggregate(pipeline).to_list(100)
        
        stats_by_municipio = {}
        total_rurales = 0
        total_urbanos = 0
        
        for r in resultados:
            mun = r["_id"]["municipio"]
            tipo = r["_id"]["tipo"]
            count = r["count"]
            
            if mun not in stats_by_municipio:
                stats_by_municipio[mun] = {"rurales": 0, "urbanos": 0, "total": 0}
            
            if tipo == "rural":
                stats_by_municipio[mun]["rurales"] = count
                total_rurales += count
            elif tipo == "urbano":
                stats_by_municipio[mun]["urbanos"] = count
                total_urbanos += count
            
            stats_by_municipio[mun]["total"] = stats_by_municipio[mun]["rurales"] + stats_by_municipio[mun]["urbanos"]
        
        return {
            "gdb_disponible": len(stats_by_municipio) > 0,
            "predios_rurales": total_rurales,
            "predios_urbanos": total_urbanos,
            "total_geometrias": total_rurales + total_urbanos,
            "municipios": stats_by_municipio,
            "municipios_disponibles": list(stats_by_municipio.keys())
        }
    except Exception as e:
        logger.error(f"Error reading GDB stats: {e}")
        raise HTTPException(status_code=500, detail=f"Error leyendo estadísticas: {str(e)}")


@api_router.get("/gdb/capas")
async def get_gdb_layers(current_user: dict = Depends(get_current_user)):
    """List all available layers in the GDB"""
    import pyogrio
    
    # Only staff can access
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    if not GDB_PATH.exists():
        raise HTTPException(status_code=404, detail="Base de datos geográfica no disponible")
    
    try:
        layers = pyogrio.list_layers(str(GDB_PATH))
        return {
            "capas": [{"nombre": layer[0], "tipo_geometria": layer[1]} for layer in layers],
            "total": len(layers)
        }
    except Exception as e:
        logger.error(f"Error listing GDB layers: {e}")
        raise HTTPException(status_code=500, detail=f"Error listando capas: {str(e)}")


# Capas estándar según normativa IGAC - SOLO estos nombres son válidos
CAPAS_ESTANDAR = {
    "terreno_rural": ["R_TERRENO"],
    "terreno_urbano": ["U_TERRENO"],
    "construccion_rural": ["R_CONSTRUCCION"],
    "construccion_urbana": ["U_CONSTRUCCION"],
    "limite_municipal": ["LIMITEMUNICIPIO", "LimiteMunicipio", "LIMITE_MUNICIPIO"]
}


@api_router.post("/gdb/analizar")
async def analizar_gdb_antes_de_cargar(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Analiza un archivo GDB/ZIP antes de cargarlo para:
    1. Listar todas las capas encontradas
    2. Identificar cuáles son reconocidas como estándar
    3. Detectar capas no estándar que necesitan renombrarse
    4. Validar códigos prediales y detectar errores de formato
    """
    import pyogrio
    import zipfile
    import tempfile
    import shutil
    
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Guardar archivo temporal
    temp_dir = Path(tempfile.mkdtemp())
    try:
        file_path = temp_dir / file.filename
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Si es ZIP, extraer
        gdb_path = None
        if file.filename.lower().endswith('.zip'):
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            # Buscar .gdb
            for item in temp_dir.rglob("*.gdb"):
                gdb_path = item
                break
        elif file.filename.lower().endswith('.gdb'):
            gdb_path = file_path
        
        if not gdb_path or not gdb_path.exists():
            raise HTTPException(status_code=400, detail="No se encontró archivo GDB válido")
        
        # Listar capas
        layers = pyogrio.list_layers(str(gdb_path))
        capas_encontradas = [{"nombre": layer[0], "tipo_geometria": layer[1]} for layer in layers]
        
        # Clasificar capas
        capas_analisis = {
            "reconocidas": [],
            "no_reconocidas": [],
            "recomendaciones": []
        }
        
        nombres_capas = [c["nombre"].upper() for c in capas_encontradas]
        
        # Verificar capas estándar
        for tipo, nombres_validos in CAPAS_ESTANDAR.items():
            encontrada = None
            for nombre in nombres_validos:
                if nombre.upper() in nombres_capas:
                    encontrada = nombre
                    break
            
            # Buscar alternativas si no se encontró estándar
            if not encontrada:
                for capa in capas_encontradas:
                    nombre_upper = capa["nombre"].upper().replace(" ", "_")
                    if tipo == "terreno_rural" and ("TERRENO" in nombre_upper and nombre_upper.startswith("R")):
                        capas_analisis["no_reconocidas"].append({
                            "capa": capa["nombre"],
                            "tipo_detectado": tipo,
                            "sugerencia": f"Renombrar a '{nombres_validos[0]}'"
                        })
                    elif tipo == "terreno_urbano" and ("TERRENO" in nombre_upper and nombre_upper.startswith("U")):
                        capas_analisis["no_reconocidas"].append({
                            "capa": capa["nombre"],
                            "tipo_detectado": tipo,
                            "sugerencia": f"Renombrar a '{nombres_validos[0]}'"
                        })
                    elif tipo == "construccion_rural" and ("CONSTRUCCION" in nombre_upper and nombre_upper.startswith("R")):
                        capas_analisis["no_reconocidas"].append({
                            "capa": capa["nombre"],
                            "tipo_detectado": tipo,
                            "sugerencia": f"Renombrar a '{nombres_validos[0]}'"
                        })
                    elif tipo == "construccion_urbana" and ("CONSTRUCCION" in nombre_upper and nombre_upper.startswith("U")):
                        capas_analisis["no_reconocidas"].append({
                            "capa": capa["nombre"],
                            "tipo_detectado": tipo,
                            "sugerencia": f"Renombrar a '{nombres_validos[0]}'"
                        })
            else:
                capas_analisis["reconocidas"].append({
                    "tipo": tipo,
                    "capa_encontrada": encontrada
                })
        
        # Analizar códigos prediales en las capas de terreno
        codigos_con_error = []
        codigos_validos = 0
        
        for tipo_capa in ["terreno_rural", "terreno_urbano"]:
            for nombre in CAPAS_ESTANDAR[tipo_capa]:
                try:
                    gdf = gpd.read_file(str(gdb_path), layer=nombre)
                    if len(gdf) > 0:
                        for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO']:
                            if col in gdf.columns:
                                for idx, row in gdf.head(100).iterrows():  # Analizar primeros 100
                                    codigo = str(row.get(col, ''))
                                    if codigo and codigo != 'nan':
                                        # Validar formato de código predial nacional (30 dígitos)
                                        codigo_limpio = codigo.strip()
                                        if len(codigo_limpio) != 30:
                                            codigos_con_error.append({
                                                "codigo": codigo_limpio,
                                                "error": f"Longitud incorrecta ({len(codigo_limpio)} caracteres, debe ser 30)",
                                                "capa": nombre
                                            })
                                        elif not codigo_limpio.isdigit():
                                            codigos_con_error.append({
                                                "codigo": codigo_limpio,
                                                "error": "Contiene caracteres no numéricos",
                                                "capa": nombre
                                            })
                                        else:
                                            codigos_validos += 1
                                break
                        break
                except:
                    continue
        
        # Generar recomendaciones
        capas_faltantes = []
        for tipo, nombres in CAPAS_ESTANDAR.items():
            if tipo in ["terreno_rural", "terreno_urbano"]:  # Solo verificar capas críticas
                encontrada = any(n.upper() in nombres_capas for n in nombres)
                if not encontrada:
                    capas_faltantes.append(f"{nombres[0]} ({tipo})")
        
        if capas_faltantes:
            capas_analisis["recomendaciones"].append(
                f"⚠️ CAPAS ESTÁNDAR FALTANTES: {', '.join(capas_faltantes)}. La GDB no puede ser procesada sin estas capas."
            )
        
        if capas_analisis["no_reconocidas"]:
            capas_analisis["recomendaciones"].append(
                "❌ Hay capas que NO siguen el estándar IGAC. DEBE renombrarlas antes de cargar."
            )
        
        if codigos_con_error:
            capas_analisis["recomendaciones"].append(
                f"⚠️ Se encontraron {len(codigos_con_error)} códigos prediales con errores de formato."
            )
        
        # Solo puede procesar si tiene al menos una capa estándar reconocida
        puede_procesar = len(capas_analisis["reconocidas"]) > 0 and len(capas_analisis["no_reconocidas"]) == 0
        
        return {
            "archivo": file.filename,
            "total_capas": len(capas_encontradas),
            "capas_encontradas": capas_encontradas,
            "analisis": capas_analisis,
            "validacion_codigos": {
                "codigos_validos": codigos_validos,
                "codigos_con_error": codigos_con_error[:20],  # Mostrar máximo 20 errores
                "total_errores": len(codigos_con_error)
            },
            "puede_procesar": puede_procesar,
            "capas_faltantes": capas_faltantes,
            "mensaje_error": "La GDB contiene capas con nombres NO estándar. Debe ajustar los nombres según el estándar IGAC antes de cargar." if not puede_procesar else None
        }
        
    finally:
        # Limpiar archivos temporales
        shutil.rmtree(temp_dir, ignore_errors=True)


# ===== NOTIFICACIONES Y ALERTAS GDB =====

@api_router.get("/notificaciones")
async def get_notificaciones(
    leidas: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las notificaciones del usuario actual"""
    query = {"usuario_id": current_user['id']}
    if leidas is not None:
        query["leida"] = leidas
    
    notificaciones = await db.notificaciones.find(query, {"_id": 0}).sort("fecha", -1).limit(50).to_list(50)
    no_leidas = await db.notificaciones.count_documents({"usuario_id": current_user['id'], "leida": False})
    
    return {
        "notificaciones": notificaciones,
        "no_leidas": no_leidas
    }

@api_router.patch("/notificaciones/{notificacion_id}/leer")
async def marcar_notificacion_leida(
    notificacion_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marca una notificación como leída"""
    result = await db.notificaciones.update_one(
        {"id": notificacion_id, "usuario_id": current_user['id']},
        {"$set": {"leida": True, "fecha_lectura": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notificación no encontrada")
    
    return {"message": "Notificación marcada como leída"}

@api_router.post("/notificaciones/marcar-todas-leidas")
async def marcar_todas_leidas(current_user: dict = Depends(get_current_user)):
    """Marca todas las notificaciones del usuario como leídas"""
    result = await db.notificaciones.update_many(
        {"usuario_id": current_user['id'], "leida": False},
        {"$set": {"leida": True, "fecha_lectura": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"{result.modified_count} notificaciones marcadas como leídas"}

async def crear_notificacion(usuario_id: str, titulo: str, mensaje: str, tipo: str = "info", enlace: str = None, enviar_email: bool = False):
    """Crea una notificación para un usuario y opcionalmente envía email"""
    notificacion = {
        "id": str(uuid.uuid4()),
        "usuario_id": usuario_id,
        "titulo": titulo,
        "mensaje": mensaje,
        "tipo": tipo,  # info, warning, success, error
        "enlace": enlace,
        "leida": False,
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    await db.notificaciones.insert_one(notificacion)
    
    # Enviar email si está habilitado
    if enviar_email:
        user = await db.users.find_one({"id": usuario_id}, {"_id": 0, "email": 1, "full_name": 1})
        if user and user.get('email'):
            try:
                await send_notification_email(user['email'], user.get('full_name', ''), titulo, mensaje)
            except Exception as e:
                logger.error(f"Error enviando email de notificación: {e}")
    
    return notificacion

async def send_notification_email(to_email: str, to_name: str, subject: str, message: str):
    """Envía un email de notificación usando la plantilla estándar"""
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        msg['Subject'] = f"[Asomunicipios] {subject}"
        
        contenido = f'''
        <p>Hola <strong>{to_name}</strong>,</p>
        <div style="background-color: #f0fdf4; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #009846;">
            <p style="margin: 0;">{message}</p>
        </div>
        '''
        
        html_body = get_email_template(
            titulo=subject,
            contenido=contenido,
            tipo_notificacion="info"
        )
        msg.attach(MIMEText(html_body, 'html'))
        
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
            
        logger.info(f"Email de notificación enviado a {to_email}")
    except Exception as e:
        logger.error(f"Error enviando email: {e}")
        raise

@api_router.post("/gdb/enviar-alertas-mensuales")
async def enviar_alertas_mensuales_gdb(current_user: dict = Depends(get_current_user)):
    """Envía alertas mensuales a los gestores con permiso GDB (ejecutar el día 1 de cada mes)"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden enviar alertas")
    
    # Buscar gestores con permiso GDB
    gestores_gdb = await db.users.find(
        {"puede_actualizar_gdb": True, "role": UserRole.GESTOR},
        {"_id": 0}
    ).to_list(100)
    
    alertas_enviadas = 0
    mes_actual = datetime.now().strftime("%B %Y")
    
    for gestor in gestores_gdb:
        titulo = "Recordatorio: Cargar Base Gráfica Mensual"
        mensaje = f"Es momento de cargar la base gráfica (GDB) correspondiente al mes de {mes_actual}. Por favor, acceda a Gestión de Predios > Base Gráfica para realizar la carga."
        
        await crear_notificacion(
            usuario_id=gestor['id'],
            titulo=titulo,
            mensaje=mensaje,
            tipo="warning",
            enviar_email=True
        )
        alertas_enviadas += 1
    
    return {
        "message": f"Alertas enviadas a {alertas_enviadas} gestores",
        "gestores_notificados": [g['full_name'] for g in gestores_gdb]
    }

@api_router.get("/gdb/verificar-alerta-mensual")
async def verificar_alerta_mensual(current_user: dict = Depends(get_current_user)):
    """Verifica si es día 1 del mes y si se debe mostrar alerta de carga GDB"""
    hoy = datetime.now()
    es_dia_1 = hoy.day == 1
    
    # Verificar si el usuario tiene permiso GDB
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    tiene_permiso_gdb = user_db.get('puede_actualizar_gdb', False) if user_db else False
    
    # Verificar si ya cargó este mes
    mes_actual = hoy.strftime("%Y-%m")
    carga_este_mes = await db.gdb_cargas.find_one({
        "mes": mes_actual,
        "uploaded_by": current_user['id']
    })
    
    mostrar_alerta = es_dia_1 and tiene_permiso_gdb and not carga_este_mes
    
    return {
        "es_dia_1": es_dia_1,
        "tiene_permiso_gdb": tiene_permiso_gdb,
        "ya_cargo_este_mes": carga_este_mes is not None,
        "mostrar_alerta": mostrar_alerta,
        "mes_actual": mes_actual
    }


# Diccionario global para almacenar el progreso de carga de GDB
gdb_upload_progress = {}


@api_router.post("/gdb/preview-layers")
async def preview_gdb_layers(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Previsualiza las capas disponibles en un archivo GDB antes de cargarlo.
    Útil para seleccionar qué capa específica cargar.
    """
    import zipfile
    import shutil
    
    # Verificar permisos
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    if not user_db:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    has_permission = await check_permission(user_db, Permission.UPLOAD_GDB)
    if not has_permission:
        raise HTTPException(status_code=403, detail="No tiene permiso para esta operación")
    
    temp_dir = None
    try:
        # Crear directorio temporal
        temp_dir = UPLOAD_DIR / f"preview_{uuid.uuid4()}"
        temp_dir.mkdir(parents=True, exist_ok=True)
        
        # Guardar archivo
        content = await file.read()
        temp_file = temp_dir / file.filename
        
        with open(temp_file, 'wb') as f:
            f.write(content)
        
        gdb_path = None
        
        # Si es ZIP, extraer
        if file.filename.lower().endswith('.zip'):
            with zipfile.ZipFile(temp_file, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Buscar carpeta .gdb
            for item in temp_dir.rglob("*.gdb"):
                if item.is_dir():
                    gdb_path = item
                    break
        elif file.filename.lower().endswith('.gdb'):
            gdb_path = temp_file
        
        if not gdb_path or not gdb_path.exists():
            raise HTTPException(status_code=400, detail="No se encontró un archivo GDB válido")
        
        # Listar capas
        capas = fiona.listlayers(str(gdb_path))
        
        # Obtener información de cada capa
        capas_info = []
        for capa in capas:
            try:
                with fiona.open(str(gdb_path), layer=capa) as src:
                    capas_info.append({
                        "nombre": capa,
                        "registros": len(src),
                        "tipo_geometria": src.schema.get('geometry', 'Unknown'),
                        "campos": list(src.schema.get('properties', {}).keys())[:10]  # Primeros 10 campos
                    })
            except Exception as e:
                capas_info.append({
                    "nombre": capa,
                    "registros": 0,
                    "error": str(e)[:50]
                })
        
        # Categorizar capas
        capas_terreno = [c for c in capas_info if any(t in c['nombre'].upper() for t in ['TERRENO', 'PREDIO', 'PARCELA'])]
        capas_construccion = [c for c in capas_info if any(t in c['nombre'].upper() for t in ['CONSTRUCCION', 'EDIFICACION', 'UNIDAD'])]
        capas_perimetro = [c for c in capas_info if any(t in c['nombre'].upper() for t in ['PERIMETRO', 'LIMITE', 'URBANO'])]
        capas_otras = [c for c in capas_info if c not in capas_terreno + capas_construccion + capas_perimetro]
        
        return {
            "gdb_nombre": gdb_path.name,
            "total_capas": len(capas),
            "capas_terreno": capas_terreno,
            "capas_construccion": capas_construccion,
            "capas_perimetro": capas_perimetro,
            "otras_capas": capas_otras,
            "todas_las_capas": capas_info
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error previsualizando GDB: {e}")
        raise HTTPException(status_code=500, detail=f"Error al analizar el archivo: {str(e)}")
    finally:
        # Limpiar archivos temporales
        if temp_dir and temp_dir.exists():
            try:
                shutil.rmtree(temp_dir)
            except:
                pass


@api_router.get("/gdb/upload-progress/{upload_id}")
async def get_gdb_upload_progress(
    upload_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el progreso de una carga de GDB"""
    if upload_id not in gdb_upload_progress:
        # Si no existe, asumimos que el proceso ya terminó exitosamente
        return {"status": "completado", "progress": 100, "message": "Proceso finalizado"}
    return gdb_upload_progress[upload_id]


@api_router.post("/gdb/upload")
async def upload_gdb_file(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    municipio: Optional[str] = Form(None),
    modo_carga: Optional[str] = Form("reemplazar"),  # "reemplazar" = borra todo, "incremental" = solo añade/actualiza
    capa_especifica: Optional[str] = Form(None),  # Si se especifica, solo carga esa capa (ej: "PERIMETRO_URBANO", "U_TERRENO")
    current_user: dict = Depends(get_current_user)
):
    """Upload GDB files (ZIP or multiple files from a GDB folder). Only authorized gestors can do this.
    
    Modos de carga:
    - reemplazar: Elimina todas las geometrías del municipio antes de cargar (comportamiento por defecto)
    - incremental: Solo añade/actualiza geometrías, manteniendo las existentes de otras capas
    
    Capa específica:
    - Si se especifica, solo se procesará esa capa del GDB (ej: PERIMETRO_URBANO, U_CONSTRUCCION)
    """
    import zipfile
    import shutil
    
    # Crear ID único para esta carga
    upload_id = str(uuid.uuid4())
    user_id = current_user['id']
    
    gdb_upload_progress[upload_id] = {
        "status": "iniciando",
        "progress": 0,
        "message": "Iniciando carga de archivos...",
        "upload_id": upload_id,
        "modo_carga": modo_carga,
        "capa_especifica": capa_especifica
    }
    
    # Check if user is an authorized gestor
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    
    if not user_db:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Check permission using the new permissions system
    has_permission = await check_permission(user_db, Permission.UPLOAD_GDB)
    if not has_permission:
        raise HTTPException(
            status_code=403, 
            detail="No tiene permiso para actualizar la base gráfica. Contacte al coordinador."
        )
    
    # Guardar archivos temporalmente para procesarlos en background
    temp_files = []
    for file in files:
        content = await file.read()
        temp_files.append({
            'filename': file.filename,
            'content': content
        })
    
    # Iniciar procesamiento en background
    background_tasks.add_task(
        process_gdb_upload_background,
        upload_id,
        user_id,
        current_user,
        temp_files,
        municipio,
        modo_carga,
        capa_especifica
    )
    
    # Responder inmediatamente con el upload_id para que el frontend pueda hacer polling
    return {
        "upload_id": upload_id,
        "status": "processing",
        "message": "Carga iniciada. Use el endpoint /api/gdb/upload-progress/{upload_id} para consultar el progreso.",
        "modo_carga": modo_carga,
        "capa_especifica": capa_especifica
    }


async def process_gdb_upload_background(
    upload_id: str,
    user_id: str,
    current_user: dict,
    temp_files: list,
    municipio: Optional[str],
    modo_carga: str = "reemplazar",
    capa_especifica: Optional[str] = None
):
    """Procesa la carga de GDB en background
    
    Args:
        modo_carga: "reemplazar" borra todo, "incremental" solo añade/actualiza
        capa_especifica: Si se especifica, solo procesa esa capa
    """
    import zipfile
    import shutil
    import geopandas as gpd
    import pandas as pd
    from pyproj import CRS, Transformer
    from shapely.ops import transform
    
    # Variable para almacenar el municipio detectado (se actualiza durante el proceso)
    municipio_para_progreso = {"nombre": municipio or "Procesando..."}
    
    async def update_progress(status: str, progress: int, message: str, **extra):
        """Actualiza el progreso y envía notificación WebSocket al usuario"""
        progress_data = {
            "status": status,
            "progress": progress,
            "message": message,
            "upload_id": upload_id,
            "municipio": municipio_para_progreso["nombre"],
            **extra
        }
        gdb_upload_progress[upload_id] = progress_data
        
        # Enviar progreso en tiempo real vía WebSocket
        await ws_manager.send_personal_message({
            "type": "gdb_upload_progress",
            "data": progress_data
        }, user_id)
    
    await update_progress("preparando", 5, "Preparando transformación de coordenadas...")
    
    # Setup coordinate transformation function (will be set based on GDB CRS)
    project = None
    
    def validate_colombia_coordinates(geom):
        """Valida que las coordenadas estén dentro de Colombia (WGS84)"""
        # Límites aproximados de Colombia en WGS84
        # Latitud: -4.23° a 12.46° (Sur a Norte)
        # Longitud: -81.73° a -66.87° (Oeste a Este)
        try:
            if geom is None:
                return False
            bounds = geom.bounds  # (minx, miny, maxx, maxy)
            min_lon, min_lat, max_lon, max_lat = bounds
            
            # Verificar que esté en el rango de Colombia
            if min_lon < -82 or max_lon > -66:
                return False
            if min_lat < -5 or max_lat > 13:
                return False
            return True
        except:
            return False
    
    def get_transformer_for_gdf(gdf):
        """Get the appropriate transformer based on GDF's CRS"""
        try:
            if gdf.crs is None:
                # Assume MAGNA-SIRGAS Colombia Bogotá zone if no CRS
                logger.info("GDB sin CRS definido, asumiendo MAGNA-SIRGAS (EPSG:3116)")
                source_crs = CRS.from_epsg(3116)
            else:
                source_crs = CRS.from_user_input(gdf.crs)
                logger.info(f"GDB CRS detectado: {source_crs.to_string()}")
            
            target_crs = CRS.from_epsg(4326)  # WGS84
            
            # Check if already in WGS84
            if source_crs.to_epsg() == 4326:
                logger.info("GDB ya está en WGS84, no se requiere transformación")
                return None
            
            transformer = Transformer.from_crs(source_crs, target_crs, always_xy=True)
            return transformer.transform
        except Exception as e:
            logger.warning(f"Error setting up CRS transformation: {e}")
            # Fallback to MAGNA-SIRGAS
            try:
                logger.info("Usando fallback MAGNA-SIRGAS (EPSG:3116)")
                source_crs = CRS.from_epsg(3116)
                target_crs = CRS.from_epsg(4326)
                transformer = Transformer.from_crs(source_crs, target_crs, always_xy=True)
                return transformer.transform
            except:
                return None
    
    try:
        gdb_data_dir = Path("/app/gdb_data")
        gdb_data_dir.mkdir(exist_ok=True)
        
        gdb_found = None
        is_zip = len(temp_files) == 1 and temp_files[0]['filename'].endswith('.zip')
        
        await update_progress("cargando", 10, "Cargando archivos GDB...")
        
        if is_zip:
            # Proceso ZIP tradicional
            file_data = temp_files[0]
            temp_zip = UPLOAD_DIR / f"temp_gdb_{uuid.uuid4()}.zip"
            with open(temp_zip, 'wb') as f:
                f.write(file_data['content'])
            
            await update_progress("extrayendo", 15, "Extrayendo archivo ZIP...")
            
            # PRIMERO: Identificar el nombre de la carpeta .gdb dentro del ZIP
            gdb_name_in_zip = None
            with zipfile.ZipFile(temp_zip, 'r') as zip_ref:
                # Buscar la carpeta .gdb en el contenido del ZIP
                for name in zip_ref.namelist():
                    if '.gdb/' in name or '.gdb\\' in name:
                        # Extraer nombre de carpeta GDB
                        parts = name.replace('\\', '/').split('/')
                        for part in parts:
                            if part.endswith('.gdb'):
                                gdb_name_in_zip = part
                                break
                        if gdb_name_in_zip:
                            break
                
                logger.info(f"GDB en ZIP detectado: {gdb_name_in_zip}")
                zip_ref.extractall(gdb_data_dir)
            
            temp_zip.unlink()
            
            # Buscar ESPECÍFICAMENTE la carpeta .gdb que estaba en el ZIP
            if gdb_name_in_zip:
                # Buscar directamente la carpeta con ese nombre
                potential_path = gdb_data_dir / gdb_name_in_zip
                if potential_path.exists() and potential_path.is_dir():
                    gdb_found = potential_path
                    logger.info(f"GDB encontrado directamente: {gdb_found}")
                else:
                    # Buscar en subdirectorios (el ZIP podría tener carpeta contenedora)
                    for item in gdb_data_dir.iterdir():
                        if item.is_dir():
                            subpath = item / gdb_name_in_zip
                            if subpath.exists() and subpath.is_dir():
                                gdb_found = subpath
                                logger.info(f"GDB encontrado en subdirectorio: {gdb_found}")
                                break
            
            # Fallback: buscar cualquier .gdb si no se pudo identificar del ZIP
            if not gdb_found:
                logger.warning("No se pudo identificar GDB del ZIP, buscando cualquier .gdb...")
                for item in gdb_data_dir.iterdir():
                    if item.suffix == '.gdb' and item.is_dir():
                        gdb_found = item
                        break
                
                if not gdb_found:
                    for item in gdb_data_dir.iterdir():
                        if item.is_dir():
                            for subitem in item.iterdir():
                                if subitem.suffix == '.gdb' and subitem.is_dir():
                                    gdb_found = subitem
                                    break
        else:
            # Proceso para archivos de carpeta GDB (múltiples archivos)
            # Determinar el nombre de la carpeta .gdb desde los archivos
            gdb_folder_name = None
            for file in files:
                # Los archivos vienen con path relativo como "54003.gdb/archivo.ext"
                parts = file.filename.split('/')
                if len(parts) > 0:
                    for part in parts:
                        if part.endswith('.gdb'):
                            gdb_folder_name = part
                            break
                if gdb_folder_name:
                    break
            
            if not gdb_folder_name:
                # Intentar extraer del nombre del archivo
                for file in files:
                    if '.gdb' in file.filename:
                        idx = file.filename.find('.gdb')
                        start = file.filename.rfind('/', 0, idx)
                        gdb_folder_name = file.filename[start+1:idx+4]
                        break
            
            if not gdb_folder_name:
                gdb_folder_name = f"{municipio or 'uploaded'}.gdb"
            
            gdb_found = gdb_data_dir / gdb_folder_name
            gdb_found.mkdir(exist_ok=True)
            
            # Guardar todos los archivos
            for file in files:
                # Extraer solo el nombre del archivo (sin la ruta de carpeta GDB)
                filename = file.filename
                if '/' in filename:
                    filename = filename.split('/')[-1]
                elif '\\' in filename:
                    filename = filename.split('\\')[-1]
                
                file_path = gdb_found / filename
                content = await file.read()
                with open(file_path, 'wb') as f:
                    f.write(content)
        
        if not gdb_found or not gdb_found.exists():
            raise HTTPException(status_code=400, detail="No se pudo crear/encontrar el archivo .gdb")
        
        # Determinar código de municipio desde el nombre del GDB
        gdb_name = gdb_found.stem  # ej: "54003"
        
        await update_progress("identificando", 20, f"GDB identificado: {gdb_name}")
        
        # Mapeo de códigos a nombres de municipio
        CODIGO_TO_MUNICIPIO = {
            '54003': 'Ábrego',
            '54109': 'Bucarasica', 
            '54128': 'Cáchira',
            '54206': 'Convención',
            '54245': 'El Carmen',
            '54250': 'El Tarra',
            '54344': 'Hacarí',
            '54398': 'La Playa',
            '54670': 'San Calixto',
            '54720': 'Sardinata',
            '54800': 'Teorama',
            '20614': 'Río de Oro',
        }
        
        # Intentar detectar municipio desde el nombre del archivo primero
        municipio_nombre_inicial = municipio or CODIGO_TO_MUNICIPIO.get(gdb_name, None)
        
        # Actualizar el nombre del municipio temprano si ya lo sabemos
        if municipio_nombre_inicial:
            municipio_para_progreso["nombre"] = municipio_nombre_inicial
        
        await update_progress("leyendo", 25, f"Leyendo capas de {gdb_name}...")
        
        # Leer capas del GDB para obtener estadísticas y relacionar con predios
        stats = {"rurales": 0, "urbanos": 0, "relacionados": 0}
        codigos_gdb = set()
        municipio_detectado_desde_codigos = None
        
        try:
            # Primero listar todas las capas disponibles para diagnóstico
            available_layers = []
            try:
                import pyogrio
                layers_info = pyogrio.list_layers(str(gdb_found))
                available_layers = [layer[0] for layer in layers_info]
                logger.info(f"GDB {gdb_name}: Capas disponibles: {available_layers}")
                await update_progress("analizando", 28, f"Analizando estructura del archivo GDB...")
            except Exception as e:
                logger.warning(f"No se pudo listar capas: {e}")
            
            # Intentar leer LIMITEMUNICIPIO para crear el límite municipal
            limite_municipal = None
            for limite_layer in ['LIMITEMUNICIPIO', 'LimiteMunicipio', 'limite_municipio', 'LIMITE_MUNICIPIO']:
                if limite_layer in available_layers:
                    try:
                        gdf_limite = gpd.read_file(str(gdb_found), layer=limite_layer)
                        if len(gdf_limite) > 0:
                            # Get transformer based on GDF's CRS
                            project = get_transformer_for_gdf(gdf_limite)
                            logger.info(f"GDB {gdb_name}: CRS del límite: {gdf_limite.crs}")
                            
                            # Guardar límite municipal (usar nombre temporal hasta detectar desde códigos)
                            temp_municipio_name = municipio_nombre_inicial or gdb_name
                            for idx, row in gdf_limite.iterrows():
                                if row.geometry:
                                    geom_wgs84 = transform(project, row.geometry) if project else row.geometry
                                    limite_municipal = geom_wgs84.__geo_interface__
                                    # Guardar en colección de límites
                                    await db.limites_municipales.update_one(
                                        {"codigo": gdb_name},
                                        {"$set": {
                                            "municipio": temp_municipio_name,
                                            "codigo": gdb_name,
                                            "geometry": limite_municipal,
                                            "fuente": "gdb",
                                            "updated_at": datetime.now(timezone.utc).isoformat()
                                        }},
                                        upsert=True
                                    )
                                    logger.info(f"GDB {gdb_name}: Límite municipal guardado desde capa {limite_layer}")
                                    break
                            break
                    except Exception as e:
                        logger.warning(f"Error leyendo límite municipal: {e}")
            
            # Intentar diferentes nombres de capas rurales - PRIORIZAR capas con TERRENO
            await update_progress("leyendo_rural", 30, "Leyendo capa rural...")
            # Lista ordenada por prioridad - TERRENO primero, evitar ZONA_HOMOGENEA
            rural_layers = ['R_TERRENO']  # SOLO nombre estándar
            
            # NO buscar dinámicamente - solo aceptar el nombre estándar
            gdf_rural = None
            rural_layer_found = None
            for rural_layer in rural_layers:
                try:
                    gdf_rural = gpd.read_file(str(gdb_found), layer=rural_layer)
                    if len(gdf_rural) > 0:
                        stats["rurales"] = len(gdf_rural)
                        rural_layer_found = rural_layer
                        logger.info(f"GDB {gdb_name}: Capa rural encontrada '{rural_layer}' con {len(gdf_rural)} registros")
                        await update_progress("leyendo_rural", 35, f"Capa rural ({rural_layer}): {len(gdf_rural)} geometrías encontradas")
                        # Extraer códigos prediales
                        for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO']:
                            if col in gdf_rural.columns:
                                codigos_gdb.update(gdf_rural[col].dropna().astype(str).tolist())
                                break
                        break
                except Exception as layer_err:
                    continue
            
            if not rural_layer_found:
                logger.warning(f"GDB {gdb_name}: No se encontró capa rural. Capas disponibles: {available_layers}")
                await update_progress("leyendo_rural", 35, "No se encontró capa rural en el GDB")
            
            await update_progress("leyendo_urbano", 40, "Leyendo capa urbana...")
            gdf_urban = None
            urban_layers = ['U_TERRENO']  # SOLO nombre estándar
            
            urban_layer_found = None
            for urban_layer in urban_layers:
                try:
                    gdf_urban = gpd.read_file(str(gdb_found), layer=urban_layer)
                    if len(gdf_urban) > 0:
                        stats["urbanos"] = len(gdf_urban)
                        urban_layer_found = urban_layer
                        logger.info(f"GDB {gdb_name}: Capa urbana encontrada '{urban_layer}' con {len(gdf_urban)} registros")
                        await update_progress("leyendo_urbano", 45, f"Capa urbana ({urban_layer}): {len(gdf_urban)} geometrías encontradas")
                        for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO']:
                            if col in gdf_urban.columns:
                                codigos_gdb.update(gdf_urban[col].dropna().astype(str).tolist())
                                break
                        break
                except:
                    continue
            
            if not urban_layer_found:
                logger.warning(f"GDB {gdb_name}: No se encontró capa urbana con TERRENO")
        except Exception as e:
            logger.warning(f"Error leyendo capas GDB: {e}")
        
        # === DETECCIÓN AUTOMÁTICA DEL MUNICIPIO DESDE LOS CÓDIGOS PREDIALES ===
        # Los códigos prediales tienen formato: DDMMMZZZ... donde DD=depto, MMM=municipio
        if codigos_gdb:
            codigo_municipio_detectado = None
            for codigo in codigos_gdb:
                if len(codigo) >= 5:
                    # Tomar los primeros 5 dígitos (departamento + municipio)
                    codigo_muni = codigo[:5]
                    if codigo_muni in CODIGO_TO_MUNICIPIO:
                        codigo_municipio_detectado = codigo_muni
                        municipio_detectado_desde_codigos = CODIGO_TO_MUNICIPIO[codigo_muni]
                        break
            
            if municipio_detectado_desde_codigos:
                logger.info(f"GDB: Municipio detectado desde códigos prediales: {municipio_detectado_desde_codigos} (código {codigo_municipio_detectado})")
                if municipio_nombre_inicial and municipio_nombre_inicial != municipio_detectado_desde_codigos:
                    logger.warning(f"GDB: Discrepancia - Nombre archivo: {municipio_nombre_inicial}, Códigos prediales: {municipio_detectado_desde_codigos}")
        
        # Usar el municipio detectado desde códigos si está disponible, sino el del nombre del archivo
        municipio_nombre = municipio_detectado_desde_codigos or municipio_nombre_inicial or gdb_name
        
        # Actualizar el nombre del municipio para el progreso
        municipio_para_progreso["nombre"] = municipio_nombre
        
        await update_progress("relacionando_capas", 50, f"Relacionando capas del GDB ({len(codigos_gdb)} geometrías)...")
        
        # Guardar geometrías en colección para búsquedas posteriores
        geometrias_guardadas = 0
        
        # Lógica de eliminación según el modo de carga
        if modo_carga == "reemplazar":
            # REEMPLAZAR COMPLETAMENTE: Limpiar TODAS las geometrías anteriores de este municipio
            deleted = await db.gdb_geometrias.delete_many({"municipio": municipio_nombre})
            logger.info(f"GDB {municipio_nombre}: Eliminadas {deleted.deleted_count} geometrías anteriores (modo: reemplazar)")
            await update_progress("limpiando", 52, f"Reemplazando geometrías anteriores ({deleted.deleted_count} eliminadas)")
        elif modo_carga == "incremental" and capa_especifica:
            # INCREMENTAL con capa específica: Solo eliminar geometrías de la misma capa
            deleted = await db.gdb_geometrias.delete_many({
                "municipio": municipio_nombre,
                "capa_origen": capa_especifica
            })
            logger.info(f"GDB {municipio_nombre}: Eliminadas {deleted.deleted_count} geometrías de capa '{capa_especifica}' (modo: incremental)")
            await update_progress("limpiando", 52, f"Actualizando capa '{capa_especifica}' ({deleted.deleted_count} reemplazadas)")
        else:
            # INCREMENTAL sin capa específica: No eliminar nada, solo añadir/actualizar
            logger.info(f"GDB {municipio_nombre}: Modo incremental - añadiendo sin eliminar existentes")
            await update_progress("limpiando", 52, f"Modo incremental: añadiendo nuevas geometrías...")
        
        # Guardar las geometrías con sus códigos
        # Inicializar diccionario de errores para el reporte de calidad
        errores_calidad = {
            'codigos_invalidos': [],
            'geometrias_rechazadas': [],
            'construcciones_huerfanas': [],
            'rurales_rechazados': 0,
            'urbanos_rechazados': 0
        }
        
        # Inicializar contadores ANTES de los try/except
        rurales_en_archivo = 0
        urbanos_en_archivo = 0
        rural_guardadas = 0
        urban_guardadas = 0
        capas_especificas_guardadas = 0
        
        # Si se especificó una capa específica, procesarla primero
        if capa_especifica:
            await update_progress("capa_especifica", 53, f"Procesando capa específica: {capa_especifica}...")
            logger.info(f"GDB {municipio_nombre}: Procesando capa específica '{capa_especifica}'")
            
            try:
                # Intentar leer la capa específica
                gdf_especifica = gpd.read_file(str(gdb_found), layer=capa_especifica)
                
                if len(gdf_especifica) > 0:
                    project = get_transformer_for_gdf(gdf_especifica)
                    logger.info(f"GDB {municipio_nombre}: Capa '{capa_especifica}' tiene {len(gdf_especifica)} registros, CRS: {gdf_especifica.crs}")
                    
                    # Determinar tipo basado en el nombre de la capa
                    tipo_capa = "otro"
                    if capa_especifica.upper().startswith("R_") or "RURAL" in capa_especifica.upper():
                        tipo_capa = "rural"
                    elif capa_especifica.upper().startswith("U_") or "URBAN" in capa_especifica.upper():
                        tipo_capa = "urbano"
                    elif "PERIMETRO" in capa_especifica.upper():
                        tipo_capa = "perimetro_urbano"
                    
                    total_especifica = len(gdf_especifica)
                    for idx, row in gdf_especifica.iterrows():
                        if idx % 200 == 0:
                            pct = 53 + int((idx / total_especifica) * 10)
                            await update_progress("guardando_especifica", pct, f"Procesando {capa_especifica}: {idx}/{total_especifica}")
                        
                        # Obtener código
                        codigo = None
                        for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO', 'CODIGO_PRED', 'ID', 'OBJECTID']:
                            if col in gdf_especifica.columns and pd.notna(row.get(col)):
                                codigo = str(row[col]).strip()
                                break
                        
                        if not codigo:
                            codigo = f"{capa_especifica}_{idx}"  # Generar código si no existe
                        
                        try:
                            geom = row.geometry
                            if geom is None or geom.is_empty:
                                continue
                            
                            # Transformar a WGS84
                            if project:
                                geom = transform(project, geom)
                            
                            # Calcular área
                            try:
                                area_m2 = geom.area if geom.area > 0 else 0
                            except:
                                area_m2 = 0
                            
                            geom_dict = mapping(geom)
                            
                            # Verificar que tenga coordenadas
                            if 'coordinates' not in geom_dict or not geom_dict['coordinates']:
                                continue
                            
                            # En modo incremental, usar upsert para actualizar si existe
                            if modo_carga == "incremental":
                                await db.gdb_geometrias.update_one(
                                    {"codigo": codigo, "capa_origen": capa_especifica, "municipio": municipio_nombre},
                                    {"$set": {
                                        "tipo": tipo_capa,
                                        "tipo_zona": tipo_capa,
                                        "capa_origen": capa_especifica,
                                        "gdb_source": gdb_name,
                                        "municipio": municipio_nombre,
                                        "area_m2": area_m2,
                                        "geometry": geom_dict
                                    }},
                                    upsert=True
                                )
                            else:
                                await db.gdb_geometrias.insert_one({
                                    "codigo": codigo,
                                    "tipo": tipo_capa,
                                    "tipo_zona": tipo_capa,
                                    "capa_origen": capa_especifica,
                                    "gdb_source": gdb_name,
                                    "municipio": municipio_nombre,
                                    "area_m2": area_m2,
                                    "geometry": geom_dict
                                })
                            
                            capas_especificas_guardadas += 1
                            geometrias_guardadas += 1
                        except Exception as geom_error:
                            logger.warning(f"Error procesando geometría de {capa_especifica}: {geom_error}")
                    
                    logger.info(f"GDB {municipio_nombre}: Guardadas {capas_especificas_guardadas} geometrías de capa '{capa_especifica}'")
                    stats['capa_especifica'] = capa_especifica
                    stats['geometrias_capa_especifica'] = capas_especificas_guardadas
                else:
                    logger.warning(f"GDB {municipio_nombre}: Capa '{capa_especifica}' está vacía")
                    
            except Exception as capa_error:
                logger.error(f"GDB {municipio_nombre}: Error leyendo capa '{capa_especifica}': {capa_error}")
                # Listar capas disponibles para ayudar al usuario
                try:
                    capas_disponibles = fiona.listlayers(str(gdb_found))
                    logger.info(f"Capas disponibles en el GDB: {capas_disponibles}")
                    await update_progress("error_capa", 55, f"Capa '{capa_especifica}' no encontrada. Disponibles: {', '.join(capas_disponibles[:10])}")
                except:
                    pass
            
            # Si solo se pidió capa específica y modo incremental, no procesar las capas estándar
            if modo_carga == "incremental":
                await update_progress("finalizando", 90, f"Carga incremental completada. {capas_especificas_guardadas} geometrías de '{capa_especifica}'")
                # Saltar directamente a las estadísticas finales
        
        try:
            # Rural - usar lista de capas específicas (SIN buscar dinámicamente para evitar ZONA_HOMOGENEA)
            
            # Usar la misma capa que se usó para leer, si se encontró
            rural_layers_to_save = ['R_TERRENO_1', 'R_TERRENO', 'TERRENO', 'R_Terreno', 'r_terreno', 'r_terreno_1', 'Terreno', 'terreno']
            
            # NO buscar dinámicamente capas que empiecen con R_ ya que pueden incluir ZONA_HOMOGENEA
            
            for rural_layer in rural_layers_to_save:
                try:
                    gdf_rural = gpd.read_file(str(gdb_found), layer=rural_layer)
                    if len(gdf_rural) == 0:
                        continue
                    
                    rurales_en_archivo = len(gdf_rural)
                    
                    # Get transformer based on this layer's CRS
                    project = get_transformer_for_gdf(gdf_rural)
                    logger.info(f"GDB {municipio_nombre}: CRS rural ({rural_layer}): {gdf_rural.crs}, Total registros: {len(gdf_rural)}")
                    
                    total_rural = len(gdf_rural)
                    for idx, row in gdf_rural.iterrows():
                        if idx % 500 == 0:
                            pct = 50 + int((idx / total_rural) * 15)
                            await update_progress("guardando_rural", pct, f"Procesando geometrías rurales: {idx}/{total_rural}")
                        
                        codigo = None
                        for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO', 'CODIGO_PRED']:
                            if col in gdf_rural.columns and pd.notna(row.get(col)):
                                codigo = str(row[col]).strip()
                                break
                        
                        if not codigo:
                            errores_calidad['rurales_rechazados'] += 1
                            continue
                        
                        # Validar longitud del código (debe ser 30 dígitos)
                        if len(codigo) != 30:
                            errores_calidad['codigos_invalidos'].append({
                                'codigo': codigo,
                                'longitud': len(codigo),
                                'capa': rural_layer
                            })
                        
                        if not row.geometry:
                            errores_calidad['rurales_rechazados'] += 1
                            errores_calidad['geometrias_rechazadas'].append({
                                'codigo': codigo,
                                'razon': 'Geometría nula',
                                'capa': rural_layer
                            })
                            continue
                        
                        try:
                            geom_wgs84 = transform(project, row.geometry) if project else row.geometry
                            
                            # Validar que las coordenadas estén en Colombia
                            if not validate_colombia_coordinates(geom_wgs84):
                                errores_calidad['rurales_rechazados'] += 1
                                errores_calidad['geometrias_rechazadas'].append({
                                    'codigo': codigo,
                                    'razon': 'Coordenadas fuera de Colombia',
                                    'capa': rural_layer
                                })
                                logger.warning(f"Geometría fuera de Colombia descartada: {codigo}")
                                continue
                            
                            # Calcular área en m2
                            area_m2 = 0
                            try:
                                # El área en grados se convierte aproximadamente a m2
                                # Factor para Colombia (~7° latitud): 1 grado ≈ 111320 m
                                area_deg = geom_wgs84.area
                                area_m2 = round(area_deg * (111320 ** 2), 2)
                            except:
                                pass
                            
                            # Validar que la geometría sea serializable
                            try:
                                geom_dict = geom_wgs84.__geo_interface__
                                # Verificar que tenga coordenadas válidas
                                if not geom_dict.get('coordinates'):
                                    raise ValueError("Geometría sin coordenadas")
                            except Exception as ser_err:
                                errores_calidad['rurales_rechazados'] += 1
                                errores_calidad['geometrias_rechazadas'].append({
                                    'codigo': codigo,
                                    'razon': f'Error serializando: {str(ser_err)[:30]}',
                                    'capa': rural_layer
                                })
                                continue
                            
                            await db.gdb_geometrias.insert_one({
                                "codigo": codigo,
                                "tipo": "rural",
                                "tipo_zona": "rural",
                                "capa_origen": rural_layer,  # Nombre de la capa de origen
                                "gdb_source": gdb_name,
                                "municipio": municipio_nombre,
                                "area_m2": area_m2,
                                "geometry": geom_dict
                            })
                            geometrias_guardadas += 1
                            rural_guardadas += 1
                        except Exception as geom_error:
                            errores_calidad['rurales_rechazados'] += 1
                            errores_calidad['geometrias_rechazadas'].append({
                                'codigo': codigo,
                                'razon': str(geom_error)[:50],
                                'capa': rural_layer
                            })
                            logger.warning(f"Error procesando geometría rural {codigo}: {geom_error}")
                    
                    logger.info(f"GDB {municipio_nombre}: Guardadas {rural_guardadas} geometrías rurales desde capa {rural_layer}")
                    break
                except Exception as layer_error:
                    logger.debug(f"Capa {rural_layer} no encontrada o error: {layer_error}")
                    continue
            
            await update_progress("guardando_urbano", 65, "Procesando geometrías urbanas...")
            # Lista de capas urbanas de terreno - priorizar nombres específicos
            urban_layers_save = [
                'U_TERRENO', 'U_TERRENO_1',  # Prioridad: nombres estándar de TERRENO
                'U_Terreno', 'u_terreno', 'u_terreno_1',
                'TERRENO_U', 'terreno_u', 'Terreno_U',
                'U_PREDIO', 'U_Predio', 'u_predio'
            ]
            # NO agregar dinámicamente otras capas U_ ya que pueden ser BARRIO, MANZANA, etc.
            
            for urban_layer in urban_layers_save:
                try:
                    gdf_urban = gpd.read_file(str(gdb_found), layer=urban_layer)
                    if len(gdf_urban) == 0:
                        continue
                    
                    urbanos_en_archivo = len(gdf_urban)
                    
                    # Get transformer based on this layer's CRS
                    project = get_transformer_for_gdf(gdf_urban)
                    logger.info(f"GDB {municipio_nombre}: CRS urbano ({urban_layer}): {gdf_urban.crs}, Total registros: {len(gdf_urban)}")
                    
                    total_urban = len(gdf_urban)
                    for idx, row in gdf_urban.iterrows():
                        if idx % 500 == 0:
                            pct = 65 + int((idx / total_urban) * 10)
                            await update_progress("guardando_urbano", pct, f"Procesando geometrías urbanas: {idx}/{total_urban}")
                        
                        codigo = None
                        for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO', 'CODIGO_PRED']:
                            if col in gdf_urban.columns and pd.notna(row.get(col)):
                                codigo = str(row[col]).strip()
                                break
                        
                        if not codigo:
                            errores_calidad['urbanos_rechazados'] += 1
                            continue
                        
                        # Validar longitud del código (debe ser 30 dígitos)
                        if len(codigo) != 30:
                            errores_calidad['codigos_invalidos'].append({
                                'codigo': codigo,
                                'longitud': len(codigo),
                                'capa': urban_layer
                            })
                        
                        if not row.geometry:
                            errores_calidad['urbanos_rechazados'] += 1
                            errores_calidad['geometrias_rechazadas'].append({
                                'codigo': codigo,
                                'razon': 'Geometría nula',
                                'capa': urban_layer
                            })
                            continue
                        
                        try:
                            geom_wgs84 = transform(project, row.geometry) if project else row.geometry
                            
                            # Validar que las coordenadas estén en Colombia
                            if not validate_colombia_coordinates(geom_wgs84):
                                errores_calidad['urbanos_rechazados'] += 1
                                errores_calidad['geometrias_rechazadas'].append({
                                    'codigo': codigo,
                                    'razon': 'Coordenadas fuera de Colombia',
                                    'capa': urban_layer
                                })
                                logger.warning(f"Geometría urbana fuera de Colombia descartada: {codigo}")
                                continue
                            
                            # Calcular área en m2
                            area_m2 = 0
                            try:
                                area_deg = geom_wgs84.area
                                area_m2 = round(area_deg * (111320 ** 2), 2)
                            except:
                                pass
                            
                            # Validar que la geometría sea serializable
                            try:
                                geom_dict = geom_wgs84.__geo_interface__
                                if not geom_dict.get('coordinates'):
                                    raise ValueError("Geometría sin coordenadas")
                            except Exception as ser_err:
                                errores_calidad['urbanos_rechazados'] += 1
                                errores_calidad['geometrias_rechazadas'].append({
                                    'codigo': codigo,
                                    'razon': f'Error serializando: {str(ser_err)[:30]}',
                                    'capa': urban_layer
                                })
                                continue
                            
                            await db.gdb_geometrias.insert_one({
                                "codigo": codigo,
                                "tipo": "urbano",
                                "tipo_zona": "urbano",
                                "capa_origen": urban_layer,  # Nombre de la capa de origen
                                "gdb_source": gdb_name,
                                "municipio": municipio_nombre,
                                "area_m2": area_m2,
                                "geometry": geom_dict
                            })
                            geometrias_guardadas += 1
                            urban_guardadas += 1
                        except Exception as geom_error:
                            errores_calidad['urbanos_rechazados'] += 1
                            errores_calidad['geometrias_rechazadas'].append({
                                'codigo': codigo,
                                'razon': str(geom_error)[:50],
                                'capa': urban_layer
                            })
                            logger.warning(f"Error procesando geometría urbana {codigo}: {geom_error}")
                    
                    logger.info(f"GDB {municipio_nombre}: Guardadas {urban_guardadas} geometrías urbanas desde capa {urban_layer}")
                    break
                except Exception as layer_error:
                    logger.debug(f"Capa {urban_layer} no encontrada o error: {layer_error}")
                    continue
        except Exception as e:
            logger.error(f"Error guardando geometrías: {e}")
        
        # Guardar estadísticas del archivo original
        stats['rurales_archivo'] = rurales_en_archivo
        stats['urbanos_archivo'] = urbanos_en_archivo
        
        # ===== PROCESAR CONSTRUCCIONES =====
        await update_progress("leyendo_construcciones", 70, "Buscando capas de construcciones...")
        
        construcciones_guardadas = 0
        construcciones_rurales = 0
        construcciones_urbanas = 0
        # SOLO nombres estándar
        construcciones_layers = ['R_CONSTRUCCION', 'U_CONSTRUCCION']
        
        try:
            # Limpiar construcciones anteriores del municipio
            await db.gdb_construcciones.delete_many({"municipio": municipio_nombre})
            
            for const_layer in construcciones_layers:
                try:
                    gdf_const = gpd.read_file(str(gdb_found), layer=const_layer)
                    if len(gdf_const) == 0:
                        continue
                    
                    # Determinar tipo (rural/urbano) basado en nombre de capa
                    tipo_construccion = "rural" if const_layer.upper().startswith('R') else "urbano"
                    
                    project = get_transformer_for_gdf(gdf_const)
                    logger.info(f"GDB {municipio_nombre}: Capa construcciones ({const_layer}): {len(gdf_const)} registros")
                    await update_progress("guardando_construcciones", 72, f"Procesando {len(gdf_const)} construcciones ({const_layer})...")
                    
                    for idx, row in gdf_const.iterrows():
                        codigo = None
                        for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO', 'CODIGO_PRED', 'CODIGO_TERRENO']:
                            if col in gdf_const.columns and pd.notna(row.get(col)):
                                codigo = str(row[col]).strip()
                                break
                        
                        if not codigo:
                            continue
                        
                        if not row.geometry:
                            errores_calidad['construcciones_huerfanas'].append({
                                'codigo': codigo,
                                'capa': const_layer,
                                'razon': 'Sin geometría'
                            })
                            continue
                        
                        try:
                            geom_wgs84 = transform(project, row.geometry) if project else row.geometry
                            
                            # Validar que las coordenadas estén en Colombia
                            if not validate_colombia_coordinates(geom_wgs84):
                                logger.warning(f"Construcción fuera de Colombia descartada: {codigo}")
                                continue
                            
                            area_m2 = round(geom_wgs84.area * (111320 ** 2), 2) if geom_wgs84.area else 0
                            
                            # Extraer atributos de construcción si existen
                            pisos = row.get('PISOS', row.get('pisos', row.get('NUM_PISOS', 1)))
                            tipo_const = row.get('TIPO_CONSTRUCCION', row.get('tipo_construccion', row.get('TIPO', '')))
                            
                            # El código del predio padre son los primeros 25 dígitos (sin la parte de construcción)
                            codigo_predio_padre = codigo[:25] + "00000" if len(codigo) >= 25 else codigo
                            
                            await db.gdb_construcciones.insert_one({
                                "codigo_construccion": codigo,  # Código completo de la construcción
                                "codigo_predio": codigo_predio_padre,  # Código del predio padre (25 dígitos + 00000)
                                "tipo_zona": tipo_construccion,
                                "gdb_source": gdb_name,
                                "municipio": municipio_nombre,
                                "area_m2": area_m2,
                                "pisos": int(pisos) if pisos and str(pisos).isdigit() else 1,
                                "tipo_construccion": str(tipo_const) if tipo_const else "",
                                "geometry": geom_wgs84.__geo_interface__
                            })
                            construcciones_guardadas += 1
                            
                            if tipo_construccion == "rural":
                                construcciones_rurales += 1
                            else:
                                construcciones_urbanas += 1
                                
                        except Exception as ce:
                            errores_calidad['construcciones_huerfanas'].append({
                                'codigo': codigo,
                                'capa': const_layer,
                                'razon': str(ce)[:30]
                            })
                    
                    logger.info(f"GDB {municipio_nombre}: Guardadas {construcciones_guardadas} construcciones desde {const_layer}")
                except Exception as layer_err:
                    logger.warning(f"Error leyendo capa {const_layer}: {layer_err}")
                    continue
            
            stats["construcciones"] = construcciones_guardadas
            stats["construcciones_rurales"] = construcciones_rurales
            stats["construcciones_urbanas"] = construcciones_urbanas
            if construcciones_guardadas > 0:
                await update_progress("construcciones_ok", 74, f"Guardadas {construcciones_guardadas} construcciones")
        except Exception as e:
            logger.warning(f"Error procesando construcciones: {e}")
        
        await update_progress("subiendo_visor", 75, f"Subiendo al Visor de predios ({len(codigos_gdb)} geometrías)...")
        
        # ========================================
        # VINCULACIÓN GDB-PREDIOS: MATCH EXACTO + ÚLTIMA VIGENCIA
        # ========================================
        if codigos_gdb:
            logger.info(f"GDB tiene {len(codigos_gdb)} códigos únicos. Iniciando vinculación con match exacto...")
            
            # 1. Obtener la ÚLTIMA VIGENCIA del municipio
            vigencias = await db.predios.distinct("vigencia", {"municipio": {"$regex": f"^{municipio_nombre}$", "$options": "i"}})
            vigencias_ordenadas = sorted([v for v in vigencias if v], key=lambda x: int(x) if isinstance(x, (int, str)) and str(x).isdigit() else 0, reverse=True)
            ultima_vigencia = vigencias_ordenadas[0] if vigencias_ordenadas else None
            
            if not ultima_vigencia:
                logger.warning(f"No se encontró vigencia para {municipio_nombre}")
                stats["relacionados"] = 0
                stats["error_vigencia"] = "No se encontró vigencia"
            else:
                logger.info(f"Vinculando con última vigencia: {ultima_vigencia}")
                await update_progress("relacionando_gestion", 78, f"Relacionando geometrías con Gestión de Predios (vigencia {ultima_vigencia})...")
                
                # 2. Crear set de códigos GDB para búsqueda rápida
                codigos_gdb_set = set(codigo.strip() for codigo in codigos_gdb if codigo)
                
                # 3. OPTIMIZACIÓN: Usar operación de update masivo en MongoDB
                # En lugar de iterar sobre cada predio, usamos updateMany para marcar todos los predios
                # que tienen código coincidente de una sola vez
                
                timestamp_now = datetime.now(timezone.utc).isoformat()
                
                # Contar cuántos predios se van a vincular (rápido con índices)
                count_query = {
                    "municipio": {"$regex": f"^{municipio_nombre}$", "$options": "i"},
                    "vigencia": ultima_vigencia,
                    "codigo_predial_nacional": {"$in": list(codigos_gdb_set)}
                }
                relacionados_total = await db.predios.count_documents(count_query)
                
                logger.info(f"Pre-conteo: {relacionados_total} predios con match exacto")
                await update_progress("relacionando_gestion", 82, f"Vinculando {relacionados_total} predios con geometrías...")
                
                # Actualizar todos los predios con match en UNA operación
                result = await db.predios.update_many(
                    count_query,
                    {"$set": {
                        "tiene_geometria": True,
                        "gdb_source": gdb_name,
                        "gdb_updated": timestamp_now,
                        "match_method": "exacto"
                    }}
                )
                
                logger.info(f"UpdateMany completado: {result.modified_count} predios actualizados")
                
                # Actualizar area_gdb para cada predio vinculado
                await update_progress("relacionando_gestion", 88, "Actualizando áreas de predios...")
                
                # Obtener las áreas de las geometrías GDB
                areas_gdb = {}
                logger.info(f"Obteniendo áreas de {len(codigos_gdb_set)} códigos GDB...")
                
                try:
                    async for geo in db.gdb_geometrias.find(
                        {"codigo": {"$in": list(codigos_gdb_set)[:5000]}, "area_m2": {"$gt": 0}},  # Limitar a 5000 para evitar timeout
                        {"_id": 0, "codigo": 1, "area_m2": 1}
                    ):
                        areas_gdb[geo["codigo"]] = geo["area_m2"]
                    
                    logger.info(f"Áreas obtenidas: {len(areas_gdb)}")
                except Exception as area_err:
                    logger.error(f"Error obteniendo áreas: {area_err}")
                    areas_gdb = {}
                
                await update_progress("relacionando_gestion", 90, f"Actualizando {len(areas_gdb)} predios con áreas GDB...")
                
                # Actualizar predios con sus áreas GDB usando bulk_write
                # Si falla, continuar sin bloquear el proceso
                areas_actualizadas = 0
                if areas_gdb:
                    try:
                        from pymongo import UpdateOne
                        bulk_ops = []
                        for codigo, area in areas_gdb.items():
                            bulk_ops.append(UpdateOne(
                                {"codigo_predial_nacional": codigo},  # Búsqueda directa sin regex
                                {"$set": {"area_gdb": area}}
                            ))
                        
                        if bulk_ops:
                            # Procesar en lotes pequeños de 100
                            total_ops = len(bulk_ops)
                            batch_size = 100
                            
                            for i in range(0, min(total_ops, 2000), batch_size):  # Máximo 2000 para evitar timeout
                                batch = bulk_ops[i:i+batch_size]
                                try:
                                    result = await db.predios.bulk_write(batch, ordered=False)
                                    areas_actualizadas += result.modified_count
                                except Exception as batch_err:
                                    logger.warning(f"Batch {i//batch_size + 1} falló: {batch_err}")
                                    continue  # Continuar con el siguiente lote
                            
                            logger.info(f"Áreas GDB actualizadas: {areas_actualizadas}")
                    except Exception as area_update_err:
                        logger.error(f"Error actualizando áreas (continuando): {area_update_err}")
                
                await update_progress("relacionando_gestion", 91, f"Áreas actualizadas: {areas_actualizadas}")
                
                await update_progress("relacionando_gestion", 92, "Finalizando vinculación con Gestión de Predios...")
                
                # Contar predios en última vigencia para el reporte
                predios_ult_vig_count = await db.predios.count_documents({
                    "municipio": {"$regex": f"^{municipio_nombre}$", "$options": "i"},
                    "vigencia": ultima_vigencia
                })
                
                stats["relacionados"] = relacionados_total
                stats["vigencia_usada"] = ultima_vigencia
                stats["predios_vigencia"] = predios_ult_vig_count
                stats["porcentaje_cobertura"] = round((relacionados_total / predios_ult_vig_count) * 100, 2) if predios_ult_vig_count else 0
                
                logger.info(f"Vinculación completada: {relacionados_total} de {predios_ult_vig_count} predios ({stats['porcentaje_cobertura']}%)")
        
        await update_progress("finalizando", 95, f"Registrando carga... {stats.get('relacionados', 0)} predios relacionados")
        
        stats["geometrias_guardadas"] = geometrias_guardadas
        stats["codigos_gdb_unicos"] = len(codigos_gdb)
        
        # Registrar carga mensual
        mes_actual = datetime.now().strftime("%Y-%m")
        await db.gdb_cargas.update_one(
            {"mes": mes_actual, "municipio": municipio_nombre},
            {"$set": {
                "id": str(uuid.uuid4()),
                "mes": mes_actual,
                "municipio": municipio_nombre,
                "gdb_file": gdb_name,
                "uploaded_by": current_user['id'],
                "uploaded_by_name": current_user['full_name'],
                "fecha": datetime.now(timezone.utc).isoformat(),
                "predios_rurales": stats["rurales"],
                "predios_urbanos": stats["urbanos"],
                "predios_relacionados": stats["relacionados"]
            }},
            upsert=True
        )
        
        # Notificar a coordinadores que se completó la carga
        coordinadores = await db.users.find(
            {"role": {"$in": [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]}},
            {"_id": 0, "id": 1, "full_name": 1}
        ).to_list(20)
        
        for coord in coordinadores:
            await crear_notificacion(
                usuario_id=coord['id'],
                titulo=f"Base Gráfica Cargada - {municipio_nombre}",
                mensaje=f"{current_user['full_name']} ha cargado la base gráfica de {municipio_nombre} para {mes_actual}. Total geometrías: {stats['rurales'] + stats['urbanos']}, predios relacionados: {stats['relacionados']}",
                tipo="success",
                enviar_email=False  # No enviar correo para cargas de GDB
            )
        
        await update_progress("completado", 100, f"¡Carga completada! {stats['relacionados']} predios vinculados de {stats['rurales'] + stats['urbanos']} geometrías")
        
        # Limpiar progreso después de 5 minutos
        # (en producción esto se haría con un scheduler)
        
        # Calcular calidad - basado en geometrías guardadas vs leídas
        total_archivo = rurales_en_archivo + urbanos_en_archivo
        total_cargadas = rural_guardadas + urban_guardadas
        
        # CALIDAD = Porcentaje de predios R1/R2 de la ÚLTIMA VIGENCIA que tienen cartografía
        # Usar los datos de la vinculación que ya calculó la última vigencia
        predios_municipio = stats.get("predios_vigencia", 0)
        vigencia_usada = stats.get("vigencia_usada", "N/A")
        
        # Contar predios que ahora tienen geometría (match exitoso)
        predios_con_cartografia = stats.get("relacionados", 0)
        
        # Calcular calidad basada en cobertura de predios de la última vigencia
        if predios_municipio > 0:
            calidad_pct = min(100.0, (predios_con_cartografia / predios_municipio * 100))
        else:
            # Si no hay predios en la última vigencia, usar el cálculo de geometrías guardadas vs archivo
            total_archivo = rurales_en_archivo + urbanos_en_archivo
            total_cargadas = rural_guardadas + urban_guardadas
            calidad_pct = min(100.0, (total_cargadas / total_archivo * 100)) if total_archivo > 0 else 100.0
        
        logger.info(f"GDB {municipio_nombre}: Calidad={calidad_pct:.1f}% (predios vigencia {vigencia_usada}:{predios_municipio}, con cartografía:{predios_con_cartografia})")
        
        # Siempre generar reporte PDF para tener registro de la carga
        reporte_path = None
        tiene_errores = (
            len(errores_calidad['codigos_invalidos']) > 0 or
            len(errores_calidad['geometrias_rechazadas']) > 0 or
            len(errores_calidad['construcciones_huerfanas']) > 0 or
            errores_calidad['rurales_rechazados'] > 0 or
            errores_calidad['urbanos_rechazados'] > 0
        )
        
        # Siempre generar reporte PDF para tener registro de la carga
        try:
            reporte_path = await generar_reporte_calidad_gdb(
                municipio=municipio_nombre,
                fecha_carga=datetime.now().strftime("%Y-%m-%d %H:%M"),
                usuario=current_user['full_name'],
                stats={
                    **stats,
                    'rurales_archivo': rurales_en_archivo,
                    'urbanos_archivo': urbanos_en_archivo,
                    'rurales_guardadas': rural_guardadas,
                    'urbanos_guardadas': urban_guardadas,
                    'predios_municipio': predios_municipio,
                    'predios_con_cartografia': predios_con_cartografia,
                    'calidad_pct': calidad_pct,
                    'vigencia_usada': vigencia_usada
                },
                errores=errores_calidad
            )
            logger.info(f"Reporte de calidad generado: {reporte_path}")
            
            # Notificar a coordinadores si hay errores significativos
            if calidad_pct < 80 or len(errores_calidad['codigos_invalidos']) > 10:
                coordinadores = await db.users.find(
                    {"role": {"$in": [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]}},
                    {"_id": 0, "id": 1, "full_name": 1}
                ).to_list(100)
                
                for coord in coordinadores:
                    await crear_notificacion(
                        usuario_id=coord['id'],
                        titulo=f"⚠️ Reporte de Calidad GDB - {municipio_nombre}",
                        mensaje=f"La carga de GDB de {municipio_nombre} tiene problemas de calidad ({calidad_pct:.1f}%). "
                               f"Códigos inválidos: {len(errores_calidad['codigos_invalidos'])}, "
                               f"Geometrías rechazadas: {len(errores_calidad['geometrias_rechazadas'])}. "
                               f"Revisar reporte PDF.",
                        tipo="warning",
                        enviar_email=False
                    )
        except Exception as report_err:
            logger.error(f"Error generando reporte de calidad: {report_err}")
        
        # Actualizar estado final con todos los datos para el polling
        final_message = f"¡Carga completada! {stats['relacionados']} predios vinculados de {rural_guardadas + urban_guardadas} geometrías"
        final_data = {
            "municipio": municipio_nombre,
            "gdb_file": gdb_name,
            "upload_id": upload_id,
            "predios_rurales": rural_guardadas,
            "predios_urbanos": urban_guardadas,
            "total_geometrias_gdb": rural_guardadas + urban_guardadas,
            "codigos_unicos_gdb": len(codigos_gdb),
            "geometrias_guardadas": rural_guardadas + urban_guardadas,
            "predios_relacionados": stats["relacionados"],
            "metodo_match": stats.get("metodo_match", "directo"),
            "calidad": {
                "porcentaje": round(calidad_pct, 1),
                "rurales_archivo": rurales_en_archivo,
                "urbanos_archivo": urbanos_en_archivo,
                "rurales_guardadas": rural_guardadas,
                "urbanos_guardadas": urban_guardadas,
                "codigos_invalidos": len(errores_calidad['codigos_invalidos']),
                "geometrias_rechazadas": len(errores_calidad['geometrias_rechazadas']),
                "construcciones_huerfanas": len(errores_calidad['construcciones_huerfanas']),
                "reporte_pdf": reporte_path.split('/')[-1] if reporte_path else None
            },
            "construcciones": {
                "total": stats.get('construcciones', 0),
                "rurales": stats.get('construcciones_rurales', 0),
                "urbanas": stats.get('construcciones_urbanas', 0)
            }
        }
        
        # Actualizar progreso final con todos los datos
        await update_progress(
            "completado", 
            100, 
            final_message,
            **final_data
        )
        
        logger.info(f"GDB {municipio_nombre} procesado correctamente. Predios relacionados: {stats['relacionados']}")
        return  # Background task no retorna al cliente
        
    except zipfile.BadZipFile:
        await update_progress("error", 0, "El archivo ZIP no es válido o está corrupto")
        logger.error("Error: ZIP inválido o corrupto")
        return  # Background task no lanza excepciones
    except Exception as e:
        error_msg = str(e)
        # Identificar errores comunes y dar mensajes más claros
        if "fiona" in error_msg.lower() or "gdal" in error_msg.lower():
            user_msg = f"Error al leer el archivo GDB. El archivo puede estar corrupto o en formato incompatible. Detalle: {error_msg[:100]}"
        elif "transform" in error_msg.lower() or "crs" in error_msg.lower():
            user_msg = f"Error en la transformación de coordenadas. El sistema de referencia del GDB puede no ser compatible. Detalle: {error_msg[:100]}"
        elif "memory" in error_msg.lower():
            user_msg = "El archivo GDB es muy grande para procesar. Intente dividirlo en partes más pequeñas."
        elif "permission" in error_msg.lower() or "access" in error_msg.lower():
            user_msg = "Error de permisos al procesar el archivo. Contacte al administrador."
        else:
            user_msg = f"Error al procesar el archivo: {error_msg[:150]}"
        
        await update_progress("error", 0, user_msg)
        logger.error(f"Error uploading GDB ({temp_files[0]['filename'] if temp_files else 'unknown'}): {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return  # Background task no lanza excepciones


# ===== FUNCIÓN PARA GENERAR REPORTE PDF DE CALIDAD GDB =====
async def generar_reporte_calidad_gdb(
    municipio: str,
    fecha_carga: str,
    usuario: str,
    stats: dict,
    errores: dict
) -> str:
    """
    Genera un PDF con el reporte de calidad de la carga GDB.
    Retorna el path del archivo generado.
    """
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    # Crear directorio si no existe
    reports_dir = Path("/app/reports/gdb_calidad")
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Nombre del archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"calidad_gdb_{municipio.replace(' ', '_')}_{timestamp}.pdf"
    filepath = reports_dir / filename
    
    # Crear documento
    doc = SimpleDocTemplate(str(filepath), pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []
    
    # Título
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1e3a5f'), spaceAfter=20)
    elements.append(Paragraph(f"Reporte de Calidad GDB - {municipio}", title_style))
    elements.append(Paragraph(f"Fecha de carga: {fecha_carga}", styles['Normal']))
    elements.append(Paragraph(f"Usuario: {usuario}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Resumen estadístico
    elements.append(Paragraph("📊 Resumen de Carga", styles['Heading2']))
    
    resumen_data = [
        ["Concepto", "Rural", "Urbano", "Total"],
        ["Geometrías en archivo GDB", str(stats.get('rurales_archivo', 0)), str(stats.get('urbanos_archivo', 0)), str(stats.get('rurales_archivo', 0) + stats.get('urbanos_archivo', 0))],
        ["Geometrías cargadas", str(stats.get('rurales_guardadas', stats.get('rurales', 0))), str(stats.get('urbanos_guardadas', stats.get('urbanos', 0))), str(stats.get('rurales_guardadas', stats.get('rurales', 0)) + stats.get('urbanos_guardadas', stats.get('urbanos', 0)))],
        ["Geometrías rechazadas", str(errores.get('rurales_rechazados', 0)), str(errores.get('urbanos_rechazados', 0)), str(errores.get('rurales_rechazados', 0) + errores.get('urbanos_rechazados', 0))],
        ["Construcciones cargadas", str(stats.get('construcciones_rurales', 0)), str(stats.get('construcciones_urbanas', 0)), str(stats.get('construcciones', 0))],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 15))
    
    # Tabla de vinculación con predios R1/R2
    elements.append(Paragraph("📋 Vinculación con Base de Datos R1/R2", styles['Heading2']))
    
    vigencia_usada = stats.get('vigencia_usada', 'N/A')
    
    vinculacion_data = [
        ["Concepto", "Cantidad"],
        [f"Vigencia analizada", f"{vigencia_usada}"],
        [f"Predios R1/R2 (vigencia {vigencia_usada})", f"{stats.get('predios_municipio', 0):,}"],
        ["Predios con cartografía (vinculados)", f"{stats.get('predios_con_cartografia', 0):,}"],
        ["Predios sin cartografía", f"{max(0, stats.get('predios_municipio', 0) - stats.get('predios_con_cartografia', 0)):,}"],
    ]
    
    vinc_table = Table(vinculacion_data, colWidths=[3.5*inch, 2*inch])
    vinc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8f5e9')]),
    ]))
    elements.append(vinc_table)
    elements.append(Spacer(1, 20))
    
    # Indicador de calidad - Basado en cobertura de predios R1/R2
    predios_municipio = stats.get('predios_municipio', 0)
    predios_con_cartografia = stats.get('predios_con_cartografia', 0)
    
    if predios_municipio > 0:
        calidad_pct = (predios_con_cartografia / predios_municipio * 100)
        calidad_base = f"Predios R1/R2 del municipio: {predios_municipio:,} | Con cartografía: {predios_con_cartografia:,}"
    else:
        # Fallback al cálculo de geometrías
        total_archivo = stats.get('rurales_archivo', 0) + stats.get('urbanos_archivo', 0)
        total_cargadas = stats.get('rurales', 0) + stats.get('urbanos', 0)
        calidad_pct = (total_cargadas / total_archivo * 100) if total_archivo > 0 else 0
        calidad_base = f"Geometrías en archivo: {total_archivo:,} | Cargadas: {total_cargadas:,}"
    
    if calidad_pct >= 95:
        calidad_color = colors.HexColor('#28a745')
        calidad_texto = "EXCELENTE"
    elif calidad_pct >= 80:
        calidad_color = colors.HexColor('#17a2b8')
        calidad_texto = "BUENA"
    elif calidad_pct >= 60:
        calidad_color = colors.HexColor('#ffc107')
        calidad_texto = "REGULAR"
    else:
        calidad_color = colors.HexColor('#dc3545')
        calidad_texto = "DEFICIENTE"
    
    calidad_style = ParagraphStyle('Calidad', parent=styles['Normal'], fontSize=12, textColor=calidad_color, fontName='Helvetica-Bold')
    elements.append(Paragraph(f"📈 Calidad de la carga: {calidad_pct:.1f}% - {calidad_texto}", calidad_style))
    elements.append(Paragraph(calidad_base, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Errores de códigos (no 30 dígitos)
    codigos_invalidos = errores.get('codigos_invalidos', [])
    if codigos_invalidos:
        elements.append(Paragraph("⚠️ Códigos con formato incorrecto (no son 30 dígitos)", styles['Heading2']))
        elements.append(Paragraph(f"Total: {len(codigos_invalidos)} códigos requieren revisión del equipo SIG", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        # Tabla de códigos inválidos (máximo 50)
        codigo_data = [["Código", "Longitud", "Capa"]]
        for item in codigos_invalidos[:50]:
            codigo_data.append([item.get('codigo', ''), str(item.get('longitud', '')), item.get('capa', '')])
        
        if len(codigos_invalidos) > 50:
            codigo_data.append([f"... y {len(codigos_invalidos) - 50} más", "", ""])
        
        codigo_table = Table(codigo_data, colWidths=[3*inch, 1*inch, 1.5*inch])
        codigo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(codigo_table)
        elements.append(Spacer(1, 20))
    
    # Geometrías rechazadas (fuera de Colombia)
    geom_rechazadas = errores.get('geometrias_rechazadas', [])
    if geom_rechazadas:
        elements.append(Paragraph("🌍 Geometrías rechazadas (coordenadas fuera de Colombia)", styles['Heading2']))
        elements.append(Paragraph(f"Total: {len(geom_rechazadas)} geometrías", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        geom_data = [["Código", "Razón", "Capa"]]
        for item in geom_rechazadas[:30]:
            geom_data.append([item.get('codigo', ''), item.get('razon', ''), item.get('capa', '')])
        
        if len(geom_rechazadas) > 30:
            geom_data.append([f"... y {len(geom_rechazadas) - 30} más", "", ""])
        
        geom_table = Table(geom_data, colWidths=[3*inch, 2*inch, 1.5*inch])
        geom_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffc107')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(geom_table)
        elements.append(Spacer(1, 20))
    
    # Construcciones huérfanas
    const_huerfanas = errores.get('construcciones_huerfanas', [])
    if const_huerfanas:
        elements.append(Paragraph("🏗️ Construcciones sin predio padre", styles['Heading2']))
        elements.append(Paragraph(f"Total: {len(const_huerfanas)} construcciones no pudieron vincularse a un predio", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        const_data = [["Código Construcción", "Capa"]]
        for item in const_huerfanas[:30]:
            const_data.append([item.get('codigo', ''), item.get('capa', '')])
        
        if len(const_huerfanas) > 30:
            const_data.append([f"... y {len(const_huerfanas) - 30} más", ""])
        
        const_table = Table(const_data, colWidths=[4*inch, 2*inch])
        const_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#17a2b8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(const_table)
    
    # Generar PDF
    doc.build(elements)
    
    return str(filepath)


@api_router.get("/gdb/reportes-calidad")
async def listar_reportes_calidad_gdb(current_user: dict = Depends(get_current_user)):
    """Lista los reportes de calidad GDB disponibles"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    reports_dir = Path("/app/reports/gdb_calidad")
    if not reports_dir.exists():
        return {"reportes": []}
    
    reportes = []
    for f in sorted(reports_dir.glob("*.pdf"), reverse=True):
        reportes.append({
            "nombre": f.name,
            "fecha": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            "tamaño": f.stat().st_size,
            "url": f"/api/gdb/reportes-calidad/{f.name}"
        })
    
    return {"reportes": reportes[:50]}  # Últimos 50


@api_router.get("/gdb/reportes-calidad/{filename}")
async def descargar_reporte_calidad_gdb(filename: str, current_user: dict = Depends(get_current_user)):
    """Descarga un reporte de calidad GDB"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    filepath = Path("/app/reports/gdb_calidad") / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    
    return FileResponse(filepath, filename=filename, media_type="application/pdf")


@api_router.get("/gdb/cargas-mensuales")
async def get_cargas_mensuales_gdb(
    mes: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las cargas de GDB del mes actual o especificado"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    mes_consulta = mes or datetime.now().strftime("%Y-%m")
    
    cargas = await db.gdb_cargas.find(
        {"mes": mes_consulta},
        {"_id": 0}
    ).to_list(50)
    
    return {
        "mes": mes_consulta,
        "total_cargas": len(cargas),
        "cargas": cargas
    }



@api_router.get("/gdb/verificar-carga-mes")
async def verificar_carga_gdb_mes(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Verifica si ya se cargó GDB este mes para un municipio o en general"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    mes_actual = datetime.now().strftime("%Y-%m")
    
    if municipio:
        # Verificar para un municipio específico
        carga = await db.gdb_cargas.find_one(
            {"mes": mes_actual, "municipio": municipio},
            {"_id": 0}
        )
        return {
            "mes": mes_actual,
            "municipio": municipio,
            "cargado": carga is not None,
            "detalle": carga
        }
    else:
        # Obtener resumen de todos los municipios
        cargas = await db.gdb_cargas.find(
            {"mes": mes_actual},
            {"_id": 0, "municipio": 1, "fecha": 1, "uploaded_by_name": 1}
        ).to_list(50)
        
        municipios_cargados = [c["municipio"] for c in cargas]
        
        # Obtener lista de todos los municipios activos
        todos_municipios = await db.predios.distinct("municipio")
        municipios_pendientes = [m for m in todos_municipios if m and m not in municipios_cargados]
        
        return {
            "mes": mes_actual,
            "total_cargados": len(municipios_cargados),
            "total_pendientes": len(municipios_pendientes),
            "municipios_cargados": cargas,
            "municipios_pendientes": municipios_pendientes
        }



@api_router.post("/gdb/revincular-predios")
async def revincular_predios_gdb(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Re-vincula las geometrías GDB existentes con predios usando MATCH EXACTO del CPN.
    IMPORTANTE: Solo vincula con predios de la ÚLTIMA VIGENCIA de cada municipio.
    
    Lógica:
    1. Obtener la última vigencia del municipio
    2. Obtener todos los códigos CPN de la GDB
    3. Buscar coincidencias EXACTAS con predios de la última vigencia
    4. Actualizar predios con tiene_geometria=true y área de GDB
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden revincular")
    
    resultados = {
        "municipios_procesados": [], 
        "total_vinculados": 0, 
        "total_sin_match": 0,
        "errores": []
    }
    
    # Obtener municipios a procesar
    if municipio:
        municipios = [municipio]
    else:
        municipios = await db.gdb_geometrias.distinct("municipio")
    
    for muni in municipios:
        try:
            # 1. Obtener la ÚLTIMA VIGENCIA del municipio
            vigencias = await db.predios.distinct("vigencia", {"municipio": {"$regex": f"^{muni}$", "$options": "i"}})
            if not vigencias:
                resultados["errores"].append(f"{muni}: Sin predios en base de datos")
                continue
            
            # Ordenar vigencias y tomar la más reciente
            vigencias_ordenadas = sorted([v for v in vigencias if v], key=lambda x: int(x) if isinstance(x, (int, str)) and str(x).isdigit() else 0, reverse=True)
            if not vigencias_ordenadas:
                resultados["errores"].append(f"{muni}: Sin vigencias válidas")
                continue
            
            ultima_vigencia = vigencias_ordenadas[0]
            
            # 2. Obtener TODOS los códigos GDB del municipio
            codigos_gdb_cursor = db.gdb_geometrias.find(
                {"municipio": muni},
                {"_id": 0, "codigo": 1, "area_m2": 1}
            )
            codigos_gdb_docs = await codigos_gdb_cursor.to_list(50000)
            
            if not codigos_gdb_docs:
                resultados["errores"].append(f"{muni}: Sin geometrías GDB")
                continue
            
            # Crear diccionario para búsqueda rápida: codigo -> area_m2
            gdb_dict = {doc["codigo"]: doc.get("area_m2", 0) for doc in codigos_gdb_docs if doc.get("codigo")}
            codigos_gdb_set = set(gdb_dict.keys())
            
            # 3. Obtener predios de la ÚLTIMA VIGENCIA (todos, para revinculación completa)
            predios_ultima_vig = await db.predios.find(
                {
                    "municipio": {"$regex": f"^{muni}$", "$options": "i"},
                    "vigencia": ultima_vigencia
                },
                {"_id": 0, "id": 1, "codigo_predial_nacional": 1, "tiene_geometria": 1}
            ).to_list(100000)
            
            if not predios_ultima_vig:
                resultados["errores"].append(f"{muni}: Sin predios en vigencia {ultima_vigencia}")
                continue
            
            vinculados_muni = 0
            sin_match_muni = 0
            
            # 4. Buscar coincidencias EXACTAS
            for predio in predios_ultima_vig:
                codigo_cpn = predio.get("codigo_predial_nacional", "")
                
                if not codigo_cpn:
                    sin_match_muni += 1
                    continue
                
                # MATCH EXACTO únicamente
                if codigo_cpn in codigos_gdb_set:
                    area_gdb = gdb_dict.get(codigo_cpn, 0)
                    
                    # Solo actualizar si no tiene geometría o si queremos refrescar
                    await db.predios.update_one(
                        {"id": predio["id"]},
                        {"$set": {
                            "tiene_geometria": True,
                            "codigo_gdb": codigo_cpn,
                            "area_gdb": area_gdb,
                            "gdb_updated": datetime.now(timezone.utc).isoformat(),
                            "match_method": "exacto"
                        }}
                    )
                    vinculados_muni += 1
                else:
                    sin_match_muni += 1
            
            # También actualizar la geometría GDB con predio_vinculado
            for codigo_cpn in codigos_gdb_set:
                predio_match = await db.predios.find_one(
                    {
                        "codigo_predial_nacional": codigo_cpn,
                        "municipio": {"$regex": f"^{muni}$", "$options": "i"},
                        "vigencia": ultima_vigencia
                    },
                    {"_id": 0, "id": 1}
                )
                if predio_match:
                    await db.gdb_geometrias.update_one(
                        {"codigo": codigo_cpn, "municipio": muni},
                        {"$set": {"predio_vinculado": predio_match["id"]}}
                    )
            
            resultados["municipios_procesados"].append({
                "municipio": muni,
                "vigencia": ultima_vigencia,
                "geometrias_gdb": len(codigos_gdb_set),
                "predios_vigencia": len(predios_ultima_vig),
                "vinculados": vinculados_muni,
                "sin_match": sin_match_muni,
                "porcentaje_cobertura": round((vinculados_muni / len(predios_ultima_vig)) * 100, 2) if predios_ultima_vig else 0
            })
            resultados["total_vinculados"] += vinculados_muni
            resultados["total_sin_match"] += sin_match_muni
            
            logger.info(f"Revinculación {muni}: {vinculados_muni} vinculados de {len(predios_ultima_vig)} predios (vigencia {ultima_vigencia})")
            
        except Exception as e:
            resultados["errores"].append(f"{muni}: {str(e)}")
            logger.error(f"Error revinculando {muni}: {e}")
    
    return resultados



@api_router.post("/gdb/recalcular-areas")
async def recalcular_areas_gdb(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Recalcula las áreas de las geometrías GDB existentes y actualiza los predios relacionados.
    Útil para geometrías que se cargaron antes de implementar el cálculo de área.
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores")
    
    from shapely.geometry import shape
    
    query = {}
    if municipio:
        query["municipio"] = municipio
    
    # Obtener geometrías sin área o con área 0
    geometrias = await db.gdb_geometrias.find(
        {**query, "$or": [{"area_m2": {"$exists": False}}, {"area_m2": 0}]},
        {"_id": 1, "codigo": 1, "geometry": 1}
    ).to_list(100000)
    
    actualizadas = 0
    errores = 0
    
    for geo in geometrias:
        try:
            geom_shape = shape(geo.get("geometry"))
            area_deg = geom_shape.area
            area_m2 = round(area_deg * (111320 ** 2), 2)
            
            await db.gdb_geometrias.update_one(
                {"_id": geo["_id"]},
                {"$set": {"area_m2": area_m2}}
            )
            
            # Actualizar predios que tengan este código GDB
            if geo.get("codigo"):
                await db.predios.update_many(
                    {"codigo_gdb": geo["codigo"]},
                    {"$set": {"area_gdb": area_m2}}
                )
            
            actualizadas += 1
        except Exception as e:
            errores += 1
    
    return {
        "mensaje": f"Áreas recalculadas",
        "geometrias_procesadas": len(geometrias),
        "actualizadas": actualizadas,
        "errores": errores
    }


@api_router.post("/gdb/sincronizar-areas-predios")
async def sincronizar_areas_predios(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Sincroniza las áreas de GDB con los predios que ya tienen geometría vinculada.
    Busca por codigo_predial_nacional para mayor cobertura.
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores")
    
    # Query base para predios con geometría
    predio_query = {"tiene_geometria": True}
    if municipio:
        predio_query["municipio"] = {"$regex": f"^{municipio}$", "$options": "i"}
    
    # Obtener códigos de predios con geometría
    predios = await db.predios.find(
        predio_query,
        {"_id": 0, "id": 1, "codigo_predial_nacional": 1}
    ).to_list(200000)
    
    if not predios:
        return {"mensaje": "No hay predios con geometría", "actualizados": 0}
    
    # Obtener códigos únicos
    codigos = list(set(p.get("codigo_predial_nacional") for p in predios if p.get("codigo_predial_nacional")))
    
    # Obtener áreas de gdb_geometrias
    areas_gdb = {}
    async for geo in db.gdb_geometrias.find(
        {"codigo": {"$in": codigos}, "area_m2": {"$gt": 0}},
        {"_id": 0, "codigo": 1, "area_m2": 1}
    ):
        areas_gdb[geo["codigo"]] = geo["area_m2"]
    
    # Actualizar predios usando bulk_write
    from pymongo import UpdateOne
    actualizados = 0
    
    if areas_gdb:
        bulk_ops = []
        for codigo, area in areas_gdb.items():
            bulk_ops.append(UpdateOne(
                {"codigo_predial_nacional": codigo},
                {"$set": {"area_gdb": area}}
            ))
        
        # Procesar en lotes
        for i in range(0, len(bulk_ops), 1000):
            batch = bulk_ops[i:i+1000]
            result = await db.predios.bulk_write(batch, ordered=False)
            actualizados += result.modified_count
    
    return {
        "mensaje": "Áreas sincronizadas",
        "predios_con_geometria": len(predios),
        "geometrias_encontradas": len(areas_gdb),
        "actualizados": actualizados
    }



@api_router.get("/gdb/predios-con-geometria")
async def get_predios_con_geometria(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene estadísticas de predios que tienen geometría GDB asociada"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {"tiene_geometria": True}
    if municipio:
        query["municipio"] = municipio
    
    total_con_geometria = await db.predios.count_documents(query)
    total_predios = await db.predios.count_documents({"municipio": municipio} if municipio else {})
    
    # Por municipio
    pipeline = [
        {"$match": {"tiene_geometria": True}},
        {"$group": {"_id": "$municipio", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    por_municipio = await db.predios.aggregate(pipeline).to_list(50)
    
    return {
        "total_con_geometria": total_con_geometria,
        "total_predios": total_predios,
        "porcentaje": round((total_con_geometria / total_predios * 100) if total_predios > 0 else 0, 2),
        "por_municipio": [{"municipio": r["_id"], "count": r["count"]} for r in por_municipio]
    }


@api_router.get("/gdb/geometrias-disponibles")
async def get_geometrias_disponibles(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las geometrías GDB guardadas por municipio"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    pipeline = [
        {"$group": {
            "_id": {"municipio": "$municipio", "tipo": "$tipo"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id.municipio": 1}}
    ]
    
    if municipio:
        pipeline.insert(0, {"$match": {"municipio": municipio}})
    
    resultados = await db.gdb_geometrias.aggregate(pipeline).to_list(100)
    
    # Organizar por municipio
    por_municipio = {}
    for r in resultados:
        mun = r["_id"]["municipio"]
        tipo = r["_id"]["tipo"]
        if mun not in por_municipio:
            por_municipio[mun] = {"rural": 0, "urbano": 0, "total": 0}
        por_municipio[mun][tipo] = r["count"]
        por_municipio[mun]["total"] += r["count"]
    
    total_geometrias = await db.gdb_geometrias.count_documents({} if not municipio else {"municipio": municipio})
    
    return {
        "total_geometrias": total_geometrias,
        "por_municipio": por_municipio,
        "municipios_con_gdb": list(por_municipio.keys())
    }


@api_router.get("/gdb/buscar-geometria/{codigo}")
async def buscar_geometria_por_codigo(
    codigo: str,
    current_user: dict = Depends(get_current_user)
):
    """Busca geometría por código predial en la colección de geometrías guardadas"""
    if current_user['role'] == UserRole.USUARIO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Buscar exacto primero
    geometria = await db.gdb_geometrias.find_one(
        {"codigo": codigo},
        {"_id": 0}
    )
    
    if not geometria:
        # Buscar por coincidencia parcial (código contenido)
        geometria = await db.gdb_geometrias.find_one(
            {"codigo": {"$regex": f".*{codigo}.*"}},
            {"_id": 0}
        )
    
    if not geometria:
        return {"encontrado": False, "mensaje": "No se encontró geometría para este código"}
    
    return {
        "encontrado": True,
        "codigo": geometria.get("codigo"),
        "tipo": geometria.get("tipo"),
        "municipio": geometria.get("municipio"),
        "geometry": geometria.get("geometry")
    }


@api_router.patch("/users/{user_id}/gdb-permission")
async def update_user_gdb_permission(
    user_id: str,
    puede_actualizar: bool,
    current_user: dict = Depends(get_current_user)
):
    """Allow coordinador to grant/revoke GDB update permission to a gestor"""
    # Only coordinador or admin can grant this permission
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden asignar este permiso")
    
    # Find user
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Only gestors can have this permission
    if user['role'] != UserRole.GESTOR:
        raise HTTPException(status_code=400, detail="Este permiso solo aplica para gestores")
    
    # Update permission
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"puede_actualizar_gdb": puede_actualizar}}
    )
    
    return {
        "message": f"Permiso {'otorgado' if puede_actualizar else 'revocado'} exitosamente",
        "user_id": user_id,
        "puede_actualizar_gdb": puede_actualizar
    }


# ===== ORTOIMÁGENES - SISTEMA DE CARGA Y TILES XYZ =====

ORTOIMAGENES_PATH = Path("/app/ortoimagenes/tiles")
ORTOIMAGENES_ORIGINALES_PATH = Path("/app/ortoimagenes/originales")

# Asegurar que los directorios existan
ORTOIMAGENES_PATH.mkdir(parents=True, exist_ok=True)
ORTOIMAGENES_ORIGINALES_PATH.mkdir(parents=True, exist_ok=True)

# Diccionario para tracking de progreso de procesamiento de ortoimágenes
ortoimagen_processing_progress = {}

@api_router.get("/ortoimagenes/disponibles")
async def listar_ortoimagenes(current_user: dict = Depends(get_current_user)):
    """Lista las ortoimágenes disponibles en el sistema (desde MongoDB)"""
    ortoimagenes = await db.ortoimagenes.find(
        {"activa": True},
        {"_id": 0}
    ).to_list(100)
    
    # Verificar que los tiles existan físicamente
    result = []
    for orto in ortoimagenes:
        tile_path = ORTOIMAGENES_PATH / orto["id"]
        if tile_path.exists():
            result.append(orto)
    
    return {"ortoimagenes": result}

@api_router.get("/ortoimagenes/todas")
async def listar_todas_ortoimagenes(current_user: dict = Depends(get_current_user)):
    """Lista todas las ortoimágenes (activas e inactivas) - solo para admin/coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    ortoimagenes = await db.ortoimagenes.find({}, {"_id": 0}).to_list(100)
    return {"ortoimagenes": ortoimagenes}

@api_router.post("/ortoimagenes/subir")
async def subir_ortoimagen(
    file: UploadFile = File(...),
    nombre: str = Form(...),
    municipio: str = Form(...),
    descripcion: str = Form(""),
    background_tasks: BackgroundTasks = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Sube una ortoimagen (GeoTIFF) y la procesa en tiles XYZ.
    Solo usuarios con permiso 'upload_gdb' pueden subir ortoimágenes.
    """
    # Verificar permisos: admin, coordinador, o gestor con permiso upload_gdb
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    permissions = user_db.get('permissions', [])
    
    has_permission = (
        current_user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR] or
        (current_user['role'] == UserRole.GESTOR and 'upload_gdb' in permissions)
    )
    
    if not has_permission:
        raise HTTPException(status_code=403, detail="No tiene permiso para subir ortoimágenes")
    
    # Validar archivo
    if not file.filename.lower().endswith(('.tif', '.tiff', '.geotiff')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos GeoTIFF (.tif, .tiff)")
    
    # Generar ID único para la ortoimagen
    orto_id = f"orto_{uuid.uuid4().hex[:8]}"
    
    # Guardar archivo original
    original_path = ORTOIMAGENES_ORIGINALES_PATH / f"{orto_id}.tif"
    
    try:
        # Guardar archivo en chunks para manejar archivos grandes
        with open(original_path, "wb") as buffer:
            while chunk := await file.read(1024 * 1024):  # 1MB chunks
                buffer.write(chunk)
        
        file_size = original_path.stat().st_size
        logger.info(f"Ortoimagen guardada: {original_path} ({file_size / (1024*1024):.2f} MB)")
        
    except Exception as e:
        logger.error(f"Error guardando ortoimagen: {e}")
        raise HTTPException(status_code=500, detail=f"Error al guardar el archivo: {str(e)}")
    
    # Inicializar progreso
    ortoimagen_processing_progress[orto_id] = {
        "status": "subido",
        "progress": 10,
        "message": "Archivo recibido, iniciando procesamiento..."
    }
    
    # Crear registro en MongoDB
    orto_doc = {
        "id": orto_id,
        "nombre": nombre,
        "municipio": municipio,
        "descripcion": descripcion,
        "archivo_original": str(original_path),
        "activa": False,  # Se activa cuando termine el procesamiento
        "procesando": True,
        "zoom_min": 14,
        "zoom_max": 20,
        "bounds": None,  # Se llena después del procesamiento
        "fecha_subida": datetime.now(timezone.utc).isoformat(),
        "subido_por": current_user['id'],
        "subido_por_nombre": current_user['full_name']
    }
    
    await db.ortoimagenes.insert_one(orto_doc)
    
    # Procesar en background
    if background_tasks:
        background_tasks.add_task(procesar_ortoimagen_background, orto_id, str(original_path), nombre)
    else:
        # Si no hay background_tasks, procesar sincrónicamente
        await procesar_ortoimagen_background(orto_id, str(original_path), nombre)
    
    return {
        "message": "Ortoimagen recibida. El procesamiento de tiles puede tomar varios minutos.",
        "id": orto_id,
        "nombre": nombre
    }

async def procesar_ortoimagen_background(orto_id: str, tiff_path: str, nombre: str):
    """Procesa un GeoTIFF y genera tiles XYZ usando gdal2tiles"""
    import subprocess
    
    def update_progress(status: str, progress: int, message: str):
        ortoimagen_processing_progress[orto_id] = {
            "status": status,
            "progress": progress,
            "message": message
        }
    
    tiles_output_path = ORTOIMAGENES_PATH / orto_id
    
    try:
        update_progress("procesando", 20, "Leyendo información del archivo GeoTIFF...")
        
        # Obtener información del GeoTIFF con gdalinfo
        result = subprocess.run(
            ["gdalinfo", "-json", tiff_path],
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if result.returncode != 0:
            raise Exception(f"Error leyendo GeoTIFF: {result.stderr}")
        
        import json
        gdal_info = json.loads(result.stdout)
        
        # Extraer bounds
        bounds = None
        if "wgs84Extent" in gdal_info:
            coords = gdal_info["wgs84Extent"]["coordinates"][0]
            # coords es un polígono, extraer SW y NE
            lngs = [c[0] for c in coords]
            lats = [c[1] for c in coords]
            bounds = [[min(lngs), min(lats)], [max(lngs), max(lats)]]
        elif "cornerCoordinates" in gdal_info:
            cc = gdal_info["cornerCoordinates"]
            bounds = [
                [cc["lowerLeft"][0], cc["lowerLeft"][1]],
                [cc["upperRight"][0], cc["upperRight"][1]]
            ]
        
        update_progress("tiles", 30, "Generando tiles XYZ (esto puede tomar varios minutos)...")
        
        # Ejecutar gdal2tiles
        tiles_output_path.mkdir(parents=True, exist_ok=True)
        
        gdal_cmd = [
            "gdal2tiles.py",
            "-z", "14-20",  # Niveles de zoom
            "-w", "none",   # Sin archivo HTML
            "-r", "average",  # Método de remuestreo
            "--processes=2",  # Usar 2 procesos
            tiff_path,
            str(tiles_output_path)
        ]
        
        logger.info(f"Ejecutando: {' '.join(gdal_cmd)}")
        
        result = subprocess.run(
            gdal_cmd,
            capture_output=True,
            text=True,
            timeout=1800  # 30 minutos máximo
        )
        
        if result.returncode != 0:
            logger.error(f"gdal2tiles error: {result.stderr}")
            raise Exception(f"Error generando tiles: {result.stderr}")
        
        update_progress("verificando", 90, "Verificando tiles generados...")
        
        # Verificar que se generaron tiles
        tile_files = list(tiles_output_path.glob("**/*.png"))
        if len(tile_files) == 0:
            raise Exception("No se generaron tiles")
        
        # Leer bounds del tilemapresource.xml si existe
        tilemap_xml = tiles_output_path / "tilemapresource.xml"
        if tilemap_xml.exists():
            import xml.etree.ElementTree as ET
            tree = ET.parse(tilemap_xml)
            root = tree.getroot()
            bbox = root.find("BoundingBox")
            if bbox is not None:
                bounds = [
                    [float(bbox.get("minx")), float(bbox.get("miny"))],
                    [float(bbox.get("maxx")), float(bbox.get("maxy"))]
                ]
        
        # Determinar zoom min/max real
        zoom_levels = [int(d.name) for d in tiles_output_path.iterdir() if d.is_dir() and d.name.isdigit()]
        zoom_min = min(zoom_levels) if zoom_levels else 14
        zoom_max = max(zoom_levels) if zoom_levels else 20
        
        # Actualizar en MongoDB
        await db.ortoimagenes.update_one(
            {"id": orto_id},
            {"$set": {
                "activa": True,
                "procesando": False,
                "bounds": bounds,
                "zoom_min": zoom_min,
                "zoom_max": zoom_max,
                "total_tiles": len(tile_files),
                "fecha_procesado": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        update_progress("completado", 100, f"¡Completado! {len(tile_files)} tiles generados")
        logger.info(f"Ortoimagen {orto_id} procesada: {len(tile_files)} tiles, bounds={bounds}")
        
    except Exception as e:
        logger.error(f"Error procesando ortoimagen {orto_id}: {e}")
        update_progress("error", 0, f"Error: {str(e)}")
        
        # Marcar como error en MongoDB
        await db.ortoimagenes.update_one(
            {"id": orto_id},
            {"$set": {
                "procesando": False,
                "error": str(e)
            }}
        )

@api_router.get("/ortoimagenes/progreso/{orto_id}")
async def obtener_progreso_ortoimagen(orto_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene el progreso del procesamiento de una ortoimagen"""
    if orto_id in ortoimagen_processing_progress:
        return ortoimagen_processing_progress[orto_id]
    
    # Verificar en MongoDB
    orto = await db.ortoimagenes.find_one({"id": orto_id}, {"_id": 0})
    if not orto:
        raise HTTPException(status_code=404, detail="Ortoimagen no encontrada")
    
    if orto.get("procesando"):
        return {"status": "procesando", "progress": 50, "message": "Procesando..."}
    elif orto.get("activa"):
        return {"status": "completado", "progress": 100, "message": "Ortoimagen lista"}
    elif orto.get("error"):
        return {"status": "error", "progress": 0, "message": orto["error"]}
    else:
        return {"status": "pendiente", "progress": 0, "message": "En espera"}

@api_router.post("/ortoimagenes/{orto_id}/reprocesar")
async def reprocesar_ortoimagen(
    orto_id: str, 
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Reprocesa una ortoimagen que falló"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    orto = await db.ortoimagenes.find_one({"id": orto_id}, {"_id": 0})
    if not orto:
        raise HTTPException(status_code=404, detail="Ortoimagen no encontrada")
    
    original_path = orto.get("archivo_original")
    if not original_path or not os.path.exists(original_path):
        raise HTTPException(status_code=400, detail="Archivo original no encontrado")
    
    # Limpiar tiles existentes si hay
    tiles_path = ORTOIMAGENES_PATH / orto_id
    if tiles_path.exists():
        shutil.rmtree(tiles_path)
    
    # Resetear estado en MongoDB
    await db.ortoimagenes.update_one(
        {"id": orto_id},
        {"$set": {"procesando": True, "error": None, "activa": False}}
    )
    
    # Iniciar reprocesamiento
    ortoimagen_processing_progress[orto_id] = {
        "status": "iniciando",
        "progress": 5,
        "message": "Iniciando reprocesamiento..."
    }
    
    background_tasks.add_task(procesar_ortoimagen_background, orto_id, original_path, orto["nombre"])
    
    return {"message": "Reprocesamiento iniciado", "id": orto_id}

@api_router.delete("/ortoimagenes/{orto_id}")
async def eliminar_ortoimagen(orto_id: str, current_user: dict = Depends(get_current_user)):
    """Elimina una ortoimagen y sus tiles"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para eliminar ortoimágenes")
    
    orto = await db.ortoimagenes.find_one({"id": orto_id}, {"_id": 0})
    if not orto:
        raise HTTPException(status_code=404, detail="Ortoimagen no encontrada")
    
    # Eliminar tiles
    tiles_path = ORTOIMAGENES_PATH / orto_id
    if tiles_path.exists():
        import shutil
        shutil.rmtree(tiles_path)
    
    # Eliminar archivo original
    original_path = Path(orto.get("archivo_original", ""))
    if original_path.exists():
        original_path.unlink()
    
    # Eliminar de MongoDB
    await db.ortoimagenes.delete_one({"id": orto_id})
    
    return {"message": f"Ortoimagen '{orto['nombre']}' eliminada"}

@api_router.get("/ortoimagenes/tiles/{orto_id}/{z}/{x}/{y}.png")
async def servir_tile_ortoimagen(orto_id: str, z: int, x: int, y: int):
    """Sirve un tile específico de una ortoimagen"""
    # Verificar que la ortoimagen existe
    tile_base = ORTOIMAGENES_PATH / orto_id
    if not tile_base.exists():
        raise HTTPException(status_code=404, detail="Ortoimagen no encontrada")
    
    # gdal2tiles genera tiles en formato TMS (y invertida)
    # Convertir de XYZ a TMS: y_tms = 2^z - 1 - y
    y_tms = (2 ** z) - 1 - y
    
    tile_path = tile_base / str(z) / str(x) / f"{y_tms}.png"
    
    if not tile_path.exists():
        # Retornar tile transparente si no existe
        raise HTTPException(status_code=204)  # No content
    
    return FileResponse(
        tile_path, 
        media_type="image/png",
        headers={
            "Cache-Control": "public, max-age=86400",  # Cache 1 día
            "Access-Control-Allow-Origin": "*"
        }
    )


# ===== MÓDULO DE ACTUALIZACIÓN - PROYECTOS =====

# Directorio para archivos de proyectos de actualización
PROYECTOS_ACTUALIZACION_PATH = Path('/app/proyectos_actualizacion')
PROYECTOS_ACTUALIZACION_PATH.mkdir(exist_ok=True)

@api_router.get("/actualizacion/proyectos")
async def listar_proyectos_actualizacion(
    estado: Optional[str] = None,
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista todos los proyectos de actualización"""
    # Admin, Coordinador y Gestor tienen acceso
    tiene_acceso = current_user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]
    
    if not tiene_acceso:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver proyectos de actualización")
    
    query = {}
    if estado:
        query["estado"] = estado
    if municipio:
        query["municipio"] = municipio
    
    proyectos = await db.proyectos_actualizacion.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    return {"proyectos": proyectos}

@api_router.get("/actualizacion/proyectos/estadisticas")
async def estadisticas_proyectos_actualizacion(current_user: dict = Depends(get_current_user)):
    """Obtiene estadísticas generales de proyectos de actualización"""
    # Admin, Coordinador y Gestor siempre tienen acceso
    tiene_acceso = current_user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]
    
    if not tiene_acceso:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver estadísticas")
    
    pipeline = [
        {
            "$group": {
                "_id": "$estado",
                "count": {"$sum": 1}
            }
        }
    ]
    
    resultados = await db.proyectos_actualizacion.aggregate(pipeline).to_list(100)
    
    stats = {
        "activos": 0,
        "pausados": 0,
        "completados": 0,
        "archivados": 0,
        "total": 0
    }
    
    for r in resultados:
        estado = r["_id"]
        count = r["count"]
        stats["total"] += count
        if estado == ProyectoActualizacionEstado.ACTIVO:
            stats["activos"] = count
        elif estado == ProyectoActualizacionEstado.PAUSADO:
            stats["pausados"] = count
        elif estado == ProyectoActualizacionEstado.COMPLETADO:
            stats["completados"] = count
        elif estado == ProyectoActualizacionEstado.ARCHIVADO:
            stats["archivados"] = count
    
    return stats

@api_router.post("/actualizacion/proyectos")
async def crear_proyecto_actualizacion(
    proyecto_data: ProyectoActualizacionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Crea un nuevo proyecto de actualización con etapas predefinidas"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden crear proyectos")
    
    # Verificar que el municipio existe
    municipio = await db.limites_municipales.find_one({"municipio": proyecto_data.municipio}, {"_id": 0})
    if not municipio:
        raise HTTPException(status_code=400, detail=f"Municipio '{proyecto_data.municipio}' no encontrado")
    
    # Verificar que no existe un proyecto activo para el mismo municipio
    proyecto_existente = await db.proyectos_actualizacion.find_one({
        "municipio": proyecto_data.municipio,
        "estado": {"$in": [ProyectoActualizacionEstado.ACTIVO, ProyectoActualizacionEstado.PAUSADO]}
    }, {"_id": 0})
    
    if proyecto_existente:
        raise HTTPException(
            status_code=400, 
            detail=f"Ya existe un proyecto activo o pausado para el municipio '{proyecto_data.municipio}'"
        )
    
    proyecto_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    proyecto = {
        "id": proyecto_id,
        "nombre": proyecto_data.nombre.strip(),
        "municipio": proyecto_data.municipio,
        "descripcion": proyecto_data.descripcion.strip() if proyecto_data.descripcion else None,
        "estado": ProyectoActualizacionEstado.ACTIVO,
        # Archivos unificados
        "base_grafica_archivo": None,
        "base_grafica_cargado_en": None,
        "base_grafica_total_predios": 0,
        "info_alfanumerica_archivo": None,
        "info_alfanumerica_cargado_en": None,
        "info_alfanumerica_total_registros": 0,
        # Fechas
        "fecha_inicio": None,
        "fecha_fin_planificada": None,
        "fecha_fin_real": None,
        # Estadísticas
        "predios_actualizados": 0,
        "predios_no_identificados": 0,
        "creado_por": current_user["id"],
        "creado_por_nombre": current_user["full_name"],
        "created_at": now,
        "updated_at": now
    }
    
    # Crear directorio para archivos del proyecto
    proyecto_dir = PROYECTOS_ACTUALIZACION_PATH / proyecto_id
    proyecto_dir.mkdir(exist_ok=True)
    
    await db.proyectos_actualizacion.insert_one(proyecto)
    
    # Crear las 3 etapas predefinidas
    etapas_default = [
        {
            "id": str(uuid.uuid4()),
            "proyecto_id": proyecto_id,
            "tipo": EtapaProyectoTipo.PREOPERATIVA,
            "nombre": "Etapa Preoperativa",
            "orden": 1,
            "descripcion": "Ejecución de las actividades de alistamiento y planeación del proyecto",
            "fecha_inicio": None,
            "fecha_fin_planificada": None,
            "fecha_fin_real": None,
            "estado": ActividadEstado.PENDIENTE,
            "created_at": now,
            "updated_at": now
        },
        {
            "id": str(uuid.uuid4()),
            "proyecto_id": proyecto_id,
            "tipo": EtapaProyectoTipo.OPERATIVA,
            "nombre": "Etapa Operativa",
            "orden": 2,
            "descripcion": "Implementación de levantamientos prediales y consolidación de componentes",
            "fecha_inicio": None,
            "fecha_fin_planificada": None,
            "fecha_fin_real": None,
            "estado": ActividadEstado.PENDIENTE,
            "created_at": now,
            "updated_at": now
        },
        {
            "id": str(uuid.uuid4()),
            "proyecto_id": proyecto_id,
            "tipo": EtapaProyectoTipo.POSTOPERATIVA,
            "nombre": "Etapa Post-Operativa",
            "orden": 3,
            "descripcion": "Construcción de memoria técnica, liquidación y cierre del proyecto",
            "fecha_inicio": None,
            "fecha_fin_planificada": None,
            "fecha_fin_real": None,
            "estado": ActividadEstado.PENDIENTE,
            "created_at": now,
            "updated_at": now
        }
    ]
    
    await db.etapas_proyecto.insert_many(etapas_default)
    
    # Remover _id de la respuesta
    proyecto.pop("_id", None)
    
    return {"message": "Proyecto creado exitosamente", "proyecto": proyecto, "etapas_creadas": 3}

@api_router.get("/actualizacion/proyectos/{proyecto_id}")
async def obtener_proyecto_actualizacion(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene los detalles de un proyecto de actualización con estadísticas de predios en tiempo real"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver este proyecto")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Calcular estadísticas de predios en tiempo real desde predios_actualizacion
    pipeline = [
        {"$match": {"proyecto_id": proyecto_id}},
        {"$group": {
            "_id": "$estado_visita",
            "count": {"$sum": 1}
        }}
    ]
    stats_cursor = db.predios_actualizacion.aggregate(pipeline)
    stats_list = await stats_cursor.to_list(length=100)
    
    # Calcular totales
    predios_total = 0
    predios_pendientes = 0
    predios_visitados = 0
    predios_actualizados = 0
    
    for stat in stats_list:
        estado = stat.get("_id") or "pendiente"
        count = stat.get("count", 0)
        predios_total += count
        if estado == "pendiente" or estado is None:
            predios_pendientes += count
        elif estado == "visitado":
            predios_visitados += count
        elif estado == "actualizado":
            predios_actualizados += count
    
    # Agregar estadísticas al proyecto
    proyecto["predios_total"] = predios_total
    proyecto["predios_pendientes"] = predios_pendientes
    proyecto["predios_visitados"] = predios_visitados
    proyecto["predios_actualizados"] = predios_actualizados
    
    return proyecto

@api_router.patch("/actualizacion/proyectos/{proyecto_id}")
async def actualizar_proyecto_actualizacion(
    proyecto_id: str,
    update_data: ProyectoActualizacionUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Actualiza un proyecto de actualización"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden actualizar proyectos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # No permitir modificar proyectos archivados
    if proyecto["estado"] == ProyectoActualizacionEstado.ARCHIVADO:
        raise HTTPException(status_code=400, detail="No se puede modificar un proyecto archivado")
    
    updates = {"updated_at": datetime.now(timezone.utc)}
    
    if update_data.nombre is not None:
        updates["nombre"] = update_data.nombre.strip()
    if update_data.descripcion is not None:
        updates["descripcion"] = update_data.descripcion.strip()
    if update_data.estado is not None:
        # Validar el nuevo estado
        valid_states = [
            ProyectoActualizacionEstado.ACTIVO,
            ProyectoActualizacionEstado.PAUSADO,
            ProyectoActualizacionEstado.COMPLETADO,
            ProyectoActualizacionEstado.ARCHIVADO
        ]
        if update_data.estado not in valid_states:
            raise HTTPException(status_code=400, detail=f"Estado inválido. Debe ser uno de: {valid_states}")
        updates["estado"] = update_data.estado
    
    await db.proyectos_actualizacion.update_one(
        {"id": proyecto_id},
        {"$set": updates}
    )
    
    proyecto_actualizado = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    return {"message": "Proyecto actualizado", "proyecto": proyecto_actualizado}

@api_router.delete("/actualizacion/proyectos/{proyecto_id}")
async def eliminar_proyecto_actualizacion(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Elimina un proyecto de actualización y todos sus archivos asociados"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden eliminar proyectos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Eliminar directorio de archivos del proyecto
    proyecto_dir = PROYECTOS_ACTUALIZACION_PATH / proyecto_id
    if proyecto_dir.exists():
        shutil.rmtree(proyecto_dir)
    
    # Eliminar datos del proyecto en MongoDB
    # Colecciones específicas del proyecto
    await db.actualizacion_predios.delete_many({"proyecto_id": proyecto_id})
    await db.actualizacion_r1.delete_many({"proyecto_id": proyecto_id})
    await db.actualizacion_r2.delete_many({"proyecto_id": proyecto_id})
    await db.geometrias_actualizacion.delete_many({"proyecto_id": proyecto_id})
    await db.construcciones_actualizacion.delete_many({"proyecto_id": proyecto_id})
    await db.predios_actualizacion.delete_many({"proyecto_id": proyecto_id})
    await db.etapas_proyecto.delete_many({"proyecto_id": proyecto_id})
    await db.actividades_proyecto.delete_many({"etapa_id": {"$in": [e["id"] for e in await db.etapas_proyecto.find({"proyecto_id": proyecto_id}, {"id": 1}).to_list(100)]}})
    
    # Eliminar el proyecto
    await db.proyectos_actualizacion.delete_one({"id": proyecto_id})
    
    return {"message": f"Proyecto '{proyecto['nombre']}' eliminado exitosamente"}

@api_router.post("/actualizacion/proyectos/{proyecto_id}/archivar")
async def archivar_proyecto_actualizacion(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Archiva un proyecto de actualización"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden archivar proyectos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if proyecto["estado"] == ProyectoActualizacionEstado.ARCHIVADO:
        raise HTTPException(status_code=400, detail="El proyecto ya está archivado")
    
    await db.proyectos_actualizacion.update_one(
        {"id": proyecto_id},
        {"$set": {
            "estado": ProyectoActualizacionEstado.ARCHIVADO,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Proyecto archivado exitosamente"}

@api_router.post("/actualizacion/proyectos/{proyecto_id}/restaurar")
async def restaurar_proyecto_actualizacion(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Restaura un proyecto archivado a estado pausado"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden restaurar proyectos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if proyecto["estado"] != ProyectoActualizacionEstado.ARCHIVADO:
        raise HTTPException(status_code=400, detail="Solo se pueden restaurar proyectos archivados")
    
    await db.proyectos_actualizacion.update_one(
        {"id": proyecto_id},
        {"$set": {
            "estado": ProyectoActualizacionEstado.PAUSADO,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Proyecto restaurado exitosamente"}

@api_router.post("/actualizacion/proyectos/{proyecto_id}/upload-base-grafica")
async def upload_base_grafica_proyecto(
    proyecto_id: str,
    file: UploadFile = File(...),
    modo_carga: Optional[str] = Form("reemplazar"),  # "reemplazar" o "incremental"
    capa_especifica: Optional[str] = Form(None),  # ej: "PERIMETRO_URBANO"
    background_tasks: BackgroundTasks = None,
    current_user: dict = Depends(get_current_user)
):
    """Sube un archivo de Base Gráfica (ZIP con GDB) para un proyecto de actualización y lo procesa
    
    Modos de carga:
    - reemplazar: Elimina todas las geometrías del proyecto antes de cargar (por defecto)
    - incremental: Solo añade/actualiza sin eliminar, ideal para capas adicionales como perímetro urbano
    
    Capa específica:
    - Si se indica, solo se procesará esa capa (ej: PERIMETRO_URBANO, U_MANZANA, etc.)
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para cargar la Base Gráfica")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if proyecto["estado"] == ProyectoActualizacionEstado.ARCHIVADO:
        raise HTTPException(status_code=400, detail="No se pueden cargar archivos a proyectos archivados")
    
    if not file.filename.endswith('.zip'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un ZIP que contenga la geodatabase (.gdb)")
    
    proyecto_dir = PROYECTOS_ACTUALIZACION_PATH / proyecto_id
    proyecto_dir.mkdir(exist_ok=True)
    
    # Guardar archivo con nombre que indique si es carga específica
    suffix = f"_{capa_especifica}" if capa_especifica else ""
    gdb_path = proyecto_dir / f"base_grafica_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}{suffix}.zip"
    
    try:
        with open(gdb_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Actualizar campo base_grafica_archivo ANTES de procesar
        update_data = {
            "base_grafica_archivo": str(gdb_path),
            "base_grafica_cargado_en": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }
        
        # Solo marcar como no procesado si es reemplazo total
        if modo_carga == "reemplazar":
            update_data["gdb_procesado"] = False
        
        await db.proyectos_actualizacion.update_one(
            {"id": proyecto_id},
            {"$set": update_data}
        )
        
        # Procesar GDB con los nuevos parámetros
        await procesar_gdb_actualizacion(proyecto_id, str(gdb_path), proyecto["municipio"], modo_carga, capa_especifica)
        
        mensaje = "Base Gráfica cargada y procesada exitosamente"
        if capa_especifica:
            mensaje = f"Capa '{capa_especifica}' cargada exitosamente (modo {modo_carga})"
        
        return {
            "message": mensaje,
            "archivo": str(gdb_path),
            "nombre_archivo": file.filename,
            "modo_carga": modo_carga,
            "capa_especifica": capa_especifica
        }
        
    except Exception as e:
        logger.error(f"Error al cargar Base Gráfica: {str(e)}")
        if gdb_path.exists():
            gdb_path.unlink()
        raise HTTPException(status_code=500, detail=f"Error al cargar el archivo: {str(e)}")


# ==================== ORTOFOTO ====================

@api_router.post("/actualizacion/proyectos/{proyecto_id}/ortofoto")
async def upload_ortofoto(
    proyecto_id: str,
    ortofoto: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Sube una ortofoto para el proyecto de actualización"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden subir ortofotos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Validar tipo de archivo
    allowed_extensions = ['.tif', '.tiff', '.png', '.jpg', '.jpeg']
    filename = ortofoto.filename.lower()
    if not any(filename.endswith(ext) for ext in allowed_extensions):
        raise HTTPException(status_code=400, detail="Formato no válido. Use TIFF, PNG o JPG")
    
    proyecto_dir = PROYECTOS_ACTUALIZACION_PATH / proyecto_id
    proyecto_dir.mkdir(exist_ok=True)
    
    # Guardar ortofoto
    extension = Path(ortofoto.filename).suffix.lower()
    ortofoto_path = proyecto_dir / f"ortofoto{extension}"
    
    try:
        content = await ortofoto.read()
        with open(ortofoto_path, "wb") as buffer:
            buffer.write(content)
        
        # Para archivos TIFF con georeferencia, extraer bounds
        bounds = None
        
        if extension in ['.tif', '.tiff']:
            try:
                import rasterio
                with rasterio.open(str(ortofoto_path)) as src:
                    # Obtener bounds del raster
                    bounds_raw = src.bounds
                    # Convertir a lat/lng si es necesario
                    if src.crs:
                        from pyproj import Transformer
                        transformer = Transformer.from_crs(src.crs, "EPSG:4326", always_xy=True)
                        
                        west, south = transformer.transform(bounds_raw.left, bounds_raw.bottom)
                        east, north = transformer.transform(bounds_raw.right, bounds_raw.top)
                        
                        bounds = {
                            "north": north,
                            "south": south,
                            "east": east,
                            "west": west
                        }
            except Exception as e:
                print(f"Error leyendo bounds de TIFF: {e}")
                # Si no se puede leer, usar bounds del GDB si existe
                if proyecto.get('gdb_bounds'):
                    bounds = proyecto['gdb_bounds']
        
        # Si no hay bounds, intentar usar los del GDB o valores por defecto del municipio
        if not bounds:
            if proyecto.get('gdb_bounds'):
                bounds = proyecto['gdb_bounds']
            else:
                # Intentar obtener bounds de las geometrías guardadas
                geometrias = await db.geometrias_actualizacion.find_one({"proyecto_id": proyecto_id})
                if geometrias and geometrias.get('bounds'):
                    bounds = geometrias['bounds']
                else:
                    # Valores aproximados de Norte de Santander
                    bounds = {
                        "north": 8.5,
                        "south": 7.0,
                        "east": -72.0,
                        "west": -73.5
                    }
        
        # Actualizar proyecto con info de ortofoto
        await db.proyectos_actualizacion.update_one(
            {"id": proyecto_id},
            {
                "$set": {
                    "ortofoto_archivo": str(ortofoto_path),
                    "ortofoto_nombre": ortofoto.filename,
                    "ortofoto_bounds": bounds,
                    "ortofoto_fecha": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        # Generar URL para el frontend
        ortofoto_url = f"/api/actualizacion/proyectos/{proyecto_id}/ortofoto/file"
        
        return {
            "message": "Ortofoto cargada exitosamente",
            "url": ortofoto_url,
            "bounds": bounds,
            "filename": ortofoto.filename
        }
        
    except Exception as e:
        if ortofoto_path.exists():
            ortofoto_path.unlink()
        raise HTTPException(status_code=500, detail=f"Error al cargar la ortofoto: {str(e)}")


@api_router.get("/actualizacion/proyectos/{proyecto_id}/ortofoto")
async def get_ortofoto_info(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene información de la ortofoto del proyecto"""
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if not proyecto.get('ortofoto_archivo'):
        raise HTTPException(status_code=404, detail="No hay ortofoto cargada")
    
    return {
        "url": f"/api/actualizacion/proyectos/{proyecto_id}/ortofoto/file",
        "bounds": proyecto.get('ortofoto_bounds'),
        "filename": proyecto.get('ortofoto_nombre'),
        "fecha": proyecto.get('ortofoto_fecha')
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/ortofoto/file")
async def get_ortofoto_file(proyecto_id: str):
    """Sirve el archivo de ortofoto"""
    from fastapi.responses import FileResponse
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id})
    if not proyecto or not proyecto.get('ortofoto_archivo'):
        raise HTTPException(status_code=404, detail="Ortofoto no encontrada")
    
    ortofoto_path = Path(proyecto['ortofoto_archivo'])
    if not ortofoto_path.exists():
        raise HTTPException(status_code=404, detail="Archivo de ortofoto no encontrado")
    
    # Determinar tipo MIME
    extension = ortofoto_path.suffix.lower()
    media_types = {
        '.tif': 'image/tiff',
        '.tiff': 'image/tiff',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg'
    }
    media_type = media_types.get(extension, 'application/octet-stream')
    
    return FileResponse(str(ortofoto_path), media_type=media_type)


async def procesar_gdb_actualizacion(proyecto_id: str, zip_path: str, municipio: str, modo_carga: str = "reemplazar", capa_especifica: Optional[str] = None):
    """Procesa el GDB de un proyecto de actualización y guarda las geometrías
    
    Args:
        modo_carga: "reemplazar" elimina todo antes, "incremental" solo añade/actualiza
        capa_especifica: Si se indica, solo procesa esa capa
    """
    import zipfile
    import tempfile
    import shutil
    
    logger.info(f"[GDB Actualizacion] Iniciando procesamiento para proyecto {proyecto_id}")
    logger.info(f"[GDB Actualizacion] Archivo: {zip_path}, Municipio: {municipio}")
    logger.info(f"[GDB Actualizacion] Modo: {modo_carga}, Capa específica: {capa_especifica}")
    
    temp_dir = tempfile.mkdtemp()
    try:
        # Extraer ZIP
        logger.info(f"[GDB Actualizacion] Extrayendo ZIP...")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Buscar .gdb o .shp
        gdb_path = None
        shp_files = []
        is_shapefile = False
        
        for root, dirs, files in os.walk(temp_dir):
            # Primero buscar .gdb (tiene prioridad)
            for d in dirs:
                if d.endswith('.gdb'):
                    gdb_path = os.path.join(root, d)
                    break
            if gdb_path:
                break
            
            # Si no hay .gdb, buscar archivos .shp
            for f in files:
                if f.endswith('.shp'):
                    shp_files.append(os.path.join(root, f))
        
        # Determinar qué tipo de archivo procesar
        if gdb_path:
            logger.info(f"[GDB Actualizacion] GDB encontrado: {gdb_path}")
            data_source = gdb_path
            is_shapefile = False
        elif shp_files:
            logger.info(f"[GDB Actualizacion] Shapefiles encontrados: {len(shp_files)}")
            for shp in shp_files:
                logger.info(f"  - {os.path.basename(shp)}")
            data_source = shp_files  # Lista de shapefiles
            is_shapefile = True
        else:
            logger.error("[GDB Actualizacion] No se encontró archivo .gdb ni .shp en el ZIP")
            raise Exception("No se encontró archivo .gdb ni .shp en el ZIP. Asegúrese de que el ZIP contenga una geodatabase (.gdb) o archivos shapefile (.shp)")
        
        
        # Leer capas según el tipo de archivo
        import pyogrio
        
        if is_shapefile:
            # Para Shapefiles, cada archivo .shp es una "capa"
            layer_names = [os.path.splitext(os.path.basename(shp))[0] for shp in shp_files]
            logger.info(f"[GDB/SHP Actualizacion] Shapefiles como capas: {layer_names}")
        else:
            # Para GDB, usar pyogrio para listar capas
            layers = pyogrio.list_layers(data_source)
            layer_names = [l[0] for l in layers]
            logger.info(f"[GDB Actualizacion] Capas encontradas: {len(layer_names)}")
        
        # Si se especificó una capa específica, verificar que exista
        if capa_especifica:
            if capa_especifica not in layer_names:
                # Buscar con case-insensitive
                capa_encontrada = None
                for ln in layer_names:
                    if ln.upper() == capa_especifica.upper():
                        capa_encontrada = ln
                        break
                if not capa_encontrada:
                    raise Exception(f"Capa '{capa_especifica}' no encontrada. Capas disponibles: {', '.join(layer_names[:15])}")
                capa_especifica = capa_encontrada
            logger.info(f"[GDB/SHP Actualizacion] Procesando solo capa específica: {capa_especifica}")
        
        # Buscar capas de terreno (usar mismos estándares que Conservación)
        rural_layers = ['R_TERRENO', 'R_TERRENO_1', 'r_terreno', 'TERRENO', 'terreno_rural', 'RURAL']
        urban_layers = ['U_TERRENO', 'U_TERRENO_1', 'u_terreno', 'terreno_urbano', 'URBANO']
        construccion_layers = ['U_CONSTRUCCION', 'R_CONSTRUCCION', 'CONSTRUCCION', 'U_UNIDAD', 'R_UNIDAD']
        perimetro_layers = ['PERIMETRO_URBANO', 'PERIMETRO', 'LIMITE_URBANO', 'perimetro_urbano']
        
        # Verificar qué capas están presentes
        rural_found = [l for l in layer_names if any(r.upper() in l.upper() for r in rural_layers)]
        urban_found = [l for l in layer_names if any(u.upper() in l.upper() for u in urban_layers)]
        perimetro_found = [l for l in layer_names if any(p.upper() in l.upper() for p in perimetro_layers)]
        
        logger.info(f"[GDB/SHP Actualizacion] Capas rurales detectadas: {rural_found}")
        logger.info(f"[GDB/SHP Actualizacion] Capas urbanas detectadas: {urban_found}")
        logger.info(f"[GDB/SHP Actualizacion] Capas perímetro detectadas: {perimetro_found}")
        
        geometrias_guardadas = 0
        construcciones_guardadas = 0
        
        # Lógica de eliminación según modo de carga
        if modo_carga == "reemplazar":
            # Eliminar TODAS las geometrías anteriores del proyecto
            logger.info(f"[GDB Actualizacion] Modo REEMPLAZAR: Eliminando todas las geometrías anteriores...")
            deleted_geom = await db.geometrias_actualizacion.delete_many({"proyecto_id": proyecto_id})
            deleted_const = await db.construcciones_actualizacion.delete_many({"proyecto_id": proyecto_id})
            logger.info(f"[GDB Actualizacion] Eliminadas {deleted_geom.deleted_count} geometrías y {deleted_const.deleted_count} construcciones")
        elif modo_carga == "incremental" and capa_especifica:
            # Solo eliminar geometrías de la capa específica
            logger.info(f"[GDB Actualizacion] Modo INCREMENTAL: Eliminando solo geometrías de capa '{capa_especifica}'...")
            deleted = await db.geometrias_actualizacion.delete_many({
                "proyecto_id": proyecto_id,
                "capa_origen": capa_especifica
            })
            logger.info(f"[GDB Actualizacion] Eliminadas {deleted.deleted_count} geometrías de capa '{capa_especifica}'")
        else:
            # Incremental sin capa específica: no eliminar nada
            logger.info(f"[GDB Actualizacion] Modo INCREMENTAL: Añadiendo sin eliminar...")
        
        # Si hay capa específica, procesar solo esa capa
        if capa_especifica:
            try:
                logger.info(f"[GDB/SHP Actualizacion] Procesando capa específica: {capa_especifica}")
                
                # Determinar la fuente de datos según el tipo
                if is_shapefile:
                    # Buscar el shapefile que coincide con la capa específica
                    shp_path = None
                    for shp in shp_files:
                        if os.path.splitext(os.path.basename(shp))[0].upper() == capa_especifica.upper():
                            shp_path = shp
                            break
                    if not shp_path:
                        raise Exception(f"Shapefile '{capa_especifica}' no encontrado")
                    gdf = pyogrio.read_dataframe(shp_path)
                else:
                    gdf = pyogrio.read_dataframe(data_source, layer=capa_especifica)
                
                if len(gdf) > 0:
                    gdf = gdf.to_crs(epsg=4326)
                    logger.info(f"[GDB/SHP Actualizacion] Capa '{capa_especifica}' tiene {len(gdf)} registros")
                    
                    # Determinar tipo de zona basado en el nombre de la capa
                    tipo_zona = "otro"
                    if capa_especifica.upper().startswith("R_") or "RURAL" in capa_especifica.upper():
                        tipo_zona = "rural"
                    elif capa_especifica.upper().startswith("U_") or "URBAN" in capa_especifica.upper():
                        tipo_zona = "urbano"
                    elif "PERIMETRO" in capa_especifica.upper():
                        tipo_zona = "perimetro_urbano"
                    elif "MANZANA" in capa_especifica.upper():
                        tipo_zona = "manzana"
                    elif "SECTOR" in capa_especifica.upper():
                        tipo_zona = "sector"
                    
                    for idx, row in gdf.iterrows():
                        geom = row.geometry.__geo_interface__
                        props = {k: (str(v) if v is not None else None) for k, v in row.items() if k != 'geometry'}
                        props['zona'] = tipo_zona
                        props['proyecto_id'] = proyecto_id
                        props['municipio'] = municipio
                        props['capa_origen'] = capa_especifica
                        
                        # Obtener código si existe
                        codigo = props.get('CODIGO', props.get('codigo', props.get('NUMERO_PREDIAL', f"{capa_especifica}_{idx}")))
                        
                        await db.geometrias_actualizacion.insert_one({
                            "proyecto_id": proyecto_id,
                            "municipio": municipio,
                            "zona": tipo_zona,
                            "capa_origen": capa_especifica,
                            "codigo_predial": codigo,
                            "numero_predial": props.get('NUMERO_PREDIAL', props.get('numero_predial', '')),
                            "geometry": geom,
                            "properties": props,
                            "created_at": datetime.now(timezone.utc)
                        })
                        geometrias_guardadas += 1
                    
                    logger.info(f"[GDB Actualizacion] Guardadas {geometrias_guardadas} geometrías de capa '{capa_especifica}'")
                else:
                    logger.warning(f"[GDB Actualizacion] Capa '{capa_especifica}' está vacía")
                    
            except Exception as e:
                logger.error(f"[GDB Actualizacion] Error procesando capa específica '{capa_especifica}': {e}")
                raise
            
            # Actualizar proyecto y finalizar
            await db.proyectos_actualizacion.update_one(
                {"id": proyecto_id},
                {"$set": {
                    "gdb_procesado": True,
                    "gdb_total_geometrias": geometrias_guardadas,
                    f"capa_{capa_especifica.lower()}_cargada": True,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            logger.info(f"[GDB/SHP Actualizacion] Carga de capa '{capa_especifica}' completada: {geometrias_guardadas} geometrías")
            return
        
        # Función auxiliar para leer un dataframe según el tipo de fuente
        def read_layer(layer_name):
            if is_shapefile:
                # Buscar el shapefile que coincide
                for shp in shp_files:
                    if os.path.splitext(os.path.basename(shp))[0].upper() == layer_name.upper():
                        return pyogrio.read_dataframe(shp)
                return None
            else:
                return pyogrio.read_dataframe(data_source, layer=layer_name)
        
        # Procesar capas estándar (comportamiento original)
        # Procesar terrenos rurales
        for layer_name in rural_found:
            try:
                logger.info(f"[GDB/SHP Actualizacion] Procesando capa rural: {layer_name}")
                gdf = read_layer(layer_name)
                if gdf is not None and len(gdf) > 0:
                    gdf = gdf.to_crs(epsg=4326)
                    for idx, row in gdf.iterrows():
                        geom = row.geometry.__geo_interface__
                        props = {k: (str(v) if v is not None else None) for k, v in row.items() if k != 'geometry'}
                        props['zona'] = 'rural'
                        props['proyecto_id'] = proyecto_id
                        props['municipio'] = municipio
                        props['capa_origen'] = layer_name
                        
                        await db.geometrias_actualizacion.insert_one({
                            "proyecto_id": proyecto_id,
                            "municipio": municipio,
                            "zona": "rural",
                            "capa_origen": layer_name,
                            "codigo_predial": props.get('CODIGO', props.get('codigo', props.get('NUMERO_PREDIAL', ''))),
                            "numero_predial": props.get('NUMERO_PREDIAL', props.get('numero_predial', '')),
                            "geometry": geom,
                            "properties": props,
                            "created_at": datetime.now(timezone.utc)
                        })
                        geometrias_guardadas += 1
            except Exception as e:
                logger.error(f"[GDB/SHP Actualizacion] Error procesando capa rural {layer_name}: {e}")
        
        # Procesar terrenos urbanos
        for layer_name in urban_found:
            try:
                logger.info(f"[GDB/SHP Actualizacion] Procesando capa urbana: {layer_name}")
                gdf = read_layer(layer_name)
                if gdf is not None and len(gdf) > 0:
                    logger.info(f"[GDB/SHP Actualizacion] Convirtiendo CRS de {len(gdf)} registros...")
                    gdf = gdf.to_crs(epsg=4326)
                    for idx, row in gdf.iterrows():
                        geom = row.geometry.__geo_interface__
                        props = {k: (str(v) if v is not None else None) for k, v in row.items() if k != 'geometry'}
                        props['zona'] = 'urbano'
                        props['proyecto_id'] = proyecto_id
                        props['municipio'] = municipio
                        props['capa_origen'] = layer_name
                        
                        await db.geometrias_actualizacion.insert_one({
                            "proyecto_id": proyecto_id,
                            "municipio": municipio,
                            "zona": "urbano",
                            "capa_origen": layer_name,
                            "codigo_predial": props.get('CODIGO', props.get('codigo', props.get('NUMERO_PREDIAL', ''))),
                            "numero_predial": props.get('NUMERO_PREDIAL', props.get('numero_predial', '')),
                            "geometry": geom,
                            "properties": props,
                            "created_at": datetime.now(timezone.utc)
                        })
                        geometrias_guardadas += 1
            except Exception as e:
                logger.error(f"[GDB/SHP Actualizacion] Error procesando capa urbana {layer_name}: {e}")
        
        # Procesar capas de perímetro si existen
        for layer_name in perimetro_found:
            try:
                logger.info(f"[GDB/SHP Actualizacion] Procesando capa perímetro: {layer_name}")
                gdf = read_layer(layer_name)
                if gdf is not None and len(gdf) > 0:
                    gdf = gdf.to_crs(epsg=4326)
                    for idx, row in gdf.iterrows():
                        geom = row.geometry.__geo_interface__
                        props = {k: (str(v) if v is not None else None) for k, v in row.items() if k != 'geometry'}
                        props['zona'] = 'perimetro_urbano'
                        props['proyecto_id'] = proyecto_id
                        props['municipio'] = municipio
                        props['capa_origen'] = layer_name
                        
                        await db.geometrias_actualizacion.insert_one({
                            "proyecto_id": proyecto_id,
                            "municipio": municipio,
                            "zona": "perimetro_urbano",
                            "capa_origen": layer_name,
                            "codigo_predial": props.get('CODIGO', props.get('codigo', f"PERIMETRO_{idx}")),
                            "numero_predial": "",
                            "geometry": geom,
                            "properties": props,
                            "created_at": datetime.now(timezone.utc)
                        })
                        geometrias_guardadas += 1
            except Exception as e:
                logger.error(f"[GDB/SHP Actualizacion] Error procesando capa perímetro {layer_name}: {e}")
        
        logger.info(f"[GDB/SHP Actualizacion] Geometrías guardadas hasta ahora: {geometrias_guardadas}")
        
        # Procesar construcciones (solo para GDB, shapefiles de construcción se procesan si existen)
        construccion_found = [l for l in layer_names if any(c.upper() in l.upper() for c in construccion_layers)]
        for layer_name in construccion_found:
            try:
                logger.info(f"[GDB/SHP Actualizacion] Procesando construcciones: {layer_name}")
                gdf = read_layer(layer_name)
                if gdf is not None and len(gdf) > 0:
                    gdf = gdf.to_crs(epsg=4326)
                    for idx, row in gdf.iterrows():
                        geom = row.geometry.__geo_interface__
                        props = {k: (str(v) if v is not None else None) for k, v in row.items() if k != 'geometry'}
                        
                        await db.construcciones_actualizacion.insert_one({
                            "proyecto_id": proyecto_id,
                            "municipio": municipio,
                            "capa_origen": layer_name,
                            "codigo_predial": props.get('CODIGO', props.get('codigo', '')),
                            "geometry": geom,
                            "properties": props,
                            "created_at": datetime.now(timezone.utc)
                        })
                        construcciones_guardadas += 1
            except Exception as e:
                logger.error(f"[GDB/SHP Actualizacion] Error procesando construcciones {layer_name}: {e}")
                import traceback
                logger.error(f"[GDB/SHP Actualizacion] Traceback: {traceback.format_exc()}")
        
        logger.info(f"[GDB/SHP Actualizacion] Total construcciones guardadas: {construcciones_guardadas}")
        
        # Actualizar proyecto con información detallada de capas procesadas
        capas_info = {
            "rural_encontradas": rural_found,
            "urbano_encontradas": urban_found,
            "perimetro_encontradas": perimetro_found,
            "construccion_encontradas": construccion_found,
            "es_shapefile": is_shapefile
        }
        
        await db.proyectos_actualizacion.update_one(
            {"id": proyecto_id},
            {"$set": {
                "gdb_procesado": True,
                "base_grafica_total_predios": geometrias_guardadas,
                "total_construcciones": construcciones_guardadas,
                "gdb_procesado_en": datetime.now(timezone.utc),
                "capas_procesadas": capas_info,
                "tiene_zona_rural": len(capas_info["rural_encontradas"]) > 0,
                "tiene_zona_urbana": len(capas_info["urbano_encontradas"]) > 0,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        logger.info(f"[GDB Actualizacion] ✅ COMPLETADO: {geometrias_guardadas} geometrías, {construcciones_guardadas} construcciones")
        logger.info(f"[GDB Actualizacion] Capas - Rural: {capas_info['rural_encontradas']}, Urbano: {capas_info['urbano_encontradas']}")
        
    except Exception as e:
        logger.error(f"[GDB Actualizacion] ❌ ERROR GENERAL: {str(e)}")
        import traceback
        logger.error(f"[GDB Actualizacion] Traceback: {traceback.format_exc()}")
        raise
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@api_router.get("/actualizacion/proyectos/{proyecto_id}/geometrias/info")
async def get_geometrias_info(
    proyecto_id: str,
    zona: str = Query(None, description="Filtrar por zona: urbano, rural"),
    current_user: dict = Depends(get_current_user)
):
    """Obtiene información sobre las geometrías del proyecto (para descarga progresiva)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {"proyecto_id": proyecto_id}
    if zona:
        query["zona"] = zona
    
    total = await db.geometrias_actualizacion.count_documents(query)
    total_construcciones = await db.construcciones_actualizacion.count_documents({"proyecto_id": proyecto_id})
    
    return {
        "total": total,
        "total_construcciones": total_construcciones,
        "proyecto_id": proyecto_id
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/geometrias")
async def get_geometrias_proyecto(
    proyecto_id: str,
    zona: str = Query(None, description="Filtrar por zona: urbano, rural"),
    offset: int = Query(0, description="Offset para paginación"),
    limit: int = Query(50000, description="Límite de resultados"),
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las geometrías procesadas del proyecto para el visor"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver geometrías")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Query de geometrías con paginación
    query = {"proyecto_id": proyecto_id}
    if zona:
        query["zona"] = zona
    
    geometrias = await db.geometrias_actualizacion.find(query, {"_id": 0}).skip(offset).limit(limit).to_list(limit)
    
    # Solo cargar construcciones en la primera página (offset=0)
    construcciones = []
    if offset == 0:
        construcciones = await db.construcciones_actualizacion.find({"proyecto_id": proyecto_id}, {"_id": 0}).to_list(50000)
    
    # Convertir a GeoJSON FeatureCollection
    geom_features = []
    for g in geometrias:
        geom_features.append({
            "type": "Feature",
            "geometry": g.get("geometry"),
            "properties": {
                "codigo_predial": g.get("codigo_predial"),
                "numero_predial": g.get("numero_predial"),
                "zona": g.get("zona"),
                **g.get("properties", {})
            }
        })
    
    constr_features = []
    for c in construcciones:
        constr_features.append({
            "type": "Feature",
            "geometry": c.get("geometry"),
            "properties": c.get("properties", {})
        })
    
    return {
        "geometrias": {
            "type": "FeatureCollection",
            "features": geom_features
        },
        "construcciones": {
            "type": "FeatureCollection",
            "features": constr_features
        },
        "total_geometrias": len(geom_features),
        "total_construcciones": len(constr_features),
        "offset": offset,
        "limit": limit
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/predios")
async def get_predios_proyecto(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene los predios R1/R2 cargados para el proyecto"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver predios")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Obtener predios del proyecto
    predios = await db.predios_actualizacion.find(
        {"proyecto_id": proyecto_id},
        {"_id": 0}
    ).to_list(50000)
    
    # Enriquecer con datos de la colección principal de predios
    if predios:
        codigos_prediales = [p.get('codigo_predial') for p in predios if p.get('codigo_predial')]
        # Buscar datos adicionales en la colección principal (puede haber múltiples vigencias)
        predios_principales = await db.predios.find(
            {"codigo_predial_nacional": {"$in": codigos_prediales}},
            {"_id": 0, "codigo_predial_nacional": 1, "codigo_homologado": 1, "r2_registros": 1}
        ).to_list(None)  # Sin límite para asegurar que traiga todos
        
        # Crear mapas de código -> datos (toma el primero encontrado)
        codigo_homologado_map = {}
        matricula_map = {}
        for p in predios_principales:
            cpn = p.get('codigo_predial_nacional')
            if cpn and cpn not in codigo_homologado_map:
                ch = p.get('codigo_homologado')
                if ch:
                    codigo_homologado_map[cpn] = ch
                # Obtener matrícula del r2_registros
                r2_registros = p.get('r2_registros', [])
                if r2_registros and len(r2_registros) > 0:
                    matricula = r2_registros[0].get('matricula_inmobiliaria')
                    if matricula:
                        matricula_map[cpn] = matricula
        
        # Agregar datos a cada predio
        for predio in predios:
            codigo = predio.get('codigo_predial')
            if codigo:
                if codigo in codigo_homologado_map:
                    predio['codigo_homologado'] = codigo_homologado_map[codigo]
                if codigo in matricula_map:
                    predio['matricula_inmobiliaria'] = matricula_map[codigo]
    
    return {
        "predios": predios,
        "total": len(predios)
    }



@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios")
async def crear_predio_actualizacion(
    proyecto_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Crea un nuevo predio dentro de un proyecto de actualización.
    El predio queda en predios_actualizacion para ser procesado durante la actualización.
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para crear predios")
    
    # Verificar que el proyecto existe y está activo
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if proyecto.get('estado') not in [ProyectoActualizacionEstado.ACTIVO, ProyectoActualizacionEstado.PAUSADO]:
        raise HTTPException(status_code=400, detail="El proyecto no está activo")
    
    municipio = proyecto.get('municipio')
    codigo_predial = data.get('codigo_predial', '')
    
    if not codigo_predial:
        raise HTTPException(status_code=400, detail="El código predial es requerido")
    
    # Verificar que no exista en el proyecto actual
    existing = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un predio con este código en el proyecto")
    
    # Crear el predio
    predio_id = str(uuid.uuid4())
    ahora = datetime.now(timezone.utc).isoformat()
    
    predio_nuevo = {
        "id": predio_id,
        "proyecto_id": proyecto_id,
        "codigo_predial": codigo_predial,
        "codigo_predial_nacional": codigo_predial,
        "numero_predial": codigo_predial,
        "municipio": municipio,
        "es_nuevo": True,
        
        # Datos R1 - Jurídicos
        "direccion": data.get('direccion', ''),
        "comuna": data.get('comuna', ''),
        "destino_economico": data.get('destino_economico', ''),
        "area_terreno": float(data.get('area_terreno') or 0),
        "area_construida": float(data.get('area_construida') or 0),
        "avaluo": float(data.get('avaluo_catastral') or data.get('avaluo') or 0),
        "avaluo_catastral": float(data.get('avaluo_catastral') or data.get('avaluo') or 0),
        "matricula_inmobiliaria": data.get('matricula_inmobiliaria', ''),
        "codigo_homologado": data.get('codigo_homologado', ''),
        
        # Propietarios con estado civil
        "propietarios": data.get('propietarios', []),
        
        # Datos R2 - Físicos
        "zonas_fisicas": data.get('zonas_fisicas', []),
        "habitaciones": data.get('habitaciones', ''),
        "banos": data.get('banos', ''),
        "locales": data.get('locales', ''),
        "pisos": data.get('pisos', ''),
        "uso": data.get('uso', ''),
        
        # Estado
        "estado_visita": "pendiente",
        
        # Metadatos
        "creado_por": current_user.get('email'),
        "creado_por_nombre": current_user.get('full_name'),
        "created_at": ahora,
        "updated_at": ahora,
        
        # Historial
        "historial_cambios": [{
            "fecha": ahora,
            "usuario": current_user.get('email'),
            "rol": current_user['role'],
            "accion": "creacion",
            "campos_modificados": list(data.keys())
        }]
    }
    
    await db.predios_actualizacion.insert_one(predio_nuevo)
    
    return {
        "message": "Predio creado correctamente",
        "id": predio_id,
        "codigo_predial": codigo_predial
    }


@api_router.patch("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}")
async def actualizar_predio_proyecto(
    proyecto_id: str,
    codigo_predial: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Actualiza un predio del proyecto (trabajo de campo)
    
    FLUJO DE ACTUALIZACIÓN:
    1. Gestor visita el predio → completa formulario de visita → estado "visitado"
    2. Gestor propone cambios a datos prediales → va a "Gestión de Propuestas"
    3. Coordinador aprueba la propuesta → cambios se aplican → estado "actualizado"
    
    - Gestores SOLO pueden: marcar visita, agregar datos de visita, completar formulario
    - Cambios a datos prediales (dirección, áreas, propietarios, etc.) 
      requieren crear una PROPUESTA que el coordinador debe aprobar
    - Todo queda registrado en historial
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para actualizar predios")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Buscar el predio por codigo_predial o numero_predial
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    })
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    es_coordinador = current_user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]
    
    # Campos que gestores pueden modificar directamente (datos de visita)
    campos_visita = [
        'estado_visita', 'observaciones_campo', 'ubicacion_gps',
        'visitado_por', 'visitado_en', 'visitado_por_nombre',
        'visita', 'datos_notificacion', 'fotos',
        'sin_cambios'  # Marcar si el predio fue visitado sin modificaciones
    ]
    
    # Campos que requieren aprobación del coordinador (datos prediales)
    campos_prediales = [
        'direccion', 'destino_economico', 'area_terreno', 'area_construida',
        'matricula_inmobiliaria', 'avaluo_catastral', 'estrato', 'comuna',
        'propietarios', 'zonas_fisicas', 'zonas_terreno', 'construcciones',
        'actualizado_por', 'actualizado_en', 'actualizado_por_nombre',
        'informacion_construcciones', 'calificacion_construccion', 
        'resumen_areas', 'es_ph', 'datos_ph', 'datos_condominio',
        'avaluo', 'codigo_homologado', 'habitaciones', 'banos', 'locales', 'pisos', 'uso',
        'codigo_predial'
    ]
    
    # Verificar si el gestor está intentando modificar datos prediales
    if not es_coordinador:
        campos_prediales_modificados = [k for k in data.keys() if k in campos_prediales]
        
        # Gestores solo pueden marcar como 'visitado', no como 'actualizado'
        if data.get('estado_visita') == 'actualizado':
            raise HTTPException(
                status_code=403, 
                detail="Los gestores no pueden marcar como 'actualizado' directamente. Debe crear una propuesta de cambio que será revisada por el coordinador."
            )
        
        # Si intenta modificar campos prediales, debe crear propuesta
        if campos_prediales_modificados:
            raise HTTPException(
                status_code=403, 
                detail=f"Los cambios a datos prediales ({', '.join(campos_prediales_modificados)}) requieren crear una propuesta de cambio. El predio debe estar visitado y luego usar 'Proponer Cambios' para enviar una solicitud al coordinador."
            )
        
        # Solo permitir campos de visita para gestores
        update_data = {k: v for k, v in data.items() if k in campos_visita}
        accion = "visita_registrada"
        
        # Agregar información del gestor que visitó
        if data.get('estado_visita') == 'visitado':
            update_data['visitado_por'] = current_user.get('email')
            update_data['visitado_por_nombre'] = current_user.get('full_name')
            update_data['visitado_en'] = datetime.now(timezone.utc).isoformat()
    else:
        # Coordinadores pueden modificar todo
        campos_permitidos = campos_visita + campos_prediales
        update_data = {k: v for k, v in data.items() if k in campos_permitidos}
        accion = "actualizacion_directa" if data.get('estado_visita') == 'actualizado' else "visita"
        
        # Si coordinador marca como actualizado, registrar quién aprobó
        if data.get('estado_visita') == 'actualizado':
            update_data['actualizado_por'] = current_user.get('email')
            update_data['actualizado_por_nombre'] = current_user.get('full_name')
            update_data['actualizado_en'] = datetime.now(timezone.utc).isoformat()
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No hay campos válidos para actualizar")
    
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    # Agregar al historial de cambios con detalle completo
    historial_entry = {
        "fecha": datetime.now(timezone.utc).isoformat(),
        "usuario": current_user.get('email'),
        "usuario_nombre": current_user.get('full_name'),
        "rol": current_user['role'],
        "accion": accion,
        "campos_modificados": list(update_data.keys()),
        "detalle": f"{'Coordinador' if es_coordinador else 'Gestor'} {current_user.get('full_name')} - {accion}"
    }
    
    await db.predios_actualizacion.update_one(
        {"_id": predio["_id"]},
        {
            "$set": update_data,
            "$push": {"historial_cambios": historial_entry}
        }
    )
    
    return {
        "message": "Predio actualizado exitosamente",
        "codigo_predial": codigo_predial
    }


# ==================== PREDIOS SIN CAMBIOS - ACTUALIZACIÓN ====================

@api_router.get("/actualizacion/proyectos/{proyecto_id}/predios-sin-cambios")
async def get_predios_sin_cambios(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene los predios visitados marcados como 'sin cambios' que esperan aprobación del coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden ver esta lista")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Obtener predios visitados marcados como sin_cambios que NO están actualizados
    predios = await db.predios_actualizacion.find(
        {
            "proyecto_id": proyecto_id,
            "sin_cambios": True,
            "estado_visita": {"$in": ["visitado", "pendiente"]},  # No incluir los ya actualizados
        },
        {"_id": 0}
    ).sort("visitado_en", -1).to_list(10000)
    
    return {
        "predios": predios,
        "total": len(predios)
    }


@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/aprobar-sin-cambios")
async def aprobar_predio_sin_cambios(
    proyecto_id: str,
    codigo_predial: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Aprueba un predio visitado sin cambios, marcándolo como actualizado"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden aprobar")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Buscar el predio
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    })
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Agregar al historial
    historial_entry = {
        "fecha": datetime.now(timezone.utc).isoformat(),
        "usuario": current_user.get('email'),
        "rol": current_user['role'],
        "accion": "aprobado_sin_cambios",
        "comentario": data.get('comentario', 'Aprobado sin cambios')
    }
    
    # Actualizar el predio a estado 'actualizado'
    await db.predios_actualizacion.update_one(
        {"_id": predio["_id"]},
        {
            "$set": {
                "estado_visita": "actualizado",
                "aprobado_sin_cambios": True,
                "aprobado_por": current_user.get('email'),
                "aprobado_en": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc)
            },
            "$push": {"historial_cambios": historial_entry}
        }
    )
    
    return {
        "message": "Predio aprobado como actualizado (sin cambios)",
        "codigo_predial": codigo_predial
    }


@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios-sin-cambios/aprobar-masivo")
async def aprobar_masivo_sin_cambios(
    proyecto_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Aprueba masivamente predios visitados sin cambios"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden aprobar")
    
    codigos = data.get('codigos_prediales', [])
    if not codigos:
        raise HTTPException(status_code=400, detail="Debe proporcionar una lista de códigos prediales")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    comentario = data.get('comentario', 'Aprobación masiva sin cambios')
    aprobados = 0
    
    for codigo in codigos:
        predio = await db.predios_actualizacion.find_one({
            "proyecto_id": proyecto_id,
            "$or": [
                {"codigo_predial": codigo},
                {"numero_predial": codigo}
            ]
        })
        
        if predio:
            historial_entry = {
                "fecha": datetime.now(timezone.utc).isoformat(),
                "usuario": current_user.get('email'),
                "rol": current_user['role'],
                "accion": "aprobado_sin_cambios_masivo",
                "comentario": comentario
            }
            
            await db.predios_actualizacion.update_one(
                {"_id": predio["_id"]},
                {
                    "$set": {
                        "estado_visita": "actualizado",
                        "aprobado_sin_cambios": True,
                        "aprobado_por": current_user.get('email'),
                        "aprobado_en": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc)
                    },
                    "$push": {"historial_cambios": historial_entry}
                }
            )
            aprobados += 1
    
    return {
        "message": f"{aprobados} predios aprobados como actualizados (sin cambios)",
        "aprobados": aprobados
    }


# ==================== VISITA DE CAMPO ====================

@api_router.get("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/visita")
async def obtener_visita_predio(
    proyecto_id: str,
    codigo_predial: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene los datos de la visita de campo de un predio"""
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial},
            {"codigo_predial_nacional": codigo_predial}
        ]
    }, {"_id": 0})
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Extraer datos de visita del formato_visita o campos separados
    visita = predio.get('formato_visita', {})
    
    # Si hay datos de visita en otros campos, incluirlos
    if predio.get('linderos'):
        visita['linderos'] = predio.get('linderos')
    if predio.get('coordenadas_visita'):
        visita['coordenadas'] = predio.get('coordenadas_visita')
    if predio.get('propietarios_visita'):
        visita['propietarios_visita'] = predio.get('propietarios_visita')
    if predio.get('construcciones_visita'):
        visita['construcciones_visita'] = predio.get('construcciones_visita')
    if predio.get('fotos_visita'):
        visita['fotos'] = predio.get('fotos_visita')
    if predio.get('firma_visita'):
        visita['firma_base64'] = predio.get('firma_visita')
    if predio.get('observaciones_generales'):
        visita['observaciones'] = predio.get('observaciones_generales')
    if predio.get('sin_cambios'):
        visita['sin_cambios'] = predio.get('sin_cambios')
    
    return {
        "visita": visita if visita else None,
        "estado_visita": predio.get('estado_visita', 'pendiente')
    }


@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/visita")
async def guardar_visita_predio(
    proyecto_id: str,
    codigo_predial: str,
    visita_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Guarda los datos de la visita de campo de un predio"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]:
        raise HTTPException(status_code=403, detail="No tiene permisos para registrar visitas")
    
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial},
            {"codigo_predial_nacional": codigo_predial}
        ]
    })
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    ahora = datetime.now(timezone.utc)
    
    # Preparar datos de actualización
    update_data = {
        "formato_visita": visita_data,
        "estado_visita": "visitado",
        "visitado_por": current_user.get('email'),
        "visitado_en": ahora.isoformat(),
        "updated_at": ahora,
        # Campos separados para facilitar consultas
        "linderos": visita_data.get('linderos'),
        "coordenadas_visita": visita_data.get('coordenadas'),
        "propietarios_visita": visita_data.get('propietarios_visita'),
        "construcciones_visita": visita_data.get('construcciones_visita'),
        "fotos_visita": visita_data.get('fotos'),
        "firma_visita": visita_data.get('firma_base64'),
        "observaciones_generales": visita_data.get('observaciones'),
        "sin_cambios": visita_data.get('sin_cambios', False),
        "fecha_visita": visita_data.get('fecha_visita'),
        "hora_visita": visita_data.get('hora_visita'),
        "persona_atiende": visita_data.get('persona_atiende'),
        "relacion_predio": visita_data.get('relacion_predio'),
        "acceso_predio": visita_data.get('acceso_predio')
    }
    
    # Registrar en historial
    historial_entry = {
        "fecha": ahora.isoformat(),
        "usuario": current_user.get('email'),
        "rol": current_user['role'],
        "accion": "visita_registrada",
        "comentario": f"Visita de campo registrada. {'Sin cambios detectados.' if visita_data.get('sin_cambios') else 'Cambios por verificar.'}"
    }
    
    await db.predios_actualizacion.update_one(
        {"_id": predio["_id"]},
        {
            "$set": update_data,
            "$push": {"historial_cambios": historial_entry}
        }
    )
    
    return {
        "message": "Visita guardada correctamente",
        "estado_visita": "visitado",
        "sin_cambios": visita_data.get('sin_cambios', False)
    }


@api_router.patch("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/estado")
async def actualizar_estado_predio(
    proyecto_id: str,
    codigo_predial: str,
    estado_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Actualiza solo el estado de visita de un predio (pendiente/visitado/actualizado)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]:
        raise HTTPException(status_code=403, detail="No tiene permisos para actualizar estado de predios")
    
    nuevo_estado = estado_data.get('estado_visita')
    if nuevo_estado not in ['pendiente', 'visitado', 'actualizado']:
        raise HTTPException(status_code=400, detail="Estado inválido. Debe ser: pendiente, visitado o actualizado")
    
    # Decodificar el código predial
    from urllib.parse import unquote
    codigo_predial = unquote(codigo_predial)
    
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial},
            {"codigo_predial_nacional": codigo_predial}
        ]
    })
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    ahora = datetime.now(timezone.utc)
    estado_anterior = predio.get('estado_visita', 'pendiente')
    
    # Preparar datos de actualización
    update_data = {
        "estado_visita": nuevo_estado,
        "updated_at": ahora
    }
    
    # Registrar quién hizo el cambio
    if nuevo_estado == 'visitado':
        update_data["visitado_por"] = current_user.get('email')
        update_data["visitado_en"] = ahora.isoformat()
    elif nuevo_estado == 'actualizado':
        update_data["actualizado_por"] = current_user.get('email')
        update_data["actualizado_en"] = ahora.isoformat()
    
    # Registrar en historial
    historial_entry = {
        "fecha": ahora.isoformat(),
        "usuario": current_user.get('email'),
        "rol": current_user['role'],
        "accion": "cambio_estado",
        "comentario": f"Estado cambiado de '{estado_anterior}' a '{nuevo_estado}'"
    }
    
    await db.predios_actualizacion.update_one(
        {"_id": predio["_id"]},
        {
            "$set": update_data,
            "$push": {"historial_cambios": historial_entry}
        }
    )
    
    return {
        "message": f"Estado actualizado a '{nuevo_estado}'",
        "estado_anterior": estado_anterior,
        "estado_nuevo": nuevo_estado
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/construcciones")
async def obtener_construcciones_proyecto(
    proyecto_id: str,
    codigo: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las construcciones del proyecto, opcionalmente filtradas por código de terreno"""
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    query = {"proyecto_id": proyecto_id, "tipo": "construccion"}
    
    if codigo:
        # Buscar construcciones que coincidan con el código del terreno (primeros 21-22 dígitos)
        codigo_base = codigo[:21] if len(codigo) >= 21 else codigo
        query["properties.terreno_codigo"] = {"$regex": f"^{codigo_base}"}
    
    construcciones = await db.geometrias_actualizacion.find(query, {"_id": 0}).to_list(length=100)
    
    # Si no encuentra por terreno_codigo, buscar por codigo directamente
    if not construcciones and codigo:
        query["properties.codigo"] = {"$regex": f"^{codigo[:17]}"}  # Buscar por manzana
        del query["properties.terreno_codigo"]
        construcciones = await db.geometrias_actualizacion.find(query, {"_id": 0}).to_list(length=100)
    
    return {
        "construcciones": construcciones,
        "total": len(construcciones)
    }


# ==================== EXPORTACIÓN EXCEL R1/R2 - ACTUALIZACIÓN ====================

@api_router.get("/actualizacion/proyectos/{proyecto_id}/exportar-excel")
async def exportar_actualizacion_excel(
    proyecto_id: str,
    solo_actualizados: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Exporta los predios del proyecto de actualización a Excel en formato R1/R2"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden exportar")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Query de predios
    query = {"proyecto_id": proyecto_id}
    if solo_actualizados:
        query["estado_visita"] = "actualizado"
    
    predios = await db.predios_actualizacion.find(query, {"_id": 0}).to_list(50000)
    
    if not predios:
        raise HTTPException(status_code=404, detail="No hay predios para exportar")
    
    # Crear workbook
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    actualizado_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === HOJA REGISTRO_R1 (Propietarios) ===
    ws_r1 = wb.active
    ws_r1.title = "REGISTRO_R1"
    
    # Headers R1
    headers_r1 = [
        "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL", 
        "CODIGO_HOMOLOGADO", "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS",
        "NOMBRE", "ESTADO_CIVIL", "TIPO_DOCUMENTO", "NUMERO_DOCUMENTO", "DIRECCION",
        "COMUNA", "DESTINO_ECONOMICO", "AREA_TERRENO", "AREA_CONSTRUIDA", "AVALUO",
        "VIGENCIA", "ESTADO_VISITA", "ACTUALIZADO_POR", "FECHA_ACTUALIZACION"
    ]
    
    for col, header in enumerate(headers_r1, 1):
        cell = ws_r1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos R1
    row = 2
    for predio in predios:
        propietarios = predio.get('propietarios', [])
        if not propietarios:
            propietarios = [{
                'nombre_propietario': predio.get('nombre_propietario', predio.get('nombre', '')),
                'tipo_documento': predio.get('tipo_documento', ''),
                'numero_documento': predio.get('numero_documento', ''),
                'estado_civil': predio.get('estado_civil', '')
            }]
        
        total_props = len(propietarios) if propietarios else 1
        es_actualizado = predio.get('estado_visita') == 'actualizado'
        
        for idx, prop in enumerate(propietarios, 1):
            ws_r1.cell(row=row, column=1, value=predio.get('departamento', proyecto.get('departamento', 'NORTE DE SANTANDER')))
            ws_r1.cell(row=row, column=2, value=predio.get('municipio', proyecto.get('municipio', '')))
            ws_r1.cell(row=row, column=3, value=predio.get('numero_predio', predio.get('numero_predial', '')))
            ws_r1.cell(row=row, column=4, value=predio.get('codigo_predial_nacional', predio.get('codigo_predial', predio.get('numero_predial', ''))))
            ws_r1.cell(row=row, column=5, value=predio.get('codigo_homologado', ''))
            ws_r1.cell(row=row, column=6, value='1')
            ws_r1.cell(row=row, column=7, value=str(idx).zfill(2))
            ws_r1.cell(row=row, column=8, value=str(total_props).zfill(2))
            ws_r1.cell(row=row, column=9, value=prop.get('nombre_propietario', prop.get('nombre', '')))
            ws_r1.cell(row=row, column=10, value=prop.get('estado_civil', ''))
            ws_r1.cell(row=row, column=11, value=prop.get('tipo_documento', ''))
            ws_r1.cell(row=row, column=12, value=prop.get('numero_documento', ''))
            ws_r1.cell(row=row, column=13, value=predio.get('direccion', ''))
            ws_r1.cell(row=row, column=14, value=predio.get('comuna', ''))
            ws_r1.cell(row=row, column=15, value=predio.get('destino_economico', ''))
            ws_r1.cell(row=row, column=16, value=predio.get('area_terreno', 0))
            ws_r1.cell(row=row, column=17, value=predio.get('area_construida', 0))
            ws_r1.cell(row=row, column=18, value=predio.get('avaluo', predio.get('avaluo_catastral', 0)))
            ws_r1.cell(row=row, column=19, value=predio.get('vigencia', datetime.now().year))
            ws_r1.cell(row=row, column=20, value=predio.get('estado_visita', 'pendiente'))
            ws_r1.cell(row=row, column=21, value=predio.get('actualizado_por', ''))
            ws_r1.cell(row=row, column=22, value=predio.get('actualizado_en', ''))
            
            # Resaltar filas actualizadas
            if es_actualizado:
                for c in range(1, 23):
                    ws_r1.cell(row=row, column=c).fill = actualizado_fill
            
            row += 1
    
    # === HOJA REGISTRO_R2 (Físico) ===
    ws_r2 = wb.create_sheet(title="REGISTRO_R2")
    
    headers_r2 = [
        "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL",
        "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS", "MATRICULA_INMOBILIARIA",
        "ZONA_FISICA_1", "ZONA_ECONOMICA_1", "AREA_TERRENO_1",
        "ZONA_FISICA_2", "ZONA_ECONOMICA_2", "AREA_TERRENO_2",
        "ZONA_FISICA_3", "ZONA_ECONOMICA_3", "AREA_TERRENO_3",
        "HABITACIONES", "BANOS", "LOCALES", "PISOS", "USO", "AREA_CONSTRUIDA",
        "VIGENCIA", "ESTADO_VISITA"
    ]
    
    for col, header in enumerate(headers_r2, 1):
        cell = ws_r2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos R2
    row = 2
    for predio in predios:
        zonas = predio.get('zonas_fisicas', predio.get('r2_registros', []))
        es_actualizado = predio.get('estado_visita') == 'actualizado'
        
        ws_r2.cell(row=row, column=1, value=predio.get('departamento', proyecto.get('departamento', 'NORTE DE SANTANDER')))
        ws_r2.cell(row=row, column=2, value=predio.get('municipio', proyecto.get('municipio', '')))
        ws_r2.cell(row=row, column=3, value=predio.get('numero_predio', predio.get('numero_predial', '')))
        ws_r2.cell(row=row, column=4, value=predio.get('codigo_predial_nacional', predio.get('codigo_predial', predio.get('numero_predial', ''))))
        ws_r2.cell(row=row, column=5, value='2')
        ws_r2.cell(row=row, column=6, value='01')
        ws_r2.cell(row=row, column=7, value='01')
        ws_r2.cell(row=row, column=8, value=predio.get('matricula_inmobiliaria', predio.get('matricula', '')))
        
        # Zonas físicas (hasta 3)
        for i in range(3):
            if i < len(zonas):
                zona = zonas[i]
                ws_r2.cell(row=row, column=9 + i*3, value=zona.get('zona_fisica', ''))
                ws_r2.cell(row=row, column=10 + i*3, value=zona.get('zona_economica', ''))
                ws_r2.cell(row=row, column=11 + i*3, value=zona.get('area_terreno', 0))
        
        # Datos de construcción
        ws_r2.cell(row=row, column=18, value=predio.get('habitaciones', ''))
        ws_r2.cell(row=row, column=19, value=predio.get('banos', ''))
        ws_r2.cell(row=row, column=20, value=predio.get('locales', ''))
        ws_r2.cell(row=row, column=21, value=predio.get('pisos', ''))
        ws_r2.cell(row=row, column=22, value=predio.get('uso', predio.get('destino_economico', '')))
        ws_r2.cell(row=row, column=23, value=predio.get('area_construida', 0))
        ws_r2.cell(row=row, column=24, value=predio.get('vigencia', datetime.now().year))
        ws_r2.cell(row=row, column=25, value=predio.get('estado_visita', 'pendiente'))
        
        # Resaltar filas actualizadas
        if es_actualizado:
            for c in range(1, 26):
                ws_r2.cell(row=row, column=c).fill = actualizado_fill
        
        row += 1
    
    # Ajustar anchos de columna
    for ws in [ws_r1, ws_r2]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column].width = adjusted_width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Actualizacion_{proyecto.get('municipio', 'Proyecto')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== PROPUESTAS DE CAMBIO ====================

@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/propuesta")
async def crear_propuesta_cambio(
    proyecto_id: str,
    codigo_predial: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Crea una propuesta de cambio para un predio (solo si está visitado)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para crear propuestas")
    
    # Buscar el predio
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    })
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Verificar que el predio esté visitado
    if predio.get('estado_visita') not in ['visitado', 'actualizado']:
        raise HTTPException(status_code=400, detail="El predio debe estar visitado antes de proponer cambios")
    
    # Obtener proyecto para el municipio
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id})
    municipio = proyecto.get('municipio', '') if proyecto else ''
    
    # Guardar snapshot completo de los datos existentes
    datos_existentes = {
        "direccion": predio.get('direccion', ''),
        "destino_economico": predio.get('destino_economico', ''),
        "area_terreno": predio.get('area_terreno', 0),
        "area_construida": predio.get('area_construida', 0),
        "avaluo": predio.get('avaluo', 0),
        "matricula": predio.get('matricula', ''),
        "estrato": predio.get('estrato', ''),
        "propietarios": predio.get('propietarios', []),
        "zonas_fisicas": predio.get('zonas_fisicas', [])
    }
    
    # Crear propuesta de cambio
    tipo_revision = data.get('tipo_revision', 'campo')  # campo, juridico, calidad
    
    # Nombres legibles para el tipo de revisión
    tipos_revision_nombres = {
        'campo': 'Revisión de Campo',
        'juridico': 'Revisión Jurídica',
        'calidad': 'Control de Calidad'
    }
    
    propuesta = {
        "id": str(uuid.uuid4()),
        "proyecto_id": proyecto_id,
        "codigo_predial": codigo_predial,
        "municipio": municipio,
        "datos_existentes": datos_existentes,
        "datos_propuestos": data.get('datos_propuestos', {}),
        "tipo_revision": tipo_revision,
        "tipo_revision_nombre": tipos_revision_nombres.get(tipo_revision, 'Revisión de Campo'),
        "justificacion": data.get('justificacion', ''),
        "estado": "pendiente",  # pendiente, aprobada, rechazada, subsanacion, reenviada, rechazada_definitiva
        "creado_por": current_user.get('email'),
        "creado_por_nombre": current_user.get('full_name', current_user.get('email')),
        "creado_en": datetime.now(timezone.utc).isoformat(),
        "revisado_por": None,
        "revisado_en": None,
        "comentario_revision": None,
        "intentos_subsanacion": 0,
        "historial_revision": []
    }
    
    await db.propuestas_cambio_actualizacion.insert_one(propuesta)
    
    # Agregar al historial del predio con tipo de revisión
    await db.predios_actualizacion.update_one(
        {"_id": predio["_id"]},
        {
            "$push": {
                "historial_cambios": {
                    "fecha": datetime.now(timezone.utc).isoformat(),
                    "usuario": current_user.get('email'),
                    "usuario_nombre": current_user.get('full_name', current_user.get('email')),
                    "tipo_revision": tipo_revision,
                    "tipo_revision_nombre": tipos_revision_nombres.get(tipo_revision, 'Revisión de Campo'),
                    "accion": "propuesta_creada",
                    "propuesta_id": propuesta["id"],
                    "campos_modificados": list(data.get('datos_propuestos', {}).keys())
                }
            }
        }
    )
    
    return {
        "message": "Propuesta de cambio creada exitosamente",
        "propuesta_id": propuesta["id"],
        "tipo_revision": tipo_revision
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/propuestas")
async def listar_propuestas_proyecto(
    proyecto_id: str,
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas las propuestas de cambio de un proyecto"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver propuestas")
    
    query = {"proyecto_id": proyecto_id}
    
    if estado:
        if estado == 'subsanacion':
            # Incluir tanto subsanacion como reenviada
            query["estado"] = {"$in": ["subsanacion", "reenviada"]}
        elif estado == 'pendiente':
            # Incluir pendiente y reenviada (ambas esperan acción del coordinador)
            query["estado"] = {"$in": ["pendiente", "reenviada"]}
        elif estado == 'rechazada':
            # Solo las rechazadas definitivas
            query["estado"] = {"$in": ["rechazada", "rechazada_definitiva"]}
        else:
            query["estado"] = estado
    
    # Obtener proyecto para agregar municipio
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0, "municipio": 1, "nombre": 1})
    
    propuestas = await db.propuestas_cambio_actualizacion.find(
        query, {"_id": 0}
    ).sort("creado_en", -1).to_list(1000)
    
    # Agregar municipio del proyecto a cada propuesta
    for prop in propuestas:
        if not prop.get('municipio') and proyecto:
            prop['municipio'] = proyecto.get('municipio', '')
    
    return {
        "propuestas": propuestas,
        "total": len(propuestas)
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/propuestas")
async def listar_propuestas_predio(
    proyecto_id: str,
    codigo_predial: str,
    current_user: dict = Depends(get_current_user)
):
    """Lista las propuestas de cambio de un predio específico"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver propuestas")
    
    propuestas = await db.propuestas_cambio_actualizacion.find(
        {"proyecto_id": proyecto_id, "codigo_predial": codigo_predial},
        {"_id": 0}
    ).sort("creado_en", -1).to_list(100)
    
    return {
        "propuestas": propuestas,
        "total": len(propuestas)
    }


# ==================== CANCELACIÓN DE PREDIOS EN ACTUALIZACIÓN ====================

@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/proponer-cancelacion")
async def proponer_cancelacion_predio(
    proyecto_id: str,
    codigo_predial: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Propone la cancelación de un predio en actualización.
    - Gestores: Solo pueden PROPONER (requiere aprobación de coordinador)
    - Coordinadores/Admin: Pueden cancelar directamente
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para cancelar predios")
    
    # Buscar el predio
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    })
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Verificar que no esté ya cancelado
    if predio.get('cancelado') or predio.get('deleted'):
        raise HTTPException(status_code=400, detail="El predio ya está cancelado")
    
    # Obtener proyecto para el municipio
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id})
    municipio = proyecto.get('municipio', '') if proyecto else ''
    
    motivo = data.get('motivo', 'Sin motivo especificado')
    
    # Si es coordinador/admin, cancelar directamente
    if current_user['role'] in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        historial_entry = {
            "accion": "Predio cancelado",
            "motivo": motivo,
            "usuario": current_user['full_name'],
            "usuario_id": current_user['id'],
            "fecha": datetime.now(timezone.utc).isoformat()
        }
        
        await db.predios_actualizacion.update_one(
            {"_id": predio["_id"]},
            {
                "$set": {
                    "cancelado": True,
                    "deleted": True,
                    "cancelado_en": datetime.now(timezone.utc).isoformat(),
                    "cancelado_por": current_user['id'],
                    "cancelado_por_nombre": current_user['full_name'],
                    "motivo_cancelacion": motivo,
                    "estado_visita": "cancelado"
                },
                "$push": {"historial": historial_entry}
            }
        )
        
        return {
            "message": "Predio cancelado exitosamente",
            "cancelado_por": current_user['full_name']
        }
    
    # Si es gestor, crear propuesta de cancelación
    propuesta = {
        "id": str(uuid.uuid4()),
        "proyecto_id": proyecto_id,
        "codigo_predial": codigo_predial,
        "municipio": municipio,
        "tipo": "cancelacion",  # Tipo de propuesta: cancelacion
        "motivo": motivo,
        "estado": "pendiente",
        "propuesto_por": current_user['id'],
        "propuesto_por_nombre": current_user['full_name'],
        "creado_en": datetime.now(timezone.utc).isoformat(),
        "datos_predio": {
            "direccion": predio.get('direccion', ''),
            "propietarios": predio.get('propietarios', []),
            "area_terreno": predio.get('area_terreno', 0),
            "estado_visita": predio.get('estado_visita', '')
        }
    }
    
    await db.propuestas_cambio_actualizacion.insert_one(propuesta)
    
    # Marcar el predio como "pendiente cancelación"
    await db.predios_actualizacion.update_one(
        {"_id": predio["_id"]},
        {
            "$set": {"cancelacion_pendiente": True},
            "$push": {"historial": {
                "accion": "Cancelación propuesta",
                "motivo": motivo,
                "usuario": current_user['full_name'],
                "fecha": datetime.now(timezone.utc).isoformat()
            }}
        }
    )
    
    return {
        "message": "Propuesta de cancelación enviada. El coordinador la revisará.",
        "propuesta_id": propuesta["id"]
    }


@api_router.delete("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}")
async def cancelar_predio_actualizacion(
    proyecto_id: str,
    codigo_predial: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Cancela un predio directamente (solo coordinador/admin).
    Para gestores usar el endpoint de proponer-cancelacion.
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden cancelar predios directamente")
    
    # Buscar el predio
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    })
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    if predio.get('cancelado') or predio.get('deleted'):
        raise HTTPException(status_code=400, detail="El predio ya está cancelado")
    
    historial_entry = {
        "accion": "Predio cancelado",
        "usuario": current_user['full_name'],
        "usuario_id": current_user['id'],
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios_actualizacion.update_one(
        {"_id": predio["_id"]},
        {
            "$set": {
                "cancelado": True,
                "deleted": True,
                "cancelado_en": datetime.now(timezone.utc).isoformat(),
                "cancelado_por": current_user['id'],
                "cancelado_por_nombre": current_user['full_name'],
                "estado_visita": "cancelado"
            },
            "$push": {"historial": historial_entry}
        }
    )
    
    return {"message": "Predio cancelado exitosamente"}


@api_router.patch("/actualizacion/propuestas/{propuesta_id}/aprobar")
async def aprobar_propuesta(
    propuesta_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Aprueba una propuesta de cambio o cancelación (solo coordinador/admin)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden aprobar propuestas")
    
    propuesta = await db.propuestas_cambio_actualizacion.find_one({"id": propuesta_id})
    if not propuesta:
        raise HTTPException(status_code=404, detail="Propuesta no encontrada")
    
    if propuesta.get('estado') != 'pendiente':
        raise HTTPException(status_code=400, detail="La propuesta ya fue revisada")
    
    # Actualizar propuesta
    await db.propuestas_cambio_actualizacion.update_one(
        {"id": propuesta_id},
        {
            "$set": {
                "estado": "aprobada",
                "revisado_por": current_user.get('email'),
                "revisado_en": datetime.now(timezone.utc).isoformat(),
                "comentario_revision": data.get('comentario', '')
            }
        }
    )
    
    # Si es propuesta de CANCELACIÓN, aplicar la cancelación
    if propuesta.get('tipo') == 'cancelacion':
        historial_entry = {
            "accion": "Cancelación aprobada",
            "motivo": propuesta.get('motivo', ''),
            "propuesto_por": propuesta.get('propuesto_por_nombre'),
            "aprobado_por": current_user['full_name'],
            "fecha": datetime.now(timezone.utc).isoformat()
        }
        
        await db.predios_actualizacion.update_one(
            {
                "proyecto_id": propuesta['proyecto_id'],
                "$or": [
                    {"codigo_predial": propuesta['codigo_predial']},
                    {"numero_predial": propuesta['codigo_predial']}
                ]
            },
            {
                "$set": {
                    "cancelado": True,
                    "deleted": True,
                    "cancelacion_pendiente": False,
                    "cancelado_en": datetime.now(timezone.utc).isoformat(),
                    "cancelado_por": current_user['id'],
                    "cancelado_por_nombre": current_user['full_name'],
                    "motivo_cancelacion": propuesta.get('motivo', ''),
                    "estado_visita": "cancelado"
                },
                "$push": {"historial": historial_entry}
            }
        )
        
        return {"message": "Cancelación aprobada. Predio cancelado exitosamente."}
    
    # Si es propuesta de CAMBIO, aplicar los cambios propuestos
    datos_propuestos = propuesta.get('datos_propuestos', {})
    
    # Preparar los datos a actualizar
    update_data = {
        **datos_propuestos,
        "estado_visita": "actualizado",  # Cambiar estado a actualizado
        "actualizado_por": current_user.get('email'),
        "actualizado_en": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.predios_actualizacion.update_one(
        {
            "proyecto_id": propuesta['proyecto_id'],
            "$or": [
                {"codigo_predial": propuesta['codigo_predial']},
                {"numero_predial": propuesta['codigo_predial']}
            ]
        },
        {
            "$set": update_data,
            "$push": {
                "historial_cambios": {
                    "fecha": datetime.now(timezone.utc).isoformat(),
                    "usuario": current_user.get('email'),
                    "accion": "propuesta_aprobada_actualizado",
                    "propuesta_id": propuesta_id,
                    "cambios_aplicados": datos_propuestos
                }
            }
        }
    )
    
    return {"message": "Propuesta aprobada y cambios aplicados. Predio marcado como ACTUALIZADO"}


@api_router.patch("/actualizacion/propuestas/{propuesta_id}/rechazar")
async def rechazar_propuesta(
    propuesta_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Rechaza una propuesta de cambio y la envía a subsanación del gestor"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden rechazar propuestas")
    
    propuesta = await db.propuestas_cambio_actualizacion.find_one({"id": propuesta_id})
    if not propuesta:
        raise HTTPException(status_code=404, detail="Propuesta no encontrada")
    
    if propuesta.get('estado') not in ['pendiente', 'reenviada']:
        raise HTTPException(status_code=400, detail="La propuesta ya fue revisada")
    
    intentos = propuesta.get('intentos_subsanacion', 0) + 1
    es_rechazo_definitivo = intentos >= 3
    
    # Actualizar propuesta
    update_data = {
        "estado": "rechazada_definitiva" if es_rechazo_definitivo else "subsanacion",
        "revisado_por": current_user.get('email'),
        "revisado_en": datetime.now(timezone.utc).isoformat(),
        "comentario_revision": data.get('comentario', ''),
        "intentos_subsanacion": intentos
    }
    
    # Agregar al historial de la propuesta
    historial_entry = {
        "fecha": datetime.now(timezone.utc).isoformat(),
        "accion": "rechazado_definitivo" if es_rechazo_definitivo else "rechazado_subsanacion",
        "usuario": current_user.get('email'),
        "comentario": data.get('comentario', ''),
        "intento": intentos
    }
    
    await db.propuestas_cambio_actualizacion.update_one(
        {"id": propuesta_id},
        {
            "$set": update_data,
            "$push": {"historial_revision": historial_entry}
        }
    )
    
    # Agregar al historial del predio
    await db.predios_actualizacion.update_one(
        {
            "proyecto_id": propuesta['proyecto_id'],
            "$or": [
                {"codigo_predial": propuesta['codigo_predial']},
                {"numero_predial": propuesta['codigo_predial']}
            ]
        },
        {
            "$push": {
                "historial_cambios": {
                    "fecha": datetime.now(timezone.utc).isoformat(),
                    "usuario": current_user.get('email'),
                    "accion": "propuesta_rechazada_definitiva" if es_rechazo_definitivo else "propuesta_enviada_subsanacion",
                    "propuesta_id": propuesta_id,
                    "motivo": data.get('comentario', ''),
                    "intento": intentos
                }
            }
        }
    )
    
    # Notificar al gestor por correo
    gestor_email = propuesta.get('creado_por')
    if gestor_email and not es_rechazo_definitivo:
        try:
            # Obtener datos del proyecto
            proyecto = await db.proyectos_actualizacion.find_one({"id": propuesta['proyecto_id']})
            municipio = proyecto.get('municipio', 'N/A') if proyecto else 'N/A'
            
            asunto = f"[Asomunicipios] Propuesta de Cambio Requiere Subsanación"
            cuerpo = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <h2 style="color: #f59e0b;">Propuesta de Cambio Rechazada - Requiere Subsanación</h2>
                <p>Su propuesta de cambio ha sido rechazada y requiere subsanación:</p>
                <div style="background: #fef3c7; padding: 15px; border-radius: 8px; margin: 15px 0; border-left: 4px solid #f59e0b;">
                    <p><strong>Código Predial:</strong> {propuesta['codigo_predial']}</p>
                    <p><strong>Municipio:</strong> {municipio}</p>
                    <p><strong>Intento:</strong> {intentos}/3</p>
                </div>
                <div style="background: #fef2f2; padding: 15px; border-radius: 8px; border-left: 4px solid #dc2626;">
                    <p><strong>Motivo del Rechazo:</strong></p>
                    <p>{data.get('comentario', 'Sin comentario')}</p>
                    <p><strong>Rechazado por:</strong> {current_user.get('email')}</p>
                </div>
                <p style="margin-top: 20px;">Por favor, ingrese al Visor de Actualización para corregir y reenviar la propuesta.</p>
                <p style="color: #666; font-size: 12px; margin-top: 20px;">
                    Tiene <strong>{3 - intentos} intentos restantes</strong>. Después del tercer rechazo, la propuesta será descartada.
                </p>
            </body>
            </html>
            """
            await enviar_email(gestor_email, asunto, cuerpo)
        except Exception as e:
            print(f"Error enviando notificación de subsanación: {e}")
    
    if es_rechazo_definitivo:
        return {"message": "Propuesta RECHAZADA DEFINITIVAMENTE (3 intentos agotados)", "definitivo": True}
    else:
        return {"message": f"Propuesta enviada a subsanación del gestor (intento {intentos}/3)", "definitivo": False}


@api_router.patch("/actualizacion/propuestas/{propuesta_id}/subsanar")
async def subsanar_propuesta(
    propuesta_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """El gestor subsana y reenvía una propuesta rechazada"""
    propuesta = await db.propuestas_cambio_actualizacion.find_one({"id": propuesta_id})
    if not propuesta:
        raise HTTPException(status_code=404, detail="Propuesta no encontrada")
    
    if propuesta.get('estado') != 'subsanacion':
        raise HTTPException(status_code=400, detail="Esta propuesta no está en estado de subsanación")
    
    # Verificar que sea el gestor que la creó
    if current_user['role'] == UserRole.GESTOR and propuesta.get('creado_por') != current_user.get('email'):
        raise HTTPException(status_code=403, detail="Solo puede subsanar sus propias propuestas")
    
    # Actualizar datos propuestos
    nuevos_datos = data.get('datos_propuestos', propuesta.get('datos_propuestos', {}))
    justificacion = data.get('justificacion_subsanacion', '')
    
    historial_entry = {
        "fecha": datetime.now(timezone.utc).isoformat(),
        "accion": "subsanado_reenviado",
        "usuario": current_user.get('email'),
        "justificacion": justificacion,
        "intento": propuesta.get('intentos_subsanacion', 1)
    }
    
    await db.propuestas_cambio_actualizacion.update_one(
        {"id": propuesta_id},
        {
            "$set": {
                "estado": "reenviada",
                "datos_propuestos": nuevos_datos,
                "justificacion_subsanacion": justificacion,
                "fecha_subsanacion": datetime.now(timezone.utc).isoformat(),
                "subsanado_por": current_user.get('email')
            },
            "$push": {"historial_revision": historial_entry}
        }
    )
    
    # Notificar al coordinador que rechazó
    coordinador_email = propuesta.get('revisado_por')
    if coordinador_email:
        try:
            asunto = f"[Asomunicipios] Propuesta Subsanada - Pendiente de Revisión"
            cuerpo = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #059669;">Propuesta Subsanada - Nueva Revisión Requerida</h2>
                <p>El gestor {current_user.get('email')} ha subsanado la propuesta:</p>
                <div style="background: #ecfdf5; padding: 15px; border-radius: 8px;">
                    <p><strong>Código Predial:</strong> {propuesta['codigo_predial']}</p>
                    <p><strong>Justificación:</strong> {justificacion}</p>
                </div>
            </body>
            </html>
            """
            await enviar_email(coordinador_email, asunto, cuerpo)
        except Exception as e:
            print(f"Error enviando notificación: {e}")
    
    return {"message": "Propuesta subsanada y reenviada para revisión"}


@api_router.get("/actualizacion/propuestas/subsanacion-pendiente")
async def get_propuestas_subsanacion(
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las propuestas pendientes de subsanación para el gestor actual"""
    query = {"estado": "subsanacion"}
    
    # Si es gestor, solo ver las suyas
    if current_user['role'] == UserRole.GESTOR:
        query["creado_por"] = current_user.get('email')
    
    propuestas = await db.propuestas_cambio_actualizacion.find(query, {"_id": 0}).sort("revisado_en", -1).to_list(500)
    
    return {
        "total": len(propuestas),
        "propuestas": propuestas
    }


@api_router.post("/actualizacion/proyectos/{proyecto_id}/propuestas/aprobar-masivo")
async def aprobar_propuestas_masivo(
    proyecto_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Aprueba múltiples propuestas de cambio (solo coordinador/admin)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden aprobar propuestas")
    
    propuesta_ids = data.get('propuesta_ids', [])
    if not propuesta_ids:
        raise HTTPException(status_code=400, detail="Debe proporcionar IDs de propuestas")
    
    aprobadas = 0
    errores = []
    
    for propuesta_id in propuesta_ids:
        try:
            propuesta = await db.propuestas_cambio_actualizacion.find_one({"id": propuesta_id, "proyecto_id": proyecto_id})
            if not propuesta:
                errores.append(f"{propuesta_id}: No encontrada")
                continue
            
            if propuesta.get('estado') != 'pendiente':
                errores.append(f"{propuesta_id}: Ya revisada")
                continue
            
            # Actualizar propuesta
            await db.propuestas_cambio_actualizacion.update_one(
                {"id": propuesta_id},
                {
                    "$set": {
                        "estado": "aprobada",
                        "revisado_por": current_user.get('email'),
                        "revisado_en": datetime.now(timezone.utc).isoformat(),
                        "comentario_revision": data.get('comentario', 'Aprobación masiva')
                    }
                }
            )
            
            # Aplicar cambios al predio
            datos_propuestos = propuesta.get('datos_propuestos', {})
            update_data = {
                **datos_propuestos,
                "estado_visita": "actualizado",
                "actualizado_por": current_user.get('email'),
                "actualizado_en": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc)
            }
            
            await db.predios_actualizacion.update_one(
                {
                    "proyecto_id": propuesta['proyecto_id'],
                    "$or": [
                        {"codigo_predial": propuesta['codigo_predial']},
                        {"numero_predial": propuesta['codigo_predial']}
                    ]
                },
                {
                    "$set": update_data,
                    "$push": {
                        "historial_cambios": {
                            "fecha": datetime.now(timezone.utc).isoformat(),
                            "usuario": current_user.get('email'),
                            "accion": "propuesta_aprobada_masivo_actualizado",
                            "propuesta_id": propuesta_id
                        }
                    }
                }
                )
            
            aprobadas += 1
        except Exception as e:
            errores.append(f"{propuesta_id}: {str(e)}")
    
    return {
        "message": f"Aprobadas {aprobadas} de {len(propuesta_ids)} propuestas",
        "aprobadas": aprobadas,
        "errores": errores
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/historial")
async def obtener_historial_predio(
    proyecto_id: str,
    codigo_predial: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el historial de cambios de un predio"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver el historial")
    
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    }, {"_id": 0, "historial_cambios": 1, "codigo_predial": 1})
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    return {
        "codigo_predial": predio.get('codigo_predial'),
        "historial": predio.get('historial_cambios', [])
    }


# ==================== GENERACIÓN DE PDF INFORME DE VISITA ====================

@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios/{codigo_predial}/generar-pdf")
async def generar_pdf_informe_visita(
    proyecto_id: str,
    codigo_predial: str,
    current_user: dict = Depends(get_current_user)
):
    """Genera el PDF del informe de visita de un predio"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para generar PDF")
    
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
    from reportlab.lib.units import inch, cm
    import io
    
    # Obtener proyecto
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Obtener predio
    predio = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "$or": [
            {"codigo_predial": codigo_predial},
            {"numero_predial": codigo_predial}
        ]
    }, {"_id": 0})
    
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        alignment=1,  # Center
        spaceAfter=12
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.darkblue,
        spaceBefore=12,
        spaceAfter=6
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9
    )
    
    elements = []
    
    # Encabezado con logo (placeholder - en producción usar logo real)
    header_data = [
        [
            Paragraph("<b>ASOMUNICIPIOS G</b><br/>Gestión y Soluciones Territoriales", normal_style),
            Paragraph(f"<b>INFORME DE VISITA</b><br/>ACTUALIZACIÓN MUNICIPIO DE {proyecto.get('municipio', '').upper()}", title_style),
            Paragraph(f"FO-FAC-PC01-02<br/>v1", normal_style)
        ]
    ]
    header_table = Table(header_data, colWidths=[2*inch, 4*inch, 1.5*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Información Básica
    elements.append(Paragraph("1. INFORMACIÓN BÁSICA", subtitle_style))
    
    info_basica = [
        ["Departamento:", proyecto.get('departamento', 'Norte de Santander'), "Municipio:", proyecto.get('municipio', '')],
        ["Código Predial:", predio.get('codigo_predial', ''), "Matrícula:", predio.get('matricula_inmobiliaria', '')],
        ["Dirección:", predio.get('direccion', ''), "Destino Econ.:", predio.get('destino_economico', '')],
        ["Área Terreno:", f"{predio.get('area_terreno', '')} m²", "Área Construida:", f"{predio.get('area_construida', '')} m²"],
    ]
    
    info_table = Table(info_basica, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Propietarios
    elements.append(Paragraph("2. INFORMACIÓN DE PROPIETARIOS", subtitle_style))
    
    propietarios = predio.get('propietarios', [])
    if propietarios:
        prop_data = [["#", "Nombre", "Tipo Doc.", "Documento", "Estado Civil"]]
        for i, prop in enumerate(propietarios, 1):
            prop_data.append([
                str(i),
                prop.get('nombre', ''),
                prop.get('tipo_documento', ''),
                prop.get('documento', ''),
                prop.get('estado_civil', '')
            ])
        prop_table = Table(prop_data, colWidths=[0.4*inch, 3*inch, 0.8*inch, 1.5*inch, 1*inch])
        prop_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elements.append(prop_table)
    else:
        elements.append(Paragraph("Sin información de propietarios", normal_style))
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Datos de Notificación
    datos_notif = predio.get('datos_notificacion', {})
    if datos_notif:
        elements.append(Paragraph("3. DATOS DE NOTIFICACIÓN", subtitle_style))
        notif_data = [
            ["Teléfono:", datos_notif.get('telefono', ''), "Correo:", datos_notif.get('correo', '')],
            ["Dirección:", datos_notif.get('direccion', ''), "Autoriza notif.:", "Sí" if datos_notif.get('autoriza_notificacion') else "No"],
        ]
        notif_table = Table(notif_data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 2.5*inch])
        notif_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
        ]))
        elements.append(notif_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Información de la Visita
    visita = predio.get('visita', {})
    elements.append(Paragraph("4. INFORMACIÓN DE LA VISITA", subtitle_style))
    
    visita_data = [
        ["Fecha visita:", visita.get('fecha_visita', predio.get('visitado_en', '')[:10] if predio.get('visitado_en') else ''), 
         "Hora:", visita.get('hora_visita', '')],
        ["Persona que atiende:", visita.get('persona_atiende', ''), "Relación:", visita.get('relacion_predio', '')],
        ["Estado predio:", visita.get('estado_predio', ''), "Acceso:", visita.get('acceso_predio', '')],
        ["Servicios:", ", ".join(visita.get('servicios_publicos', [])), "", ""],
    ]
    visita_table = Table(visita_data, colWidths=[1.5*inch, 2.5*inch, 1*inch, 2.5*inch])
    visita_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
    ]))
    elements.append(visita_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Zonas Físicas (Construcciones)
    zonas_fisicas = predio.get('zonas_fisicas', [])
    if zonas_fisicas:
        elements.append(Paragraph("5. ZONAS FÍSICAS (CONSTRUCCIONES)", subtitle_style))
        
        zonas_data = [["#", "Tipo Const.", "Uso", "Área (m²)", "Pisos", "Puntos", "Año", "Observaciones"]]
        for i, zona in enumerate(zonas_fisicas, 1):
            zonas_data.append([
                str(i),
                zona.get('tipo_construccion', ''),
                zona.get('uso_construccion', ''),
                str(zona.get('area_construida', '')),
                str(zona.get('total_pisos', '')),
                str(zona.get('puntos_construccion', '')),
                str(zona.get('anio_construccion', '')),
                zona.get('observaciones', '')[:30] + ('...' if len(zona.get('observaciones', '')) > 30 else '')
            ])
        
        zonas_table = Table(zonas_data, colWidths=[0.3*inch, 1*inch, 0.8*inch, 0.7*inch, 0.5*inch, 0.6*inch, 0.5*inch, 2.1*inch])
        zonas_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elements.append(zonas_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Linderos
    elements.append(Paragraph("6. LINDEROS DEL PREDIO", subtitle_style))
    
    linderos_data = [
        ["NORTE:", predio.get('lindero_norte', 'Sin información')],
        ["SUR:", predio.get('lindero_sur', 'Sin información')],
        ["ESTE:", predio.get('lindero_este', 'Sin información')],
        ["OESTE:", predio.get('lindero_oeste', 'Sin información')],
    ]
    
    linderos_table = Table(linderos_data, colWidths=[1*inch, 6.5*inch])
    linderos_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(linderos_table)
    
    if predio.get('observaciones_linderos'):
        elements.append(Paragraph(f"<b>Observaciones linderos:</b> {predio.get('observaciones_linderos')}", normal_style))
    
    verificado_text = "Sí" if predio.get('linderos_verificados') else "No"
    fecha_verif = predio.get('fecha_verificacion_linderos', '')
    if fecha_verif:
        verificado_text += f" (Fecha: {fecha_verif})"
    elements.append(Paragraph(f"<b>Linderos verificados en campo:</b> {verificado_text}", normal_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Coordenadas
    elements.append(Paragraph("7. GEORREFERENCIACIÓN", subtitle_style))
    
    coord_data = [
        ["Sistema Referencia:", predio.get('sistema_referencia', 'MAGNA-SIRGAS'), 
         "Precisión GPS:", predio.get('precision_gps', '-').upper() if predio.get('precision_gps') else '-'],
        ["Latitud (Y):", str(predio.get('latitud_centroide', '-')), 
         "Longitud (X):", str(predio.get('longitud_centroide', '-'))],
        ["Área Calculada:", f"{predio.get('area_calculada', '-')} m²" if predio.get('area_calculada') else '-', 
         "Equipo GPS:", predio.get('equipo_gps', '-')],
    ]
    
    coord_table = Table(coord_data, colWidths=[1.5*inch, 2.25*inch, 1.5*inch, 2.25*inch])
    coord_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
    ]))
    elements.append(coord_table)
    
    # Vértices del polígono si existen
    vertices = predio.get('vertices_poligono', [])
    if vertices and len(vertices) > 0:
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"<b>Vértices del polígono ({len(vertices)} puntos):</b>", normal_style))
        
        # Mostrar máximo 8 vértices en el PDF para no hacerlo muy largo
        vertices_data = [["#", "X (Longitud)", "Y (Latitud)"]]
        for i, v in enumerate(vertices[:8], 1):
            vertices_data.append([
                str(i),
                f"{float(v.get('x', 0)):.6f}",
                f"{float(v.get('y', 0)):.6f}"
            ])
        if len(vertices) > 8:
            vertices_data.append(["...", f"(+{len(vertices) - 8} vértices más)", ""])
        
        vert_table = Table(vertices_data, colWidths=[0.5*inch, 2*inch, 2*inch])
        vert_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(vert_table)
    
    if predio.get('fecha_captura_coordenadas'):
        elements.append(Paragraph(f"<i>Fecha de captura: {predio.get('fecha_captura_coordenadas')}</i>", normal_style))
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Observaciones
    elements.append(Paragraph("8. OBSERVACIONES GENERALES", subtitle_style))
    obs_text = predio.get('observaciones_campo', visita.get('observaciones', 'Sin observaciones'))
    elements.append(Paragraph(obs_text if obs_text else "Sin observaciones", normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Firmas
    elements.append(Paragraph("9. FIRMAS", subtitle_style))
    
    firma_data = [
        ["FUNCIONARIO", "QUIEN ATENDIÓ LA VISITA"],
        [Paragraph(f"Nombre: {visita.get('realizada_por', current_user.get('email', ''))}", normal_style),
         Paragraph(f"Nombre: {visita.get('persona_atiende', '')}", normal_style)],
        ["Firma: ____________________", "Firma: ____________________"],
    ]
    
    # Si hay firma digital, mostrar indicador
    if visita.get('firma_base64'):
        firma_data[2][1] = "Firma: [FIRMA DIGITAL CAPTURADA]"
    
    firma_table = Table(firma_data, colWidths=[3.75*inch, 3.75*inch])
    firma_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ]))
    elements.append(firma_table)
    
    # Ubicación GPS si existe
    if predio.get('ubicacion_gps'):
        gps = predio['ubicacion_gps']
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(f"<i>Ubicación GPS: {gps.get('lat', '')}, {gps.get('lng', '')} (±{gps.get('accuracy', '')}m)</i>", normal_style))
    
    # Generar PDF
    doc.build(elements)
    
    # Guardar PDF
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Guardar en disco
    pdf_dir = PROYECTOS_ACTUALIZACION_PATH / proyecto_id / "pdfs"
    pdf_dir.mkdir(exist_ok=True)
    pdf_filename = f"informe_visita_{codigo_predial}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf_path = pdf_dir / pdf_filename
    
    with open(pdf_path, 'wb') as f:
        f.write(pdf_content)
    
    # Devolver como base64 para descarga
    import base64
    pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
    
    return {
        "message": "PDF generado exitosamente",
        "filename": pdf_filename,
        "pdf_base64": pdf_base64
    }


@api_router.post("/actualizacion/proyectos/{proyecto_id}/upload-info-alfanumerica")
async def upload_info_alfanumerica_proyecto(
    proyecto_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Sube un archivo de Información Alfanumérica (R1/R2 Excel) para un proyecto de actualización y lo procesa"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para cargar Información Alfanumérica")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if proyecto["estado"] == ProyectoActualizacionEstado.ARCHIVADO:
        raise HTTPException(status_code=400, detail="No se pueden cargar archivos a proyectos archivados")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")
    
    proyecto_dir = PROYECTOS_ACTUALIZACION_PATH / proyecto_id
    proyecto_dir.mkdir(exist_ok=True)
    
    # Guardar archivo
    ext = Path(file.filename).suffix
    file_path = proyecto_dir / f"info_alfanumerica_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}{ext}"
    
    try:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Procesar Excel R1/R2
        registros_procesados = await procesar_r1r2_actualizacion(proyecto_id, str(file_path), proyecto["municipio"])
        
        return {
            "message": f"Información Alfanumérica cargada y procesada ({registros_procesados} registros)",
            "archivo": str(file_path),
            "nombre_archivo": file.filename,
            "registros_procesados": registros_procesados
        }
        
    except Exception as e:
        if file_path.exists():
            file_path.unlink()
        raise HTTPException(status_code=500, detail=f"Error al cargar el archivo: {str(e)}")


async def procesar_r1r2_actualizacion(proyecto_id: str, file_path: str, municipio: str):
    """Procesa el Excel R1/R2 y guarda los predios en la colección del proyecto"""
    import pandas as pd
    
    # Leer Excel
    try:
        # Intentar leer con openpyxl (xlsx)
        df_r1 = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')  # Primera hoja (R1)
        try:
            df_r2 = pd.read_excel(file_path, sheet_name=1, engine='openpyxl')  # Segunda hoja (R2)
        except:
            df_r2 = pd.DataFrame()  # R2 opcional
    except:
        # Fallback a xlrd (xls)
        df_r1 = pd.read_excel(file_path, sheet_name=0, engine='xlrd')
        try:
            df_r2 = pd.read_excel(file_path, sheet_name=1, engine='xlrd')
        except:
            df_r2 = pd.DataFrame()
    
    # Normalizar nombres de columnas
    df_r1.columns = [str(c).strip().upper() for c in df_r1.columns]
    if len(df_r2) > 0:
        df_r2.columns = [str(c).strip().upper() for c in df_r2.columns]
    
    # Eliminar predios anteriores del proyecto
    await db.predios_actualizacion.delete_many({"proyecto_id": proyecto_id})
    
    registros = 0
    
    # Mapear columnas R1 a campos estándar (incluye variantes de nombres de columnas)
    col_mapping_r1 = {
        # Código predial
        'CODIGO_PREDIAL_NACIONAL': 'codigo_predial',
        'CODIGO PREDIAL NACIONAL': 'codigo_predial',
        'CODIGO_PREDIAL': 'codigo_predial',
        'CODIGO PREDIAL': 'codigo_predial',
        'CODIGO': 'codigo_predial',
        # Número predial
        'NUMERO_DEL_PREDIO': 'numero_predial',
        'NUMERO DEL PREDIO': 'numero_predial',
        'NUMERO_PREDIAL': 'numero_predial',
        'NUMERO PREDIAL': 'numero_predial',
        # Datos básicos
        'DIRECCION': 'direccion',
        'DESTINO_ECONOMICO': 'destino_economico',
        'DESTINO ECONOMICO': 'destino_economico',
        'AREA_TERRENO': 'area_terreno',
        'AREA TERRENO': 'area_terreno',
        'AREA_CONSTRUIDA': 'area_construida',
        'AREA CONSTRUIDA': 'area_construida',
        'AVALUO': 'avaluo_catastral',
        'AVALUO_CATASTRAL': 'avaluo_catastral',
        'AVALUO CATASTRAL': 'avaluo_catastral',
        'VIGENCIA': 'vigencia',
        'ESTRATO': 'estrato',
        'TIPO_PREDIO': 'tipo_predio',
        'TIPO PREDIO': 'tipo_predio',
        'COMUNA': 'comuna',
        # Matrícula inmobiliaria
        'MATRICULA_INMOBILIARIA': 'matricula_inmobiliaria',
        'MATRICULA INMOBILIARIA': 'matricula_inmobiliaria',
        'MATRICULA': 'matricula_inmobiliaria',
        # Propietario (del R1)
        'NOMBRE': 'propietario_nombre',
        'TIPO_DOCUMENTO': 'propietario_tipo_doc',
        'TIPO DOCUMENTO': 'propietario_tipo_doc',
        'NUMERO_DOCUMENTO': 'propietario_documento',
        'NUMERO DOCUMENTO': 'propietario_documento',
        'ESTADO_CIVIL': 'propietario_estado_civil',
        'ESTADO CIVIL': 'propietario_estado_civil'
    }
    
    # Agrupar propietarios por código predial
    predios_dict = {}
    
    # Procesar R1
    for _, row in df_r1.iterrows():
        predio_data = {
            "proyecto_id": proyecto_id,
            "municipio": municipio,
            "created_at": datetime.now(timezone.utc),
            "propietarios": []
        }
        
        propietario = {}
        codigo_predial = None
        
        for col, field in col_mapping_r1.items():
            if col in df_r1.columns:
                val = row[col]
                if pd.notna(val):
                    if field in ['area_terreno', 'area_construida', 'avaluo_catastral']:
                        try:
                            predio_data[field] = float(val)
                        except:
                            predio_data[field] = None
                    elif field.startswith('propietario_'):
                        # Datos del propietario
                        prop_field = field.replace('propietario_', '')
                        propietario[prop_field] = str(val).strip()
                    elif field == 'codigo_predial':
                        codigo_predial = str(val).strip()
                        predio_data[field] = codigo_predial
                    else:
                        predio_data[field] = str(val).strip()
        
        # Solo procesar si tiene código predial
        if codigo_predial:
            if codigo_predial not in predios_dict:
                predios_dict[codigo_predial] = predio_data
            
            # Agregar propietario si tiene nombre
            if propietario.get('nombre'):
                predios_dict[codigo_predial]['propietarios'].append(propietario)
    
    # Guardar predios únicos con sus propietarios
    for codigo, predio in predios_dict.items():
        await db.predios_actualizacion.insert_one(predio)
        registros += 1
    
    # Actualizar proyecto
    await db.proyectos_actualizacion.update_one(
        {"id": proyecto_id},
        {"$set": {
            "info_alfanumerica_archivo": file_path,
            "info_alfanumerica_cargado_en": datetime.now(timezone.utc),
            "info_alfanumerica_total_registros": registros,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return registros


@api_router.get("/actualizacion/proyectos/{proyecto_id}/descargar-base-grafica")
async def descargar_base_grafica_proyecto(
    proyecto_id: str,
    token: str = Query(None, description="JWT token para autenticación vía query param")
):
    """Descarga el archivo de Base Gráfica de un proyecto"""
    # Autenticación via query param para permitir window.open
    if not token:
        raise HTTPException(status_code=401, detail="Token requerido")
    
    try:
        payload = decode_token(token)
        current_user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not current_user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
    except Exception:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para descargar archivos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if not proyecto.get("base_grafica_archivo"):
        raise HTTPException(status_code=404, detail="No hay Base Gráfica cargada")
    
    file_path = Path(proyecto["base_grafica_archivo"])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en el servidor")
    
    return FileResponse(
        file_path,
        media_type="application/zip",
        filename=file_path.name,
        headers={"Content-Disposition": f'attachment; filename="{file_path.name}"'}
    )

@api_router.get("/actualizacion/proyectos/{proyecto_id}/descargar-info-alfanumerica")
async def descargar_info_alfanumerica_proyecto(
    proyecto_id: str,
    token: str = Query(None, description="JWT token para autenticación vía query param")
):
    """Descarga el archivo de Información Alfanumérica (R1/R2) de un proyecto"""
    # Autenticación via query param para permitir window.open
    if not token:
        raise HTTPException(status_code=401, detail="Token requerido")
    
    try:
        payload = decode_token(token)
        current_user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not current_user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
    except Exception:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para descargar archivos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if not proyecto.get("info_alfanumerica_archivo"):
        raise HTTPException(status_code=404, detail="No hay Información Alfanumérica cargada")
    
    file_path = Path(proyecto["info_alfanumerica_archivo"])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en el servidor")
    
    # Determinar el tipo de archivo
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if file_path.suffix.lower() == ".xls":
        media_type = "application/vnd.ms-excel"
    
    return FileResponse(
        file_path,
        media_type=media_type,
        filename=file_path.name,
        headers={"Content-Disposition": f'attachment; filename="{file_path.name}"'}
    )


# ===== ENDPOINTS DE CRONOGRAMA - ETAPAS =====

@api_router.get("/actualizacion/proyectos/{proyecto_id}/etapas")
async def listar_etapas_proyecto(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Lista las etapas de un proyecto con sus actividades"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver las etapas")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    etapas = await db.etapas_proyecto.find({"proyecto_id": proyecto_id}, {"_id": 0}).sort("orden", 1).to_list(10)
    
    # Para cada etapa, obtener sus actividades
    for etapa in etapas:
        actividades = await db.actividades_proyecto.find(
            {"etapa_id": etapa["id"]}, 
            {"_id": 0}
        ).sort("orden", 1).to_list(1000)
        
        # Obtener nombres de responsables
        for act in actividades:
            if act.get("responsables_ids"):
                responsables = await db.users.find(
                    {"id": {"$in": act["responsables_ids"]}},
                    {"_id": 0, "id": 1, "full_name": 1}
                ).to_list(100)
                act["responsables"] = responsables
            else:
                act["responsables"] = []
        
        etapa["actividades"] = actividades
        
        # Calcular progreso de la etapa
        total_actividades = len(actividades)
        completadas = sum(1 for a in actividades if a.get("estado") == ActividadEstado.COMPLETADA)
        etapa["progreso"] = round((completadas / total_actividades * 100) if total_actividades > 0 else 0)
    
    return {"etapas": etapas}

@api_router.patch("/actualizacion/etapas/{etapa_id}")
async def actualizar_etapa(
    etapa_id: str,
    update_data: EtapaProyectoUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Actualiza una etapa del proyecto"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden actualizar etapas")
    
    etapa = await db.etapas_proyecto.find_one({"id": etapa_id}, {"_id": 0})
    if not etapa:
        raise HTTPException(status_code=404, detail="Etapa no encontrada")
    
    updates = {"updated_at": datetime.now(timezone.utc)}
    
    if update_data.nombre is not None:
        updates["nombre"] = update_data.nombre.strip()
    if update_data.fecha_inicio is not None:
        updates["fecha_inicio"] = update_data.fecha_inicio
    if update_data.fecha_fin_planificada is not None:
        updates["fecha_fin_planificada"] = update_data.fecha_fin_planificada
    if update_data.fecha_fin_real is not None:
        updates["fecha_fin_real"] = update_data.fecha_fin_real
    if update_data.estado is not None:
        updates["estado"] = update_data.estado
    
    await db.etapas_proyecto.update_one({"id": etapa_id}, {"$set": updates})
    
    etapa_actualizada = await db.etapas_proyecto.find_one({"id": etapa_id}, {"_id": 0})
    return {"message": "Etapa actualizada", "etapa": etapa_actualizada}


# ===== ENDPOINTS DE CRONOGRAMA - ACTIVIDADES =====

@api_router.post("/actualizacion/etapas/{etapa_id}/actividades")
async def crear_actividad(
    etapa_id: str,
    actividad_data: ActividadCreate,
    current_user: dict = Depends(get_current_user)
):
    """Crea una nueva actividad en una etapa"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden crear actividades")
    
    etapa = await db.etapas_proyecto.find_one({"id": etapa_id}, {"_id": 0})
    if not etapa:
        raise HTTPException(status_code=404, detail="Etapa no encontrada")
    
    # Obtener el orden máximo actual
    ultima_actividad = await db.actividades_proyecto.find_one(
        {"etapa_id": etapa_id},
        {"_id": 0, "orden": 1},
        sort=[("orden", -1)]
    )
    nuevo_orden = (ultima_actividad.get("orden", 0) if ultima_actividad else 0) + 1
    
    now = datetime.now(timezone.utc)
    actividad_id = str(uuid.uuid4())
    
    actividad = {
        "id": actividad_id,
        "etapa_id": etapa_id,
        "proyecto_id": etapa["proyecto_id"],
        "nombre": actividad_data.nombre.strip(),
        "descripcion": actividad_data.descripcion.strip() if actividad_data.descripcion else None,
        "fase": actividad_data.fase.strip() if actividad_data.fase else None,
        "orden": nuevo_orden,
        "fecha_inicio": actividad_data.fecha_inicio,
        "fecha_fin_planificada": actividad_data.fecha_fin_planificada,
        "fecha_fin_real": None,
        "estado": ActividadEstado.PENDIENTE,
        "prioridad": actividad_data.prioridad or ActividadPrioridad.MEDIA,
        "porcentaje_avance": 0,
        "responsables_ids": actividad_data.responsables_ids or [],
        "actividad_padre_id": actividad_data.actividad_padre_id,
        "notas": None,
        "creado_por": current_user["id"],
        "created_at": now,
        "updated_at": now
    }
    
    await db.actividades_proyecto.insert_one(actividad)
    actividad.pop("_id", None)
    
    return {"message": "Actividad creada", "actividad": actividad}

@api_router.patch("/actualizacion/actividades/{actividad_id}")
async def actualizar_actividad(
    actividad_id: str,
    update_data: ActividadUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Actualiza una actividad"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden actualizar actividades")
    
    actividad = await db.actividades_proyecto.find_one({"id": actividad_id}, {"_id": 0})
    if not actividad:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    
    updates = {"updated_at": datetime.now(timezone.utc)}
    
    if update_data.nombre is not None:
        updates["nombre"] = update_data.nombre.strip()
    if update_data.descripcion is not None:
        updates["descripcion"] = update_data.descripcion.strip()
    if update_data.fase is not None:
        updates["fase"] = update_data.fase.strip()
    if update_data.fecha_inicio is not None:
        updates["fecha_inicio"] = update_data.fecha_inicio
    if update_data.fecha_fin_planificada is not None:
        updates["fecha_fin_planificada"] = update_data.fecha_fin_planificada
    if update_data.fecha_fin_real is not None:
        updates["fecha_fin_real"] = update_data.fecha_fin_real
    if update_data.estado is not None:
        updates["estado"] = update_data.estado
    if update_data.prioridad is not None:
        updates["prioridad"] = update_data.prioridad
    if update_data.porcentaje_avance is not None:
        updates["porcentaje_avance"] = min(100, max(0, update_data.porcentaje_avance))
    if update_data.notas is not None:
        updates["notas"] = update_data.notas
    if update_data.actividad_padre_id is not None:
        updates["actividad_padre_id"] = update_data.actividad_padre_id
    
    await db.actividades_proyecto.update_one({"id": actividad_id}, {"$set": updates})
    
    actividad_actualizada = await db.actividades_proyecto.find_one({"id": actividad_id}, {"_id": 0})
    return {"message": "Actividad actualizada", "actividad": actividad_actualizada}

@api_router.delete("/actualizacion/actividades/{actividad_id}")
async def eliminar_actividad(
    actividad_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Elimina una actividad"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden eliminar actividades")
    
    actividad = await db.actividades_proyecto.find_one({"id": actividad_id}, {"_id": 0})
    if not actividad:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    
    await db.actividades_proyecto.delete_one({"id": actividad_id})
    
    return {"message": "Actividad eliminada"}

@api_router.post("/actualizacion/actividades/{actividad_id}/asignar")
async def asignar_responsable(
    actividad_id: str,
    user_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Asigna un responsable a una actividad"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden asignar responsables")
    
    actividad = await db.actividades_proyecto.find_one({"id": actividad_id}, {"_id": 0})
    if not actividad:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    
    usuario = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    responsables = actividad.get("responsables_ids", [])
    if user_id not in responsables:
        responsables.append(user_id)
        await db.actividades_proyecto.update_one(
            {"id": actividad_id},
            {"$set": {"responsables_ids": responsables, "updated_at": datetime.now(timezone.utc)}}
        )
    
    return {"message": f"Usuario {usuario['full_name']} asignado a la actividad"}

@api_router.delete("/actualizacion/actividades/{actividad_id}/asignar/{user_id}")
async def desasignar_responsable(
    actividad_id: str,
    user_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Quita un responsable de una actividad"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo administradores y coordinadores pueden desasignar responsables")
    
    actividad = await db.actividades_proyecto.find_one({"id": actividad_id}, {"_id": 0})
    if not actividad:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    
    responsables = actividad.get("responsables_ids", [])
    if user_id in responsables:
        responsables.remove(user_id)
        await db.actividades_proyecto.update_one(
            {"id": actividad_id},
            {"$set": {"responsables_ids": responsables, "updated_at": datetime.now(timezone.utc)}}
        )
    
    return {"message": "Responsable removido de la actividad"}


# ===== ENDPOINTS DE ALERTAS =====

@api_router.get("/actualizacion/alertas-proximas")
async def obtener_alertas_proximas(current_user: dict = Depends(get_current_user)):
    """Obtiene las actividades próximas a vencer (7, 3, 1 día) para mostrar alertas"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para ver alertas")
    
    hoy = datetime.now(timezone.utc).date()
    alertas = []
    
    # Obtener actividades pendientes o en progreso con fecha límite
    actividades = await db.actividades_proyecto.find(
        {
            "estado": {"$in": [ActividadEstado.PENDIENTE, ActividadEstado.EN_PROGRESO]},
            "fecha_fin_planificada": {"$ne": None}
        },
        {"_id": 0}
    ).to_list(1000)
    
    for act in actividades:
        fecha_fin_str = act.get("fecha_fin_planificada")
        if not fecha_fin_str:
            continue
        
        try:
            # Parsear fecha (puede ser string o datetime)
            if isinstance(fecha_fin_str, str):
                fecha_fin = datetime.fromisoformat(fecha_fin_str.replace('Z', '+00:00')).date()
            else:
                fecha_fin = fecha_fin_str.date() if hasattr(fecha_fin_str, 'date') else fecha_fin_str
            
            dias_restantes = (fecha_fin - hoy).days
            
            if dias_restantes < 0:
                # Vencida
                alertas.append({
                    "actividad_id": act["id"],
                    "actividad_nombre": act["nombre"],
                    "proyecto_id": act["proyecto_id"],
                    "etapa_id": act["etapa_id"],
                    "fecha_fin": str(fecha_fin),
                    "dias_restantes": dias_restantes,
                    "tipo_alerta": "vencida",
                    "prioridad": "critica"
                })
            elif dias_restantes <= 1:
                alertas.append({
                    "actividad_id": act["id"],
                    "actividad_nombre": act["nombre"],
                    "proyecto_id": act["proyecto_id"],
                    "etapa_id": act["etapa_id"],
                    "fecha_fin": str(fecha_fin),
                    "dias_restantes": dias_restantes,
                    "tipo_alerta": "urgente",
                    "prioridad": "alta"
                })
            elif dias_restantes <= 3:
                alertas.append({
                    "actividad_id": act["id"],
                    "actividad_nombre": act["nombre"],
                    "proyecto_id": act["proyecto_id"],
                    "etapa_id": act["etapa_id"],
                    "fecha_fin": str(fecha_fin),
                    "dias_restantes": dias_restantes,
                    "tipo_alerta": "proximo",
                    "prioridad": "media"
                })
            elif dias_restantes <= 7:
                alertas.append({
                    "actividad_id": act["id"],
                    "actividad_nombre": act["nombre"],
                    "proyecto_id": act["proyecto_id"],
                    "etapa_id": act["etapa_id"],
                    "fecha_fin": str(fecha_fin),
                    "dias_restantes": dias_restantes,
                    "tipo_alerta": "recordatorio",
                    "prioridad": "baja"
                })
        except Exception:
            continue
    
    # Ordenar por días restantes (más urgentes primero)
    alertas.sort(key=lambda x: x["dias_restantes"])
    
    # Obtener nombres de proyectos
    proyecto_ids = list(set(a["proyecto_id"] for a in alertas))
    if proyecto_ids:
        proyectos = await db.proyectos_actualizacion.find(
            {"id": {"$in": proyecto_ids}},
            {"_id": 0, "id": 1, "nombre": 1, "municipio": 1}
        ).to_list(100)
        proyecto_map = {p["id"]: p for p in proyectos}
        
        for alerta in alertas:
            proyecto = proyecto_map.get(alerta["proyecto_id"], {})
            alerta["proyecto_nombre"] = proyecto.get("nombre", "")
            alerta["municipio"] = proyecto.get("municipio", "")
    
    return {
        "alertas": alertas,
        "total": len(alertas),
        "vencidas": sum(1 for a in alertas if a["tipo_alerta"] == "vencida"),
        "urgentes": sum(1 for a in alertas if a["tipo_alerta"] == "urgente"),
        "proximas": sum(1 for a in alertas if a["tipo_alerta"] == "proximo")
    }

@api_router.get("/actualizacion/municipios-disponibles")
async def municipios_disponibles_para_proyecto(current_user: dict = Depends(get_current_user)):
    """Lista los municipios que no tienen un proyecto activo - ordenados alfabéticamente (español)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para esta operación")
    
    # Municipios que NO deben estar disponibles para proyectos de actualización
    MUNICIPIOS_EXCLUIDOS = ["Ocaña", "La Esperanza", "González", "Tibú"]
    
    # Obtener todos los municipios
    municipios = await db.limites_municipales.find({}, {"_id": 0, "municipio": 1}).to_list(100)
    todos_municipios = [m["municipio"] for m in municipios if m["municipio"] not in MUNICIPIOS_EXCLUIDOS]
    
    # Ordenar alfabéticamente con soporte para acentos español
    import unicodedata
    
    def normalize_for_sort(s):
        """Normaliza el texto para ordenar correctamente en español (Á = A, etc.)"""
        # Normalizar NFD y luego quitar los diacríticos para el ordenamiento
        normalized = unicodedata.normalize('NFD', s)
        # Remover los caracteres diacríticos (acentos) para ordenar
        ascii_str = ''.join(c for c in normalized if not unicodedata.combining(c))
        return ascii_str.lower()
    
    todos_municipios.sort(key=normalize_for_sort)
    
    # Obtener municipios con proyectos activos o pausados
    proyectos_activos = await db.proyectos_actualizacion.find(
        {"estado": {"$in": [ProyectoActualizacionEstado.ACTIVO, ProyectoActualizacionEstado.PAUSADO]}},
        {"_id": 0, "municipio": 1}
    ).to_list(100)
    
    municipios_ocupados = set(p["municipio"] for p in proyectos_activos)
    
    # Municipios disponibles (ya vienen ordenados)
    disponibles = [m for m in todos_municipios if m not in municipios_ocupados]
    
    return {
        "disponibles": disponibles,
        "ocupados": sorted(list(municipios_ocupados)),
        "total": len(todos_municipios)
    }


# ==========================================
# ACTUALIZACIÓN - PREDIOS NUEVOS Y FINALIZAR PROYECTO
# ==========================================

@api_router.post("/actualizacion/proyectos/{proyecto_id}/predios-nuevos")
async def crear_predio_nuevo_actualizacion(
    proyecto_id: str,
    predio_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Crea un predio nuevo dentro de un proyecto de actualización.
    Asigna código homologado automáticamente.
    El predio queda en predios_actualizacion con es_nuevo=true hasta que se finalice el proyecto.
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.GESTOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para crear predios")
    
    # Verificar que el proyecto existe y está activo
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if proyecto.get('estado') not in [ProyectoActualizacionEstado.ACTIVO, ProyectoActualizacionEstado.PAUSADO]:
        raise HTTPException(status_code=400, detail="El proyecto no está activo")
    
    municipio = proyecto.get('municipio')
    
    # Obtener datos del formulario
    r1 = predio_data.get('r1', {})
    r2 = predio_data.get('r2', {})
    propietarios = predio_data.get('propietarios', [])
    zonas_fisicas = predio_data.get('zonas_fisicas', [])
    formato_visita = predio_data.get('formato_visita', {})
    
    # Construir código predial nacional
    divipola = MUNICIPIOS_DIVIPOLA.get(municipio)
    if not divipola:
        raise HTTPException(status_code=400, detail=f"Municipio {municipio} no válido")
    
    prefijo = divipola["departamento"] + divipola["municipio"]
    
    # Construir el código completo de 30 dígitos
    codigo_predial = (
        f"{prefijo}"
        f"{str(r1.get('zona', '0')).zfill(2)}"
        f"{str(r1.get('sector', '0')).zfill(2)}"
        f"{str(r1.get('comuna', '0')).zfill(2)}"
        f"{str(r1.get('barrio', '0')).zfill(2)}"
        f"{str(r1.get('manzana_vereda', '0')).zfill(4)}"
        f"{str(r1.get('terreno', '0')).zfill(4)}"
        f"{str(r1.get('condicion_predio', '0')).zfill(4)}"
        f"{str(r1.get('predio_horizontal', '0')).zfill(5)}"
    )
    codigo_predial = codigo_predial[:30].ljust(30, '0')
    
    # Verificar que no exista en predios principal
    existing = await db.predios.find_one({"codigo_predial_nacional": codigo_predial})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un predio con este código predial en conservación")
    
    # Verificar que no exista en el proyecto actual
    existing_proyecto = await db.predios_actualizacion.find_one({
        "proyecto_id": proyecto_id,
        "codigo_predial": codigo_predial
    })
    if existing_proyecto:
        raise HTTPException(status_code=400, detail="Ya existe un predio con este código en el proyecto")
    
    # Asignar código homologado automáticamente
    codigo_homologado = await asignar_codigo_homologado(municipio, None)
    if not codigo_homologado:
        # Fallback: generar código si no hay disponibles
        codigo_homologado, _ = await generate_codigo_homologado(municipio)
    
    # Crear el predio nuevo
    predio_id = str(uuid.uuid4())
    ahora = datetime.now(timezone.utc).isoformat()
    
    predio_nuevo = {
        "id": predio_id,
        "proyecto_id": proyecto_id,
        "codigo_predial": codigo_predial,
        "codigo_homologado": codigo_homologado,
        "municipio": municipio,
        "es_nuevo": True,  # Marca que es un predio nuevo creado en actualización
        
        # Datos R1
        "direccion": r1.get('direccion', ''),
        "destino_economico": r1.get('destino_economico', ''),
        "area_terreno": float(r1.get('area_terreno', 0)),
        "area_construida": float(r1.get('area_construida', 0)),
        "avaluo": float(r1.get('avaluo', 0)),
        "matricula_inmobiliaria": r2.get('matricula_inmobiliaria', ''),
        
        # Propietarios
        "propietarios": propietarios,
        
        # Zonas físicas y construcciones (R2)
        "zonas_fisicas": zonas_fisicas,
        "r2": r2,
        
        # Formato de visita
        "formato_visita": formato_visita,
        "estado_visita": "visitado",
        
        # Metadatos
        "creado_por": current_user.get('email'),
        "creado_por_nombre": current_user.get('full_name'),
        "created_at": ahora,
        "updated_at": ahora,
        
        # Historial
        "historial_cambios": [{
            "fecha": ahora,
            "usuario": current_user.get('email'),
            "accion": "predio_nuevo_creado",
            "descripcion": f"Predio nuevo creado en campo. Código homologado asignado: {codigo_homologado}"
        }]
    }
    
    await db.predios_actualizacion.insert_one(predio_nuevo)
    
    # Actualizar contador del proyecto
    await db.proyectos_actualizacion.update_one(
        {"id": proyecto_id},
        {"$inc": {"predios_nuevos_count": 1}}
    )
    
    return {
        "message": "Predio nuevo creado exitosamente",
        "predio_id": predio_id,
        "codigo_predial": codigo_predial,
        "codigo_homologado": codigo_homologado
    }


@api_router.get("/actualizacion/proyectos/{proyecto_id}/resumen-finalizacion")
async def get_resumen_finalizacion(
    proyecto_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene un resumen de lo que se migrará a conservación al finalizar el proyecto.
    Incluye validaciones de requisitos.
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden finalizar proyectos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Contar predios por estado
    predios = await db.predios_actualizacion.find({"proyecto_id": proyecto_id}, {"_id": 0}).to_list(50000)
    
    total_predios = len(predios)
    predios_nuevos = [p for p in predios if p.get('es_nuevo')]
    predios_existentes = [p for p in predios if not p.get('es_nuevo')]
    predios_visitados = [p for p in predios if p.get('estado_visita') in ['visitado', 'actualizado']]
    predios_pendientes = [p for p in predios if p.get('estado_visita') == 'pendiente']
    
    # Contar propuestas por estado
    propuestas = await db.propuestas_cambio_actualizacion.find(
        {"proyecto_id": proyecto_id},
        {"_id": 0, "estado": 1, "codigo_predial": 1}
    ).to_list(50000)
    
    propuestas_pendientes = [p for p in propuestas if p.get('estado') == 'pendiente']
    propuestas_aprobadas = [p for p in propuestas if p.get('estado') == 'aprobada']
    propuestas_subsanacion = [p for p in propuestas if p.get('estado') == 'subsanacion']
    propuestas_rechazadas = [p for p in propuestas if p.get('estado') in ['rechazada', 'rechazada_definitiva']]
    
    # Validaciones
    validaciones = {
        "predios_pendientes": len(predios_pendientes) == 0,
        "propuestas_pendientes": len(propuestas_pendientes) == 0,
        "propuestas_subsanacion": len(propuestas_subsanacion) == 0,
        "puede_finalizar": len(predios_pendientes) == 0 and len(propuestas_pendientes) == 0 and len(propuestas_subsanacion) == 0
    }
    
    return {
        "proyecto": {
            "id": proyecto_id,
            "nombre": proyecto.get('nombre'),
            "municipio": proyecto.get('municipio'),
            "estado": proyecto.get('estado')
        },
        "resumen": {
            "total_predios": total_predios,
            "predios_nuevos": len(predios_nuevos),
            "predios_existentes": len(predios_existentes),
            "predios_visitados": len(predios_visitados),
            "predios_pendientes": len(predios_pendientes)
        },
        "propuestas": {
            "total": len(propuestas),
            "aprobadas": len(propuestas_aprobadas),
            "pendientes": len(propuestas_pendientes),
            "subsanacion": len(propuestas_subsanacion),
            "rechazadas": len(propuestas_rechazadas)
        },
        "migracion": {
            "predios_nuevos_a_crear": len(predios_nuevos),
            "cambios_a_aplicar": len(propuestas_aprobadas)
        },
        "validaciones": validaciones
    }


@api_router.post("/actualizacion/proyectos/{proyecto_id}/finalizar")
async def finalizar_proyecto_actualizacion(
    proyecto_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Finaliza un proyecto de actualización y migra los datos a conservación.
    1. Crea los predios nuevos en la colección 'predios'
    2. Aplica los cambios aprobados a los predios existentes
    3. Marca el proyecto como COMPLETADO
    """
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden finalizar proyectos")
    
    proyecto = await db.proyectos_actualizacion.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if proyecto.get('estado') == ProyectoActualizacionEstado.COMPLETADO:
        raise HTTPException(status_code=400, detail="El proyecto ya está finalizado")
    
    if proyecto.get('estado') == ProyectoActualizacionEstado.ARCHIVADO:
        raise HTTPException(status_code=400, detail="El proyecto está archivado")
    
    municipio = proyecto.get('municipio')
    ahora = datetime.now(timezone.utc).isoformat()
    # Permitir seleccionar vigencia, por defecto año actual
    vigencia_seleccionada = data.get('vigencia')
    vigencia_actual = int(vigencia_seleccionada) if vigencia_seleccionada else datetime.now(timezone.utc).year
    
    # Obtener resumen para validaciones
    predios = await db.predios_actualizacion.find({"proyecto_id": proyecto_id}, {"_id": 0}).to_list(50000)
    propuestas = await db.propuestas_cambio_actualizacion.find({"proyecto_id": proyecto_id}, {"_id": 0}).to_list(50000)
    
    predios_pendientes = [p for p in predios if p.get('estado_visita') == 'pendiente']
    propuestas_pendientes = [p for p in propuestas if p.get('estado') == 'pendiente']
    propuestas_subsanacion = [p for p in propuestas if p.get('estado') == 'subsanacion']
    
    # Validar que se puede finalizar (a menos que se fuerce)
    forzar = data.get('forzar', False)
    if not forzar:
        if predios_pendientes:
            raise HTTPException(status_code=400, detail=f"Hay {len(predios_pendientes)} predios sin visitar")
        if propuestas_pendientes:
            raise HTTPException(status_code=400, detail=f"Hay {len(propuestas_pendientes)} propuestas pendientes de revisión")
        if propuestas_subsanacion:
            raise HTTPException(status_code=400, detail=f"Hay {len(propuestas_subsanacion)} propuestas en subsanación")
    
    resultados = {
        "predios_nuevos_creados": 0,
        "cambios_aplicados": 0,
        "errores": []
    }
    
    # 1. Migrar predios NUEVOS a la colección principal
    predios_nuevos = [p for p in predios if p.get('es_nuevo')]
    for predio in predios_nuevos:
        try:
            # Preparar predio para colección principal
            predio_principal = {
                "id": predio.get('id'),
                "codigo_predial_nacional": predio.get('codigo_predial'),
                "codigo_homologado": predio.get('codigo_homologado'),
                "municipio": municipio,
                "departamento": "Norte de Santander",
                "direccion": predio.get('direccion', ''),
                "destino_economico": predio.get('destino_economico', ''),
                "area_terreno": predio.get('area_terreno', 0),
                "area_construida": predio.get('area_construida', 0),
                "avaluo": predio.get('avaluo', 0),
                "matricula_inmobiliaria": predio.get('matricula_inmobiliaria', ''),
                "vigencia": vigencia_actual,
                "propietarios": predio.get('propietarios', []),
                
                # R1 format
                "r1": {
                    "numero_orden": "1",
                    "area_total_terreno": predio.get('area_terreno', 0),
                    "valor_referencia": predio.get('avaluo', 0)
                },
                
                # R2 registros
                "r2_registros": [],
                
                # Formato de visita
                "formato_visita": predio.get('formato_visita', {}),
                
                # Historial
                "historial": [{
                    "accion": "Migrado desde proyecto de actualización",
                    "usuario": current_user.get('full_name'),
                    "fecha": ahora,
                    "proyecto_id": proyecto_id,
                    "proyecto_nombre": proyecto.get('nombre')
                }],
                
                # Metadatos
                "created_at": predio.get('created_at', ahora),
                "updated_at": ahora,
                "origen": "actualizacion",
                "proyecto_origen_id": proyecto_id
            }
            
            # Convertir zonas físicas a r2_registros
            zonas = predio.get('zonas_fisicas', [])
            if zonas:
                for i, zona in enumerate(zonas):
                    r2_registro = {
                        "numero_orden": str(i + 1),
                        "tipo_construccion": zona.get('tipo_construccion', 'C'),
                        "numero_pisos": zona.get('pisos', 0),
                        "numero_habitaciones": zona.get('habitaciones', 0),
                        "numero_banios": zona.get('banos', 0),
                        "numero_locales": zona.get('locales', 0),
                        "anio_construccion": zona.get('anio_construccion', ''),
                        "area_construida": zona.get('area_construida', 0),
                        "puntaje": zona.get('puntaje', 0),
                        "matricula_inmobiliaria": predio.get('matricula_inmobiliaria', '')
                    }
                    predio_principal["r2_registros"].append(r2_registro)
            
            # Insertar en colección principal
            await db.predios.insert_one(predio_principal)
            resultados["predios_nuevos_creados"] += 1
            
        except Exception as e:
            resultados["errores"].append({
                "tipo": "predio_nuevo",
                "codigo": predio.get('codigo_predial'),
                "error": str(e)
            })
    
    # 2. Aplicar cambios APROBADOS a predios existentes
    propuestas_aprobadas = [p for p in propuestas if p.get('estado') == 'aprobada']
    for propuesta in propuestas_aprobadas:
        try:
            codigo_predial = propuesta.get('codigo_predial')
            datos_propuestos = propuesta.get('datos_propuestos', {})
            
            # Buscar el predio en la colección principal
            predio_existente = await db.predios.find_one({"codigo_predial_nacional": codigo_predial})
            if not predio_existente:
                resultados["errores"].append({
                    "tipo": "cambio_aprobado",
                    "codigo": codigo_predial,
                    "error": "Predio no encontrado en conservación"
                })
                continue
            
            # Preparar actualización
            update_data = {
                **datos_propuestos,
                "updated_at": ahora
            }
            
            # Agregar al historial
            historial_entry = {
                "accion": "Cambios aplicados desde proyecto de actualización",
                "usuario": current_user.get('full_name'),
                "fecha": ahora,
                "proyecto_id": proyecto_id,
                "propuesta_id": propuesta.get('id'),
                "cambios": datos_propuestos
            }
            
            await db.predios.update_one(
                {"codigo_predial_nacional": codigo_predial},
                {
                    "$set": update_data,
                    "$push": {"historial": historial_entry}
                }
            )
            resultados["cambios_aplicados"] += 1
            
        except Exception as e:
            resultados["errores"].append({
                "tipo": "cambio_aprobado",
                "codigo": propuesta.get('codigo_predial'),
                "error": str(e)
            })
    
    # 3. Marcar proyecto como COMPLETADO
    await db.proyectos_actualizacion.update_one(
        {"id": proyecto_id},
        {
            "$set": {
                "estado": ProyectoActualizacionEstado.COMPLETADO,
                "finalizado_por": current_user.get('email'),
                "finalizado_por_nombre": current_user.get('full_name'),
                "fecha_finalizacion": ahora,
                "resultados_finalizacion": resultados
            }
        }
    )
    
    return {
        "message": "Proyecto finalizado exitosamente",
        "resultados": resultados
    }


# ==========================================
# GESTOR DE BASE DE DATOS
# ==========================================

@api_router.get("/database/status")
async def get_database_status(current_user: dict = Depends(get_current_user)):
    """Obtener estado de la base de datos - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para acceder a esta sección")
    
    # Obtener estadísticas de la base de datos
    collections = await db.list_collection_names()
    
    stats = {
        "db_name": db.name,
        "collections_count": len(collections),
        "collections": []
    }
    
    total_size = 0
    for coll_name in collections:
        coll_stats = await db.command("collStats", coll_name)
        count = await db[coll_name].count_documents({})
        size_mb = coll_stats.get('size', 0) / (1024 * 1024)
        total_size += coll_stats.get('size', 0)
        stats["collections"].append({
            "name": coll_name,
            "count": count,
            "size_mb": round(size_mb, 2)
        })
    
    stats["total_size_mb"] = round(total_size / (1024 * 1024), 2)
    
    # Obtener último backup del historial
    ultimo_backup = await db.backup_history.find_one(
        {}, 
        sort=[("fecha", -1)],
        projection={"_id": 0, "fecha": 1, "tipo": 1}
    )
    stats["ultimo_backup"] = ultimo_backup
    
    return stats


@api_router.get("/database/backups")
async def get_backup_history(current_user: dict = Depends(get_current_user)):
    """Obtener historial de backups - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para acceder a esta sección")
    
    backups = await db.backup_history.find(
        {},
        {"_id": 0}
    ).sort("fecha", -1).limit(50).to_list(50)
    
    return {"backups": backups}


# Variable global para trackear backups en progreso
backup_progress = {}

async def run_backup_task(backup_id: str, backup_path: str, backup_info: dict, collections_to_backup: list):
    """Tarea de backup que corre en segundo plano"""
    import zipfile
    import json
    from bson import json_util
    
    global backup_progress
    backup_progress[backup_id] = {"status": "running", "progress": 0, "current_collection": "", "error": None}
    
    try:
        total_registros = 0
        total_collections = len(collections_to_backup)
        
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Crear archivo de metadata
            metadata = {
                "backup_id": backup_id,
                "db_name": db.name,
                "fecha": datetime.now(timezone.utc).isoformat(),
                "colecciones": collections_to_backup,
                "version": "1.0"
            }
            zipf.writestr("_metadata.json", json.dumps(metadata, indent=2))
            
            # Exportar cada colección
            for idx, coll_name in enumerate(collections_to_backup):
                backup_progress[backup_id]["current_collection"] = coll_name
                backup_progress[backup_id]["progress"] = int((idx / total_collections) * 100)
                
                # Exportar en lotes para colecciones grandes
                docs = []
                cursor = db[coll_name].find({})
                async for doc in cursor:
                    docs.append(doc)
                
                total_registros += len(docs)
                
                # Convertir a JSON compatible (maneja ObjectId, datetime, etc.)
                json_data = json_util.dumps(docs, indent=2)
                zipf.writestr(f"{coll_name}.json", json_data)
        
        # Obtener tamaño del archivo
        file_size = os.path.getsize(backup_path)
        backup_info["size_mb"] = round(file_size / (1024 * 1024), 2)
        backup_info["registros_total"] = total_registros
        backup_info["estado"] = "completado"
        
        # Guardar en historial
        await db.backup_history.insert_one(backup_info)
        
        backup_progress[backup_id] = {"status": "completed", "progress": 100, "current_collection": "", "error": None}
        
    except Exception as e:
        backup_progress[backup_id] = {"status": "error", "progress": 0, "current_collection": "", "error": str(e)}
        # Limpiar archivo si hubo error
        if os.path.exists(backup_path):
            os.remove(backup_path)

@api_router.post("/database/backup")
async def create_backup(
    background_tasks: BackgroundTasks,
    tipo: str = "completo",
    colecciones: List[str] = Query(default=[]),
    current_user: dict = Depends(get_current_user)
):
    """Crear backup de la base de datos - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para crear backups")
    
    backup_id = str(uuid.uuid4())
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    backup_filename = f"backup_{db.name}_{timestamp}.zip"
    backup_path = f"/app/backend/static/backups/{backup_filename}"
    
    # Crear directorio si no existe
    os.makedirs(os.path.dirname(backup_path), exist_ok=True)
    
    # Obtener colecciones a respaldar
    if tipo == "completo":
        collections_to_backup = await db.list_collection_names()
        # Excluir la colección de historial de backups
        collections_to_backup = [c for c in collections_to_backup if c != 'backup_history']
    else:
        collections_to_backup = colecciones if colecciones else []
    
    if not collections_to_backup:
        raise HTTPException(status_code=400, detail="No hay colecciones para respaldar")
    
    backup_info = {
        "id": backup_id,
        "filename": backup_filename,
        "fecha": datetime.now(timezone.utc).isoformat(),
        "tipo": tipo,
        "colecciones": collections_to_backup,
        "colecciones_count": len(collections_to_backup),
        "creado_por_id": current_user['id'],
        "creado_por": current_user['full_name'],
        "size_mb": 0,
        "registros_total": 0,
        "estado": "en_progreso"
    }
    
    # Iniciar backup en segundo plano
    background_tasks.add_task(run_backup_task, backup_id, backup_path, backup_info, collections_to_backup)
    
    return {
        "success": True,
        "message": "Backup iniciado en segundo plano",
        "backup_id": backup_id,
        "backup": {
            "id": backup_id,
            "filename": backup_filename,
            "tipo": tipo,
            "colecciones_count": len(collections_to_backup),
            "estado": "en_progreso"
        }
    }

@api_router.get("/database/backup/{backup_id}/status")
async def get_backup_status(backup_id: str, current_user: dict = Depends(get_current_user)):
    """Obtener estado de un backup en progreso"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    global backup_progress
    if backup_id in backup_progress:
        return backup_progress[backup_id]
    
    # Verificar si ya está completado en la BD
    backup = await db.backup_history.find_one({"id": backup_id}, {"_id": 0})
    if backup:
        return {"status": "completed", "progress": 100, "current_collection": "", "error": None}
    
    return {"status": "not_found", "progress": 0, "current_collection": "", "error": "Backup no encontrado"}


@api_router.get("/database/backup/{backup_id}/download")
async def download_backup(backup_id: str, current_user: dict = Depends(get_current_user)):
    """Descargar un backup - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso para descargar backups")
    
    backup = await db.backup_history.find_one({"id": backup_id}, {"_id": 0})
    if not backup:
        raise HTTPException(status_code=404, detail="Backup no encontrado")
    
    backup_path = f"/app/backend/static/backups/{backup['filename']}"
    if not os.path.exists(backup_path):
        raise HTTPException(status_code=404, detail="Archivo de backup no encontrado")
    
    return FileResponse(
        path=backup_path,
        filename=backup['filename'],
        media_type="application/zip"
    )


@api_router.delete("/database/backup/{backup_id}")
async def delete_backup(backup_id: str, current_user: dict = Depends(get_current_user)):
    """Eliminar un backup - Solo admin"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar backups")
    
    backup = await db.backup_history.find_one({"id": backup_id}, {"_id": 0})
    if not backup:
        raise HTTPException(status_code=404, detail="Backup no encontrado")
    
    # Eliminar archivo
    backup_path = f"/app/backend/static/backups/{backup['filename']}"
    if os.path.exists(backup_path):
        os.remove(backup_path)
    
    # Eliminar del historial
    await db.backup_history.delete_one({"id": backup_id})
    
    return {"success": True, "message": "Backup eliminado"}


@api_router.post("/database/restore/{backup_id}")
async def restore_backup(
    backup_id: str, 
    colecciones: List[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Restaurar un backup - Solo admin"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden restaurar backups")
    
    import zipfile
    from bson import json_util
    
    backup = await db.backup_history.find_one({"id": backup_id}, {"_id": 0})
    if not backup:
        raise HTTPException(status_code=404, detail="Backup no encontrado")
    
    backup_path = f"/app/backend/static/backups/{backup['filename']}"
    if not os.path.exists(backup_path):
        raise HTTPException(status_code=404, detail="Archivo de backup no encontrado")
    
    try:
        restored_collections = []
        total_restored = 0
        
        with zipfile.ZipFile(backup_path, 'r') as zipf:
            # Leer metadata
            metadata = json.loads(zipf.read("_metadata.json"))
            
            # Determinar qué colecciones restaurar
            collections_to_restore = colecciones if colecciones else backup['colecciones']
            
            for coll_name in collections_to_restore:
                if f"{coll_name}.json" not in zipf.namelist():
                    continue
                
                # Leer datos de la colección
                json_data = zipf.read(f"{coll_name}.json")
                docs = json_util.loads(json_data)
                
                if docs:
                    # Eliminar colección existente y restaurar
                    await db[coll_name].drop()
                    await db[coll_name].insert_many(docs)
                    total_restored += len(docs)
                    restored_collections.append({
                        "name": coll_name,
                        "count": len(docs)
                    })
        
        # Registrar la restauración en el historial
        await db.backup_history.update_one(
            {"id": backup_id},
            {"$push": {
                "restauraciones": {
                    "fecha": datetime.now(timezone.utc).isoformat(),
                    "usuario": current_user['full_name'],
                    "colecciones": [c['name'] for c in restored_collections]
                }
            }}
        )
        
        return {
            "success": True,
            "message": f"Backup restaurado exitosamente",
            "colecciones_restauradas": restored_collections,
            "total_registros": total_restored
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error restaurando backup: {str(e)}")


@api_router.get("/database/backup/{backup_id}/preview")
async def preview_backup(backup_id: str, current_user: dict = Depends(get_current_user)):
    """Vista previa del contenido de un backup - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    import zipfile
    from bson import json_util
    
    backup = await db.backup_history.find_one({"id": backup_id}, {"_id": 0})
    if not backup:
        raise HTTPException(status_code=404, detail="Backup no encontrado")
    
    backup_path = f"/app/backend/static/backups/{backup['filename']}"
    if not os.path.exists(backup_path):
        raise HTTPException(status_code=404, detail="Archivo de backup no encontrado")
    
    preview = {
        "backup_info": backup,
        "colecciones": []
    }
    
    try:
        with zipfile.ZipFile(backup_path, 'r') as zipf:
            metadata = json.loads(zipf.read("_metadata.json"))
            
            for coll_name in backup['colecciones']:
                if f"{coll_name}.json" not in zipf.namelist():
                    continue
                
                json_data = zipf.read(f"{coll_name}.json")
                docs = json_util.loads(json_data)
                
                # Convert sample document to JSON-serializable format (exclude _id)
                sample = None
                if docs:
                    sample_doc = docs[0]
                    if isinstance(sample_doc, dict):
                        sample = {k: str(v) if hasattr(v, '__str__') and not isinstance(v, (str, int, float, bool, list, dict, type(None))) else v 
                                  for k, v in sample_doc.items() if k != '_id'}
                
                preview["colecciones"].append({
                    "name": coll_name,
                    "count": len(docs)
                })
        
        return preview
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error leyendo backup: {str(e)}")


# ===== CONFIGURACIÓN DE BACKUPS AUTOMÁTICOS =====

# Variable global para el scheduler de backups
backup_scheduler_task = None

@api_router.get("/database/config")
async def get_backup_config(current_user: dict = Depends(get_current_user)):
    """Obtener configuración de backups - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    config = await db.system_config.find_one({"type": "backup_config"}, {"_id": 0})
    
    if not config:
        # Configuración por defecto
        config = {
            "type": "backup_config",
            "modo": "manual",  # "manual" o "automatico"
            "frecuencia": "diario",  # "diario", "semanal", "mensual"
            "hora": "02:00",  # Hora de ejecución (formato 24h)
            "dia_semana": 0,  # 0=Lunes, 6=Domingo (para semanal)
            "dia_mes": 1,  # 1-28 (para mensual)
            "tipo_backup": "completo",  # "completo" o "selectivo"
            "colecciones_seleccionadas": [],  # Para backup selectivo
            "retener_ultimos": 7,  # Número de backups a retener
            "notificar_email": True,
            "ultimo_backup_auto": None,
            "proximo_backup": None,
            "activo": False
        }
        await db.system_config.insert_one(config)
    
    return config

@api_router.put("/database/config")
async def update_backup_config(
    modo: str = "manual",
    frecuencia: str = "diario",
    hora: str = "02:00",
    dia_semana: int = 0,
    dia_mes: int = 1,
    tipo_backup: str = "completo",
    colecciones_seleccionadas: List[str] = Query(default=[]),
    retener_ultimos: int = 7,
    notificar_email: bool = True,
    current_user: dict = Depends(get_current_user)
):
    """Actualizar configuración de backups - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Validar parámetros
    if modo not in ["manual", "automatico"]:
        raise HTTPException(status_code=400, detail="Modo debe ser 'manual' o 'automatico'")
    if frecuencia not in ["diario", "semanal", "mensual"]:
        raise HTTPException(status_code=400, detail="Frecuencia inválida")
    if tipo_backup not in ["completo", "selectivo"]:
        raise HTTPException(status_code=400, detail="Tipo de backup inválido")
    if tipo_backup == "selectivo" and not colecciones_seleccionadas:
        raise HTTPException(status_code=400, detail="Debe seleccionar colecciones para backup selectivo")
    
    # Calcular próximo backup si es automático
    proximo_backup = None
    if modo == "automatico":
        proximo_backup = calcular_proximo_backup(frecuencia, hora, dia_semana, dia_mes)
    
    config = {
        "type": "backup_config",
        "modo": modo,
        "frecuencia": frecuencia,
        "hora": hora,
        "dia_semana": dia_semana,
        "dia_mes": dia_mes,
        "tipo_backup": tipo_backup,
        "colecciones_seleccionadas": colecciones_seleccionadas,
        "retener_ultimos": retener_ultimos,
        "notificar_email": notificar_email,
        "activo": modo == "automatico",
        "proximo_backup": proximo_backup,
        "modificado_por": current_user['full_name'],
        "fecha_modificacion": datetime.now(timezone.utc).isoformat()
    }
    
    await db.system_config.update_one(
        {"type": "backup_config"},
        {"$set": config},
        upsert=True
    )
    
    # Reconfigurar el scheduler con la nueva configuración
    configurar_scheduler_backup(config)
    
    return {"message": "Configuración actualizada", "config": config}

def calcular_proximo_backup(frecuencia: str, hora: str, dia_semana: int, dia_mes: int) -> str:
    """Calcula la fecha/hora del próximo backup"""
    from datetime import datetime, timezone, timedelta
    
    now = datetime.now(timezone.utc)
    hora_parts = hora.split(":")
    hora_backup = int(hora_parts[0])
    minuto_backup = int(hora_parts[1]) if len(hora_parts) > 1 else 0
    
    if frecuencia == "diario":
        # Próximo día a la hora indicada
        proximo = now.replace(hour=hora_backup, minute=minuto_backup, second=0, microsecond=0)
        if proximo <= now:
            proximo += timedelta(days=1)
    
    elif frecuencia == "semanal":
        # Próximo día de la semana a la hora indicada
        dias_hasta = (dia_semana - now.weekday()) % 7
        if dias_hasta == 0:
            proximo = now.replace(hour=hora_backup, minute=minuto_backup, second=0, microsecond=0)
            if proximo <= now:
                proximo += timedelta(days=7)
        else:
            proximo = now + timedelta(days=dias_hasta)
            proximo = proximo.replace(hour=hora_backup, minute=minuto_backup, second=0, microsecond=0)
    
    elif frecuencia == "mensual":
        # Próximo día del mes a la hora indicada
        try:
            proximo = now.replace(day=dia_mes, hour=hora_backup, minute=minuto_backup, second=0, microsecond=0)
            if proximo <= now:
                # Siguiente mes
                if now.month == 12:
                    proximo = proximo.replace(year=now.year + 1, month=1)
                else:
                    proximo = proximo.replace(month=now.month + 1)
        except ValueError:
            # Si el día no existe en el mes, usar el último día
            proximo = (now.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            proximo = proximo.replace(hour=hora_backup, minute=minuto_backup, second=0, microsecond=0)
    
    return proximo.isoformat()


# ===== BACKUP SCHEDULER FUNCTIONS =====
async def ejecutar_backup_automatico():
    """Ejecuta un backup automático según la configuración guardada"""
    try:
        logging.info("=" * 60)
        logging.info("🔄 INICIANDO BACKUP AUTOMÁTICO PROGRAMADO")
        logging.info("=" * 60)
        
        config = await db.system_config.find_one({"type": "backup_config"}, {"_id": 0})
        if not config or config.get("modo") != "automatico":
            logging.warning("Backup automático cancelado: modo no es automático")
            return
        
        # Crear backup
        backup_id = str(uuid.uuid4())
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_auto_{db.name}_{timestamp}.zip"
        backup_path = f"/app/backend/static/backups/{backup_filename}"
        
        os.makedirs(os.path.dirname(backup_path), exist_ok=True)
        
        # Obtener colecciones
        if config.get("tipo_backup") == "completo":
            collections_to_backup = await db.list_collection_names()
            collections_to_backup = [c for c in collections_to_backup if c not in ['backup_history', 'system_config']]
        else:
            collections_to_backup = config.get("colecciones_seleccionadas", [])
        
        if not collections_to_backup:
            logging.error("No hay colecciones para respaldar")
            return
        
        backup_info = {
            "id": backup_id,
            "filename": backup_filename,
            "fecha": datetime.now(timezone.utc).isoformat(),
            "tipo": config.get("tipo_backup", "completo"),
            "colecciones": collections_to_backup,
            "colecciones_count": len(collections_to_backup),
            "creado_por_id": "sistema",
            "creado_por": "Backup automático programado",
            "size_mb": 0,
            "registros_total": 0,
            "estado": "en_progreso",
            "es_automatico": True
        }
        
        # Ejecutar backup directamente (no en background)
        await run_backup_task(backup_id, backup_path, backup_info, collections_to_backup)
        
        # Actualizar último backup y próximo
        proximo = calcular_proximo_backup(
            config.get("frecuencia", "diario"),
            config.get("hora", "02:00"),
            config.get("dia_semana", 0),
            config.get("dia_mes", 1)
        )
        
        await db.system_config.update_one(
            {"type": "backup_config"},
            {"$set": {
                "ultimo_backup_auto": datetime.now(timezone.utc).isoformat(),
                "proximo_backup": proximo
            }}
        )
        
        # Limpiar backups antiguos según retención
        await limpiar_backups_por_retencion(config.get("retener_ultimos", 7))
        
        logging.info("=" * 60)
        logging.info(f"✅ BACKUP AUTOMÁTICO COMPLETADO: {backup_filename}")
        logging.info(f"📅 Próximo backup: {proximo}")
        logging.info("=" * 60)
        
    except Exception as e:
        logging.error(f"❌ Error en backup automático: {str(e)}")


async def limpiar_backups_por_retencion(retener: int):
    """Limpia backups automáticos que excedan el límite de retención"""
    try:
        # Obtener backups automáticos ordenados por fecha
        backups = await db.backup_history.find(
            {"es_automatico": True, "estado": "completado"}
        ).sort("fecha", -1).to_list(None)
        
        if len(backups) > retener:
            backups_a_eliminar = backups[retener:]
            for backup in backups_a_eliminar:
                # Eliminar archivo físico
                backup_path = f"/app/backend/static/backups/{backup.get('filename', '')}"
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                    logging.info(f"🗑️ Eliminado backup antiguo: {backup.get('filename')}")
                
                # Eliminar registro de BD
                await db.backup_history.delete_one({"id": backup["id"]})
            
            logging.info(f"🧹 Limpieza de retención: eliminados {len(backups_a_eliminar)} backups antiguos")
    except Exception as e:
        logging.error(f"Error en limpieza de backups: {str(e)}")


def configurar_scheduler_backup(config: dict):
    """Configura el scheduler según la configuración de backup"""
    global backup_scheduler
    
    # Remover job existente si existe
    try:
        backup_scheduler.remove_job('backup_automatico')
        logging.info("Job de backup anterior removido")
    except:
        pass
    
    if config.get("modo") != "automatico" or not config.get("activo", False):
        logging.info("Scheduler de backup desactivado (modo manual o inactivo)")
        return
    
    # Parsear hora
    hora_parts = config.get("hora", "02:00").split(":")
    hora = int(hora_parts[0])
    minuto = int(hora_parts[1]) if len(hora_parts) > 1 else 0
    
    frecuencia = config.get("frecuencia", "diario")
    
    # Crear trigger según frecuencia
    if frecuencia == "diario":
        trigger = CronTrigger(hour=hora, minute=minuto)
        logging.info(f"📅 Scheduler configurado: DIARIO a las {hora:02d}:{minuto:02d}")
    
    elif frecuencia == "semanal":
        dia_semana = config.get("dia_semana", 0)  # 0=Lunes
        dias_semana = ['lun', 'mar', 'mié', 'jue', 'vie', 'sáb', 'dom']
        trigger = CronTrigger(day_of_week=dia_semana, hour=hora, minute=minuto)
        logging.info(f"📅 Scheduler configurado: SEMANAL los {dias_semana[dia_semana]} a las {hora:02d}:{minuto:02d}")
    
    elif frecuencia == "mensual":
        dia_mes = config.get("dia_mes", 1)
        trigger = CronTrigger(day=dia_mes, hour=hora, minute=minuto)
        logging.info(f"📅 Scheduler configurado: MENSUAL el día {dia_mes} a las {hora:02d}:{minuto:02d}")
    
    else:
        logging.warning(f"Frecuencia desconocida: {frecuencia}")
        return
    
    # Agregar job
    backup_scheduler.add_job(
        ejecutar_backup_automatico,
        trigger,
        id='backup_automatico',
        name='Backup Automático de Base de Datos',
        replace_existing=True
    )
    
    logging.info("✅ Scheduler de backup configurado correctamente")


@api_router.post("/database/backup/ejecutar-programado")
async def ejecutar_backup_programado(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Ejecutar backup programado manualmente - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    config = await db.system_config.find_one({"type": "backup_config"}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="No hay configuración de backup")
    
    # Crear backup según la configuración
    backup_id = str(uuid.uuid4())
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    backup_filename = f"backup_auto_{db.name}_{timestamp}.zip"
    backup_path = f"/app/backend/static/backups/{backup_filename}"
    
    os.makedirs(os.path.dirname(backup_path), exist_ok=True)
    
    # Obtener colecciones
    if config.get("tipo_backup") == "completo":
        collections_to_backup = await db.list_collection_names()
        collections_to_backup = [c for c in collections_to_backup if c not in ['backup_history', 'system_config']]
    else:
        collections_to_backup = config.get("colecciones_seleccionadas", [])
    
    if not collections_to_backup:
        raise HTTPException(status_code=400, detail="No hay colecciones para respaldar")
    
    backup_info = {
        "id": backup_id,
        "filename": backup_filename,
        "fecha": datetime.now(timezone.utc).isoformat(),
        "tipo": config.get("tipo_backup", "completo"),
        "colecciones": collections_to_backup,
        "colecciones_count": len(collections_to_backup),
        "creado_por_id": "sistema",
        "creado_por": f"Backup automático (iniciado por {current_user['full_name']})",
        "size_mb": 0,
        "registros_total": 0,
        "estado": "en_progreso",
        "es_automatico": True
    }
    
    background_tasks.add_task(run_backup_task, backup_id, backup_path, backup_info, collections_to_backup)
    
    # Actualizar último backup y próximo
    proximo = calcular_proximo_backup(
        config.get("frecuencia", "diario"),
        config.get("hora", "02:00"),
        config.get("dia_semana", 0),
        config.get("dia_mes", 1)
    )
    
    await db.system_config.update_one(
        {"type": "backup_config"},
        {"$set": {
            "ultimo_backup_auto": datetime.now(timezone.utc).isoformat(),
            "proximo_backup": proximo
        }}
    )
    
    return {
        "message": "Backup programado iniciado",
        "backup_id": backup_id,
        "tipo": config.get("tipo_backup"),
        "colecciones": len(collections_to_backup)
    }

@api_router.get("/database/scheduler/status")
async def get_scheduler_status(current_user: dict = Depends(get_current_user)):
    """Obtener estado del scheduler de backups - Solo admin y coordinador"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    config = await db.system_config.find_one({"type": "backup_config"}, {"_id": 0})
    
    # Obtener jobs del scheduler
    jobs = []
    if backup_scheduler.running:
        for job in backup_scheduler.get_jobs():
            jobs.append({
                "id": job.id,
                "name": job.name,
                "next_run_time": job.next_run_time.isoformat() if job.next_run_time else None,
                "trigger": str(job.trigger)
            })
    
    return {
        "scheduler_running": backup_scheduler.running,
        "modo": config.get("modo", "manual") if config else "manual",
        "activo": config.get("activo", False) if config else False,
        "frecuencia": config.get("frecuencia", "diario") if config else "diario",
        "hora": config.get("hora", "02:00") if config else "02:00",
        "proximo_backup": config.get("proximo_backup") if config else None,
        "ultimo_backup_auto": config.get("ultimo_backup_auto") if config else None,
        "jobs": jobs
    }

@api_router.post("/database/backup/limpiar-antiguos")
async def limpiar_backups_antiguos(current_user: dict = Depends(get_current_user)):
    """Eliminar backups antiguos según la configuración de retención - Solo admin"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden limpiar backups")
    
    config = await db.system_config.find_one({"type": "backup_config"}, {"_id": 0})
    retener = config.get("retener_ultimos", 7) if config else 7
    
    # Obtener todos los backups ordenados por fecha
    all_backups = await db.backup_history.find(
        {},
        {"_id": 0}
    ).sort("fecha", -1).to_list(1000)
    
    if len(all_backups) <= retener:
        return {"message": "No hay backups para eliminar", "eliminados": 0}
    
    backups_a_eliminar = all_backups[retener:]
    eliminados = 0
    
    for backup in backups_a_eliminar:
        backup_path = f"/app/backend/static/backups/{backup['filename']}"
        if os.path.exists(backup_path):
            os.remove(backup_path)
        await db.backup_history.delete_one({"id": backup["id"]})
        eliminados += 1
    
    return {
        "message": f"Limpieza completada",
        "eliminados": eliminados,
        "retenidos": retener
    }


@api_router.post("/database/backup/analizar")
async def analizar_backup_archivo(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Analiza un archivo de backup y compara con los datos actuales"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    if not file.filename.endswith('.zip'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .zip")
    
    import zipfile
    import tempfile
    from bson import json_util
    
    # Guardar archivo temporal
    with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        comparacion = []
        total_registros = 0
        registros_actuales = 0
        metadata = {}
        
        with zipfile.ZipFile(tmp_path, 'r') as zipf:
            # Leer metadata si existe
            if '_metadata.json' in zipf.namelist():
                with zipf.open('_metadata.json') as f:
                    metadata = json.loads(f.read().decode('utf-8'))
            
            # Analizar cada colección
            for name in zipf.namelist():
                if name.endswith('.json') and name != '_metadata.json':
                    coll_name = name.replace('.json', '')
                    
                    with zipf.open(name) as f:
                        data = json_util.loads(f.read().decode('utf-8'))
                        registros_backup = len(data) if isinstance(data, list) else 0
                        total_registros += registros_backup
                    
                    # Contar registros actuales
                    actual = await db[coll_name].count_documents({})
                    registros_actuales += actual
                    
                    comparacion.append({
                        "coleccion": coll_name,
                        "en_backup": registros_backup,
                        "actual": actual,
                        "diferencia": registros_backup - actual
                    })
        
        # Ordenar por diferencia (más cambios primero)
        comparacion.sort(key=lambda x: abs(x['diferencia']), reverse=True)
        
        return {
            "archivo": file.filename,
            "db_name": metadata.get("db_name", "N/A"),
            "fecha_backup": metadata.get("fecha"),
            "total_colecciones": len(comparacion),
            "total_registros": total_registros,
            "registros_actuales": registros_actuales,
            "comparacion": comparacion
        }
        
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Archivo ZIP inválido o corrupto")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analizando backup: {str(e)}")
    finally:
        os.unlink(tmp_path)


@api_router.post("/database/backup/restaurar-archivo")
async def restaurar_backup_archivo(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Restaura la base de datos desde un archivo de backup"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden restaurar backups")
    
    if not file.filename.endswith('.zip'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .zip")
    
    import zipfile
    import tempfile
    from bson import json_util
    
    # Guardar archivo temporal
    with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        colecciones_restauradas = 0
        registros_restaurados = 0
        
        with zipfile.ZipFile(tmp_path, 'r') as zipf:
            for name in zipf.namelist():
                if name.endswith('.json') and name != '_metadata.json':
                    coll_name = name.replace('.json', '')
                    
                    with zipf.open(name) as f:
                        data = json_util.loads(f.read().decode('utf-8'))
                    
                    if isinstance(data, list) and len(data) > 0:
                        # Limpiar colección existente
                        await db[coll_name].delete_many({})
                        
                        # Insertar datos del backup
                        await db[coll_name].insert_many(data)
                        
                        colecciones_restauradas += 1
                        registros_restaurados += len(data)
        
        # Registrar la restauración
        await db.backup_history.insert_one({
            "id": str(uuid.uuid4()),
            "filename": f"restaurado_{file.filename}",
            "fecha": datetime.now(timezone.utc).isoformat(),
            "tipo": "restauracion_archivo",
            "colecciones_count": colecciones_restauradas,
            "registros_total": registros_restaurados,
            "creado_por_id": current_user['id'],
            "creado_por": current_user['full_name'],
            "estado": "restaurado"
        })
        
        return {
            "message": "Backup restaurado exitosamente",
            "colecciones_restauradas": colecciones_restauradas,
            "registros_restaurados": registros_restaurados
        }
        
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Archivo ZIP inválido o corrupto")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error restaurando backup: {str(e)}")
    finally:
        os.unlink(tmp_path)


# ===== GESTIÓN DE COLECCIONES =====

@api_router.post("/database/collection/{collection_name}/empty")
async def empty_collection(
    collection_name: str,
    current_user: dict = Depends(get_current_user)
):
    """Vaciar todos los documentos de una colección - Solo admin"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden vaciar colecciones")
    
    # Verificar que la colección existe
    collections = await db.list_collection_names()
    if collection_name not in collections:
        raise HTTPException(status_code=404, detail=f"Colección '{collection_name}' no encontrada")
    
    # Proteger colecciones del sistema
    protected_collections = ['system_config']
    if collection_name in protected_collections:
        raise HTTPException(status_code=403, detail=f"La colección '{collection_name}' está protegida y no puede ser vaciada")
    
    # Contar documentos antes de eliminar
    count_before = await db[collection_name].count_documents({})
    
    # Eliminar todos los documentos
    result = await db[collection_name].delete_many({})
    
    # Registrar la acción
    await db.audit_log.insert_one({
        "action": "empty_collection",
        "collection": collection_name,
        "deleted_count": result.deleted_count,
        "user_id": current_user['id'],
        "user_name": current_user['full_name'],
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": f"Colección '{collection_name}' vaciada exitosamente",
        "deleted_count": result.deleted_count,
        "collection": collection_name
    }


@api_router.delete("/database/collection/{collection_name}")
async def drop_collection(
    collection_name: str,
    current_user: dict = Depends(get_current_user)
):
    """Eliminar una colección completamente - Solo admin"""
    if current_user['role'] != UserRole.ADMINISTRADOR:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar colecciones")
    
    # Verificar que la colección existe
    collections = await db.list_collection_names()
    if collection_name not in collections:
        raise HTTPException(status_code=404, detail=f"Colección '{collection_name}' no encontrada")
    
    # Proteger colecciones críticas del sistema
    critical_collections = ['users', 'system_config', 'backup_history', 'audit_log']
    if collection_name in critical_collections:
        raise HTTPException(status_code=403, detail=f"La colección '{collection_name}' es crítica y no puede ser eliminada")
    
    # Contar documentos antes de eliminar
    count_before = await db[collection_name].count_documents({})
    
    # Eliminar la colección
    await db[collection_name].drop()
    
    # Registrar la acción
    await db.audit_log.insert_one({
        "action": "drop_collection",
        "collection": collection_name,
        "documents_lost": count_before,
        "user_id": current_user['id'],
        "user_name": current_user['full_name'],
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": f"Colección '{collection_name}' eliminada completamente",
        "documents_lost": count_before,
        "collection": collection_name
    }


# Endpoint para descargar archivos de cambios del servidor
@api_router.get("/descargas/{filename}")
async def descargar_archivo_servidor(filename: str):
    """Endpoint para descargar archivos de cambios"""
    filepath = Path(f"/app/backend/static/descargas/{filename}")
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/octet-stream"
    )


# Include the router in the main app
app.include_router(api_router)


# ===== WEBSOCKET ENDPOINT FOR REAL-TIME NOTIFICATIONS =====
@app.websocket("/ws/{user_id}")
async def websocket_endpoint(websocket: WebSocket, user_id: str):
    """WebSocket endpoint for real-time notifications"""
    await ws_manager.connect(websocket, user_id)
    try:
        while True:
            # Keep connection alive, listen for messages
            data = await websocket.receive_text()
            # Echo back for ping/pong
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        ws_manager.disconnect(user_id)
    except Exception as e:
        logging.error(f"WebSocket error for {user_id}: {e}")
        ws_manager.disconnect(user_id)


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# Importar sistema de migraciones
from migrations import ejecutar_migraciones

@app.on_event("startup")
async def ejecutar_migraciones_automaticas():
    """
    Ejecuta las migraciones automáticas al iniciar el servidor.
    Las migraciones solo se ejecutan una vez (se registran en la colección 'migraciones').
    """
    try:
        logger.info("=" * 60)
        logger.info("🔄 VERIFICANDO MIGRACIONES PENDIENTES...")
        logger.info("=" * 60)
        
        # Ejecutar migraciones en un hilo separado para no bloquear el servidor
        import threading
        migration_thread = threading.Thread(target=ejecutar_migraciones)
        migration_thread.start()
        
        # No esperamos a que termine para no bloquear el inicio del servidor
        # Las migraciones se ejecutan en segundo plano
        logger.info("📋 Migraciones iniciadas en segundo plano")
        
    except Exception as e:
        logger.error(f"Error al iniciar migraciones: {str(e)}")


@app.on_event("startup")
async def create_default_admin():
    """
    Creates a default admin user if no users exist in the database.
    This ensures fresh deployments have a way to access the system.
    """
    try:
        # Check if any users exist
        user_count = await db.users.count_documents({})
        
        if user_count == 0:
            logger.warning("=" * 60)
            logger.warning("NO USERS FOUND IN DATABASE - Creating default admin user")
            logger.warning("=" * 60)
            
            # Create default admin user
            default_admin = {
                "id": str(uuid.uuid4()),
                "email": "catastro@asomunicipios.gov.co",
                "password": hash_password("Asm*123*"),
                "full_name": "Administrador Catastro",
                "role": UserRole.ADMINISTRADOR,
                "email_verified": True,
                "is_active": True,
                "verification_code": None,
                "verification_code_expires": None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "permissions": [
                    Permission.UPLOAD_GDB,
                    Permission.IMPORT_R1R2,
                    Permission.APPROVE_CHANGES,
                    Permission.ACCESO_ACTUALIZACION
                ],
                "municipios_asignados": []
            }
            
            await db.users.insert_one(default_admin)
            
            logger.warning("-" * 60)
            logger.warning("DEFAULT ADMIN USER CREATED SUCCESSFULLY")
            logger.warning(f"  Email: catastro@asomunicipios.gov.co")
            logger.warning(f"  Password: Asm*123*")
            logger.warning(f"  Role: {UserRole.ADMINISTRADOR}")
            logger.warning("-" * 60)
            logger.warning("⚠️  SECURITY WARNING: Please change this password immediately!")
            logger.warning("=" * 60)
        else:
            logger.info(f"Database has {user_count} existing user(s). Skipping default admin creation.")
            
    except Exception as e:
        logger.error(f"Error during startup admin check: {str(e)}")


@app.on_event("startup")
async def iniciar_backup_scheduler():
    """Inicializa el scheduler de backups automáticos"""
    try:
        logger.info("=" * 60)
        logger.info("🚀 INICIANDO SCHEDULER DE BACKUPS AUTOMÁTICOS")
        logger.info("=" * 60)
        
        # Obtener configuración de backup
        config = await db.system_config.find_one({"type": "backup_config"}, {"_id": 0})
        
        if config and config.get("modo") == "automatico" and config.get("activo"):
            configurar_scheduler_backup(config)
            backup_scheduler.start()
            logger.info("✅ Scheduler de backups iniciado correctamente")
            
            # Mostrar próximo backup programado
            if config.get("proximo_backup"):
                logger.info(f"📅 Próximo backup programado: {config.get('proximo_backup')}")
        else:
            logger.info("ℹ️ Backups automáticos desactivados (modo manual)")
            # Iniciar scheduler de todas formas para poder activarlo después
            backup_scheduler.start()
            
    except Exception as e:
        logger.error(f"❌ Error al iniciar scheduler de backups: {str(e)}")


@app.on_event("startup")
async def crear_indices_mongodb():
    """Crea índices necesarios en MongoDB para optimizar consultas"""
    try:
        logger.info("=" * 60)
        logger.info("📊 CREANDO ÍNDICES EN MONGODB")
        logger.info("=" * 60)
        
        # Índices para colección predios
        await db.predios.create_index("codigo_predial_nacional", background=True)
        await db.predios.create_index([("municipio", 1), ("vigencia", 1)], background=True)
        await db.predios.create_index("tiene_geometria", background=True)
        await db.predios.create_index([("municipio", 1), ("vigencia", 1), ("tiene_geometria", 1)], background=True)
        logger.info("✅ Índices de 'predios' creados")
        
        # Índices para colección gdb_geometrias
        await db.gdb_geometrias.create_index("codigo", background=True)
        await db.gdb_geometrias.create_index("municipio", background=True)
        await db.gdb_geometrias.create_index([("municipio", 1), ("tipo", 1)], background=True)
        logger.info("✅ Índices de 'gdb_geometrias' creados")
        
        # Índices para colección predios_nuevos
        await db.predios_nuevos.create_index("municipio", background=True)
        await db.predios_nuevos.create_index("estado_flujo", background=True)
        await db.predios_nuevos.create_index("gestor_apoyo_id", background=True)
        logger.info("✅ Índices de 'predios_nuevos' creados")
        
        # Índices para colección cambios_pendientes
        await db.cambios_pendientes.create_index("municipio", background=True)
        await db.cambios_pendientes.create_index("estado", background=True)
        await db.cambios_pendientes.create_index("gestor_apoyo_id", background=True)
        logger.info("✅ Índices de 'cambios_pendientes' creados")
        
        # Índices para colección petitions
        await db.petitions.create_index("user_id", background=True)
        await db.petitions.create_index("estado", background=True)
        await db.petitions.create_index("gestores_asignados", background=True)
        logger.info("✅ Índices de 'petitions' creados")
        
        logger.info("=" * 60)
        logger.info("✅ TODOS LOS ÍNDICES CREADOS CORRECTAMENTE")
        logger.info("=" * 60)
        
    except Exception as e:
        logger.error(f"❌ Error creando índices (no crítico): {str(e)}")


@app.on_event("shutdown")
async def shutdown_db_client():
    # Detener scheduler de backups
    if backup_scheduler.running:
        backup_scheduler.shutdown(wait=False)
        logger.info("Scheduler de backups detenido")
    client.close()
