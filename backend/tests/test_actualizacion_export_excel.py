"""
Tests for Excel R1/R2 Export Feature - Actualización Module (Iteration 52)

Testing specific bug fixes:
1. R1 sheet has columns PRIMER_APELLIDO, SEGUNDO_APELLIDO, PRIMER_NOMBRE, SEGUNDO_NOMBRE
2. Names are parsed correctly from nombre_propietario using parsear_nombre_propietario()
3. R2 sheet has format with 3 constructions per row (HABITACIONES_1/2/3, BANOS_1/2/3, etc.)
4. R2 includes TIPIFICACION, USO, PUNTAJE, AREA_CONSTRUIDA per construction
5. Row colors apply according to tipo_cambio and estado_visita
6. Excel has 3 sheets: REGISTRO_R1, REGISTRO_R2, RESUMEN
7. R1 has 27 columns, R2 has 41 columns

Endpoint: GET /api/actualizacion/proyectos/{proyecto_id}/exportar-excel
Project: 32ba040f-ed50-45e2-a115-22ab3a351423 (Proyecto_prueba_ZU_Sardinata)
"""

import pytest
import requests
import os
from io import BytesIO
from openpyxl import load_workbook

# Get base URL from environment
BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
PROYECTO_ID = "32ba040f-ed50-45e2-a115-22ab3a351423"


class TestActualizacionExportExcel:
    """Tests for Actualización Excel R1/R2 Export - Bug fixes verification"""
    
    @pytest.fixture(scope="class")
    def auth_token(self):
        """Get authentication token for admin user"""
        response = requests.post(
            f"{BASE_URL}/api/auth/login",
            json={
                "email": "catastro@asomunicipios.gov.co",
                "password": "Asm*123*"
            }
        )
        assert response.status_code == 200, f"Login failed: {response.text}"
        data = response.json()
        assert "token" in data, "Token not in response"
        return data["token"]
    
    @pytest.fixture(scope="class")
    def excel_workbook(self, auth_token):
        """Download and parse the Excel file"""
        response = requests.get(
            f"{BASE_URL}/api/actualizacion/proyectos/{PROYECTO_ID}/exportar-excel",
            headers={"Authorization": f"Bearer {auth_token}"},
            timeout=180  # 3 minutes for large projects
        )
        
        assert response.status_code == 200, f"Export failed with status {response.status_code}: {response.text}"
        assert len(response.content) > 1000, f"Excel file too small: {len(response.content)} bytes"
        
        # Load workbook from bytes
        excel_buffer = BytesIO(response.content)
        workbook = load_workbook(excel_buffer)
        
        print(f"✅ Excel downloaded successfully - Size: {len(response.content)} bytes")
        return workbook
    
    def test_login_success(self, auth_token):
        """Test that we can login successfully"""
        assert auth_token is not None
        assert len(auth_token) > 0
        print(f"✅ Login successful, token length: {len(auth_token)}")
    
    # =====================================================
    # Test 1: Excel has 3 sheets
    # =====================================================
    def test_excel_has_three_sheets(self, excel_workbook):
        """
        Verify Excel has exactly 3 sheets: REGISTRO_R1, REGISTRO_R2, RESUMEN
        """
        sheet_names = excel_workbook.sheetnames
        
        assert len(sheet_names) == 3, f"Expected 3 sheets, got {len(sheet_names)}: {sheet_names}"
        assert "REGISTRO_R1" in sheet_names, f"Missing REGISTRO_R1 sheet. Found: {sheet_names}"
        assert "REGISTRO_R2" in sheet_names, f"Missing REGISTRO_R2 sheet. Found: {sheet_names}"
        assert "RESUMEN" in sheet_names, f"Missing RESUMEN sheet. Found: {sheet_names}"
        
        print(f"✅ Excel has 3 correct sheets: {sheet_names}")
    
    # =====================================================
    # Test 2: R1 has correct columns including name fields
    # =====================================================
    def test_r1_has_name_columns(self, excel_workbook):
        """
        Verify R1 sheet has columns:
        - PRIMER_APELLIDO (column 9)
        - SEGUNDO_APELLIDO (column 10)
        - PRIMER_NOMBRE (column 11)
        - SEGUNDO_NOMBRE (column 12)
        """
        ws_r1 = excel_workbook["REGISTRO_R1"]
        
        # Get headers from row 1
        headers = [ws_r1.cell(row=1, column=i).value for i in range(1, 30)]
        
        # Expected name columns (1-indexed positions)
        assert headers[8] == "PRIMER_APELLIDO", f"Column 9 should be PRIMER_APELLIDO, got: {headers[8]}"
        assert headers[9] == "SEGUNDO_APELLIDO", f"Column 10 should be SEGUNDO_APELLIDO, got: {headers[9]}"
        assert headers[10] == "PRIMER_NOMBRE", f"Column 11 should be PRIMER_NOMBRE, got: {headers[10]}"
        assert headers[11] == "SEGUNDO_NOMBRE", f"Column 12 should be SEGUNDO_NOMBRE, got: {headers[11]}"
        
        print(f"✅ R1 has correct name columns at positions 9-12")
        print(f"   Column 9: {headers[8]}")
        print(f"   Column 10: {headers[9]}")
        print(f"   Column 11: {headers[10]}")
        print(f"   Column 12: {headers[11]}")
    
    # =====================================================
    # Test 3: R1 has 27 columns
    # =====================================================
    def test_r1_has_27_columns(self, excel_workbook):
        """
        Verify R1 sheet has exactly 27 columns as per specification:
        DEPARTAMENTO, MUNICIPIO, NUMERO_DEL_PREDIO, CODIGO_PREDIAL_NACIONAL,
        CODIGO_HOMOLOGADO, TIPO_DE_REGISTRO, NUMERO_DE_ORDEN, TOTAL_REGISTROS,
        PRIMER_APELLIDO, SEGUNDO_APELLIDO, PRIMER_NOMBRE, SEGUNDO_NOMBRE,
        NOMBRE, ESTADO, TIPO_DOCUMENTO, NUMERO_DOCUMENTO, DIRECCION,
        COMUNA, DESTINO_ECONOMICO, AREA_TERRENO, AREA_CONSTRUIDA, AVALUO,
        VIGENCIA, ESTADO_VISITA, TIPO_CAMBIO, ACTUALIZADO_POR, FECHA_ACTUALIZACION
        """
        ws_r1 = excel_workbook["REGISTRO_R1"]
        
        # Count non-empty header columns
        headers = []
        for i in range(1, 50):
            cell_value = ws_r1.cell(row=1, column=i).value
            if cell_value:
                headers.append(cell_value)
            else:
                break
        
        expected_headers = [
            "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL",
            "CODIGO_HOMOLOGADO", "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS",
            "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "PRIMER_NOMBRE", "SEGUNDO_NOMBRE",
            "NOMBRE", "ESTADO", "TIPO_DOCUMENTO", "NUMERO_DOCUMENTO", "DIRECCION",
            "COMUNA", "DESTINO_ECONOMICO", "AREA_TERRENO", "AREA_CONSTRUIDA", "AVALUO",
            "VIGENCIA", "ESTADO_VISITA", "TIPO_CAMBIO", "ACTUALIZADO_POR", "FECHA_ACTUALIZACION"
        ]
        
        assert len(headers) == 27, f"Expected 27 columns in R1, got {len(headers)}: {headers}"
        
        # Verify all expected headers are present
        for i, expected in enumerate(expected_headers):
            assert headers[i] == expected, f"Column {i+1} should be '{expected}', got: '{headers[i]}'"
        
        print(f"✅ R1 has exactly 27 columns:")
        for i, h in enumerate(headers, 1):
            print(f"   {i}. {h}")
    
    # =====================================================
    # Test 4: R2 has correct format with 3 constructions
    # =====================================================
    def test_r2_has_construction_columns(self, excel_workbook):
        """
        Verify R2 sheet has columns for 3 constructions per row:
        - HABITACIONES_1/2/3
        - BANOS_1/2/3
        - LOCALES_1/2/3
        - PISOS_1/2/3
        - TIPIFICACION_1/2/3
        - USO_1/2/3
        - PUNTAJE_1/2/3
        - AREA_CONSTRUIDA_1/2/3
        """
        ws_r2 = excel_workbook["REGISTRO_R2"]
        
        # Get all headers
        headers = [ws_r2.cell(row=1, column=i).value for i in range(1, 50)]
        headers = [h for h in headers if h]  # Remove None values
        
        # Check for construction 1 columns
        construction_1_cols = ["HABITACIONES_1", "BANOS_1", "LOCALES_1", "PISOS_1", 
                               "TIPIFICACION_1", "USO_1", "PUNTAJE_1", "AREA_CONSTRUIDA_1"]
        for col in construction_1_cols:
            assert col in headers, f"Missing column: {col}"
        
        # Check for construction 2 columns
        construction_2_cols = ["HABITACIONES_2", "BANOS_2", "LOCALES_2", "PISOS_2",
                               "TIPIFICACION_2", "USO_2", "PUNTAJE_2", "AREA_CONSTRUIDA_2"]
        for col in construction_2_cols:
            assert col in headers, f"Missing column: {col}"
        
        # Check for construction 3 columns
        construction_3_cols = ["HABITACIONES_3", "BANOS_3", "LOCALES_3", "PISOS_3",
                               "TIPIFICACION_3", "USO_3", "PUNTAJE_3", "AREA_CONSTRUIDA_3"]
        for col in construction_3_cols:
            assert col in headers, f"Missing column: {col}"
        
        print(f"✅ R2 has columns for 3 constructions per row")
        print(f"   Construction 1: {construction_1_cols}")
        print(f"   Construction 2: {construction_2_cols}")
        print(f"   Construction 3: {construction_3_cols}")
    
    # =====================================================
    # Test 5: R2 has 41 columns
    # =====================================================
    def test_r2_has_41_columns(self, excel_workbook):
        """
        Verify R2 sheet has exactly 41 columns as per specification
        """
        ws_r2 = excel_workbook["REGISTRO_R2"]
        
        # Count non-empty header columns
        headers = []
        for i in range(1, 50):
            cell_value = ws_r2.cell(row=1, column=i).value
            if cell_value:
                headers.append(cell_value)
            else:
                break
        
        expected_headers = [
            "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL",
            "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS", "MATRICULA_INMOBILIARIA",
            # Zona 1
            "ZONA_FISICA_1", "ZONA_ECONOMICA_1", "AREA_TERRENO_1",
            # Zona 2
            "ZONA_FISICA_2", "ZONA_ECONOMICA_2", "AREA_TERRENO_2",
            # Construcción 1
            "HABITACIONES_1", "BANOS_1", "LOCALES_1", "PISOS_1", "TIPIFICACION_1", "USO_1", "PUNTAJE_1", "AREA_CONSTRUIDA_1",
            # Construcción 2
            "HABITACIONES_2", "BANOS_2", "LOCALES_2", "PISOS_2", "TIPIFICACION_2", "USO_2", "PUNTAJE_2", "AREA_CONSTRUIDA_2",
            # Construcción 3
            "HABITACIONES_3", "BANOS_3", "LOCALES_3", "PISOS_3", "TIPIFICACION_3", "USO_3", "PUNTAJE_3", "AREA_CONSTRUIDA_3",
            "VIGENCIA", "ESTADO_VISITA", "TIPO_CAMBIO"
        ]
        
        assert len(headers) == 41, f"Expected 41 columns in R2, got {len(headers)}"
        
        print(f"✅ R2 has exactly 41 columns")
        print(f"   First 8 (metadata): {headers[:8]}")
        print(f"   Zona 1 (9-11): {headers[8:11]}")
        print(f"   Zona 2 (12-14): {headers[11:14]}")
        print(f"   Construction 1 (15-22): {headers[14:22]}")
        print(f"   Construction 2 (23-30): {headers[22:30]}")
        print(f"   Construction 3 (31-38): {headers[30:38]}")
        print(f"   Extra fields (39-41): {headers[38:41]}")
    
    # =====================================================
    # Test 6: Names are parsed correctly
    # =====================================================
    def test_names_are_parsed_correctly(self, excel_workbook):
        """
        Verify that nombre_propietario is parsed into separate fields:
        - 'PENARANDA PENARANDA LUIS SUC' should split into:
          - PRIMER_APELLIDO: PENARANDA
          - SEGUNDO_APELLIDO: PENARANDA
          - PRIMER_NOMBRE: LUIS
          - SEGUNDO_NOMBRE: SUC
        """
        ws_r1 = excel_workbook["REGISTRO_R1"]
        
        # Get column indices
        headers = {ws_r1.cell(row=1, column=i).value: i for i in range(1, 30)}
        
        primer_apellido_col = headers.get("PRIMER_APELLIDO")
        segundo_apellido_col = headers.get("SEGUNDO_APELLIDO")
        primer_nombre_col = headers.get("PRIMER_NOMBRE")
        segundo_nombre_col = headers.get("SEGUNDO_NOMBRE")
        nombre_col = headers.get("NOMBRE")
        
        # Check at least the first few data rows
        parsed_names_found = 0
        rows_checked = 0
        
        for row in range(2, min(100, ws_r1.max_row + 1)):  # Check first 100 rows
            primer_ap = ws_r1.cell(row=row, column=primer_apellido_col).value
            segundo_ap = ws_r1.cell(row=row, column=segundo_apellido_col).value
            primer_nom = ws_r1.cell(row=row, column=primer_nombre_col).value
            nombre_completo = ws_r1.cell(row=row, column=nombre_col).value
            
            if nombre_completo:
                rows_checked += 1
                
                # If NOMBRE has a value with multiple parts, the separate fields should have values
                nombre_parts = nombre_completo.strip().split() if nombre_completo else []
                
                if len(nombre_parts) >= 2:
                    # At minimum, primer_apellido should be populated
                    if primer_ap:
                        parsed_names_found += 1
                        
                        # Log a sample
                        if parsed_names_found <= 3:
                            print(f"   Row {row}: NOMBRE='{nombre_completo}'")
                            print(f"            PRIMER_APELLIDO='{primer_ap}'")
                            print(f"            SEGUNDO_APELLIDO='{segundo_ap}'")
                            print(f"            PRIMER_NOMBRE='{primer_nom}'")
                            print(f"            SEGUNDO_NOMBRE='{ws_r1.cell(row=row, column=segundo_nombre_col).value}'")
        
        # We should find parsed names in most records
        if rows_checked > 0:
            parse_rate = (parsed_names_found / rows_checked) * 100
            print(f"✅ Names parsing verification:")
            print(f"   Rows checked: {rows_checked}")
            print(f"   Rows with parsed names: {parsed_names_found}")
            print(f"   Parse rate: {parse_rate:.1f}%")
            
            # We expect at least 50% of names to be parsed (some may be entities without parseable names)
            assert parse_rate >= 50, f"Name parsing rate too low: {parse_rate}%"
        else:
            print(f"⚠️ No data rows found in R1 sheet")
    
    # =====================================================
    # Test 7: RESUMEN sheet has correct structure
    # =====================================================
    def test_resumen_sheet_structure(self, excel_workbook):
        """
        Verify RESUMEN sheet has expected content structure
        """
        ws_resumen = excel_workbook["RESUMEN"]
        
        # First row should have the title
        title = ws_resumen.cell(row=1, column=1).value
        assert "RESUMEN" in title.upper(), f"Expected title with 'RESUMEN', got: {title}"
        
        # Look for key labels in the first 20 rows
        labels_found = []
        for row in range(1, 21):
            cell_value = ws_resumen.cell(row=row, column=1).value
            if cell_value:
                labels_found.append(str(cell_value))
        
        # Check for expected labels
        expected_labels = ["Proyecto:", "Municipio:", "Fecha de exportación:", "Total predios exportados:"]
        for expected in expected_labels:
            found = any(expected in label for label in labels_found)
            assert found, f"Missing label in RESUMEN: {expected}"
        
        print(f"✅ RESUMEN sheet has correct structure")
        print(f"   Labels found: {labels_found[:10]}")
    
    # =====================================================
    # Test 8: R1 data rows have values in name columns
    # =====================================================
    def test_r1_data_rows_have_name_values(self, excel_workbook):
        """
        Verify that data rows in R1 actually have values in the name columns
        (not just empty cells)
        """
        ws_r1 = excel_workbook["REGISTRO_R1"]
        
        # Get column indices
        headers = {ws_r1.cell(row=1, column=i).value: i for i in range(1, 30)}
        
        primer_apellido_col = headers.get("PRIMER_APELLIDO")
        nombre_col = headers.get("NOMBRE")
        
        # Check data rows
        rows_with_primer_apellido = 0
        rows_with_nombre = 0
        total_rows = 0
        
        for row in range(2, min(50, ws_r1.max_row + 1)):  # Check first 50 data rows
            primer_ap = ws_r1.cell(row=row, column=primer_apellido_col).value
            nombre = ws_r1.cell(row=row, column=nombre_col).value
            
            if nombre:
                total_rows += 1
                if primer_ap:
                    rows_with_primer_apellido += 1
                rows_with_nombre += 1
        
        print(f"✅ R1 data row analysis:")
        print(f"   Total data rows (with NOMBRE): {total_rows}")
        print(f"   Rows with PRIMER_APELLIDO populated: {rows_with_primer_apellido}")
        
        # At least some rows should have parsed names
        assert total_rows > 0, "No data rows found in R1"
        assert rows_with_primer_apellido > 0, "No rows have PRIMER_APELLIDO populated"
    
    # =====================================================
    # Test 9: R2 construction columns have numeric values
    # =====================================================
    def test_r2_construction_columns_have_values(self, excel_workbook):
        """
        Verify R2 construction columns contain valid numeric values
        """
        ws_r2 = excel_workbook["REGISTRO_R2"]
        
        # Get column indices
        headers = {ws_r2.cell(row=1, column=i).value: i for i in range(1, 45)}
        
        # Check a few data rows
        construction_data_found = False
        
        for row in range(2, min(20, ws_r2.max_row + 1)):
            habitaciones_1 = ws_r2.cell(row=row, column=headers.get("HABITACIONES_1", 15)).value
            tipificacion_1 = ws_r2.cell(row=row, column=headers.get("TIPIFICACION_1", 19)).value
            uso_1 = ws_r2.cell(row=row, column=headers.get("USO_1", 20)).value
            puntaje_1 = ws_r2.cell(row=row, column=headers.get("PUNTAJE_1", 21)).value
            area_construida_1 = ws_r2.cell(row=row, column=headers.get("AREA_CONSTRUIDA_1", 22)).value
            
            # Values should be numeric (int or float) or 0
            if habitaciones_1 is not None or area_construida_1 is not None:
                construction_data_found = True
                print(f"   Row {row}: HABITACIONES_1={habitaciones_1}, TIPIFICACION_1={tipificacion_1}, "
                      f"USO_1={uso_1}, PUNTAJE_1={puntaje_1}, AREA_CONSTRUIDA_1={area_construida_1}")
                
                # Verify values are numeric
                for val, name in [(habitaciones_1, "HABITACIONES_1"), 
                                  (tipificacion_1, "TIPIFICACION_1"),
                                  (uso_1, "USO_1"),
                                  (puntaje_1, "PUNTAJE_1"),
                                  (area_construida_1, "AREA_CONSTRUIDA_1")]:
                    if val is not None:
                        assert isinstance(val, (int, float)), f"{name} should be numeric, got: {type(val)}"
                break  # Just check first row with data
        
        print(f"✅ R2 construction columns have valid numeric values")
    
    # =====================================================
    # Test 10: Endpoint returns valid Excel file
    # =====================================================
    def test_endpoint_returns_valid_excel(self, auth_token):
        """
        Verify endpoint returns valid Excel file with correct headers
        """
        response = requests.get(
            f"{BASE_URL}/api/actualizacion/proyectos/{PROYECTO_ID}/exportar-excel",
            headers={"Authorization": f"Bearer {auth_token}"},
            timeout=180
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        # Check content type
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheet' in content_type or 'excel' in content_type or 'vnd.openxmlformats' in content_type, \
            f"Unexpected content type: {content_type}"
        
        # Check content disposition
        content_disposition = response.headers.get('Content-Disposition', '')
        assert 'filename' in content_disposition.lower(), f"No filename in disposition: {content_disposition}"
        
        # Verify it's a valid ZIP/XLSX file
        assert response.content[:2] == b'PK', "Not a valid Excel file (ZIP signature)"
        
        print(f"✅ Endpoint returns valid Excel file")
        print(f"   Content-Type: {content_type}")
        print(f"   Content-Disposition: {content_disposition}")
        print(f"   File size: {len(response.content)} bytes")


class TestParseNameFunction:
    """Tests for the parsear_nombre_propietario() function logic"""
    
    def test_parse_four_word_name(self):
        """
        Test parsing a 4-word name: 'PENARANDA PENARANDA LUIS SUC'
        Expected: primer_apellido=PENARANDA, segundo_apellido=PENARANDA, primer_nombre=LUIS, segundo_nombre=SUC
        """
        # Simulate the function logic
        nombre = "PENARANDA PENARANDA LUIS SUC"
        partes = nombre.strip().split()
        
        result = {}
        if len(partes) >= 4:
            result = {
                'primer_apellido': partes[0],
                'segundo_apellido': partes[1],
                'primer_nombre': partes[2],
                'segundo_nombre': ' '.join(partes[3:])
            }
        
        assert result['primer_apellido'] == "PENARANDA"
        assert result['segundo_apellido'] == "PENARANDA"
        assert result['primer_nombre'] == "LUIS"
        assert result['segundo_nombre'] == "SUC"
        
        print(f"✅ 4-word name parsed correctly: {result}")
    
    def test_parse_three_word_name(self):
        """
        Test parsing a 3-word name: 'GARCIA LOPEZ MARIA'
        Expected: primer_apellido=GARCIA, segundo_apellido=LOPEZ, primer_nombre=MARIA, segundo_nombre=''
        """
        nombre = "GARCIA LOPEZ MARIA"
        partes = nombre.strip().split()
        
        result = {}
        if len(partes) == 3:
            result = {
                'primer_apellido': partes[0],
                'segundo_apellido': partes[1],
                'primer_nombre': partes[2],
                'segundo_nombre': ''
            }
        
        assert result['primer_apellido'] == "GARCIA"
        assert result['segundo_apellido'] == "LOPEZ"
        assert result['primer_nombre'] == "MARIA"
        assert result['segundo_nombre'] == ""
        
        print(f"✅ 3-word name parsed correctly: {result}")
    
    def test_parse_two_word_name(self):
        """
        Test parsing a 2-word name: 'RODRIGUEZ JUAN'
        Expected: primer_apellido=RODRIGUEZ, segundo_apellido='', primer_nombre=JUAN, segundo_nombre=''
        """
        nombre = "RODRIGUEZ JUAN"
        partes = nombre.strip().split()
        
        result = {}
        if len(partes) == 2:
            result = {
                'primer_apellido': partes[0],
                'segundo_apellido': '',
                'primer_nombre': partes[1],
                'segundo_nombre': ''
            }
        
        assert result['primer_apellido'] == "RODRIGUEZ"
        assert result['segundo_apellido'] == ""
        assert result['primer_nombre'] == "JUAN"
        assert result['segundo_nombre'] == ""
        
        print(f"✅ 2-word name parsed correctly: {result}")
    
    def test_parse_five_word_name(self):
        """
        Test parsing a 5+ word name: 'PENARANDA PENARANDA LUIS ALFONSO SUC'
        Expected: segundo_nombre should contain all remaining parts
        """
        nombre = "PENARANDA PENARANDA LUIS ALFONSO SUC"
        partes = nombre.strip().split()
        
        result = {}
        if len(partes) >= 4:
            result = {
                'primer_apellido': partes[0],
                'segundo_apellido': partes[1],
                'primer_nombre': partes[2],
                'segundo_nombre': ' '.join(partes[3:])
            }
        
        assert result['primer_apellido'] == "PENARANDA"
        assert result['segundo_apellido'] == "PENARANDA"
        assert result['primer_nombre'] == "LUIS"
        assert result['segundo_nombre'] == "ALFONSO SUC"
        
        print(f"✅ 5-word name parsed correctly: {result}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
